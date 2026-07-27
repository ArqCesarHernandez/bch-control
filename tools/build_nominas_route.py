"""Genera el Blueprint de Nóminas desde la entrega original de PythonAnywhere.

Este script de mantenimiento conserva los cálculos y rutas operativas del
archivo original, retirando únicamente su aplicación Flask, autenticación y
modelos duplicados. No se ejecuta al iniciar el ERP.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path


HEADER = '''"""Módulo de nóminas integrado desde la versión PythonAnywhere 2026-07-17.

La lógica de cálculo de este archivo proviene del sistema original validado.
La integración sustituye únicamente autenticación, usuarios y centros de costo
por los componentes compartidos del ERP V2.
"""

from __future__ import annotations

import hmac
import re
import secrets
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps
from io import BytesIO
from typing import Iterable

from flask import (
    Blueprint,
    Response,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template as flask_render_template,
    request,
    session,
    url_for as flask_url_for,
)
from flask_login import current_user, login_required
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from werkzeug.wsgi import FileWrapper

from models import (
    BitacoraAuditoria,
    CentroCosto,
    Usuario,
    db,
    usuario_centros_nomina as user_projects,
    utc_now,
)
from nominas_models import (
    AdditionalPayment,
    BudgetItem,
    Company,
    Contractor,
    Employee,
    Loan,
    LoanPayment,
    OfficeExpense,
    Payroll,
    PayrollLine,
    Subcontract,
    SubcontractPayment,
    WeeklyResourceAvailability,
)


Project = CentroCosto
User = Usuario
nominas_bp = Blueprint("nominas", __name__)

MONEY_STEP = Decimal("0.01")
IVA_RATE = Decimal("0.16")
WORKDAYS = 5
WEEKDAY_FIELDS = ("lunes", "martes", "miercoles", "jueves", "viernes")


def url_for(endpoint: str, **values):
    """Resuelve endpoints originales dentro del Blueprint de Nóminas."""

    if endpoint == "static":
        return flask_url_for("static", **values)
    if "." not in endpoint:
        endpoint = f"nominas.{endpoint}"
    return flask_url_for(endpoint, **values)


def render_template(template_name: str, **context):
    """Mantiene las rutas de plantilla originales dentro de su namespace."""

    return flask_render_template(f"nominas/{template_name}", **context)


@nominas_bp.get("/modulo-nominas")
@login_required
def index():
    return redirect(url_for("dashboard"))

'''

FOOTER = '''

@nominas_bp.errorhandler(400)
def bad_request(error):
    return render_template("errors/error.html", code=400, title="Solicitud no válida", message=str(error)), 400


@nominas_bp.errorhandler(404)
def not_found(_error):
    return render_template("errors/error.html", code=404, title="No encontrado", message="El registro o página no existe."), 404


@nominas_bp.errorhandler(409)
def conflict(_error):
    return render_template("errors/error.html", code=409, title="Acción no disponible", message="El estado actual del registro no permite esa acción."), 409
'''


def between(source: str, start: str, end: str) -> str:
    start_index = source.index(start)
    end_index = source.index(end, start_index)
    return source[start_index:end_index]


def build(source: str) -> str:
    constants = between(source, "IMPORT_DEFINITIONS =", "\n\n\ndef utc_now")
    helpers = between(
        source,
        "def decimal_value(value) -> Decimal:",
        "# ---------------------------------------------------------------------------\n# Autenticación, usuarios y panel principal",
    )
    helpers = helpers[: helpers.index("def initialize_database():")]

    routes = between(
        source,
        '@app.route("/panel")',
        "# ---------------------------------------------------------------------------\n# Errores e inicio de la aplicación",
    )

    body = constants + "\n\n\n" + helpers + routes
    body = body.replace("@app.route", "@nominas_bp.route")
    body = body.replace("@app.context_processor", "@nominas_bp.context_processor")
    body = body.replace("@app.template_filter", "@nominas_bp.app_template_filter")
    body = body.replace("@app.before_request", "@nominas_bp.before_request")
    body = body.replace(
        '        "today": date.today(),\n    }',
        '        "today": date.today(),\n        "url_for": url_for,\n    }',
    )
    body = body.replace(
        '    if request.method in {"POST", "PUT", "PATCH", "DELETE"}:',
        '    if (\n'
        '        current_app.config.get("WTF_CSRF_ENABLED", True)\n'
        '        and request.method in {"POST", "PUT", "PATCH", "DELETE"}\n'
        '    ):',
    )

    # Flask-Login ya carga Usuario desde app.py; el cargador duplicado del
    # sistema independiente no forma parte del módulo integrado.
    body = re.sub(
        r"\n@login_manager\.user_loader\ndef load_user\(user_id\):\n"
        r"    return db\.session\.get\(User, int\(user_id\)\)\n",
        "\n",
        body,
    )

    # La bitácora del ERP reemplaza a audit_logs sin cambiar las llamadas de
    # negocio del código original.
    body = re.sub(
        r"def audit\(action: str, entity: str, entity_id=None, detail: str \| None = None\):\n"
        r"    db\.session\.add\(\n"
        r"        AuditLog\(\n"
        r"            user_id=current_user\.id if current_user\.is_authenticated else None,\n"
        r"            accion=action,\n"
        r"            entidad=entity,\n"
        r"            entidad_id=entity_id,\n"
        r"            detalle=detail,\n"
        r"        \)\n"
        r"    \)\n",
        "def audit(action: str, entity: str, entity_id=None, detail: str | None = None):\n"
        "    db.session.add(\n"
        "        BitacoraAuditoria(\n"
        "            usuario_id=current_user.id if current_user.is_authenticated else None,\n"
        "            accion=action,\n"
        "            tabla_afectada=entity,\n"
        "            registro_id=entity_id,\n"
        "            detalle=detail,\n"
        "        )\n"
        "    )\n",
        body,
    )

    body = body.replace(
        "AuditLog.query.options(joinedload(AuditLog.user)).order_by(AuditLog.created_at.desc())",
        "BitacoraAuditoria.query.options(joinedload(BitacoraAuditoria.usuario)).order_by(BitacoraAuditoria.fecha_hora.desc())",
    )

    # CentroCosto conserva valores minúsculos de la Fase 2.
    body = body.replace('project.tipo != "OBRA"', 'project.tipo != "obra"')
    body = body.replace('project.tipo == "OBRA"', 'project.tipo == "obra"')
    body = body.replace('project.tipo != "OFICINA"', 'project.tipo != "oficina"')
    body = body.replace('project.tipo == "OFICINA"', 'project.tipo == "oficina"')
    body = body.replace('p.tipo == "OFICINA"', 'p.tipo == "oficina"')
    body = body.replace('tipo="OBRA", activa=True', 'tipo="obra", activa=True')
    body = body.replace(
        'project_type = request.form.get("tipo", "OBRA")',
        'project_type = request.form.get("tipo", "obra").lower()',
    )
    body = body.replace(
        'if project_type not in {"OBRA", "OFICINA"}:',
        'if project_type not in {"obra", "oficina"}:',
    )

    # El correo es el identificador de acceso del ERP y nunca puede quedar nulo.
    body = body.replace(
        'email = request.form.get("email", "").strip().lower() or None',
        'email = request.form.get("email", "").strip().lower()',
    )
    body = body.replace(
        'duplicate = User.query.filter(func.lower(User.email) == email, User.id != user.id).first() if email else None',
        'if not email:\n                raise ValueError("El correo es obligatorio.")\n'
        '            duplicate = User.query.filter(func.lower(User.email) == email, User.id != user.id).first()',
    )
    body = body.replace(
        'user.projects = Project.query.filter(Project.id.in_(selected)).all() if role == "capturista" and selected else []',
        'user.projects = Project.query.filter(Project.id.in_(selected)).all() if role == "capturista" and selected else []\n'
        '            user.centro_costo_id = selected[0] if role == "capturista" and selected else None',
    )

    return HEADER + body.rstrip() + FOOTER + "\n"


def main() -> int:
    if len(sys.argv) != 3:
        print("Uso: build_nominas_route.py ORIGINAL_APP.PY DESTINO.PY")
        return 2
    source_path = Path(sys.argv[1]).resolve()
    destination = Path(sys.argv[2]).resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(build(source_path.read_text(encoding="utf-8")), encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
