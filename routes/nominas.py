"""Módulo de nóminas integrado desde la versión PythonAnywhere 2026-07-17.

La lógica de cálculo de este archivo proviene del sistema original validado.
La integración sustituye únicamente autenticación, usuarios y centros de costo
por los componentes compartidos del ERP V2.
"""

from __future__ import annotations

import re
from calendar import monthrange
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps
from io import BytesIO
from typing import Iterable

from flask import (
    Blueprint,
    Response,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template as flask_render_template,
    request,
    url_for as flask_url_for,
)
from flask_login import current_user, login_required
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from werkzeug.wsgi import FileWrapper

from models import (
    BitacoraAuditoria,
    CentroCosto,
    Usuario,
    db,
    usuario_centros_nomina as user_projects,
    utc_now,
)
from nominas_models import (
    AdditionalPayment,
    BudgetItem,
    Company,
    Contractor,
    Employee,
    Loan,
    LoanPayment,
    OfficeExpense,
    Payroll,
    PayrollLine,
    Subcontract,
    SubcontractPayment,
    WeeklyResourceAvailability,
)
from compras_models import (
    ACTIVE_ORDER_STATES,
    BudgetExplosionItem,
    PaymentMethod,
    PurchaseNotification,
    PurchaseOrder,
    Supplier,
    SupplyItem,
)
from utils.access import verificar_acceso_obra
from utils.decorators import permission_required
from utils.privacy import nss_para_usuario, puede_ver_nss_completo
from utils.user_permissions import verificar_cambio_rol
from utils.project_scope import obra_activa_id
from services.actualizacion_operativa import (
    asignar_obra_a_compradores,
    asignar_todas_las_obras_comprador,
)
from services.weekly_resources import (
    payment_channel,
    week_start_for,
    week_starts_between,
    weekly_resource_breakdown as calculate_weekly_resource_breakdown,
)


Project = CentroCosto
User = Usuario
nominas_bp = Blueprint("nominas", __name__)

MONEY_STEP = Decimal("0.01")
IVA_RATE = Decimal("0.16")
WORKDAYS = 5
WEEKDAY_FIELDS = ("lunes", "martes", "miercoles", "jueves", "viernes")
SHARED_PAYMENT_METHOD_DEFAULTS = (
    ("TRANSFERENCIA", "Transferencia bancaria"),
    ("CHEQUE", "Cheque"),
    ("EFECTIVO", "Efectivo"),
)
PAYROLL_STATES = ("borrador", "enviada", "aprobada", "pagada", "conciliada")
FINALIZED_PAYROLL_STATES = ("aprobada", "pagada", "conciliada")
LOAN_STATES = ("pendiente", "aprobado", "rechazado", "activo", "liquidado")
LOAN_DELIVERY_METHODS = ("EFECTIVO", "TRANSFERENCIA")
PAYROLL_PARTIDA_REQUIRED_MESSAGE = (
    "Debe asignar una partida a cada trabajador antes de guardar."
)


def url_for(endpoint: str, **values):
    """Resuelve endpoints originales dentro del Blueprint de Nóminas."""

    if endpoint == "static":
        return flask_url_for("static", **values)
    if "." not in endpoint:
        endpoint = f"nominas.{endpoint}"
    return flask_url_for(endpoint, **values)


def render_template(template_name: str, **context):
    """Mantiene las rutas de plantilla originales dentro de su namespace."""

    return flask_render_template(f"nominas/{template_name}", **context)


def shared_payment_methods() -> list[PaymentMethod]:
    """Catálogo común usado por Compras, préstamos y demás salidas."""

    for name, description in SHARED_PAYMENT_METHOD_DEFAULTS:
        if not PaymentMethod.query.filter(func.upper(PaymentMethod.nombre) == name).first():
            db.session.add(
                PaymentMethod(nombre=name, descripcion=description, activo=True)
            )
    db.session.flush()
    return PaymentMethod.query.filter_by(activo=True).order_by(PaymentMethod.nombre).all()


def selected_payment_method(legacy_field: str) -> PaymentMethod:
    method_id = request.form.get("payment_method_id", type=int)
    method = db.session.get(PaymentMethod, method_id) if method_id else None
    if not method:
        legacy_name = (request.form.get(legacy_field) or "").strip().upper()
        if legacy_name:
            method = PaymentMethod.query.filter(
                func.upper(PaymentMethod.nombre) == legacy_name
            ).first()
    if not method or not method.activo:
        raise ValueError("Selecciona un método de pago activo del catálogo.")
    return method


def loan_delivery_methods() -> list[PaymentMethod]:
    """Métodos que representan una entrega válida de préstamo."""

    return [
        method
        for method in shared_payment_methods()
        if method.nombre.strip().upper() in LOAN_DELIVERY_METHODS
    ]


def selected_loan_delivery_method() -> PaymentMethod:
    """Valida el método obligatorio y conserva EFECTIVO como fallback legado."""

    method_id = request.form.get("payment_method_id", type=int)
    method = db.session.get(PaymentMethod, method_id) if method_id else None
    legacy_name = (request.form.get("metodo_entrega") or "").strip().upper()
    if method is None and legacy_name:
        method = PaymentMethod.query.filter(
            func.upper(PaymentMethod.nombre) == legacy_name
        ).first()
    if method is None and not method_id and not legacy_name:
        method = PaymentMethod.query.filter(
            func.upper(PaymentMethod.nombre) == "EFECTIVO"
        ).first()
    normalized = (method.nombre if method else "").strip().upper()
    if (
        not method
        or not method.activo
        or normalized not in LOAN_DELIVERY_METHODS
    ):
        raise ValueError(
            "El método de entrega del préstamo debe ser Efectivo o Transferencia."
        )
    return method


@nominas_bp.get("/modulo-nominas")
@login_required
def index():
    if current_user.rol == "costos":
        return redirect(url_for("projects_list"))
    return redirect(url_for("dashboard"))

IMPORT_DEFINITIONS = {
    "trabajadores": {
        "title": "Importar trabajadores",
        "entity": "trabajadores",
        "columns": [
            ("NOMBRE_COMPLETO", True, "Nombre y al menos un apellido.", "JUAN PÉREZ LÓPEZ"),
            ("FECHA_INGRESO", True, "Fecha AAAA-MM-DD.", "2026-07-20"),
            ("PUESTO", True, "Categoría o puesto.", "OFICIAL ALBAÑIL"),
            ("OBRA_CODIGO", True, "Código de una obra existente.", "OBRA-01"),
            ("SALARIO_SEMANAL", True, "Sueldo libre semanal.", 5000),
            ("CUADRILLA", False, "Cuadrilla.", "ALEX"),
            ("SUPERVISOR", False, "Supervisor.", "AMIR"),
            ("EMPRESA_OPERATIVA", False, "Empresa operativa.", "RGOVC"),
            ("ACTIVO", False, "SI o NO. Vacío equivale a SI.", "SI"),
            ("FECHA_BAJA", False, "Fecha AAAA-MM-DD, si aplica.", ""),
            ("INFONAVIT", False, "Deducción semanal.", 0),
            ("TIENE_IMSS", False, "SI o NO.", "NO"),
            ("NSS", False, "Obligatorio cuando TIENE_IMSS=SI.", ""),
            ("EMPRESA_IMSS", False, "Código de empresa de registro.", ""),
            ("IMSS_TIPO", False, "FIJO o PORCENTAJE.", "FIJO"),
            ("IMSS_VALOR", False, "Pesos o porcentaje según tipo.", 0),
            ("TRANSFERENCIA_PREDETERMINADA", False, "Monto semanal por banco.", 0),
            ("EMPRESA_TRANSFERENCIA", False, "Código de empresa bancaria.", ""),
            ("EMPRESA_EFECTIVO", False, "Código de empresa para efectivo.", "BCH"),
            ("NOTAS", False, "Observaciones.", "ALTA MASIVA"),
        ],
    },
    "proveedores": {
        "title": "Importar pagos de proveedores",
        "entity": "pagos a proveedores",
        "columns": [
            ("FECHA", True, "Fecha del pago AAAA-MM-DD.", "2026-07-20"),
            ("OBRA_CODIGO", True, "Código de obra u oficina.", "OBRA-01"),
            ("PARTIDA_CODIGO", True, "Código de partida/subpartida.", "ADI-01"),
            ("INSUMO_CLAVE", True, "Clave existente en la explosión de esa partida.", "CEM-GRIS-50"),
            ("PROVEEDOR_CODIGO", True, "Código del proveedor en Compras.", "PROV-001"),
            ("PROVEEDOR", True, "Proveedor o beneficiario.", "MATERIALES DEL PACÍFICO"),
            ("CONCEPTO", True, "Descripción del gasto.", "MATERIAL EXTRA"),
            ("MONTO", True, "Importe capturado.", 2500),
            ("TIPO_MONTO", True, "SIN_IVA o CON_IVA.", "SIN_IVA"),
            ("METODO_PAGO", True, "EFECTIVO, TRANSFERENCIA o CHEQUE.", "TRANSFERENCIA"),
            ("EMPRESA_CODIGO", True, "Código de empresa que paga.", "RGOVC"),
            ("NOTAS", False, "Observaciones.", "PAGO DEL VIERNES"),
        ],
    },
    "subcontratistas": {
        "title": "Importar subcontratistas",
        "entity": "subcontratos",
        "columns": [
            ("CONTRATISTA", True, "Nombre o razón social.", "CONSTRUCTORA EJEMPLO"),
            ("ESPECIALIDAD", True, "Alcance del subcontrato.", "INSTALACIÓN ELÉCTRICA"),
            ("OBRA_CODIGO", True, "Código de una obra existente.", "OBRA-01"),
            ("PARTIDA_CODIGO", True, "Partida de categoría SUBCONTRATO.", "SUB-01"),
            ("PRESUPUESTO_SIN_IVA", True, "Monto contratado sin IVA.", 100000),
            ("AVANCE_PORCENTAJE", False, "Número de 0 a 100.", 0),
            ("UMBRAL_ALERTA_PORCENTAJE", False, "Número de 0 a 100.", 15),
            ("TELEFONO", False, "Teléfono del contratista.", "6121234567"),
            ("EMAIL", False, "Correo del contratista.", "pagos@ejemplo.com"),
            ("ACTIVO", False, "SI o NO. Vacío equivale a SI.", "SI"),
            ("OBSERVACIONES", False, "Notas del contrato.", "IMPORTADO DESDE EXCEL"),
        ],
    },
}


def decimal_value(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError, ValueError):
        return Decimal("0")


def money(value) -> Decimal:
    return decimal_value(value).quantize(MONEY_STEP, rounding=ROUND_HALF_UP)


def parse_money(field_name: str, default="0", required=False) -> Decimal:
    raw = request.form.get(field_name, "").strip().replace(",", "")
    if not raw:
        if required:
            raise ValueError(f"El campo {field_name} es obligatorio.")
        raw = str(default)
    try:
        value = Decimal(raw).quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"El valor de {field_name} no es un importe válido.") from exc
    if value < 0:
        raise ValueError(f"El valor de {field_name} no puede ser negativo.")
    return value


def parse_date(field_name: str, required=True) -> date | None:
    raw = request.form.get(field_name, "").strip()
    if not raw and not required:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"La fecha de {field_name} no es válida.") from exc


def excel_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def excel_money(value, field: str, required=False) -> Decimal:
    raw = excel_text(value).replace("$", "").replace(",", "").replace("%", "")
    if not raw:
        if required:
            raise ValueError(f"{field} es obligatorio.")
        return Decimal("0")
    try:
        amount = Decimal(raw).quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field} no es un importe válido.") from exc
    if amount < 0:
        raise ValueError(f"{field} no puede ser negativo.")
    return amount


def excel_date(value, field: str, required=False) -> date | None:
    if value in (None, ""):
        if required:
            raise ValueError(f"{field} es obligatorio.")
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = excel_text(value)
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"{field} debe ser una fecha AAAA-MM-DD.")


def excel_bool(value, default=False) -> bool:
    raw = excel_text(value).upper()
    if not raw:
        return default
    if raw in {"SI", "SÍ", "1", "TRUE", "VERDADERO", "ACTIVO"}:
        return True
    if raw in {"NO", "0", "FALSE", "FALSO", "INACTIVO"}:
        return False
    raise ValueError(f"El valor '{raw}' debe ser SI o NO.")


def import_project(code) -> Project:
    raw = excel_text(code).upper()
    project = Project.query.filter(func.upper(Project.codigo) == raw).first()
    if not project:
        raise ValueError(f"No existe la obra/oficina con código {raw or '(vacío)' }.")
    return project


def import_item(project: Project, code, required_category: str | None = None) -> BudgetItem:
    raw = excel_text(code).upper()
    item = BudgetItem.query.filter(
        BudgetItem.project_id == project.id,
        func.upper(BudgetItem.codigo) == raw,
        BudgetItem.activa.is_(True),
    ).first()
    if not item:
        raise ValueError(f"No existe la partida activa {raw or '(vacío)'} en {project.codigo}.")
    if required_category and item.categoria != required_category:
        raise ValueError(
            f"La partida {item.codigo} debe tener categoría {required_category}."
        )
    return item


def import_company(code, field: str, required=False) -> Company | None:
    raw = excel_text(code).upper()
    if not raw and not required:
        return None
    company = Company.query.filter(func.upper(Company.codigo) == raw).first()
    if not company:
        raise ValueError(f"{field}: no existe la empresa con código {raw or '(vacío)' }.")
    return company


def import_worker_row(data: dict, existing_action: str) -> tuple[str, str]:
    name = validate_full_name(excel_text(data["NOMBRE_COMPLETO"]))
    project = import_project(data["OBRA_CODIGO"])
    existing = Employee.query.filter_by(
        nombre_completo=name, project_id=project.id
    ).first()
    if existing and existing_action == "OMITIR":
        return "OMITIDO", f"{name} ya existe en {project.codigo}."

    employee = existing or Employee()
    position = excel_text(data["PUESTO"]).upper()
    if not position:
        raise ValueError("PUESTO es obligatorio.")
    salary = excel_money(data["SALARIO_SEMANAL"], "SALARIO_SEMANAL", required=True)
    if salary <= 0:
        raise ValueError("SALARIO_SEMANAL debe ser mayor que cero.")
    transfer = excel_money(data.get("TRANSFERENCIA_PREDETERMINADA"), "TRANSFERENCIA_PREDETERMINADA")
    transfer_company = import_company(
        data.get("EMPRESA_TRANSFERENCIA"), "EMPRESA_TRANSFERENCIA", required=transfer > 0
    )
    cash_company = import_company(data.get("EMPRESA_EFECTIVO"), "EMPRESA_EFECTIVO")
    has_imss = excel_bool(data.get("TIENE_IMSS"), default=False)
    imss_type = excel_text(data.get("IMSS_TIPO")).upper() or "FIJO"
    if imss_type not in {"FIJO", "PORCENTAJE"}:
        raise ValueError("IMSS_TIPO debe ser FIJO o PORCENTAJE.")
    imss_value = excel_money(data.get("IMSS_VALOR"), "IMSS_VALOR")
    if imss_type == "PORCENTAJE" and imss_value > 100:
        raise ValueError("IMSS_VALOR no puede superar 100 cuando es porcentaje.")
    nss = excel_text(data.get("NSS")) or None
    imss_company = import_company(
        data.get("EMPRESA_IMSS"), "EMPRESA_IMSS", required=has_imss
    )
    if has_imss and not nss:
        raise ValueError("NSS es obligatorio cuando TIENE_IMSS=SI.")

    employee.nombre_completo = name
    employee.fecha_ingreso = excel_date(data["FECHA_INGRESO"], "FECHA_INGRESO", required=True)
    employee.fecha_baja = excel_date(data.get("FECHA_BAJA"), "FECHA_BAJA")
    employee.activo = excel_bool(data.get("ACTIVO"), default=True) and not employee.fecha_baja
    employee.puesto = position
    employee.cuadrilla = excel_text(data.get("CUADRILLA")).upper() or None
    employee.supervisor = excel_text(data.get("SUPERVISOR")).upper() or None
    employee.empresa_operativa = excel_text(data.get("EMPRESA_OPERATIVA")).upper() or None
    employee.project_id = project.id
    # La partida no pertenece al maestro. Se selecciona por trabajador y
    # semana dentro de la nómina.
    employee.budget_item_id = None
    employee.salario_semanal = salary
    employee.descuento_infonavit = excel_money(data.get("INFONAVIT"), "INFONAVIT")
    employee.registrado_imss = has_imss
    employee.nss = nss if has_imss else None
    employee.empresa_imss_id = imss_company.id if imss_company and has_imss else None
    employee.imss_tipo = imss_type
    employee.descuento_imss = imss_value if has_imss else Decimal("0")
    employee.transferencia_predeterminada = transfer
    employee.empresa_transferencia_id = transfer_company.id if transfer_company else None
    employee.empresa_efectivo_id = cash_company.id if cash_company else None
    employee.notas = excel_text(data.get("NOTAS")) or None
    db.session.add(employee)
    db.session.flush()
    action = "ACTUALIZADO" if existing else "CREADO"
    audit(action, "TRABAJADOR", employee.id, f"Importación Excel: {name}")
    return action, f"{name} · {project.codigo}"


def import_supplier_row(data: dict, existing_action: str) -> tuple[str, str]:
    raise ValueError(
        "La carga directa de pagos de obra quedó cerrada. Registra el pago desde Compras contra una OC recepcionada."
    )
    project = import_project(data["OBRA_CODIGO"])
    item = import_item(project, data["PARTIDA_CODIGO"])
    supply_key = excel_text(data["INSUMO_CLAVE"]).upper()
    explosion = (
        BudgetExplosionItem.query.join(SupplyItem)
        .filter(
            BudgetExplosionItem.project_id == project.id,
            BudgetExplosionItem.budget_item_id == item.id,
            BudgetExplosionItem.activo.is_(True),
            func.upper(SupplyItem.clave) == supply_key,
        )
        .first()
    )
    if not explosion:
        raise ValueError(
            f"No existe el insumo activo {supply_key or '(vacío)'} en la explosión de {item.codigo}."
        )
    payment_date = excel_date(data["FECHA"], "FECHA", required=True)
    provider = normalize_name(excel_text(data["PROVEEDOR"]))
    provider_code = excel_text(data["PROVEEDOR_CODIGO"]).upper()
    concept = normalize_name(excel_text(data["CONCEPTO"]))
    if not provider_code or not provider or not concept:
        raise ValueError("PROVEEDOR_CODIGO, PROVEEDOR y CONCEPTO son obligatorios.")
    supplier = Supplier.query.filter(func.upper(Supplier.codigo) == provider_code).first()
    if not supplier:
        supplier = Supplier(codigo=provider_code, nombre=provider, moneda="MXN", activo=True)
        db.session.add(supplier)
        db.session.flush()
    elif supplier.nombre != provider:
        raise ValueError(
            f"El código {provider_code} pertenece a {supplier.nombre}, no a {provider}."
        )
    amount = excel_money(data["MONTO"], "MONTO", required=True)
    if amount <= 0:
        raise ValueError("MONTO debe ser mayor que cero.")
    amount_type = excel_text(data["TIPO_MONTO"]).upper()
    method = excel_text(data["METODO_PAGO"]).upper()
    if amount_type not in {"SIN_IVA", "CON_IVA"}:
        raise ValueError("TIPO_MONTO debe ser SIN_IVA o CON_IVA.")
    if method not in {"EFECTIVO", "TRANSFERENCIA", "CHEQUE"}:
        raise ValueError("METODO_PAGO debe ser EFECTIVO, TRANSFERENCIA o CHEQUE.")
    company = import_company(data["EMPRESA_CODIGO"], "EMPRESA_CODIGO", required=True)
    existing = AdditionalPayment.query.filter_by(
        fecha=payment_date,
        project_id=project.id,
        beneficiario=provider,
        concepto=concept,
    ).first()
    if existing and existing_action == "OMITIR":
        return "OMITIDO", f"{provider} · {concept} ya existe para esa fecha."
    payment = existing or AdditionalPayment(created_by_id=current_user.id)
    payment.fecha = payment_date
    payment.project_id = project.id
    payment.budget_item_id = item.id
    payment.explosion_item_id = explosion.id
    payment.supplier_id = supplier.id
    payment.purchase_order_id = None
    payment.beneficiario = provider
    payment.concepto = concept
    payment.monto_capturado = amount
    payment.tipo_monto = amount_type
    payment.monto_sin_iva = amount_without_vat(amount, amount_type)
    payment.metodo_pago = method
    payment.company_id = company.id
    payment.notas = excel_text(data.get("NOTAS")) or None
    db.session.add(payment)
    db.session.flush()
    action = "ACTUALIZADO" if existing else "CREADO"
    audit(action, "PAGO_ADICIONAL", payment.id, f"Importación Excel: {provider}")
    return action, f"{provider} · {concept} · {project.codigo}"


def import_subcontract_row(data: dict, existing_action: str) -> tuple[str, str]:
    contractor_name = normalize_name(excel_text(data["CONTRATISTA"]))
    specialty = normalize_name(excel_text(data["ESPECIALIDAD"]))
    if not contractor_name or not specialty:
        raise ValueError("CONTRATISTA y ESPECIALIDAD son obligatorios.")
    project = import_project(data["OBRA_CODIGO"])
    if project.tipo != "obra":
        raise ValueError("Los subcontratos solo pueden ligarse a centros tipo OBRA.")
    item = import_item(project, data["PARTIDA_CODIGO"], "SUBCONTRATO")
    contractor = Contractor.query.filter_by(nombre=contractor_name).first()
    if not contractor:
        contractor = Contractor(nombre=contractor_name, activo=True)
        db.session.add(contractor)
        db.session.flush()
    existing = Subcontract.query.filter_by(
        project_id=project.id,
        contractor_id=contractor.id,
        especialidad=specialty,
    ).first()
    if existing and existing_action == "OMITIR":
        return "OMITIDO", f"{contractor_name} · {specialty} ya existe en {project.codigo}."

    budget = excel_money(data["PRESUPUESTO_SIN_IVA"], "PRESUPUESTO_SIN_IVA", required=True)
    if budget <= 0:
        raise ValueError("PRESUPUESTO_SIN_IVA debe ser mayor que cero.")
    progress = excel_money(data.get("AVANCE_PORCENTAJE"), "AVANCE_PORCENTAJE")
    threshold_raw = excel_text(data.get("UMBRAL_ALERTA_PORCENTAJE"))
    threshold = excel_money(threshold_raw or "15", "UMBRAL_ALERTA_PORCENTAJE")
    if progress > 100 or threshold > 100:
        raise ValueError("Avance y umbral deben estar entre 0 y 100.")
    contractor.telefono = excel_text(data.get("TELEFONO")) or contractor.telefono
    contractor.email = excel_text(data.get("EMAIL")).lower() or contractor.email
    contractor.activo = excel_bool(data.get("ACTIVO"), default=True)
    subcontract = existing or Subcontract()
    subcontract.project_id = project.id
    subcontract.budget_item_id = item.id
    subcontract.contractor_id = contractor.id
    subcontract.especialidad = specialty
    subcontract.presupuesto_sin_iva = budget
    subcontract.avance_fisico = progress / Decimal("100")
    subcontract.umbral_alerta = threshold / Decimal("100")
    subcontract.observaciones = excel_text(data.get("OBSERVACIONES")) or None
    subcontract.activo = excel_bool(data.get("ACTIVO"), default=True)
    db.session.add(subcontract)
    db.session.flush()
    action = "ACTUALIZADO" if existing else "CREADO"
    audit(action, "SUBCONTRATO", subcontract.id, f"Importación Excel: {contractor_name}")
    return action, f"{contractor_name} · {specialty} · {project.codigo}"


def process_excel_import(kind: str, uploaded_file, existing_action: str) -> dict:
    definition = IMPORT_DEFINITIONS[kind]
    try:
        workbook = load_workbook(uploaded_file.stream, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("No fue posible leer el archivo. Usa la plantilla .xlsx sin proteger.") from exc
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_values = next(rows)
    except StopIteration as exc:
        raise ValueError("El archivo está vacío.") from exc
    headers = [excel_text(value).upper().replace(" ", "_") for value in header_values]
    expected = [column[0] for column in definition["columns"]]
    missing = [header for header in expected if header not in headers]
    if missing:
        raise ValueError("Faltan encabezados: " + ", ".join(missing))
    index = {header: headers.index(header) for header in expected}
    importer = {
        "trabajadores": import_worker_row,
        "proveedores": import_supplier_row,
        "subcontratistas": import_subcontract_row,
    }[kind]
    result = {"processed": 0, "created": 0, "updated": 0, "omitted": 0, "successes": [], "errors": []}
    for row_number, values in enumerate(rows, start=2):
        if row_number > 5001:
            result["errors"].append({"row": row_number, "message": "Se alcanzó el límite de 5,000 registros."})
            break
        data = {header: values[position] if position < len(values) else None for header, position in index.items()}
        if not any(value not in (None, "") for value in data.values()):
            continue
        result["processed"] += 1
        try:
            with db.session.begin_nested():
                status, message = importer(data, existing_action)
            if status == "CREADO":
                result["created"] += 1
            elif status == "ACTUALIZADO":
                result["updated"] += 1
            else:
                result["omitted"] += 1
            result["successes"].append({"row": row_number, "status": status, "message": message})
        except (ValueError, IntegrityError) as exc:
            result["errors"].append({"row": row_number, "message": str(exc).split("[SQL:")[0].strip()})
    db.session.commit()
    return result


def current_week_start(today: date | None = None) -> date:
    today = today or date.today()
    return today - timedelta(days=today.weekday())


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).upper()


def validate_full_name(value: str) -> str:
    name = normalize_name(value)
    words = [word for word in name.split(" ") if len(re.sub(r"[^A-ZÁÉÍÓÚÜÑ]", "", word)) >= 2]
    if len(words) < 2:
        raise ValueError("Captura el nombre completo del trabajador (mínimo nombre y apellido).")
    return name


def amount_without_vat(amount: Decimal, amount_type: str) -> Decimal:
    return money(amount / (Decimal("1") + IVA_RATE)) if amount_type == "CON_IVA" else money(amount)


def audit(action: str, entity: str, entity_id=None, detail: str | None = None):
    db.session.add(
        BitacoraAuditoria(
            usuario_id=current_user.id if current_user.is_authenticated else None,
            accion=action,
            tabla_afectada=entity,
            registro_id=entity_id,
            detalle=detail,
        )
    )


def nominas_permission_context() -> tuple[str, str]:
    endpoint = (request.endpoint or "").removeprefix("nominas.").lower()
    if endpoint in {"index", "dashboard"}:
        module = "nomina_dashboard"
    elif endpoint == "weekly_resource_save":
        module = "nominas_semanales"
    elif endpoint.startswith(("user",)):
        module = "usuarios"
    elif endpoint.startswith(("compan",)):
        module = "empresas_pago"
    elif endpoint.startswith(("project", "budget_item")):
        module = "obras_partidas"
    elif endpoint.startswith(("employee",)):
        module = "trabajadores"
    elif endpoint.startswith(("loan",)):
        module = "prestamos"
    elif endpoint.startswith(("payroll",)):
        module = "nominas_semanales"
    elif endpoint.startswith(("additional_payment",)):
        module = "pagos_adicionales"
    elif endpoint.startswith(("office_expense",)):
        module = "gastos_oficina"
    elif endpoint.startswith(("contractor",)):
        module = "contratistas"
    elif endpoint.startswith(("subcontract",)):
        module = "subcontratos"
    elif endpoint in {
        "weekly_closing_report",
        "reports_index",
        "reports_export",
    }:
        module = "reportes_nomina"
    elif endpoint == "audit_list":
        module = "seguridad"
    elif endpoint.startswith("import_"):
        kind = (request.view_args or {}).get("kind", "")
        module = {
            "trabajadores": "trabajadores",
            "proveedores": "pagos_proveedores",
            "subcontratos": "subcontratos",
        }.get(kind, "seguridad")
    else:
        module = "nomina_dashboard"
    explicit_actions = {
        "user_edit": "editar",
        "project_new": "crear",
        "project_edit": "editar",
        "budget_item_new": "crear",
        "budget_item_edit": "editar",
        "employee_new": "crear",
        "employee_edit": "editar",
        "loan_new": "crear",
        "payroll_new": "crear",
        "payroll_add_employee": "editar",
        "additional_payment_new": "crear",
        "office_expense_new": "crear",
        "subcontract_new": "crear",
        "subcontract_edit": "editar",
        "subcontract_payment_new": "crear",
        "import_data": "crear",
        "import_template_download": "crear",
    }
    if endpoint in explicit_actions:
        return module, explicit_actions[endpoint]
    if request.method == "GET":
        return module, "ver"
    if endpoint in {"companies_list", "contractors_list"}:
        return module, "crear"
    if endpoint in {"loan_approve", "loan_reject"}:
        return module, "aprobar"
    if any(term in endpoint for term in ("eliminar", "delete", "baja")):
        action = "eliminar"
    elif "cancel" in endpoint:
        action = "cancelar"
    elif any(term in endpoint for term in ("approve", "reject", "aprobar", "rechazar")):
        action = "aprobar"
    elif "reconcile" in endpoint or "conciliar" in endpoint:
        action = "conciliar"
    elif endpoint.endswith("_pay") or "_payment_new" in endpoint:
        action = "pagar"
    elif any(term in endpoint for term in ("nuevo", "new", "agregar", "import")):
        action = "crear"
    else:
        action = "editar"
    return module, action


def admin_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        module, action = nominas_permission_context()
        if not current_user.tiene_permiso(module, action):
            abort(403)
        return view(*args, **kwargs)

    return wrapped


def accessible_projects_query(include_inactive=False):
    query = Project.query
    if not include_inactive:
        query = query.filter_by(activa=True)
    if not current_user.acceso_global_obras:
        query = query.join(user_projects).filter(user_projects.c.user_id == current_user.id)
    if current_user.rol == "supervisor":
        selected_id = obra_activa_id(
            current_user,
            incluir_inactivas=include_inactive,
        )
        query = query.filter(Project.id == (selected_id or -1))
    return query.order_by(Project.nombre)


def can_access_project(project: Project) -> bool:
    if current_user.acceso_global_obras:
        return True
    if project not in current_user.projects:
        return False
    return (
        current_user.rol != "supervisor"
        or obra_activa_id(current_user, incluir_inactivas=False) == project.id
    )


def require_project_access(project: Project):
    verificar_acceso_obra(current_user, project.id)


@nominas_bp.context_processor
def inject_helpers():
    return {
        "money": money,
        "current_week_start": current_week_start,
        "today": date.today(),
        "url_for": url_for,
        "nss_display": lambda value: nss_para_usuario(current_user, value),
        "can_view_full_nss": puede_ver_nss_completo(current_user),
    }


@nominas_bp.app_template_filter("currency")
def currency_filter(value):
    return f"${money(value):,.2f}"


@nominas_bp.app_template_filter("date_es")
def date_filter(value):
    return value.strftime("%d/%m/%Y") if value else "—"


@nominas_bp.before_request
def enforce_nominas_role_scope():
    """Aplica la matriz configurable antes de cualquier vista de Nóminas."""

    if not current_user.is_authenticated:
        return None
    # El listado y detalle de las obras asignadas forman parte de la operación
    # cotidiana de Nómina, aunque la administración del catálogo pertenezca al
    # módulo ``centros_costo``. Así un capturista puede abrir su obra sin poder
    # crearla, editarla ni cerrarla.
    if request.method == "GET" and request.endpoint in {
        "nominas.projects_list",
        "nominas.project_detail",
    }:
        if not current_user.tiene_permiso("obras_partidas", "ver"):
            abort(403)
        return None
    module, action = nominas_permission_context()
    if not current_user.tiene_permiso(module, action):
        abort(403)
    return None


def descendant_item_ids(item: BudgetItem) -> list[int]:
    ids = [item.id]
    for child in item.children:
        ids.extend(descendant_item_ids(child))
    return ids


def payroll_line_item_filter(item: BudgetItem):
    """Filtra una partida o subpartida usando la asignación semanal.

    El fallback por ``budget_item_id`` solo cubre filas históricas todavía no
    normalizadas; nunca consulta el maestro ``employees``.
    """

    if item.parent_id is None:
        return or_(
            PayrollLine.partida_id == item.id,
            and_(
                PayrollLine.partida_id.is_(None),
                PayrollLine.budget_item_id.in_(descendant_item_ids(item)),
            ),
        )
    return or_(
        PayrollLine.subpartida_id == item.id,
        and_(
            PayrollLine.subpartida_id.is_(None),
            PayrollLine.budget_item_id == item.id,
        ),
    )


def payroll_line_items_filter(item_ids: Iterable[int]):
    """Expresión de compatibilidad para un conjunto de ítems presupuestarios."""

    ids = list(item_ids)
    return or_(
        PayrollLine.partida_id.in_(ids or [-1]),
        PayrollLine.subpartida_id.in_(ids or [-1]),
        and_(
            PayrollLine.partida_id.is_(None),
            PayrollLine.subpartida_id.is_(None),
            PayrollLine.budget_item_id.in_(ids or [-1]),
        ),
    )


def infer_payroll_line_allocation(
    line: PayrollLine,
) -> tuple[BudgetItem | None, BudgetItem | None]:
    """Resuelve partida/subpartida, incluyendo una fila histórica."""

    partida = line.partida
    subpartida = line.subpartida
    if partida is not None:
        return partida, subpartida
    item = line.budget_item
    if item is None:
        return None, None
    if item.parent_id is None:
        return item, None
    return item.parent, item


def assign_payroll_line_budget(
    line: PayrollLine,
    payroll: Payroll,
    partida_id: int | None,
    subpartida_id: int | None,
) -> None:
    """Valida y guarda la selección presupuestaria de una línea."""

    if not partida_id:
        raise ValueError(PAYROLL_PARTIDA_REQUIRED_MESSAGE)
    partida = db.session.get(BudgetItem, partida_id)
    if (
        not partida
        or partida.project_id != payroll.project_id
        or partida.parent_id is not None
        or not partida.activa
    ):
        raise ValueError("Selecciona una partida activa de la obra.")

    subpartidas_activas = [
        item
        for item in partida.children
        if item.activa and item.project_id == payroll.project_id
    ]
    subpartida = (
        db.session.get(BudgetItem, subpartida_id)
        if subpartida_id
        else None
    )
    if subpartidas_activas and subpartida is None:
        raise ValueError(PAYROLL_PARTIDA_REQUIRED_MESSAGE)
    if subpartida is not None and (
        subpartida.parent_id != partida.id
        or subpartida.project_id != payroll.project_id
        or not subpartida.activa
    ):
        raise ValueError("Selecciona una subpartida válida de la partida indicada.")
    if not subpartidas_activas and subpartida is not None:
        raise ValueError("La partida seleccionada no requiere subpartida.")

    line.partida = partida
    line.subpartida = subpartida
    line.partida_id = partida.id
    line.subpartida_id = subpartida.id if subpartida else None
    line.budget_item = subpartida or partida
    line.budget_item_id = subpartida.id if subpartida else partida.id


def validate_payroll_line_budget(line: PayrollLine) -> None:
    """Normaliza una fila histórica y aplica la validación de guardado/cierre."""

    partida, subpartida = infer_payroll_line_allocation(line)
    assign_payroll_line_budget(
        line,
        line.payroll,
        partida.id if partida else None,
        subpartida.id if subpartida else None,
    )


def item_consumption(item: BudgetItem) -> dict[str, Decimal]:
    """Consumo real sin IVA de una partida y todas sus subpartidas."""

    ids = descendant_item_ids(item)
    payroll_value = (
        db.session.query(
            func.coalesce(
                func.sum(
                    PayrollLine.monto_devengado
                    + PayrollLine.pago_extra
                    + PayrollLine.descuento_imss
                ),
                0,
            )
        )
        .join(Payroll)
        .filter(
            Payroll.estado.in_(FINALIZED_PAYROLL_STATES),
            payroll_line_item_filter(item),
        )
        .scalar()
    )
    imss_value = (
        db.session.query(func.coalesce(func.sum(PayrollLine.descuento_imss), 0))
        .join(Payroll)
        .filter(
            Payroll.estado.in_(FINALIZED_PAYROLL_STATES),
            payroll_line_item_filter(item),
        )
        .scalar()
    )
    additional_value = (
        db.session.query(func.coalesce(func.sum(AdditionalPayment.monto_sin_iva), 0))
        .filter(AdditionalPayment.budget_item_id.in_(ids))
        .scalar()
    )
    subcontract_value = (
        db.session.query(func.coalesce(func.sum(SubcontractPayment.monto_sin_iva), 0))
        .join(Subcontract)
        .filter(Subcontract.budget_item_id.in_(ids))
        .scalar()
    )
    office_value = (
        db.session.query(func.coalesce(func.sum(OfficeExpense.monto_sin_iva), 0))
        .filter(OfficeExpense.budget_item_id.in_(ids))
        .scalar()
    )
    result = {
        "nomina": money(payroll_value),
        "imss": money(imss_value),
        "adicionales": money(additional_value),
        "subcontratos": money(subcontract_value),
        "oficina": money(office_value),
    }
    result["total"] = money(
        result["nomina"] + result["adicionales"] + result["subcontratos"] + result["oficina"]
    )
    result["restante"] = money(decimal_value(item.presupuesto) - result["total"])
    return result


def project_consumption(project: Project) -> dict[str, Decimal]:
    """Resume presupuesto, asignaciones y consumo real de una obra.

    Los cinco indicadores del dashboard se concilian así::

        presupuesto total = disponible real + comprometido
        comprometido = consumido real + disponible comprometido

    ``comprometido`` es exclusivamente el presupuesto capturado en
    subpartidas. ``consumido real`` considera únicamente nóminas cerradas;
    los borradores todavía no afectan el presupuesto ejecutado.
    """

    payroll_value = (
        db.session.query(
            func.coalesce(
                func.sum(
                    PayrollLine.monto_devengado
                    + PayrollLine.pago_extra
                    + PayrollLine.descuento_imss
                ),
                0,
            )
        )
        .join(Payroll)
        .filter(Payroll.estado.in_(FINALIZED_PAYROLL_STATES), Payroll.project_id == project.id)
        .scalar()
    )
    imss_value = (
        db.session.query(func.coalesce(func.sum(PayrollLine.descuento_imss), 0))
        .join(Payroll)
        .filter(Payroll.estado.in_(FINALIZED_PAYROLL_STATES), Payroll.project_id == project.id)
        .scalar()
    )
    additional_value = (
        db.session.query(func.coalesce(func.sum(AdditionalPayment.monto_sin_iva), 0))
        .filter(AdditionalPayment.project_id == project.id)
        .scalar()
    )
    subcontract_value = (
        db.session.query(func.coalesce(func.sum(SubcontractPayment.monto_sin_iva), 0))
        .join(Subcontract)
        .filter(Subcontract.project_id == project.id)
        .scalar()
    )
    office_value = (
        db.session.query(func.coalesce(func.sum(OfficeExpense.monto_sin_iva), 0))
        .filter(OfficeExpense.project_id == project.id)
        .scalar()
    )
    labor_ids = [i.id for i in project.budget_items if i.categoria == "MANO_OBRA"]
    labor_value = Decimal("0")
    if labor_ids:
        labor_value = money(
            db.session.query(
                func.coalesce(
                    func.sum(
                        PayrollLine.monto_devengado
                        + PayrollLine.pago_extra
                        + PayrollLine.descuento_imss
                    ),
                    0,
                )
            )
            .join(Payroll)
            .filter(
                Payroll.estado.in_(FINALIZED_PAYROLL_STATES),
                payroll_line_items_filter(labor_ids),
            )
            .scalar()
        )
    result = {
        "nomina": money(payroll_value),
        "imss": money(imss_value),
        "adicionales": money(additional_value),
        "subcontratos": money(subcontract_value),
        "oficina": money(office_value),
        "mano_obra": money(labor_value),
    }
    # Para este dashboard, consumo real es únicamente lo aplicado a nóminas
    # cerradas. Los demás movimientos conservan sus reportes operativos, pero
    # no se mezclan con los cinco indicadores solicitados para la obra.
    result["total"] = result["nomina"]
    total_budget = decimal_value(project.presupuesto_total)
    top_items = [item for item in project.budget_items if item.parent_id is None]
    subitems = [item for item in project.budget_items if item.parent_id is not None]
    parent_budget = sum(
        (decimal_value(item.presupuesto) for item in top_items), Decimal("0")
    )
    committed_budget = sum(
        (decimal_value(item.presupuesto) for item in subitems), Decimal("0")
    )
    if total_budget == 0:
        total_budget = parent_budget
    result["presupuesto_base"] = money(total_budget)
    result["presupuesto_partidas"] = money(parent_budget)
    result["presupuesto_subpartidas"] = money(committed_budget)

    # Disponible real = saldo todavía libre dentro de las partidas + la parte
    # del presupuesto general que aún no se distribuye a ninguna partida.
    result["disponible_partidas"] = money(parent_budget - committed_budget)
    result["sin_asignar_partidas"] = money(total_budget - parent_budget)
    result["restante_total"] = money(
        result["disponible_partidas"] + result["sin_asignar_partidas"]
    )

    # Disponible comprometido = lo asignado en subpartidas que todavía no se
    # ha consumido mediante nóminas cerradas.
    result["total_comprometido"] = money(committed_budget)
    result["restante_comprometido"] = money(
        result["total_comprometido"] - result["total"]
    )
    result["restante_mano_obra"] = money(
        decimal_value(project.presupuesto_mano_obra) - result["mano_obra"]
    )
    if total_budget > 0:
        result["porcentaje_consumido"] = money(result["total"] / total_budget * Decimal("100"))
        result["porcentaje_comprometido"] = money(
            result["total_comprometido"] / total_budget * Decimal("100")
        )
    else:
        result["porcentaje_consumido"] = Decimal("0")
        result["porcentaje_comprometido"] = Decimal("0")
    result["alerta_consumido"] = result["porcentaje_consumido"] >= Decimal("80")
    result["alerta_comprometido"] = result["porcentaje_comprometido"] >= Decimal("80")
    return result


def weekly_resource_breakdown(
    week_start: date, project_ids: Iterable[int] | None = None
) -> dict:
    """Adaptador histórico hacia la única fórmula de recurso semanal."""

    return calculate_weekly_resource_breakdown(
        week_start,
        project_ids,
        draft_refresher=refresh_draft_payroll,
    )


def loan_remaining(loan: Loan) -> Decimal:
    paid = (
        db.session.query(func.coalesce(func.sum(LoanPayment.monto), 0))
        .filter(LoanPayment.loan_id == loan.id)
        .scalar()
    )
    return max(Decimal("0"), money(loan.total_pagar) - money(paid))


def add_calendar_months(value: date, months: int) -> date:
    """Suma meses sin convertir la regla laboral en un número fijo de días."""

    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    day = min(value.day, monthrange(year, month)[1])
    return date(year, month, day)


def employee_weekly_net(employee: Employee) -> Decimal:
    """Neto base disponible antes de préstamos y descuentos extraordinarios."""

    return max(
        Decimal("0"),
        money(employee.salario_semanal) - money(employee.descuento_infonavit),
    )


def validate_new_loan(employee: Employee, loan_date: date, amount: Decimal) -> Decimal:
    """Aplica antigüedad, interés obligatorio y tope sobre el salario neto."""

    if loan_date <= add_calendar_months(employee.fecha_ingreso, 6):
        raise ValueError(
            "El trabajador no cumple con la antigüedad mínima de 6 meses."
        )
    total = money(amount * Decimal("1.05"))
    if total > employee_weekly_net(employee):
        raise ValueError(
            "El préstamo no puede exceder el salario semanal del trabajador."
        )
    return total


def notify_loan_admins(loan: Loan) -> None:
    """Crea una notificación interna para cada administrador activo."""

    admins = Usuario.query.filter_by(rol="admin", activo=True).all()
    link = url_for("loans_list", estado="pendiente")
    for admin in admins:
        db.session.add(
            PurchaseNotification(
                user_id=admin.id,
                tipo="PRESTAMO_PENDIENTE",
                mensaje=(
                    f"Solicitud de préstamo de {loan.employee.nombre_completo} "
                    f"por {currency_filter(loan.monto)} más 5% de interés."
                )[:500],
                enlace=link,
            )
        )


def loan_plan(employee_id: int, week_start: date, available: Decimal) -> list[tuple[Loan, Decimal]]:
    """Retenciones elegibles: inician la semana posterior al préstamo."""

    available = max(Decimal("0"), money(available))
    plan: list[tuple[Loan, Decimal]] = []
    loans = (
        Loan.query.filter(
            Loan.employee_id == employee_id,
            Loan.estado == "activo",
            Loan.fecha_prestamo < week_start,
        )
        .order_by(Loan.fecha_prestamo, Loan.id)
        .all()
    )
    for loan in loans:
        if available <= 0:
            break
        remaining = loan_remaining(loan)
        if remaining <= 0:
            continue
        deduction = min(money(loan.retencion_semanal), remaining, available)
        if deduction > 0:
            plan.append((loan, deduction))
            available -= deduction
    return plan


def scheduled_loan_deduction(employee_id: int, week_start: date) -> Decimal:
    """Máximo semanal elegible, antes de limitarlo por el neto disponible."""

    plan = loan_plan(employee_id, week_start, Decimal("9999999999.99"))
    return money(sum((amount for _loan, amount in plan), Decimal("0")))


def calculate_imss_cost(employee: Employee | None, gross: Decimal) -> Decimal:
    """Calcula la prestación IMSS patronal sin descontarla del sueldo libre."""

    if not employee or not employee.registrado_imss:
        return Decimal("0")
    configured = decimal_value(employee.descuento_imss)
    if (employee.imss_tipo or "FIJO") == "PORCENTAJE":
        return money(decimal_value(gross) * configured / Decimal("100"))
    return money(configured)


def recalculate_line(line: PayrollLine, transfer_requested=None, strict_transfer=False):
    """Centraliza todas las fórmulas que antes estaban repartidas en Excel."""

    weekly = money(line.salario_semanal)
    days = Decimal(sum(1 for field in WEEKDAY_FIELDS if getattr(line, field)))
    daily = money(weekly / Decimal(WORKDAYS))
    gross = money(daily * days)
    absence = money(weekly - gross)

    line.dias_trabajados = days
    line.numero_faltas = Decimal(WORKDAYS) - days
    line.sueldo_diario = daily
    line.descuento_faltas = absence
    line.monto_devengado = gross
    line.descuento_imss = calculate_imss_cost(line.employee, gross)

    before_loans = money(
        gross
        + decimal_value(line.pago_extra)
        - decimal_value(line.descuento_infonavit)
        - decimal_value(line.otro_descuento)
    )
    before_loans = max(Decimal("0"), before_loans)
    plan = loan_plan(line.employee_id, line.payroll.semana_inicio, before_loans)
    line.descuento_prestamo = money(sum((amount for _loan, amount in plan), Decimal("0")))
    net = money(max(Decimal("0"), before_loans - decimal_value(line.descuento_prestamo)))
    line.neto_pagar = net

    requested = money(line.pago_transferencia if transfer_requested is None else transfer_requested)
    if strict_transfer and requested > net:
        difference = money(requested - net)
        raise ValueError(
            f"La transferencia de {line.nombre_trabajador} es {currency_filter(requested)}, "
            f"pero su neto actual es {currency_filter(net)} "
            f"(diferencia: {currency_filter(difference)}). "
            f"El préstamo automático aplicado es {currency_filter(line.descuento_prestamo)}. "
            "Actualiza la transferencia y vuelve a guardar."
        )
    requested = min(max(Decimal("0"), requested), net)
    line.pago_transferencia = requested
    line.pago_efectivo = money(net - requested)
    return plan


def refresh_draft_payroll(payroll: Payroll) -> int:
    """Actualiza en memoria préstamos y netos de una nómina aún editable."""

    if payroll.estado != "borrador":
        return 0
    for line in payroll.lines:
        recalculate_line(line)
    return len(payroll.lines)


def refresh_employee_draft_lines(employee_id: int) -> int:
    """Sincroniza borradores si un préstamo se registra o cancela después de crearlos."""

    lines = (
        PayrollLine.query.join(Payroll)
        .filter(
            PayrollLine.employee_id == employee_id,
            Payroll.estado == "borrador",
        )
        .all()
    )
    changed = 0
    for line in lines:
        previous = (
            money(line.descuento_prestamo),
            money(line.neto_pagar),
            money(line.pago_transferencia),
            money(line.pago_efectivo),
        )
        recalculate_line(line)
        current = (
            money(line.descuento_prestamo),
            money(line.neto_pagar),
            money(line.pago_transferencia),
            money(line.pago_efectivo),
        )
        if current != previous:
            changed += 1
    return changed


def validate_line_for_close(line: PayrollLine):
    effective_item = line.budget_item_efectivo
    if not effective_item or effective_item.project_id != line.payroll.project_id:
        raise ValueError(PAYROLL_PARTIDA_REQUIRED_MESSAGE)
    validate_full_name(line.nombre_trabajador)
    if decimal_value(line.pago_transferencia) > 0 and not line.empresa_transferencia_id:
        raise ValueError(f"Selecciona la empresa de transferencia para {line.nombre_trabajador}.")
    if decimal_value(line.pago_efectivo) > 0 and not line.empresa_efectivo_id:
        raise ValueError(f"Selecciona la empresa de efectivo para {line.nombre_trabajador}.")


@nominas_bp.route("/panel")
@login_required
def dashboard():
    week_start = current_week_start()
    week_end = week_start + timedelta(days=4)
    projects = accessible_projects_query().all()
    project_ids = [p.id for p in projects]

    payroll_query = Payroll.query.filter(
        Payroll.semana_inicio == week_start, Payroll.project_id.in_(project_ids or [-1])
    )
    current_payrolls = payroll_query.options(joinedload(Payroll.lines), joinedload(Payroll.project)).all()
    total_current = money(sum((p.total_neto for p in current_payrolls), Decimal("0")))
    total_gross = money(sum((p.total_devengado for p in current_payrolls), Decimal("0")))
    active_workers = Employee.query.filter(
        Employee.activo.is_(True), Employee.project_id.in_(project_ids or [-1])
    ).count()
    active_loans = Loan.query.join(Employee).filter(
        Loan.estado == "activo", Employee.project_id.in_(project_ids or [-1])
    ).all()
    loan_balance = money(sum((loan_remaining(loan) for loan in active_loans), Decimal("0")))
    project_cards = [(project, project_consumption(project)) for project in projects]
    budget_alerts = [
        (project, values)
        for project, values in project_cards
        if values["alerta_consumido"] or values["alerta_comprometido"]
    ]

    capture_progress = []
    payroll_by_project = {payroll.project_id: payroll for payroll in current_payrolls}
    for project in projects:
        active_count = Employee.query.filter_by(project_id=project.id, activo=True).count()
        payroll = payroll_by_project.get(project.id)
        captured = len(payroll.lines) if payroll else 0
        progress = (
            Decimal("100")
            if payroll and active_count == 0
            else money(min(Decimal("100"), Decimal(captured) / Decimal(active_count) * 100))
            if active_count
            else Decimal("0")
        )
        capture_progress.append(
            {
                "project": project,
                "payroll": payroll,
                "active_workers": active_count,
                "captured_workers": captured,
                "progress": progress,
                "pending": not payroll or payroll.estado in {"borrador", "enviada"},
            }
        )

    budget_totals = {
        "presupuesto": money(
            sum((values["presupuesto_base"] for _project, values in project_cards), Decimal("0"))
        ),
        "consumido": money(sum((values["total"] for _project, values in project_cards), Decimal("0"))),
        "comprometido": money(
            sum((values["total_comprometido"] for _project, values in project_cards), Decimal("0"))
        ),
    }
    budget_totals["disponible"] = money(
        sum((values["restante_total"] for _project, values in project_cards), Decimal("0"))
    )
    budget_totals["disponible_comprometido"] = money(
        sum(
            (values["restante_comprometido"] for _project, values in project_cards),
            Decimal("0"),
        )
    )
    budget_totals["porcentaje_consumido"] = (
        money(budget_totals["consumido"] / budget_totals["presupuesto"] * 100)
        if budget_totals["presupuesto"]
        else Decimal("0")
    )
    budget_totals["porcentaje_comprometido"] = (
        money(budget_totals["comprometido"] / budget_totals["presupuesto"] * 100)
        if budget_totals["presupuesto"]
        else Decimal("0")
    )

    labor_status = []
    all_payrolls = (
        Payroll.query.options(joinedload(Payroll.lines))
        .filter(Payroll.project_id.in_(project_ids or [-1]))
        .all()
    )
    payroll_history: dict[int, list[Payroll]] = {}
    for payroll in all_payrolls:
        payroll_history.setdefault(payroll.project_id, []).append(payroll)
    for project in projects:
        records = payroll_history.get(project.id, [])
        paid = money(
            sum((p.total_neto for p in records if p.estado in FINALIZED_PAYROLL_STATES), Decimal("0"))
        )
        pending = money(
            sum((p.total_neto for p in records if p.estado in {"borrador", "enviada"}), Decimal("0"))
        )
        labor_status.append(
            {
                "project": project,
                "devengado": money(paid + pending),
                "pagado": paid,
                "pendiente": pending,
            }
        )

    subcontract_status = (
        Subcontract.query.options(
            joinedload(Subcontract.project),
            joinedload(Subcontract.contractor),
            joinedload(Subcontract.payments),
        )
        .filter(Subcontract.project_id.in_(project_ids or [-1]), Subcontract.activo.is_(True))
        .order_by(Subcontract.project_id, Subcontract.especialidad)
        .all()
    )
    resource_summary = weekly_resource_breakdown(
        week_start, None if current_user.acceso_global_obras else project_ids
    )

    return render_template(
        "dashboard.html",
        week_start=week_start,
        week_end=week_end,
        payrolls=current_payrolls,
        total_current=total_current,
        total_gross=total_gross,
        active_workers=active_workers,
        active_loans=len(active_loans),
        loan_balance=loan_balance,
        project_cards=project_cards,
        budget_alerts=budget_alerts,
        budget_totals=budget_totals,
        resource_summary=resource_summary,
        subcontract_status=subcontract_status,
        labor_status=labor_status,
        capture_progress=capture_progress,
        drafts=sum(1 for p in current_payrolls if p.estado in {"borrador", "enviada"}),
        closed=sum(1 for p in current_payrolls if p.estado in FINALIZED_PAYROLL_STATES),
    )


@nominas_bp.route("/recursos-semanales/guardar", methods=["POST"])
@admin_required
def weekly_resource_save():
    try:
        selected_date = parse_date("semana_inicio")
        week_start = current_week_start(selected_date)
        for method, field in (
            ("EFECTIVO", "disponible_efectivo"),
            ("TRANSFERENCIA", "disponible_transferencia"),
        ):
            amount = parse_money(field, required=True)
            row = WeeklyResourceAvailability.query.filter_by(
                semana_inicio=week_start, metodo=method
            ).first()
            if not row:
                row = WeeklyResourceAvailability(
                    semana_inicio=week_start,
                    metodo=method,
                    updated_by_id=current_user.id,
                )
                db.session.add(row)
            row.monto_disponible = amount
            row.updated_by_id = current_user.id
            row.updated_at = utc_now()
        audit("ACTUALIZAR", "RECURSO_SEMANAL", detail=str(week_start))
        db.session.commit()
        flash("Disponibilidad semanal guardada por método de pago.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    destination = request.form.get("return_to", "")
    return redirect(destination if destination.startswith("/") and not destination.startswith("//") else url_for("dashboard"))


@nominas_bp.route("/usuarios")
@admin_required
def users_list():
    return render_template("users/list.html", users=User.query.order_by(User.nombre_completo).all())


@nominas_bp.route("/usuarios/<int:user_id>/editar", methods=["GET", "POST"])
@admin_required
def user_edit(user_id):
    user = db.get_or_404(User, user_id)
    projects = accessible_projects_query().all()
    if request.method == "POST":
        try:
            user.nombre_completo = validate_full_name(request.form.get("nombre_completo", ""))
            email = request.form.get("email", "").strip().lower()
            if not email:
                raise ValueError("El correo es obligatorio.")
            duplicate = User.query.filter(func.lower(User.email) == email, User.id != user.id).first()
            if duplicate:
                raise ValueError("Ese correo ya está registrado.")
            user.email = email
            role = request.form.get("role", "capturista")
            if role not in {
                "administrador",
                "admin_financiero",
                "capturista",
                "supervisor",
                "comprador",
                "almacenista",
                "ceo",
                "costos",
            }:
                raise ValueError("Rol inválido.")
            normalized_role = verificar_cambio_rol(current_user, user, role)
            selected = [int(value) for value in request.form.getlist("project_ids") if value.isdigit()]
            editing_self = user.id == current_user.id
            current_project_ids = {project.id for project in user.projects}
            if editing_self and set(selected) != current_project_ids:
                abort(403, description="No puedes modificar tu propio rol o permisos.")
            if user.id == 1 and normalized_role != "admin":
                abort(403, description="El administrador principal debe conservar acceso total.")
            if not editing_self:
                for project_id in selected:
                    verificar_acceso_obra(current_user, project_id)
                user.role = normalized_role
                scoped_role = normalized_role in {
                    "capturista",
                    "supervisor",
                    "comprador",
                    "almacenista",
                }
                selected_projects = (
                    Project.query.filter(Project.id.in_(selected)).all()
                    if scoped_role and selected
                    else []
                )
                if normalized_role in {"comprador", "almacenista"} and any(
                    project.tipo != "obra" for project in selected_projects
                ):
                    raise ValueError(
                        "Comprador y Almacenista solo pueden asignarse a obras."
                    )
                user.projects = selected_projects
                if normalized_role == "comprador":
                    asignar_todas_las_obras_comprador(user)
                user.centro_costo_id = (
                    selected[0]
                    if normalized_role in {"capturista", "supervisor"} and selected
                    else None
                )
            new_password = request.form.get("password", "")
            if new_password:
                if len(new_password) < 8:
                    raise ValueError("La nueva contraseña debe tener al menos 8 caracteres.")
                user.set_password(new_password)
            audit("EDITAR", "USUARIO", user.id)
            db.session.commit()
            flash("Usuario actualizado.", "success")
            return redirect(url_for("users_list"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "users/form.html",
        user=user,
        projects=projects,
        editing_self=user.id == current_user.id,
    )


@nominas_bp.route("/usuarios/<int:user_id>/estado", methods=["POST"])
@admin_required
def user_toggle(user_id):
    user = db.get_or_404(User, user_id)
    if user.id == 1:
        flash("El administrador principal (id=1) debe permanecer activo.", "danger")
    elif user.id == current_user.id:
        flash("No puedes desactivar tu propia cuenta.", "danger")
    else:
        user.activo = not user.activo
        audit("ACTIVAR" if user.activo else "DESACTIVAR", "USUARIO", user.id)
        db.session.commit()
        flash("Estado del usuario actualizado.", "success")
    return redirect(url_for("users_list"))


@nominas_bp.route("/administracion/importar/<kind>", methods=["GET", "POST"])
@admin_required
def import_data(kind):
    if kind not in IMPORT_DEFINITIONS:
        abort(404)
    definition = IMPORT_DEFINITIONS[kind]
    result = None
    if request.method == "POST":
        uploaded = request.files.get("archivo")
        existing_action = request.form.get("existing_action", "OMITIR").upper()
        if existing_action not in {"OMITIR", "ACTUALIZAR"}:
            existing_action = "OMITIR"
        if not uploaded or not uploaded.filename:
            flash("Selecciona un archivo Excel .xlsx.", "danger")
        elif not uploaded.filename.lower().endswith(".xlsx"):
            flash("El archivo debe tener extensión .xlsx.", "danger")
        else:
            try:
                result = process_excel_import(kind, uploaded, existing_action)
                flash(
                    f"Importación terminada: {result['created']} creados, "
                    f"{result['updated']} actualizados, {result['omitted']} omitidos y "
                    f"{len(result['errors'])} errores.",
                    "success" if not result["errors"] else "warning",
                )
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
    return render_template(
        "imports/form.html", kind=kind, definition=definition, result=result
    )


@nominas_bp.route("/administracion/importar/<kind>/plantilla.xlsx")
@admin_required
def import_template_download(kind):
    if kind not in IMPORT_DEFINITIONS:
        abort(404)
    definition = IMPORT_DEFINITIONS[kind]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Datos"
    headers = [column[0] for column in definition["columns"]]
    examples = [column[3] for column in definition["columns"]]
    worksheet.append(headers)
    worksheet.append(examples)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="24557A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, column in enumerate(definition["columns"], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(column[0]) + 2, 16), 30
        )
    instructions = workbook.create_sheet("Instrucciones")
    instructions.append(["ENCABEZADO", "OBLIGATORIO", "DESCRIPCIÓN", "EJEMPLO"])
    for header, required, description, example in definition["columns"]:
        instructions.append([header, "SÍ" if required else "NO", description, example])
    for cell in instructions[1]:
        cell.fill = PatternFill("solid", fgColor="24557A")
        cell.font = Font(color="FFFFFF", bold=True)
    instructions.freeze_panes = "A2"
    instructions.column_dimensions["A"].width = 34
    instructions.column_dimensions["B"].width = 14
    instructions.column_dimensions["C"].width = 55
    instructions.column_dimensions["D"].width = 28
    buffer = BytesIO()
    workbook.save(buffer)
    filename = f"plantilla_importar_{kind}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Empresas, obras y partidas
# ---------------------------------------------------------------------------


@nominas_bp.route("/empresas", methods=["GET", "POST"])
@admin_required
def companies_list():
    if request.method == "POST":
        name = request.form.get("nombre", "").strip()
        code = request.form.get("codigo", "").strip().upper()
        if not name or not code:
            flash("Nombre y código son obligatorios.", "danger")
        elif Company.query.filter(or_(Company.codigo == code, func.lower(Company.nombre) == name.lower())).first():
            flash("Ya existe una empresa con ese nombre o código.", "danger")
        else:
            company = Company(nombre=name, codigo=code)
            db.session.add(company)
            db.session.flush()
            audit("CREAR", "EMPRESA", company.id)
            db.session.commit()
            flash("Empresa agregada.", "success")
            return redirect(url_for("companies_list"))
    return render_template("companies/list.html", companies=Company.query.order_by(Company.nombre).all())


@nominas_bp.route("/empresas/<int:company_id>/estado", methods=["POST"])
@admin_required
def company_toggle(company_id):
    company = db.get_or_404(Company, company_id)
    company.activa = not company.activa
    audit("ACTIVAR" if company.activa else "DESACTIVAR", "EMPRESA", company.id)
    db.session.commit()
    flash("Estado de la empresa actualizado.", "success")
    return redirect(url_for("companies_list"))


@nominas_bp.route("/obras")
@login_required
def projects_list():
    projects = accessible_projects_query(
        include_inactive=current_user.tiene_permiso("obras_partidas", "editar")
    ).all()
    cards = [(project, project_consumption(project)) for project in projects]
    return render_template("projects/list.html", project_cards=cards)


@nominas_bp.route("/obras/nueva", methods=["GET", "POST"])
@admin_required
def project_new():
    return _project_form(Project())


@nominas_bp.route("/obras/<int:project_id>/editar", methods=["GET", "POST"])
@admin_required
def project_edit(project_id):
    return _project_form(db.get_or_404(Project, project_id))


def _project_form(project: Project):
    is_new = project.id is None
    if request.method == "POST":
        try:
            name = request.form.get("nombre", "").strip().upper()
            code = request.form.get("codigo", "").strip().upper()
            project_type = request.form.get("tipo", "obra").lower()
            if not name or not code:
                raise ValueError("Nombre y código son obligatorios.")
            if project_type not in {"obra", "oficina"}:
                raise ValueError("Tipo de centro de costo inválido.")
            duplicate = Project.query.filter(
                Project.id != (project.id or 0),
                or_(func.lower(Project.nombre) == name.lower(), func.lower(Project.codigo) == code.lower()),
            ).first()
            if duplicate:
                raise ValueError("Ya existe una obra/oficina con ese nombre o código.")
            project.nombre = name
            project.codigo = code
            project.tipo = project_type
            project.presupuesto_total = parse_money("presupuesto_total")
            project.presupuesto_mano_obra = parse_money("presupuesto_mano_obra")
            project.descripcion = request.form.get("descripcion", "").strip() or None
            delivery_address = " ".join(
                (request.form.get("direccion_entrega") or "").split()
            )
            if len(delivery_address) > 500:
                raise ValueError(
                    "La dirección de entrega no puede superar 500 caracteres."
                )
            project.direccion_entrega = delivery_address or None
            project.activa = request.form.get("activa") == "on"
            db.session.add(project)
            db.session.flush()
            if is_new:
                asignar_obra_a_compradores(project)
            audit("CREAR" if is_new else "EDITAR", "OBRA", project.id)
            db.session.commit()
            if project_type == "obra" and not delivery_address:
                flash(
                    "La obra quedó guardada sin dirección de entrega. "
                    "Compras mostrará un aviso hasta que se complete.",
                    "warning",
                )
            flash("Obra/oficina guardada. Ahora agrega sus partidas.", "success")
            return redirect(url_for("project_detail", project_id=project.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template("projects/form.html", project=project, is_new=is_new)


@nominas_bp.route("/obras/<int:project_id>")
@login_required
def project_detail(project_id):
    project = db.get_or_404(Project, project_id)
    require_project_access(project)
    top_items = sorted((i for i in project.budget_items if i.parent_id is None), key=lambda x: x.codigo)
    item_rows = [(item, item_consumption(item)) for item in top_items]
    payrolls = Payroll.query.filter_by(project_id=project.id).order_by(Payroll.semana_inicio.desc()).limit(12).all()
    return render_template(
        "projects/detail.html",
        project=project,
        consumption=project_consumption(project),
        item_rows=item_rows,
        payrolls=payrolls,
        resource_summary=weekly_resource_breakdown(
            current_week_start(), [project.id]
        ),
    )


@nominas_bp.route("/obras/<int:project_id>/partidas/nueva", methods=["GET", "POST"])
@admin_required
def budget_item_new(project_id):
    project = db.get_or_404(Project, project_id)
    return _budget_item_form(BudgetItem(project_id=project.id), project)


@nominas_bp.route("/partidas/<int:item_id>/editar", methods=["GET", "POST"])
@admin_required
def budget_item_edit(item_id):
    item = db.get_or_404(BudgetItem, item_id)
    return _budget_item_form(item, item.project)


@nominas_bp.route("/api/obras/<int:project_id>/partidas")
@login_required
def budget_items_api(project_id):
    project = db.get_or_404(Project, project_id)
    require_project_access(project)
    category = request.args.get("categoria", "").strip().upper()
    query = BudgetItem.query.filter_by(project_id=project.id, activa=True)
    if category:
        query = query.filter(BudgetItem.categoria == category)
    items = query.order_by(BudgetItem.codigo).all()
    return jsonify(
        [
            {
                "id": item.id,
                "label": item.etiqueta,
                "category": item.categoria,
                "parent_id": item.parent_id,
            }
            for item in items
        ]
    )


def _budget_item_form(item: BudgetItem, project: Project):
    is_new = item.id is None
    parents = [i for i in project.budget_items if i.id != item.id and i.parent_id is None]
    if request.method == "POST":
        try:
            code = request.form.get("codigo", "").strip().upper()
            name = request.form.get("nombre", "").strip().upper()
            category = request.form.get("categoria", "MANO_OBRA")
            if not code or not name:
                raise ValueError("Código y nombre de la partida son obligatorios.")
            if category not in {"MANO_OBRA", "SUBCONTRATO", "INDIRECTO", "ADICIONAL"}:
                raise ValueError("Categoría inválida.")
            duplicate = BudgetItem.query.filter(
                BudgetItem.project_id == project.id,
                BudgetItem.codigo == code,
                BudgetItem.id != (item.id or 0),
            ).first()
            if duplicate:
                raise ValueError("Ese código de partida ya existe en la obra.")
            parent_id = request.form.get("parent_id", "")
            parent = db.session.get(BudgetItem, int(parent_id)) if parent_id.isdigit() else None
            if parent and parent.project_id != project.id:
                raise ValueError("La partida padre no pertenece a esta obra.")
            item.codigo = code
            item.nombre = name
            item.categoria = category
            item.presupuesto = parse_money("presupuesto")
            item.cantidad_objetivo = decimal_value(
                request.form.get("cantidad_objetivo", "0")
            )
            item.unidad_medida = (
                request.form.get("unidad_medida", "").strip().upper() or None
            )
            if item.cantidad_objetivo < 0:
                raise ValueError("La cantidad objetivo no puede ser negativa.")
            if item.cantidad_objetivo > 0 and not item.unidad_medida:
                raise ValueError(
                    "Captura la unidad de medida cuando exista una cantidad objetivo."
                )
            item.parent = parent
            item.activa = request.form.get("activa") == "on"
            db.session.add(item)
            db.session.flush()
            audit("CREAR" if is_new else "EDITAR", "PARTIDA", item.id, project.codigo)
            db.session.commit()
            flash("Partida guardada.", "success")
            return redirect(url_for("project_detail", project_id=project.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "budget_items/form.html", item=item, project=project, parents=parents, is_new=is_new
    )


# ---------------------------------------------------------------------------
# Trabajadores y préstamos
# ---------------------------------------------------------------------------


@nominas_bp.route("/trabajadores")
@permission_required("nomina", "ver")
def employees_list():
    project_id = request.args.get("project_id", type=int)
    status = request.args.get("estado", "activos")
    project_ids = [p.id for p in accessible_projects_query().all()]
    query = Employee.query.options(joinedload(Employee.project)).filter(
        Employee.project_id.in_(project_ids or [-1])
    )
    if project_id:
        if project_id not in project_ids:
            abort(404)
        query = query.filter(Employee.project_id == project_id)
    if status == "activos":
        query = query.filter(Employee.activo.is_(True))
    elif status == "bajas":
        query = query.filter(Employee.activo.is_(False))
    search = request.args.get("q", "").strip()
    if search:
        query = query.filter(Employee.nombre_completo.ilike(f"%{search}%"))
    employees = query.order_by(Employee.nombre_completo).all()
    return render_template(
        "employees/list.html",
        employees=employees,
        projects=accessible_projects_query().all(),
        selected_project=project_id,
        status=status,
        search=search,
    )


@nominas_bp.route("/trabajadores/nuevo", methods=["GET", "POST"])
@permission_required("nomina", "crear")
def employee_new():
    return _employee_form(Employee(fecha_ingreso=date.today(), activo=True))


@nominas_bp.route("/trabajadores/<int:employee_id>/editar", methods=["GET", "POST"])
@permission_required("nomina", "editar")
def employee_edit(employee_id):
    employee = db.get_or_404(Employee, employee_id)
    if employee.project:
        require_project_access(employee.project)
    return _employee_form(employee)


def _employee_form(employee: Employee):
    is_new = employee.id is None
    projects = accessible_projects_query().all()
    project_ids = [p.id for p in projects]
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    selected_project_id = request.form.get("project_id", type=int) or employee.project_id
    if selected_project_id and selected_project_id not in project_ids:
        abort(404)

    if request.method == "POST":
        try:
            name = validate_full_name(request.form.get("nombre_completo", ""))
            project_id = request.form.get("project_id", type=int)
            if not project_id or project_id not in project_ids:
                raise ValueError("Selecciona una obra/oficina autorizada.")
            position = request.form.get("puesto", "").strip().upper()
            if not position:
                raise ValueError("El puesto es obligatorio.")
            employee.nombre_completo = name
            employee.fecha_ingreso = parse_date("fecha_ingreso")
            employee.fecha_baja = parse_date("fecha_baja", required=False)
            employee.activo = request.form.get("activo") == "on"
            if employee.fecha_baja:
                employee.activo = False
            employee.puesto = position
            employee.cuadrilla = request.form.get("cuadrilla", "").strip().upper() or None
            employee.supervisor = request.form.get("supervisor", "").strip().upper() or None
            employee.empresa_operativa = (
                request.form.get("empresa_operativa", "").strip().upper() or None
            )
            employee.project_id = project_id
            # La asignación de partida es semanal y se captura en PayrollLine.
            employee.budget_item_id = None
            employee.salario_semanal = parse_money("salario_semanal", required=True)
            if employee.salario_semanal <= 0:
                raise ValueError("El salario semanal debe ser mayor que cero.")
            employee.descuento_infonavit = parse_money("descuento_infonavit")
            # Solo el administrador determina quién recibe la prestación IMSS y
            # si su costo se calcula como cuota fija o porcentaje del devengado.
            if current_user.tiene_permiso("seguridad", "editar"):
                employee.registrado_imss = request.form.get("registrado_imss") == "on"
                if puede_ver_nss_completo(current_user):
                    employee.nss = request.form.get("nss", "").strip() or None
                elif is_new:
                    employee.nss = None
                employee.empresa_imss_id = request.form.get("empresa_imss_id", type=int)
                employee.imss_tipo = request.form.get("imss_tipo", "FIJO")
                if employee.imss_tipo not in {"FIJO", "PORCENTAJE"}:
                    raise ValueError("El tipo de costo IMSS no es válido.")
                employee.descuento_imss = parse_money("descuento_imss")
                if employee.imss_tipo == "PORCENTAJE" and employee.descuento_imss > 100:
                    raise ValueError("El porcentaje de IMSS no puede ser mayor a 100%.")
                if employee.registrado_imss and (
                    not employee.nss or not employee.empresa_imss_id
                ):
                    raise ValueError(
                        "Para un trabajador con IMSS captura NSS y empresa de registro."
                    )
            elif is_new:
                employee.registrado_imss = False
                employee.nss = None
                employee.empresa_imss_id = None
                employee.imss_tipo = "FIJO"
                employee.descuento_imss = Decimal("0")
            employee.transferencia_predeterminada = parse_money("transferencia_predeterminada")
            employee.empresa_transferencia_id = request.form.get("empresa_transferencia_id", type=int)
            employee.empresa_efectivo_id = request.form.get("empresa_efectivo_id", type=int)
            if employee.transferencia_predeterminada > 0 and not employee.empresa_transferencia_id:
                raise ValueError("Selecciona la empresa de la transferencia predeterminada.")
            employee.notas = request.form.get("notas", "").strip() or None
            db.session.add(employee)
            db.session.flush()
            audit("CREAR" if is_new else "EDITAR", "TRABAJADOR", employee.id, name)
            db.session.commit()
            flash("Trabajador guardado correctamente.", "success")
            return_to = request.form.get("return_to", "")
            if return_to.startswith("/") and not return_to.startswith("//"):
                return redirect(return_to)
            return redirect(url_for("employee_detail", employee_id=employee.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            selected_project_id = request.form.get("project_id", type=int)
    return render_template(
        "employees/form.html",
        employee=employee,
        is_new=is_new,
        projects=projects,
        companies=companies,
        selected_project_id=selected_project_id,
        return_to=request.args.get("return_to", ""),
    )


@nominas_bp.route("/trabajadores/<int:employee_id>")
@permission_required("nomina", "ver")
def employee_detail(employee_id):
    employee = db.get_or_404(Employee, employee_id)
    if employee.project:
        require_project_access(employee.project)
    loans = Loan.query.options(joinedload(Loan.payments)).filter_by(employee_id=employee.id).order_by(
        Loan.fecha_prestamo.desc()
    ).all()
    lines = (
        PayrollLine.query.join(Payroll)
        .options(
            joinedload(PayrollLine.payroll).joinedload(Payroll.project),
            joinedload(PayrollLine.budget_item),
            joinedload(PayrollLine.partida),
            joinedload(PayrollLine.subpartida),
        )
        .filter(PayrollLine.employee_id == employee.id)
        .order_by(Payroll.semana_inicio.desc())
        .limit(20)
        .all()
    )
    return render_template("employees/detail.html", employee=employee, loans=loans, lines=lines)


@nominas_bp.route("/trabajadores/<int:employee_id>/baja", methods=["POST"])
@permission_required("nomina", "editar")
def employee_deactivate(employee_id):
    employee = db.get_or_404(Employee, employee_id)
    if employee.project:
        require_project_access(employee.project)
    employee.activo = False
    employee.fecha_baja = parse_date("fecha_baja", required=False) or date.today()
    audit("BAJA", "TRABAJADOR", employee.id, employee.nombre_completo)
    db.session.commit()
    flash("Trabajador dado de baja. Su historial se conserva y ya no se precargará.", "success")
    return redirect(url_for("employees_list", project_id=employee.project_id))


@nominas_bp.route("/prestamos")
@permission_required("nomina", "ver")
def loans_list():
    project_ids = [p.id for p in accessible_projects_query().all()]
    status = request.args.get("estado", "pendiente").lower()
    query = Loan.query.join(Employee).options(joinedload(Loan.employee), joinedload(Loan.payments)).filter(
        Employee.project_id.in_(project_ids or [-1])
    )
    if status in LOAN_STATES:
        query = query.filter(Loan.estado == status)
    loans = query.order_by(Loan.fecha_prestamo.desc(), Loan.id.desc()).all()
    return render_template("loans/list.html", loans=loans, status=status)


@nominas_bp.route("/prestamos/nuevo", methods=["GET", "POST"])
@permission_required("nomina", "crear")
def loan_new():
    project_ids = [p.id for p in accessible_projects_query().all()]
    employees = Employee.query.filter(
        Employee.activo.is_(True), Employee.project_id.in_(project_ids or [-1])
    ).order_by(Employee.nombre_completo).all()
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    payment_methods = loan_delivery_methods()
    # El catálogo compartido debe sobrevivir aun cuando la solicitud posterior
    # falle una regla de antigüedad o monto y la transacción se revierta.
    db.session.commit()
    preselected_employee = request.args.get("employee_id", type=int)
    if request.method == "POST":
        try:
            employee = db.session.get(Employee, request.form.get("employee_id", type=int))
            if not employee or employee.project_id not in project_ids:
                raise ValueError("Trabajador inválido o fuera de tus obras asignadas.")
            method = selected_loan_delivery_method()
            company = db.session.get(
                Company, request.form.get("company_id", type=int)
            )
            if not company or not company.activa:
                raise ValueError(
                    "Selecciona una empresa activa para entregar el préstamo."
                )
            loan_date = parse_date("fecha_prestamo")
            amount = parse_money("monto", required=True)
            weekly_withholding = parse_money("retencion_semanal", required=True)
            total_to_pay = validate_new_loan(employee, loan_date, amount)
            loan = Loan(
                employee_id=employee.id,
                project_id=employee.project_id,
                fecha_prestamo=loan_date,
                monto=amount,
                tasa_interes=5.0,
                total_pagar=total_to_pay,
                retencion_semanal=weekly_withholding,
                metodo_entrega=method.nombre.strip().upper(),
                payment_method_id=method.id,
                company_id=company.id,
                concepto=request.form.get("concepto", "").strip() or None,
                estado="pendiente",
                solicitante_id=current_user.id,
                created_by_id=current_user.id,
            )
            if loan.monto <= 0 or loan.retencion_semanal <= 0:
                raise ValueError("Monto y retención semanal deben ser mayores que cero.")
            if money(loan.retencion_semanal) > money(loan.total_pagar):
                raise ValueError("La retención semanal no puede exceder el total a pagar.")
            db.session.add(loan)
            db.session.flush()
            notify_loan_admins(loan)
            audit(
                "SOLICITAR",
                "PRESTAMO",
                loan.id,
                f"{employee.nombre_completo}: capital {loan.monto}; total {loan.total_pagar}",
            )
            db.session.commit()
            flash(
                "Solicitud enviada a los administradores. No se aplicará ninguna "
                "retención hasta que sea aprobada.",
                "success",
            )
            return redirect(url_for("employee_detail", employee_id=employee.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "loans/form.html",
        employees=employees,
        companies=companies,
        payment_methods=payment_methods,
        preselected_employee=preselected_employee,
    )


def require_general_admin() -> None:
    if not current_user.tiene_permiso("prestamos", "aprobar"):
        abort(403)


@nominas_bp.route("/prestamos/<int:loan_id>/aprobar", methods=["POST"])
@permission_required("nomina", "editar")
def loan_approve(loan_id):
    require_general_admin()
    loan = db.get_or_404(Loan, loan_id)
    if loan.employee.project:
        require_project_access(loan.employee.project)
    if loan.estado != "pendiente":
        flash("Solo puede aprobarse una solicitud pendiente.", "danger")
        return redirect(request.referrer or url_for("loans_list"))

    if loan.project_id is None:
        loan.project_id = loan.employee.project_id
    loan.estado = "activo"
    loan.aprobador_id = current_user.id
    loan.fecha_aprobacion = utc_now()
    loan.motivo_rechazo = None
    db.session.flush()
    updated_drafts = refresh_employee_draft_lines(loan.employee_id)
    audit(
        "APROBAR",
        "PRESTAMO",
        loan.id,
        f"Capital {loan.monto}; tasa {loan.tasa_interes}%; total {loan.total_pagar}",
    )
    db.session.commit()
    first_week = loan.fecha_prestamo + timedelta(days=7 - loan.fecha_prestamo.weekday())
    flash(
        f"Préstamo aprobado y activado. La primera retención elegible inicia el "
        f"{first_week.strftime('%d/%m/%Y')}. Se actualizaron {updated_drafts} "
        "nómina(s) en borrador.",
        "success",
    )
    return redirect(request.referrer or url_for("loans_list"))


@nominas_bp.route("/prestamos/<int:loan_id>/rechazar", methods=["POST"])
@permission_required("nomina", "editar")
def loan_reject(loan_id):
    require_general_admin()
    loan = db.get_or_404(Loan, loan_id)
    if loan.employee.project:
        require_project_access(loan.employee.project)
    if loan.estado != "pendiente":
        flash("Solo puede rechazarse una solicitud pendiente.", "danger")
        return redirect(request.referrer or url_for("loans_list"))
    reason = request.form.get("motivo", "").strip()
    if not reason:
        flash("Captura el motivo del rechazo.", "danger")
        return redirect(request.referrer or url_for("loans_list"))

    loan.estado = "rechazado"
    loan.aprobador_id = current_user.id
    loan.fecha_aprobacion = utc_now()
    loan.motivo_rechazo = reason[:500]
    audit("RECHAZAR", "PRESTAMO", loan.id, loan.motivo_rechazo)
    db.session.commit()
    flash("Solicitud de préstamo rechazada.", "warning")
    return redirect(request.referrer or url_for("loans_list"))


@nominas_bp.route("/prestamos/<int:loan_id>/cancelar", methods=["POST"])
@permission_required("nomina", "editar")
def loan_cancel(loan_id):
    """Alias seguro para marcadores antiguos; no cancela deuda ya activa."""

    return loan_reject(loan_id)


# ---------------------------------------------------------------------------
# Nóminas semanales
# ---------------------------------------------------------------------------


def form_money_value(key: str, default="0") -> Decimal:
    raw = request.form.get(key, str(default)).strip().replace(",", "")
    try:
        value = Decimal(raw or str(default)).quantize(MONEY_STEP, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Importe inválido en {key}.") from exc
    if value < 0:
        raise ValueError(f"El importe {key} no puede ser negativo.")
    return value


def employee_eligible(employee: Employee, week_start: date) -> bool:
    if not employee.activo:
        return False
    if employee.fecha_ingreso and employee.fecha_ingreso > week_start + timedelta(days=4):
        return False
    if employee.fecha_baja and employee.fecha_baja < week_start:
        return False
    return True


def worker_already_in_week(employee_id: int, week_start: date, exclude_payroll_id=None) -> bool:
    query = PayrollLine.query.join(Payroll).filter(
        PayrollLine.employee_id == employee_id, Payroll.semana_inicio == week_start
    )
    if exclude_payroll_id:
        query = query.filter(Payroll.id != exclude_payroll_id)
    return query.first() is not None


def build_line_from_employee(
    payroll: Payroll, employee: Employee, previous_line: PayrollLine | None = None
) -> PayrollLine | None:
    if not employee_eligible(employee, payroll.semana_inicio):
        return None
    if worker_already_in_week(employee.id, payroll.semana_inicio, payroll.id):
        return None

    line = PayrollLine(
        employee=employee,
        budget_item_id=None,
        partida_id=None,
        subpartida_id=None,
        nombre_trabajador=validate_full_name(employee.nombre_completo),
        puesto=(previous_line.puesto if previous_line else employee.puesto) or "SIN PUESTO",
        cuadrilla=previous_line.cuadrilla if previous_line else employee.cuadrilla,
        supervisor=previous_line.supervisor if previous_line else employee.supervisor,
        empresa_operativa=(
            previous_line.empresa_operativa if previous_line else employee.empresa_operativa
        ),
        salario_semanal=money(previous_line.salario_semanal if previous_line else employee.salario_semanal),
        lunes=True,
        martes=True,
        miercoles=True,
        jueves=True,
        viernes=True,
        pago_extra=0,
        descuento_infonavit=money(employee.descuento_infonavit),
        descuento_imss=0,
        otro_descuento=0,
        vales_gasolina=0,
        empresa_transferencia_id=(
            previous_line.empresa_transferencia_id if previous_line else employee.empresa_transferencia_id
        ),
        empresa_efectivo_id=(previous_line.empresa_efectivo_id if previous_line else employee.empresa_efectivo_id),
        pago_transferencia=money(
            previous_line.pago_transferencia if previous_line else employee.transferencia_predeterminada
        ),
    )
    if previous_line is not None:
        previous_partida, previous_subpartida = infer_payroll_line_allocation(
            previous_line
        )
        try:
            assign_payroll_line_budget(
                line,
                payroll,
                previous_partida.id if previous_partida else None,
                previous_subpartida.id if previous_subpartida else None,
            )
        except ValueError:
            # Una sugerencia obsoleta nunca bloquea la creación del borrador.
            # El usuario deberá elegir una asignación válida antes de guardar.
            line.partida = None
            line.subpartida = None
            line.partida_id = None
            line.subpartida_id = None
            line.budget_item_id = None
    payroll.lines.append(line)
    recalculate_line(line)
    return line


def create_payroll_with_preload(project: Project, week_start: date, creator_id: int) -> Payroll:
    payroll = Payroll(
        project_id=project.id,
        semana_inicio=week_start,
        semana_fin=week_start + timedelta(days=4),
        created_by_id=creator_id,
        estado="borrador",
    )
    db.session.add(payroll)
    db.session.flush()

    previous = (
        Payroll.query.filter(
            Payroll.project_id == project.id,
            Payroll.semana_inicio == week_start - timedelta(days=7),
        )
        .first()
    )
    added: set[int] = set()
    if previous:
        for previous_line in previous.lines:
            employee = previous_line.employee
            if employee and employee.id not in added:
                line = build_line_from_employee(payroll, employee, previous_line)
                if line:
                    added.add(employee.id)

    # También incluye altas nuevas que no existían en la semana anterior.
    active_employees = Employee.query.filter_by(project_id=project.id, activo=True).order_by(
        Employee.nombre_completo
    ).all()
    for employee in active_employees:
        if employee.id not in added:
            line = build_line_from_employee(payroll, employee)
            if line:
                added.add(employee.id)
    return payroll


@nominas_bp.route("/nominas")
@permission_required("nomina", "ver")
def payrolls_list():
    project_ids = [p.id for p in accessible_projects_query().all()]
    query = Payroll.query.options(joinedload(Payroll.project), joinedload(Payroll.lines)).filter(
        Payroll.project_id.in_(project_ids or [-1])
    )
    project_id = request.args.get("project_id", type=int)
    status = request.args.get("estado", "")
    week = request.args.get("semana", "")
    if project_id:
        if project_id not in project_ids:
            abort(404)
        query = query.filter(Payroll.project_id == project_id)
    status_aliases = {"BORRADOR": "borrador", "CERRADA": "aprobada"}
    status = status_aliases.get(status, status.lower())
    if status in PAYROLL_STATES:
        query = query.filter(Payroll.estado == status)
    if week:
        try:
            query = query.filter(Payroll.semana_inicio == datetime.strptime(week, "%Y-%m-%d").date())
        except ValueError:
            flash("La semana del filtro no es válida.", "warning")
    payrolls = query.order_by(Payroll.semana_inicio.desc(), Payroll.project_id).all()
    return render_template(
        "payrolls/list.html",
        payrolls=payrolls,
        projects=accessible_projects_query().all(),
        selected_project=project_id,
        selected_status=status,
        selected_week=week,
    )


@nominas_bp.route("/nominas/nueva", methods=["GET", "POST"])
@permission_required("nomina", "crear")
def payroll_new():
    projects = accessible_projects_query().all()
    if request.method == "POST":
        try:
            project = db.session.get(Project, request.form.get("project_id", type=int))
            if not project or not can_access_project(project):
                raise ValueError("Selecciona una obra/oficina autorizada.")
            week_start = parse_date("semana_inicio")
            if week_start.weekday() != 0:
                raise ValueError("La fecha de inicio debe ser lunes.")
            existing = Payroll.query.filter_by(project_id=project.id, semana_inicio=week_start).first()
            if existing:
                flash("Esa nómina ya existe; te llevé a su captura.", "info")
                return redirect(url_for("payroll_detail", payroll_id=existing.id))
            if not BudgetItem.query.filter_by(
                project_id=project.id,
                activa=True,
                parent_id=None,
            ).first():
                raise ValueError("La obra debe tener al menos una partida activa antes de crear su nómina.")
            payroll = create_payroll_with_preload(project, week_start, current_user.id)
            audit("CREAR", "NOMINA", payroll.id, f"{project.codigo} {week_start}")
            db.session.commit()
            flash(
                f"Nómina creada con {len(payroll.lines)} trabajadores precargados. Revisa asistencia y pagos.",
                "success",
            )
            return redirect(url_for("payroll_detail", payroll_id=payroll.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "payrolls/new.html", projects=projects, default_week=current_week_start().isoformat()
    )


@nominas_bp.route("/nominas/<int:payroll_id>")
@permission_required("nomina", "ver")
def payroll_detail(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    # Seguridad adicional: al abrir un borrador siempre se muestran el préstamo,
    # el neto y la distribución de pago calculados con la información vigente.
    refresh_draft_payroll(payroll)
    loan_limits = {
        line.id: scheduled_loan_deduction(line.employee_id, payroll.semana_inicio)
        for line in payroll.lines
    }
    items = BudgetItem.query.filter_by(
        project_id=payroll.project_id,
        activa=True,
        parent_id=None,
    ).order_by(BudgetItem.codigo).all()
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    consumption = project_consumption(payroll.project)
    item_rows = [
        (item, item_consumption(item))
        for item in sorted(
            (i for i in payroll.project.budget_items if i.parent_id is None), key=lambda i: i.codigo
        )
    ]
    resource_summary = weekly_resource_breakdown(
        payroll.semana_inicio,
        None
        if current_user.is_admin
        else [project.id for project in accessible_projects_query().all()],
    )
    return render_template(
        "payrolls/detail.html",
        payroll=payroll,
        items=items,
        companies=companies,
        consumption=consumption,
        item_rows=item_rows,
        loan_limits=loan_limits,
        resource_summary=resource_summary,
    )


@nominas_bp.route("/nominas/<int:payroll_id>/guardar", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_save(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    if payroll.estado != "borrador":
        flash("La nómina ya no está en borrador y no puede editarse.", "warning")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    try:
        for line in payroll.lines:
            prefix = f"line_{line.id}_"
            partida_id = request.form.get(prefix + "partida_id", type=int)
            subpartida_id = request.form.get(prefix + "subpartida_id", type=int)
            # Compatibilidad con formularios del paquete anterior.
            if partida_id is None:
                legacy_item_id = request.form.get(
                    prefix + "budget_item_id", type=int
                )
                legacy_item = (
                    db.session.get(BudgetItem, legacy_item_id)
                    if legacy_item_id
                    else None
                )
                if legacy_item is not None:
                    if legacy_item.parent_id is None:
                        partida_id = legacy_item.id
                    else:
                        partida_id = legacy_item.parent_id
                        subpartida_id = legacy_item.id
            assign_payroll_line_budget(
                line,
                payroll,
                partida_id,
                subpartida_id,
            )
            line.puesto = request.form.get(prefix + "puesto", "").strip().upper()
            line.cuadrilla = request.form.get(prefix + "cuadrilla", "").strip().upper() or None
            line.supervisor = request.form.get(prefix + "supervisor", "").strip().upper() or None
            line.empresa_operativa = (
                request.form.get(prefix + "empresa_operativa", "").strip().upper() or None
            )
            if not line.puesto:
                raise ValueError(f"El puesto de {line.nombre_trabajador} es obligatorio.")
            line.salario_semanal = form_money_value(prefix + "salario_semanal")
            if line.salario_semanal <= 0:
                raise ValueError(f"El salario de {line.nombre_trabajador} debe ser mayor que cero.")
            for field in WEEKDAY_FIELDS:
                setattr(line, field, request.form.get(prefix + field) == "on")
            line.pago_extra = form_money_value(prefix + "pago_extra")
            line.descuento_infonavit = form_money_value(prefix + "descuento_infonavit")
            line.otro_descuento = form_money_value(prefix + "otro_descuento")
            line.concepto_otro_descuento = (
                request.form.get(prefix + "concepto_otro_descuento", "").strip() or None
            )
            line.vales_gasolina = form_money_value(prefix + "vales_gasolina")
            transfer = form_money_value(prefix + "pago_transferencia")
            line.empresa_transferencia_id = request.form.get(prefix + "empresa_transferencia_id", type=int)
            line.empresa_efectivo_id = request.form.get(prefix + "empresa_efectivo_id", type=int)
            line.notas = request.form.get(prefix + "notas", "").strip() or None
            recalculate_line(line, transfer_requested=transfer, strict_transfer=True)
            validate_line_for_close(line)

        payroll.notas = request.form.get("notas", "").strip() or None
        audit("GUARDAR", "NOMINA", payroll.id, f"{len(payroll.lines)} líneas")
        db.session.commit()
        flash("Nómina guardada y cálculos actualizados.", "success")
        if (
            request.form.get("next_action") == "weekly_report"
            and current_user.tiene_permiso("reportes_nomina", "ver")
        ):
            return redirect(
                url_for("weekly_closing_report", semana=payroll.semana_inicio.isoformat())
            )
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


@nominas_bp.route("/nominas/<int:payroll_id>/agregar", methods=["GET", "POST"])
@permission_required("nomina", "editar")
def payroll_add_employee(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    if payroll.estado != "borrador":
        abort(409)
    existing_ids = {line.employee_id for line in payroll.lines}
    accessible_ids = [project.id for project in accessible_projects_query().all()]
    employees = Employee.query.filter(
        Employee.activo.is_(True),
        Employee.id.notin_(existing_ids or [-1]),
        Employee.project_id.in_(accessible_ids or [-1]),
    ).order_by(Employee.nombre_completo).all()
    # Se permiten movimientos temporales entre obras autorizadas, sin mostrar personal
    # de centros de costo a los que el capturista no tiene acceso.
    employees = [e for e in employees if not worker_already_in_week(e.id, payroll.semana_inicio)]
    if request.method == "POST":
        try:
            employee = db.session.get(Employee, request.form.get("employee_id", type=int))
            if not employee or employee.id not in {e.id for e in employees}:
                raise ValueError("Selecciona un trabajador disponible.")
            line = build_line_from_employee(payroll, employee)
            if not line:
                raise ValueError("No se pudo agregar al trabajador a esta semana.")
            # La asignación queda vacía si no existe sugerencia de la semana
            # anterior; debe elegirse antes de guardar.
            db.session.flush()
            audit("AGREGAR_TRABAJADOR", "NOMINA", payroll.id, employee.nombre_completo)
            db.session.commit()
            flash("Trabajador agregado a la nómina.", "success")
            return redirect(url_for("payroll_detail", payroll_id=payroll.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template("payrolls/add_employee.html", payroll=payroll, employees=employees)


@nominas_bp.route("/nominas/<int:payroll_id>/lineas/<int:line_id>/eliminar", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_delete_line(payroll_id, line_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    if payroll.estado != "borrador":
        abort(409)
    line = db.get_or_404(PayrollLine, line_id)
    if line.payroll_id != payroll.id:
        abort(404)
    name = line.nombre_trabajador
    db.session.delete(line)
    audit("QUITAR_TRABAJADOR", "NOMINA", payroll.id, name)
    db.session.commit()
    flash("Trabajador quitado de esta semana. No se eliminó su historial ni su alta.", "success")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


def require_financial_admin() -> None:
    """Comprueba la acción financiera granular de la ruta actual."""

    module, action = nominas_permission_context()
    if not current_user.tiene_permiso(module, action):
        abort(403)


def validate_payroll_transition(payroll: Payroll):
    if not payroll.lines:
        raise ValueError("No se puede cerrar una nómina sin trabajadores.")
    plans: list[tuple[PayrollLine, list[tuple[Loan, Decimal]]]] = []
    for line in payroll.lines:
        validate_payroll_line_budget(line)
        plan = recalculate_line(line)
        validate_line_for_close(line)
        plans.append((line, plan))
    return plans


def submit_payroll(payroll: Payroll):
    if payroll.estado != "borrador":
        flash("Solo una nómina en borrador puede enviarse a aprobación.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    try:
        validate_payroll_transition(payroll)
        payroll.estado = "enviada"
        audit("ENVIAR", "NOMINA", payroll.id, f"Neto propuesto: {payroll.total_neto}")
        db.session.commit()
        flash("Nómina enviada. Quedó bloqueada hasta la revisión administrativa.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


def approve_payroll(payroll: Payroll):
    require_financial_admin()
    if payroll.estado not in {"borrador", "enviada"}:
        flash("Solo una nómina en borrador o enviada puede aprobarse.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    try:
        plans = validate_payroll_transition(payroll)
        touched_loans: set[int] = set()
        for line, plan in plans:
            for loan, amount in plan:
                db.session.add(
                    LoanPayment(loan_id=loan.id, payroll_line=line, monto=amount)
                )
                touched_loans.add(loan.id)
        payroll.estado = "aprobada"
        payroll.closed_by_id = current_user.id
        payroll.closed_at = utc_now()
        db.session.flush()
        for loan_id in touched_loans:
            loan = db.session.get(Loan, loan_id)
            if loan and loan_remaining(loan) <= 0:
                loan.estado = "liquidado"

        next_start = payroll.semana_inicio + timedelta(days=7)
        next_payroll = Payroll.query.filter_by(
            project_id=payroll.project_id, semana_inicio=next_start
        ).first()
        if not next_payroll:
            next_payroll = create_payroll_with_preload(
                payroll.project, next_start, current_user.id
            )
            audit("PRECARGAR", "NOMINA", next_payroll.id, f"Desde nómina {payroll.id}")
        audit("APROBAR", "NOMINA", payroll.id, f"Neto: {payroll.total_neto}")
        db.session.commit()
        consumption = project_consumption(payroll.project)
        flash(
            f"Nómina aprobada. Costo de mano de obra con IMSS: "
            f"{currency_filter(payroll.total_costo_mano_obra)}. "
            f"Disponible de mano de obra: {currency_filter(consumption['restante_mano_obra'])}. "
            "La semana siguiente quedó precargada.",
            "success",
        )
        if consumption["alerta_consumido"] or consumption["alerta_comprometido"]:
            flash(
                f"Alerta presupuestal: consumido {consumption['porcentaje_consumido']}% y "
                f"comprometido {consumption['porcentaje_comprometido']}% del presupuesto total.",
                "warning",
            )
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


@nominas_bp.route("/nominas/<int:payroll_id>/enviar", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_submit(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    return submit_payroll(payroll)


@nominas_bp.route("/nominas/<int:payroll_id>/aprobar", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_approve(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    return approve_payroll(payroll)


@nominas_bp.route("/nominas/<int:payroll_id>/cerrar", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_close(payroll_id):
    """Compatibilidad: el capturista envía y Administración aprueba."""

    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    if current_user.tiene_permiso("nominas_semanales", "aprobar"):
        return approve_payroll(payroll)
    return submit_payroll(payroll)


@nominas_bp.route("/nominas/<int:payroll_id>/reabrir", methods=["POST"])
@admin_required
def payroll_reopen(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    require_financial_admin()
    if payroll.estado in {"pagada", "conciliada"}:
        flash(
            "Una nómina pagada no puede reabrirse. La corrección requiere un ajuste versionado.",
            "danger",
        )
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    if payroll.estado != "aprobada":
        flash("La reapertura solo se permite de aprobada a borrador.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    reason = request.form.get("motivo", "").strip()
    if not reason:
        flash("Indica el motivo de la reapertura.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    line_ids = [line.id for line in payroll.lines]
    payments = LoanPayment.query.filter(LoanPayment.payroll_line_id.in_(line_ids or [-1])).all()
    loan_ids = {payment.loan_id for payment in payments}
    for payment in payments:
        db.session.delete(payment)
    db.session.flush()
    for loan_id in loan_ids:
        loan = db.session.get(Loan, loan_id)
        if loan and loan.estado != "rechazado" and loan_remaining(loan) > 0:
            loan.estado = "activo"
    payroll.estado = "borrador"
    payroll.closed_by_id = None
    payroll.closed_at = None
    for line in payroll.lines:
        recalculate_line(line)
    audit(
        "REABRIR",
        "NOMINA",
        payroll.id,
        f"Motivo: {reason[:500]}. Retenciones de préstamos revertidas.",
    )
    db.session.commit()
    flash("Nómina reabierta. Las retenciones de préstamos quedaron revertidas hasta el nuevo cierre.", "warning")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


@nominas_bp.route("/nominas/<int:payroll_id>/devolver", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_return_to_draft(payroll_id):
    """Devuelve una captura enviada sin tratarla como reapertura financiera."""

    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    require_financial_admin()
    if payroll.estado != "enviada":
        flash("Solo una nómina enviada puede devolverse a captura.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    reason = request.form.get("motivo", "").strip()
    if not reason:
        flash("Indica el motivo de la devolución.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    payroll.estado = "borrador"
    audit("DEVOLVER", "NOMINA", payroll.id, f"Motivo: {reason[:500]}")
    db.session.commit()
    flash("Nómina devuelta a borrador para corrección.", "warning")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


@nominas_bp.route("/nominas/<int:payroll_id>/pagar", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_pay(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    require_financial_admin()
    if payroll.estado != "aprobada":
        flash("Solo una nómina aprobada puede marcarse como pagada.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    payroll.estado = "pagada"
    payroll.paid_by_id = current_user.id
    payroll.paid_at = utc_now()
    audit("PAGAR", "NOMINA", payroll.id, f"Neto ejecutado: {payroll.total_neto}")
    db.session.commit()
    flash("Pago de nómina registrado. Esta nómina ya no puede reabrirse.", "success")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


@nominas_bp.route("/nominas/<int:payroll_id>/conciliar", methods=["POST"])
@permission_required("nomina", "editar")
def payroll_reconcile(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    require_financial_admin()
    if payroll.estado != "pagada":
        flash("Solo una nómina pagada puede conciliarse.", "danger")
        return redirect(url_for("payroll_detail", payroll_id=payroll.id))
    payroll.estado = "conciliada"
    payroll.reconciled_by_id = current_user.id
    payroll.reconciled_at = utc_now()
    audit("CONCILIAR", "NOMINA", payroll.id, f"Neto: {payroll.total_neto}")
    db.session.commit()
    flash("Nómina conciliada correctamente.", "success")
    return redirect(url_for("payroll_detail", payroll_id=payroll.id))


@nominas_bp.route("/nominas/<int:payroll_id>/eliminar", methods=["POST"])
@admin_required
def payroll_delete(payroll_id):
    payroll = db.get_or_404(Payroll, payroll_id)
    require_project_access(payroll.project)
    if payroll.estado != "borrador":
        flash("Solo se puede eliminar una nómina en borrador.", "danger")
    else:
        detail = f"{payroll.project.codigo} {payroll.semana_inicio}"
        db.session.delete(payroll)
        audit("ELIMINAR", "NOMINA", payroll.id, detail)
        db.session.commit()
        flash("Borrador de nómina eliminado.", "success")
    return redirect(url_for("payrolls_list"))


# ---------------------------------------------------------------------------
# Pagos adicionales y gastos de oficina
# ---------------------------------------------------------------------------


@nominas_bp.route("/pagos-adicionales")
@login_required
def additional_payments_list():
    project_ids = [p.id for p in accessible_projects_query().all()]
    query = AdditionalPayment.query.options(
        joinedload(AdditionalPayment.project),
        joinedload(AdditionalPayment.budget_item),
        joinedload(AdditionalPayment.explosion_item).joinedload(
            BudgetExplosionItem.supply_item
        ),
    ).filter(AdditionalPayment.project_id.in_(project_ids or [-1]))
    project_id = request.args.get("project_id", type=int)
    if project_id:
        if project_id not in project_ids:
            abort(404)
        query = query.filter(AdditionalPayment.project_id == project_id)
    payments = query.order_by(AdditionalPayment.fecha.desc(), AdditionalPayment.id.desc()).all()
    eligible_orders = PurchaseOrder.query.filter(
        PurchaseOrder.project_id.in_(project_ids or [-1]),
        PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES - {"CERRADA"}),
    ).all()
    has_payable = any(order.saldo_pagable > 0 for order in eligible_orders)
    return render_template(
        "additional_payments/list.html",
        payments=payments,
        projects=accessible_projects_query().all(),
        selected_project=project_id,
        has_payable=has_payable,
    )


@nominas_bp.route("/pagos-adicionales/nuevo", methods=["GET", "POST"])
@login_required
def additional_payment_new():
    """Conserva la URL anterior, pero la captura real vive en Compras."""

    flash(
        "Los pagos de obra ahora se registran desde Compras para exigir Obra → Partida → Insumo.",
        "info",
    )
    return redirect(flask_url_for("compras.supplier_payment_new"))


@nominas_bp.route("/pagos-adicionales/<int:payment_id>/eliminar", methods=["POST"])
@admin_required
def additional_payment_delete(payment_id):
    payment = db.get_or_404(AdditionalPayment, payment_id)
    if payment.purchase_order_id:
        flash(
            "Los pagos vinculados a una OC se conservan en el estado de cuenta; no pueden eliminarse desde Nóminas.",
            "danger",
        )
        return redirect(url_for("additional_payments_list", project_id=payment.project_id))
    project_id = payment.project_id
    audit("ELIMINAR", "PAGO_ADICIONAL", payment.id, str(payment.monto_capturado))
    db.session.delete(payment)
    db.session.commit()
    flash("Pago adicional eliminado del control.", "success")
    return redirect(url_for("additional_payments_list", project_id=project_id))


@nominas_bp.route("/gastos-oficina")
@login_required
def office_expenses_list():
    office_projects = [p for p in accessible_projects_query().all() if p.tipo == "oficina"]
    project_ids = [p.id for p in office_projects]
    expenses = (
        OfficeExpense.query.options(joinedload(OfficeExpense.project), joinedload(OfficeExpense.budget_item))
        .filter(OfficeExpense.project_id.in_(project_ids or [-1]))
        .order_by(OfficeExpense.fecha.desc(), OfficeExpense.id.desc())
        .all()
    )
    return render_template("office_expenses/list.html", expenses=expenses, projects=office_projects)


@nominas_bp.route("/gastos-oficina/nuevo", methods=["GET", "POST"])
@login_required
def office_expense_new():
    projects = [p for p in accessible_projects_query().all() if p.tipo == "oficina"]
    project_ids = [p.id for p in projects]
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    payment_methods = shared_payment_methods()
    selected_project_id = request.form.get("project_id", type=int) or request.args.get("project_id", type=int)
    items = (
        BudgetItem.query.filter_by(
            project_id=selected_project_id, categoria="INDIRECTO", activa=True
        ).order_by(BudgetItem.codigo).all()
        if selected_project_id in project_ids
        else []
    )
    if request.method == "POST":
        try:
            project = db.session.get(Project, request.form.get("project_id", type=int))
            if not project or project.id not in project_ids or project.tipo != "oficina":
                raise ValueError("Selecciona un centro de costo tipo oficina.")
            item = db.session.get(BudgetItem, request.form.get("budget_item_id", type=int))
            if not item or item.project_id != project.id or item.categoria != "INDIRECTO":
                raise ValueError("El gasto de oficina debe ir a una partida de indirectos.")
            amount = parse_money("monto_capturado", required=True)
            amount_type = request.form.get("tipo_monto", "SIN_IVA")
            if amount <= 0 or amount_type not in {"SIN_IVA", "CON_IVA"}:
                raise ValueError("Revisa el monto y el IVA.")
            method = selected_payment_method("metodo_pago")
            expense = OfficeExpense(
                fecha=parse_date("fecha"),
                project_id=project.id,
                budget_item_id=item.id,
                proveedor=request.form.get("proveedor", "").strip().upper(),
                concepto=request.form.get("concepto", "").strip().upper(),
                monto_capturado=amount,
                tipo_monto=amount_type,
                monto_sin_iva=amount_without_vat(amount, amount_type),
                metodo_pago=method.nombre,
                payment_method_id=method.id,
                company_id=request.form.get("company_id", type=int),
                notas=request.form.get("notas", "").strip() or None,
                created_by_id=current_user.id,
            )
            if not expense.proveedor or not expense.concepto or not expense.company_id:
                raise ValueError("Proveedor, concepto y empresa son obligatorios.")
            db.session.add(expense)
            db.session.flush()
            audit("CREAR", "GASTO_OFICINA", expense.id, str(amount))
            db.session.commit()
            flash("Gasto de oficina registrado en indirectos.", "success")
            return redirect(url_for("office_expenses_list"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "office_expenses/form.html",
        projects=projects,
        companies=companies,
        payment_methods=payment_methods,
        items=items,
        selected_project_id=selected_project_id,
    )


@nominas_bp.route("/gastos-oficina/<int:expense_id>/eliminar", methods=["POST"])
@admin_required
def office_expense_delete(expense_id):
    expense = db.get_or_404(OfficeExpense, expense_id)
    audit("ELIMINAR", "GASTO_OFICINA", expense.id, str(expense.monto_capturado))
    db.session.delete(expense)
    db.session.commit()
    flash("Gasto de oficina eliminado.", "success")
    return redirect(url_for("office_expenses_list"))


# ---------------------------------------------------------------------------
# Contratistas, subcontratos, avance y pagos
# ---------------------------------------------------------------------------


@nominas_bp.route("/contratistas", methods=["GET", "POST"])
@login_required
def contractors_list():
    if request.method == "POST":
        name = request.form.get("nombre", "").strip().upper()
        if not name:
            flash("El nombre del contratista es obligatorio.", "danger")
        elif Contractor.query.filter(func.lower(Contractor.nombre) == name.lower()).first():
            flash("Ese contratista ya existe.", "danger")
        else:
            contractor = Contractor(
                nombre=name,
                especialidad=request.form.get("especialidad", "").strip().upper() or None,
                telefono=request.form.get("telefono", "").strip() or None,
                email=request.form.get("email", "").strip().lower() or None,
            )
            db.session.add(contractor)
            db.session.flush()
            audit("CREAR", "CONTRATISTA", contractor.id, name)
            db.session.commit()
            flash("Contratista agregado.", "success")
            return redirect(url_for("contractors_list"))
    return render_template(
        "contractors/list.html", contractors=Contractor.query.order_by(Contractor.nombre).all()
    )


@nominas_bp.route("/subcontratos")
@login_required
def subcontracts_list():
    project_ids = [p.id for p in accessible_projects_query().all()]
    query = Subcontract.query.options(
        joinedload(Subcontract.project), joinedload(Subcontract.contractor), joinedload(Subcontract.payments)
    ).filter(Subcontract.project_id.in_(project_ids or [-1]))
    project_id = request.args.get("project_id", type=int)
    if project_id:
        if project_id not in project_ids:
            abort(404)
        query = query.filter(Subcontract.project_id == project_id)
    subcontracts = query.order_by(Subcontract.project_id, Subcontract.especialidad).all()
    return render_template(
        "subcontracts/list.html",
        subcontracts=subcontracts,
        projects=accessible_projects_query().all(),
        selected_project=project_id,
    )


@nominas_bp.route("/subcontratos/nuevo", methods=["GET", "POST"])
@admin_required
def subcontract_new():
    return _subcontract_form(Subcontract(avance_fisico=0, umbral_alerta=Decimal("0.15"), activo=True))


@nominas_bp.route("/subcontratos/<int:subcontract_id>/editar", methods=["GET", "POST"])
@admin_required
def subcontract_edit(subcontract_id):
    return _subcontract_form(db.get_or_404(Subcontract, subcontract_id))


def _subcontract_form(subcontract: Subcontract):
    is_new = subcontract.id is None
    projects = Project.query.filter_by(tipo="obra", activa=True).order_by(Project.nombre).all()
    project_ids = [p.id for p in projects]
    contractors = Contractor.query.filter_by(activo=True).order_by(Contractor.nombre).all()
    selected_project_id = request.form.get("project_id", type=int) or subcontract.project_id
    items = (
        BudgetItem.query.filter_by(project_id=selected_project_id, activa=True).order_by(BudgetItem.codigo).all()
        if selected_project_id in project_ids
        else []
    )
    if request.method == "POST":
        try:
            project = db.session.get(Project, request.form.get("project_id", type=int))
            item = db.session.get(BudgetItem, request.form.get("budget_item_id", type=int))
            contractor = db.session.get(Contractor, request.form.get("contractor_id", type=int))
            if not project or project.tipo != "obra" or not item or item.project_id != project.id:
                raise ValueError("Selecciona una obra y una partida válidas.")
            if not contractor:
                raise ValueError("Selecciona un contratista.")
            specialty = request.form.get("especialidad", "").strip().upper()
            if not specialty:
                raise ValueError("La especialidad/subcontrato es obligatoria.")
            progress = decimal_value(request.form.get("avance_fisico", "0")) / Decimal("100")
            threshold = decimal_value(request.form.get("umbral_alerta", "15")) / Decimal("100")
            if progress < 0 or progress > 1 or threshold < 0 or threshold > 1:
                raise ValueError("Avance y umbral deben estar entre 0% y 100%.")
            subcontract.project_id = project.id
            subcontract.budget_item_id = item.id
            subcontract.contractor_id = contractor.id
            subcontract.especialidad = specialty
            subcontract.presupuesto_sin_iva = parse_money("presupuesto_sin_iva", required=True)
            subcontract.avance_fisico = progress
            subcontract.umbral_alerta = threshold
            subcontract.observaciones = request.form.get("observaciones", "").strip() or None
            subcontract.activo = request.form.get("activo") == "on"
            db.session.add(subcontract)
            db.session.flush()
            audit("CREAR" if is_new else "EDITAR", "SUBCONTRATO", subcontract.id, specialty)
            db.session.commit()
            flash("Subcontrato guardado.", "success")
            return redirect(url_for("subcontract_detail", subcontract_id=subcontract.id))
        except (ValueError, IntegrityError) as exc:
            db.session.rollback()
            flash("Ese subcontrato ya existe o contiene datos inválidos: " + str(exc), "danger")
    return render_template(
        "subcontracts/form.html",
        subcontract=subcontract,
        is_new=is_new,
        projects=projects,
        items=items,
        contractors=contractors,
        selected_project_id=selected_project_id,
    )


@nominas_bp.route("/subcontratos/<int:subcontract_id>")
@login_required
def subcontract_detail(subcontract_id):
    subcontract = db.get_or_404(Subcontract, subcontract_id)
    require_project_access(subcontract.project)
    return render_template("subcontracts/detail.html", subcontract=subcontract)


@nominas_bp.route("/subcontratos/<int:subcontract_id>/pagos/nuevo", methods=["GET", "POST"])
@login_required
def subcontract_payment_new(subcontract_id):
    subcontract = db.get_or_404(Subcontract, subcontract_id)
    require_project_access(subcontract.project)
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    payment_methods = shared_payment_methods()
    if request.method == "POST":
        try:
            amount = parse_money("monto_capturado", required=True)
            amount_type = request.form.get("tipo_monto", "SIN_IVA")
            if amount <= 0 or amount_type not in {"SIN_IVA", "CON_IVA"}:
                raise ValueError("Revisa el monto y el IVA.")
            method = selected_payment_method("metodo_pago")
            payment = SubcontractPayment(
                subcontract_id=subcontract.id,
                fecha=parse_date("fecha"),
                concepto=request.form.get("concepto", "").strip().upper(),
                monto_capturado=amount,
                tipo_monto=amount_type,
                monto_sin_iva=amount_without_vat(amount, amount_type),
                metodo_pago=method.nombre,
                payment_method_id=method.id,
                company_id=request.form.get("company_id", type=int),
                notas=request.form.get("notas", "").strip() or None,
                created_by_id=current_user.id,
            )
            if not payment.concepto or not payment.company_id:
                raise ValueError("Concepto y empresa son obligatorios.")
            db.session.add(payment)
            db.session.flush()
            audit("CREAR", "PAGO_SUBCONTRATO", payment.id, f"{subcontract.id}: {amount}")
            db.session.commit()
            flash("Pago de subcontrato registrado sin duplicar el IVA en el costo.", "success")
            return redirect(url_for("subcontract_detail", subcontract_id=subcontract.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "subcontracts/payment_form.html",
        subcontract=subcontract,
        companies=companies,
        payment_methods=payment_methods,
    )


@nominas_bp.route("/subcontratos/pagos/<int:payment_id>/eliminar", methods=["POST"])
@admin_required
def subcontract_payment_delete(payment_id):
    payment = db.get_or_404(SubcontractPayment, payment_id)
    subcontract_id = payment.subcontract_id
    audit("ELIMINAR", "PAGO_SUBCONTRATO", payment.id, str(payment.monto_capturado))
    db.session.delete(payment)
    db.session.commit()
    flash("Pago de subcontrato eliminado.", "success")
    return redirect(url_for("subcontract_detail", subcontract_id=subcontract_id))


# ---------------------------------------------------------------------------
# Reportes y exportación a Excel
# ---------------------------------------------------------------------------


def build_weekly_closing_report(week_start: date) -> dict:
    """Consolida el recurso y el costo de todas las obras para revisión del CEO."""

    week_end = week_start + timedelta(days=4)
    payrolls = (
        Payroll.query.join(Project)
        .options(
            joinedload(Payroll.project),
            joinedload(Payroll.lines).joinedload(PayrollLine.employee),
            joinedload(Payroll.lines).joinedload(PayrollLine.budget_item),
            joinedload(Payroll.lines).joinedload(PayrollLine.empresa_transferencia),
            joinedload(Payroll.lines).joinedload(PayrollLine.empresa_efectivo),
        )
        .filter(Payroll.semana_inicio == week_start)
        .order_by(Project.nombre)
        .all()
    )

    # Los borradores se recalculan solo para presentar datos vigentes. El reporte
    # no cierra ni registra movimientos contables.
    weekly_loan_deductions: dict[int, Decimal] = {}
    actual_loan_ids: set[int] = set()
    for payroll in payrolls:
        if payroll.estado in {"borrador", "enviada"}:
            for line in payroll.lines:
                for loan, amount in recalculate_line(line):
                    weekly_loan_deductions[loan.id] = money(
                        weekly_loan_deductions.get(loan.id, Decimal("0")) + amount
                    )

    line_ids = [line.id for payroll in payrolls for line in payroll.lines]
    if line_ids:
        applied_payments = LoanPayment.query.filter(
            LoanPayment.payroll_line_id.in_(line_ids)
        ).all()
        for payment in applied_payments:
            actual_loan_ids.add(payment.loan_id)
            weekly_loan_deductions[payment.loan_id] = money(
                weekly_loan_deductions.get(payment.loan_id, Decimal("0"))
                + decimal_value(payment.monto)
            )

    additional_payments = (
        AdditionalPayment.query.options(
            joinedload(AdditionalPayment.project),
            joinedload(AdditionalPayment.budget_item),
            joinedload(AdditionalPayment.company),
        )
        .filter(AdditionalPayment.fecha.between(week_start, week_end))
        .order_by(AdditionalPayment.fecha, AdditionalPayment.id)
        .all()
    )
    office_expenses = (
        OfficeExpense.query.options(
            joinedload(OfficeExpense.project),
            joinedload(OfficeExpense.budget_item),
            joinedload(OfficeExpense.company),
        )
        .filter(OfficeExpense.fecha.between(week_start, week_end))
        .order_by(OfficeExpense.fecha, OfficeExpense.id)
        .all()
    )
    subcontract_payments = (
        SubcontractPayment.query.join(Subcontract)
        .options(
            joinedload(SubcontractPayment.subcontract).joinedload(Subcontract.project),
            joinedload(SubcontractPayment.subcontract).joinedload(Subcontract.budget_item),
            joinedload(SubcontractPayment.subcontract).joinedload(Subcontract.contractor),
            joinedload(SubcontractPayment.company),
        )
        .filter(SubcontractPayment.fecha.between(week_start, week_end))
        .order_by(SubcontractPayment.fecha, SubcontractPayment.id)
        .all()
    )

    loan_filters = [
        # Al consultar una semana histórica no deben aparecer préstamos que
        # todavía no existían en esa fecha.
        Loan.estado.in_(("activo", "liquidado")) & (Loan.fecha_prestamo <= week_end),
        Loan.fecha_prestamo.between(week_start, week_end),
    ]
    if weekly_loan_deductions:
        loan_filters.append(Loan.id.in_(list(weekly_loan_deductions)))
    loans = (
        Loan.query.options(
            joinedload(Loan.employee).joinedload(Employee.project),
            joinedload(Loan.project),
            joinedload(Loan.company),
            joinedload(Loan.payments),
        )
        .filter(or_(*loan_filters))
        .order_by(Loan.estado, Loan.fecha_prestamo, Loan.id)
        .all()
    )

    project_rows: dict[int, dict] = {}

    def ensure_project(project: Project | None) -> dict:
        key = project.id if project else 0
        if key not in project_rows:
            project_rows[key] = {
                "project": project,
                "codigo": project.codigo if project else "SIN OBRA",
                "nombre": project.nombre if project else "Sin obra asignada",
                "payroll_status": "SIN NÓMINA",
                "trabajadores": 0,
                "nomina_devengada": Decimal("0"),
                "nomina_imss": Decimal("0"),
                "nomina_costo": Decimal("0"),
                "nomina_neto": Decimal("0"),
                "nomina_transferencia": Decimal("0"),
                "nomina_efectivo": Decimal("0"),
                "nomina_infonavit": Decimal("0"),
                "nomina_prestamos": Decimal("0"),
                "nomina_otros": Decimal("0"),
                "proveedores_recurso": Decimal("0"),
                "proveedores_costo": Decimal("0"),
                "subcontratos_recurso": Decimal("0"),
                "subcontratos_costo": Decimal("0"),
                "prestamos_entregados": Decimal("0"),
                "prestamos_efectivo": Decimal("0"),
                "prestamos_transferencia": Decimal("0"),
                "recurso_total": Decimal("0"),
                "costo_total": Decimal("0"),
            }
        return project_rows[key]

    funding: dict[tuple[str, str, str], dict] = {}

    def add_funding(category: str, method: str, company: Company | None, amount) -> None:
        amount = money(amount)
        if amount <= 0:
            return
        company_code = company.codigo if company else "SIN EMPRESA"
        company_name = company.nombre if company else "Sin empresa asignada"
        key = (category, method or "SIN MÉTODO", company_code)
        row = funding.setdefault(
            key,
            {
                "categoria": category,
                "metodo": method or "SIN MÉTODO",
                "empresa_codigo": company_code,
                "empresa_nombre": company_name,
                "monto": Decimal("0"),
            },
        )
        row["monto"] = money(row["monto"] + amount)

    for payroll in payrolls:
        row = ensure_project(payroll.project)
        row["payroll_status"] = payroll.estado
        row["trabajadores"] = len(payroll.lines)
        row["nomina_devengada"] = money(payroll.total_devengado)
        row["nomina_imss"] = money(
            sum((decimal_value(line.descuento_imss) for line in payroll.lines), Decimal("0"))
        )
        row["nomina_costo"] = money(payroll.total_costo_mano_obra)
        row["nomina_neto"] = money(payroll.total_neto)
        row["nomina_transferencia"] = money(
            sum((decimal_value(line.pago_transferencia) for line in payroll.lines), Decimal("0"))
        )
        row["nomina_efectivo"] = money(
            sum((decimal_value(line.pago_efectivo) for line in payroll.lines), Decimal("0"))
        )
        row["nomina_infonavit"] = money(
            sum((decimal_value(line.descuento_infonavit) for line in payroll.lines), Decimal("0"))
        )
        row["nomina_prestamos"] = money(
            sum((decimal_value(line.descuento_prestamo) for line in payroll.lines), Decimal("0"))
        )
        row["nomina_otros"] = money(
            sum((decimal_value(line.otro_descuento) for line in payroll.lines), Decimal("0"))
        )
        for line in payroll.lines:
            add_funding(
                "NÓMINA",
                "TRANSFERENCIA",
                line.empresa_transferencia,
                line.pago_transferencia,
            )
            add_funding("NÓMINA", "EFECTIVO", line.empresa_efectivo, line.pago_efectivo)

    supplier_rows: list[dict] = []
    for payment in additional_payments:
        row = ensure_project(payment.project)
        row["proveedores_recurso"] = money(
            row["proveedores_recurso"] + decimal_value(payment.monto_capturado)
        )
        row["proveedores_costo"] = money(
            row["proveedores_costo"] + decimal_value(payment.monto_sin_iva)
        )
        supplier_rows.append(
            {
                "tipo": "PAGO ADICIONAL",
                "fecha": payment.fecha,
                "project": payment.project,
                "budget_item": payment.budget_item,
                "beneficiario": payment.beneficiario,
                "concepto": payment.concepto,
                "monto_capturado": money(payment.monto_capturado),
                "tipo_monto": payment.tipo_monto,
                "monto_sin_iva": money(payment.monto_sin_iva),
                "metodo_pago": payment.metodo_pago,
                "company": payment.company,
                "notas": payment.notas,
            }
        )
        add_funding("PROVEEDORES", payment.metodo_pago, payment.company, payment.monto_capturado)

    for expense in office_expenses:
        row = ensure_project(expense.project)
        row["proveedores_recurso"] = money(
            row["proveedores_recurso"] + decimal_value(expense.monto_capturado)
        )
        row["proveedores_costo"] = money(
            row["proveedores_costo"] + decimal_value(expense.monto_sin_iva)
        )
        supplier_rows.append(
            {
                "tipo": "GASTO DE OFICINA",
                "fecha": expense.fecha,
                "project": expense.project,
                "budget_item": expense.budget_item,
                "beneficiario": expense.proveedor,
                "concepto": expense.concepto,
                "monto_capturado": money(expense.monto_capturado),
                "tipo_monto": expense.tipo_monto,
                "monto_sin_iva": money(expense.monto_sin_iva),
                "metodo_pago": expense.metodo_pago,
                "company": expense.company,
                "notas": expense.notas,
            }
        )
        add_funding("PROVEEDORES", expense.metodo_pago, expense.company, expense.monto_capturado)

    for payment in subcontract_payments:
        project = payment.subcontract.project
        row = ensure_project(project)
        row["subcontratos_recurso"] = money(
            row["subcontratos_recurso"] + decimal_value(payment.monto_capturado)
        )
        row["subcontratos_costo"] = money(
            row["subcontratos_costo"] + decimal_value(payment.monto_sin_iva)
        )
        add_funding(
            "SUBCONTRATISTAS", payment.metodo_pago, payment.company, payment.monto_capturado
        )

    employee_week_project = {
        line.employee_id: payroll.project
        for payroll in payrolls
        for line in payroll.lines
    }
    loan_rows: list[dict] = []
    for loan in loans:
        weekly_deduction = money(weekly_loan_deductions.get(loan.id, Decimal("0")))
        remaining = money(loan_remaining(loan))
        actual_this_week = loan.id in actual_loan_ids
        projected_remaining = (
            remaining
            if actual_this_week
            else money(max(Decimal("0"), remaining - weekly_deduction))
        )
        projected_paid = money(
            decimal_value(loan.total_pagar) - projected_remaining
        )
        if loan.estado == "rechazado":
            week_status = "RECHAZADO"
        elif loan.estado in {"pendiente", "aprobado"}:
            week_status = loan.estado.upper()
        elif projected_remaining <= 0:
            week_status = "LIQUIDADO" if actual_this_week else "SE LIQUIDA ESTA SEMANA"
        elif weekly_deduction > 0:
            week_status = "RETENCIÓN APLICADA" if actual_this_week else "RETENCIÓN PREVISTA"
        else:
            week_status = loan.estado
        delivered_this_week = (
            loan.estado in {"activo", "liquidado"}
            and week_start <= loan.fecha_prestamo <= week_end
        )
        delivery_project = loan.obra_entrega
        project = (
            delivery_project
            if delivered_this_week
            else employee_week_project.get(loan.employee_id) or delivery_project
        )
        if delivered_this_week:
            row = ensure_project(delivery_project)
            row["prestamos_entregados"] = money(
                row["prestamos_entregados"] + decimal_value(loan.monto)
            )
            channel = payment_channel(loan.metodo_entrega)
            method_key = (
                "prestamos_efectivo"
                if channel == "EFECTIVO"
                else "prestamos_transferencia"
            )
            row[method_key] = money(
                row[method_key] + decimal_value(loan.monto)
            )
            add_funding(
                "PRÉSTAMOS",
                channel,
                loan.company,
                loan.monto,
            )
        first_eligible_week = loan.fecha_prestamo + timedelta(
            days=7 - loan.fecha_prestamo.weekday()
        )
        loan_rows.append(
            {
                "loan": loan,
                "project": project,
                "weekly_deduction": weekly_deduction,
                "remaining": remaining,
                "projected_paid": projected_paid,
                "projected_remaining": projected_remaining,
                "week_status": week_status,
                "delivered_this_week": delivered_this_week,
                "first_eligible_week": first_eligible_week,
            }
        )

    for row in project_rows.values():
        row["recurso_total"] = money(
            row["nomina_neto"]
            + row["proveedores_recurso"]
            + row["subcontratos_recurso"]
            + row["prestamos_entregados"]
        )
        row["costo_total"] = money(
            row["nomina_costo"]
            + row["proveedores_costo"]
            + row["subcontratos_costo"]
        )

    ordered_projects = sorted(
        project_rows.values(), key=lambda row: (row["codigo"] == "SIN OBRA", row["nombre"])
    )
    totals = {
        key: money(sum((decimal_value(row[key]) for row in ordered_projects), Decimal("0")))
        for key in (
            "nomina_devengada",
            "nomina_imss",
            "nomina_costo",
            "nomina_neto",
            "nomina_transferencia",
            "nomina_efectivo",
            "nomina_infonavit",
            "nomina_prestamos",
            "nomina_otros",
            "proveedores_recurso",
            "proveedores_costo",
            "subcontratos_recurso",
            "subcontratos_costo",
            "prestamos_entregados",
            "prestamos_efectivo",
            "prestamos_transferencia",
            "recurso_total",
            "costo_total",
        )
    }
    totals["trabajadores"] = sum(row["trabajadores"] for row in ordered_projects)
    totals["nominas"] = len(payrolls)
    totals["borradores"] = sum(
        1 for payroll in payrolls if payroll.estado in {"borrador", "enviada"}
    )
    totals["cerradas"] = sum(
        1 for payroll in payrolls if payroll.estado in FINALIZED_PAYROLL_STATES
    )
    totals["retencion_prestamos"] = money(
        sum(weekly_loan_deductions.values(), Decimal("0"))
    )

    active_employee_project_ids = {
        project_id
        for (project_id,) in db.session.query(Employee.project_id)
        .filter(Employee.activo.is_(True), Employee.project_id.is_not(None))
        .distinct()
        .all()
    }
    payroll_project_ids = {payroll.project_id for payroll in payrolls}
    missing_payroll_projects = (
        Project.query.filter(Project.id.in_(active_employee_project_ids - payroll_project_ids))
        .order_by(Project.nombre)
        .all()
        if active_employee_project_ids - payroll_project_ids
        else []
    )

    supplier_rows.sort(key=lambda row: (row["fecha"], row["project"].nombre, row["beneficiario"]))
    funding_rows = sorted(
        funding.values(),
        key=lambda row: (row["categoria"], row["metodo"], row["empresa_codigo"]),
    )
    return {
        "week_start": week_start,
        "week_end": week_end,
        "payrolls": payrolls,
        "projects": ordered_projects,
        "totals": totals,
        "supplier_rows": supplier_rows,
        "subcontract_payments": subcontract_payments,
        "loan_rows": loan_rows,
        "funding_rows": funding_rows,
        "resource_summary": weekly_resource_breakdown(week_start),
        "missing_payroll_projects": missing_payroll_projects,
    }


@nominas_bp.route("/reportes/cierre-semanal")
@admin_required
def weekly_closing_report():
    if not current_user.acceso_global_obras:
        abort(403)
    raw_week = request.args.get("semana", "").strip()
    try:
        selected_date = (
            datetime.strptime(raw_week, "%Y-%m-%d").date()
            if raw_week
            else current_week_start()
        )
        week_start = current_week_start(selected_date)
    except ValueError:
        week_start = current_week_start()
        flash("La semana indicada no es válida; se mostró la semana actual.", "warning")
    report = build_weekly_closing_report(week_start)
    return render_template(
        "reports/weekly_closing.html",
        report=report,
        generated_at=datetime.now(),
    )


def filtered_report_lines():
    allowed_project_ids = [project.id for project in accessible_projects_query().all()]
    query = PayrollLine.query.join(Payroll).join(Project, Payroll.project_id == Project.id).options(
        joinedload(PayrollLine.payroll).joinedload(Payroll.project),
        joinedload(PayrollLine.employee),
        joinedload(PayrollLine.budget_item),
        joinedload(PayrollLine.partida),
        joinedload(PayrollLine.subpartida),
        joinedload(PayrollLine.empresa_transferencia),
        joinedload(PayrollLine.empresa_efectivo),
    ).filter(Payroll.project_id.in_(allowed_project_ids or [-1]))
    project_id = request.args.get("project_id", type=int)
    item_id = request.args.get("budget_item_id", type=int)
    status = request.args.get("estado", "")
    start = request.args.get("desde", "")
    end = request.args.get("hasta", "")
    absences = request.args.get("solo_faltas") == "1"
    if project_id:
        if project_id not in allowed_project_ids:
            abort(404)
        query = query.filter(Payroll.project_id == project_id)
    if item_id:
        item = db.session.get(BudgetItem, item_id)
        if not item or item.project_id not in allowed_project_ids:
            abort(404)
        if project_id and item.project_id != project_id:
            abort(404)
        query = query.filter(payroll_line_item_filter(item))
    status = {"BORRADOR": "borrador", "CERRADA": "aprobada"}.get(
        status, status.lower()
    )
    if status in PAYROLL_STATES:
        query = query.filter(Payroll.estado == status)
    if start:
        try:
            query = query.filter(Payroll.semana_inicio >= datetime.strptime(start, "%Y-%m-%d").date())
        except ValueError:
            pass
    if end:
        try:
            query = query.filter(Payroll.semana_inicio <= datetime.strptime(end, "%Y-%m-%d").date())
        except ValueError:
            pass
    if absences:
        query = query.filter(PayrollLine.numero_faltas > 0)
    return query.order_by(Payroll.semana_inicio.desc(), Project.nombre, PayrollLine.nombre_trabajador)


@nominas_bp.route("/reportes")
@admin_required
def reports_index():
    lines = filtered_report_lines().all()
    summaries: dict[tuple[str, date], dict[str, Decimal]] = {}
    for line in lines:
        key = (line.payroll.project.nombre, line.payroll.semana_inicio)
        row = summaries.setdefault(
            key,
            {
                "devengado": Decimal("0"),
                "neto": Decimal("0"),
                "transferencia": Decimal("0"),
                "efectivo": Decimal("0"),
                "faltas": Decimal("0"),
                "prestamos": Decimal("0"),
                "imss": Decimal("0"),
                "costo_mano_obra": Decimal("0"),
                "trabajadores": Decimal("0"),
            },
        )
        row["devengado"] += decimal_value(line.monto_devengado) + decimal_value(line.pago_extra)
        row["neto"] += decimal_value(line.neto_pagar)
        row["transferencia"] += decimal_value(line.pago_transferencia)
        row["efectivo"] += decimal_value(line.pago_efectivo)
        row["faltas"] += decimal_value(line.numero_faltas)
        row["prestamos"] += decimal_value(line.descuento_prestamo)
        row["imss"] += decimal_value(line.descuento_imss)
        row["costo_mano_obra"] += (
            decimal_value(line.monto_devengado)
            + decimal_value(line.pago_extra)
            + decimal_value(line.descuento_imss)
        )
        row["trabajadores"] += 1
    projects = accessible_projects_query().all()
    selected_project = request.args.get("project_id", type=int)
    items = (
        BudgetItem.query.filter_by(project_id=selected_project).order_by(BudgetItem.codigo).all()
        if selected_project
        else []
    )
    totals = {
        "devengado": money(sum((decimal_value(l.monto_devengado) + decimal_value(l.pago_extra) for l in lines), Decimal("0"))),
        "neto": money(sum((decimal_value(l.neto_pagar) for l in lines), Decimal("0"))),
        "imss": money(sum((decimal_value(l.descuento_imss) for l in lines), Decimal("0"))),
        "costo_mano_obra": money(
            sum(
                (
                    decimal_value(l.monto_devengado)
                    + decimal_value(l.pago_extra)
                    + decimal_value(l.descuento_imss)
                    for l in lines
                ),
                Decimal("0"),
            )
        ),
        "transferencia": money(sum((decimal_value(l.pago_transferencia) for l in lines), Decimal("0"))),
        "efectivo": money(sum((decimal_value(l.pago_efectivo) for l in lines), Decimal("0"))),
        "trabajadores": len(lines),
    }
    resource_week = current_week_start()
    if request.args.get("desde"):
        try:
            resource_week = current_week_start(
                datetime.strptime(request.args["desde"], "%Y-%m-%d").date()
            )
        except ValueError:
            pass
    return render_template(
        "reports/index.html",
        lines=lines,
        summaries=summaries,
        totals=totals,
        projects=projects,
        items=items,
        filters=request.args,
        # Si se filtra una obra, el requerido corresponde solo a esa obra. La
        # disponibilidad no se prorratea: es un fondo semanal global confirmado
        # por Administración y solo se muestra en la consulta conjunta.
        resource_summary=weekly_resource_breakdown(
            resource_week, [selected_project] if selected_project else None
        ),
    )


def style_excel_sheet(ws, currency_columns: Iterable[int] = (), percent_columns: Iterable[int] = ()):
    navy = "24557A"
    light = "DCE6F1"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34
    for column in currency_columns:
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = '$#,##0.00;[Red]-$#,##0.00'
    for column in percent_columns:
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = "0.0%"
    for index, column_cells in enumerate(ws.columns, start=1):
        length = max((len(str(cell.value or "")) for cell in column_cells[: min(ws.max_row, 100)]), default=8)
        ws.column_dimensions[get_column_letter(index)].width = min(max(length + 2, 11), 34)
    for row in ws.iter_rows(min_row=2):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=light)


def resource_weeks_for_export(lines: list[PayrollLine]) -> list[date]:
    """Resuelve las semanas del Excel sin perder periodos sin nómina."""

    parsed: dict[str, date] = {}
    for key in ("desde", "hasta"):
        raw = (request.args.get(key) or "").strip()
        if raw:
            try:
                parsed[key] = datetime.strptime(raw, "%Y-%m-%d").date()
            except ValueError:
                pass
    if parsed:
        start = parsed.get("desde") or parsed["hasta"]
        end = parsed.get("hasta") or parsed["desde"]
        if start > end:
            start, end = end, start
        return week_starts_between(start, end)
    weeks = sorted({line.payroll.semana_inicio for line in lines})
    return weeks or [current_week_start()]


def append_weekly_resource_sheet(
    workbook: Workbook,
    weeks: Iterable[date],
    project_ids: Iterable[int] | None,
    scope_label: str,
) -> None:
    """Agrega al Excel los mismos totales que consumen los dashboards."""

    sheet = workbook.create_sheet("Recurso semanal")
    sheet.append(
        [
            "Semana",
            "Alcance",
            "Nómina efectivo",
            "Préstamos nuevos efectivo",
            "Gastos operativos efectivo",
            "Pagos adicionales efectivo",
            "Subcontratos efectivo",
            "Efectivo requerido",
            "Nómina transferencia",
            "Préstamos nuevos transferencia",
            "Gastos operativos transferencia",
            "Pagos adicionales transferencia",
            "Subcontratos transferencia",
            "Transferencias requeridas",
            "Recurso total requerido",
        ]
    )
    for week in weeks:
        resource = weekly_resource_breakdown(week, project_ids)
        cash = resource["methods"]["EFECTIVO"]
        bank = resource["methods"]["TRANSFERENCIA"]
        sheet.append(
            [
                resource["week_start"],
                scope_label,
                float(cash["nomina"]),
                float(cash["prestamos"]),
                float(cash["gastos_operativos"]),
                float(cash["pagos_adicionales"]),
                float(cash["subcontratos"]),
                float(cash["requerido"]),
                float(bank["nomina"]),
                float(bank["prestamos"]),
                float(bank["gastos_operativos"]),
                float(bank["pagos_adicionales"]),
                float(bank["subcontratos"]),
                float(bank["requerido"]),
                float(resource["requerido_total"]),
            ]
        )
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "dd/mm/yyyy"
    style_excel_sheet(sheet, currency_columns=range(3, 16))


@nominas_bp.route("/reportes/exportar.xlsx")
@admin_required
def reports_export():
    lines = filtered_report_lines().all()
    allowed_project_ids = [project.id for project in accessible_projects_query().all()]
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Nóminas"
    headers = [
        "Semana inicio",
        "Semana fin",
        "Estado",
        "Obra / oficina",
        "Partida",
        "Subpartida",
        "Trabajador",
        "Fecha ingreso",
        "NSS",
        "Empresa IMSS",
        "Empresa operativa",
        "Cuadrilla",
        "Supervisor",
        "Puesto",
        "Salario semanal proyectado",
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Días trabajados",
        "No. faltas",
        "Sueldo diario",
        "Descuento faltas",
        "Monto devengado",
        "Pago extra",
        "Infonavit",
        "Costo IMSS patronal (no se descuenta)",
        "Préstamos",
        "Otros descuentos",
        "Concepto otro descuento",
        "Vales gasolina",
        "Transferencia / cheque",
        "Empresa transferencia",
        "Efectivo",
        "Empresa efectivo",
        "Neto a pagar",
        "Notas",
    ]
    ws.append(headers)
    for line in lines:
        employee = line.employee
        ws.append(
            [
                line.payroll.semana_inicio,
                line.payroll.semana_fin,
                line.payroll.estado,
                line.payroll.project.nombre,
                (
                    line.partida_resuelta.etiqueta
                    if line.partida_resuelta
                    else None
                ),
                (
                    line.subpartida_resuelta.etiqueta
                    if line.subpartida_resuelta
                    else None
                ),
                line.nombre_trabajador,
                employee.fecha_ingreso if employee else None,
                nss_para_usuario(current_user, employee.nss) if employee else None,
                employee.empresa_imss.codigo if employee and employee.empresa_imss else None,
                line.empresa_operativa,
                line.cuadrilla,
                line.supervisor,
                line.puesto,
                float(decimal_value(line.salario_semanal)),
                "Sí" if line.lunes else "No",
                "Sí" if line.martes else "No",
                "Sí" if line.miercoles else "No",
                "Sí" if line.jueves else "No",
                "Sí" if line.viernes else "No",
                float(decimal_value(line.dias_trabajados)),
                float(decimal_value(line.numero_faltas)),
                float(decimal_value(line.sueldo_diario)),
                float(decimal_value(line.descuento_faltas)),
                float(decimal_value(line.monto_devengado)),
                float(decimal_value(line.pago_extra)),
                float(decimal_value(line.descuento_infonavit)),
                float(decimal_value(line.descuento_imss)),
                float(decimal_value(line.descuento_prestamo)),
                float(decimal_value(line.otro_descuento)),
                line.concepto_otro_descuento,
                float(decimal_value(line.vales_gasolina)),
                float(decimal_value(line.pago_transferencia)),
                line.empresa_transferencia.codigo if line.empresa_transferencia else None,
                float(decimal_value(line.pago_efectivo)),
                line.empresa_efectivo.codigo if line.empresa_efectivo else None,
                float(decimal_value(line.neto_pagar)),
                line.notas,
            ]
        )
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).number_format = "dd/mm/yyyy"
        ws.cell(row, 2).number_format = "dd/mm/yyyy"
        ws.cell(row, 8).number_format = "dd/mm/yyyy"
    style_excel_sheet(
        ws,
        currency_columns=[
            15,
            23,
            24,
            25,
            26,
            27,
            28,
            29,
            30,
            32,
            33,
            35,
            37,
        ],
    )

    summary_ws = workbook.create_sheet("Resumen por obra")
    summary_ws.append(
        [
            "Obra / oficina",
            "Semana",
            "Estado",
            "Trabajadores",
            "Devengado + extras",
            "Transferencia",
            "Efectivo",
            "Neto",
            "Infonavit",
            "Costo IMSS patronal",
            "Costo mano de obra",
            "Préstamos",
            "Faltas",
        ]
    )
    grouped: dict[tuple[int, date, str], list[PayrollLine]] = {}
    for line in lines:
        grouped.setdefault(
            (line.payroll.project_id, line.payroll.semana_inicio, line.payroll.estado), []
        ).append(line)
    for (project_id, week, status), group_lines in sorted(grouped.items(), key=lambda x: x[0][1], reverse=True):
        project = db.session.get(Project, project_id)
        summary_ws.append(
            [
                project.nombre,
                week,
                status,
                len(group_lines),
                float(sum((decimal_value(x.monto_devengado) + decimal_value(x.pago_extra) for x in group_lines), Decimal("0"))),
                float(sum((decimal_value(x.pago_transferencia) for x in group_lines), Decimal("0"))),
                float(sum((decimal_value(x.pago_efectivo) for x in group_lines), Decimal("0"))),
                float(sum((decimal_value(x.neto_pagar) for x in group_lines), Decimal("0"))),
                float(sum((decimal_value(x.descuento_infonavit) for x in group_lines), Decimal("0"))),
                float(sum((decimal_value(x.descuento_imss) for x in group_lines), Decimal("0"))),
                float(
                    sum(
                        (
                            decimal_value(x.monto_devengado)
                            + decimal_value(x.pago_extra)
                            + decimal_value(x.descuento_imss)
                            for x in group_lines
                        ),
                        Decimal("0"),
                    )
                ),
                float(sum((decimal_value(x.descuento_prestamo) for x in group_lines), Decimal("0"))),
                float(sum((decimal_value(x.numero_faltas) for x in group_lines), Decimal("0"))),
            ]
        )
    for row in range(2, summary_ws.max_row + 1):
        summary_ws.cell(row, 2).number_format = "dd/mm/yyyy"
    style_excel_sheet(summary_ws, currency_columns=range(5, 13))

    selected_resource_project_id = request.args.get("project_id", type=int)
    if selected_resource_project_id:
        resource_project_ids: Iterable[int] | None = [
            selected_resource_project_id
        ]
        resource_scope_label = db.session.get(
            Project, selected_resource_project_id
        ).nombre
    elif current_user.acceso_global_obras:
        resource_project_ids = None
        resource_scope_label = "Consolidado"
    else:
        resource_project_ids = allowed_project_ids
        resource_scope_label = "Obras accesibles"
    append_weekly_resource_sheet(
        workbook,
        resource_weeks_for_export(lines),
        resource_project_ids,
        resource_scope_label,
    )

    budget_ws = workbook.create_sheet("Presupuesto por partida")
    budget_ws.append(
        [
            "Obra / oficina",
            "Código",
            "Partida / subpartida",
            "Categoría",
            "Presupuesto",
            "Nómina con IMSS",
            "IMSS patronal incluido",
            "Adicionales",
            "Subcontratos pagados",
            "Gastos oficina",
            "Consumido total",
            "Disponible",
        ]
    )
    selected_project_ids = sorted({line.payroll.project_id for line in lines})
    if not selected_project_ids:
        selected_project_ids = allowed_project_ids
    for project in Project.query.filter(Project.id.in_(selected_project_ids)).order_by(Project.nombre):
        for item in sorted(project.budget_items, key=lambda value: value.codigo):
            values = item_consumption(item)
            budget_ws.append(
                [
                    project.nombre,
                    item.codigo,
                    item.nombre,
                    item.categoria,
                    float(decimal_value(item.presupuesto)),
                    float(values["nomina"]),
                    float(values["imss"]),
                    float(values["adicionales"]),
                    float(values["subcontratos"]),
                    float(values["oficina"]),
                    float(values["total"]),
                    float(values["restante"]),
                ]
            )
    style_excel_sheet(budget_ws, currency_columns=range(5, 13))

    control_ws = workbook.create_sheet("Control presupuestal")
    control_ws.append(
        [
            "Obra / oficina",
            "Presupuesto total",
            "Consumido real (nóminas cerradas)",
            "% consumido",
            "Comprometido (subpartidas)",
            "% comprometido",
            "Disponible dentro de partidas",
            "Presupuesto aún fuera de partidas",
            "Disponible real",
            "Disponible comprometido",
            "Alerta",
        ]
    )
    for project in Project.query.filter(Project.id.in_(selected_project_ids)).order_by(Project.nombre):
        values = project_consumption(project)
        alerts = []
        if values["alerta_consumido"]:
            alerts.append("CONSUMIDO >= 80%")
        if values["alerta_comprometido"]:
            alerts.append("COMPROMETIDO >= 80%")
        control_ws.append(
            [
                project.nombre,
                float(values["presupuesto_base"]),
                float(values["total"]),
                float(values["porcentaje_consumido"] / Decimal("100")),
                float(values["total_comprometido"]),
                float(values["porcentaje_comprometido"] / Decimal("100")),
                float(values["disponible_partidas"]),
                float(values["sin_asignar_partidas"]),
                float(values["restante_total"]),
                float(values["restante_comprometido"]),
                " · ".join(alerts) if alerts else "EN CONTROL",
            ]
        )
    style_excel_sheet(
        control_ws,
        currency_columns=[2, 3, 5, 7, 8, 9, 10],
        percent_columns=[4, 6],
    )

    loans_ws = workbook.create_sheet("Préstamos")
    loans_ws.append(
        [
            "Trabajador",
            "Obra de entrega",
            "Fecha préstamo",
            "Capital",
            "Tasa de interés %",
            "Total a pagar",
            "Retención semanal",
            "Abonado",
            "Restante",
            "Estado",
            "Método entrega",
            "Empresa",
            "Concepto",
        ]
    )
    loan_query = (
        Loan.query.join(Employee)
        .options(
            joinedload(Loan.employee),
            joinedload(Loan.project),
            joinedload(Loan.payments),
        )
        .filter(
            func.coalesce(Loan.project_id, Employee.project_id).in_(
                (
                    [selected_resource_project_id]
                    if selected_resource_project_id
                    else allowed_project_ids
                )
                or [-1]
            )
        )
        .order_by(Loan.fecha_prestamo.desc())
    )
    for loan in loan_query:
        loans_ws.append(
            [
                loan.employee.nombre_completo,
                loan.obra_entrega.nombre if loan.obra_entrega else None,
                loan.fecha_prestamo,
                float(decimal_value(loan.monto)),
                float(loan.tasa_interes),
                float(decimal_value(loan.total_pagar)),
                float(decimal_value(loan.retencion_semanal)),
                float(loan.abonado),
                float(loan.restante),
                loan.estado,
                loan.metodo_entrega,
                loan.company.codigo if loan.company else None,
                loan.concepto,
            ]
        )
    for row in range(2, loans_ws.max_row + 1):
        loans_ws.cell(row, 3).number_format = "dd/mm/yyyy"
    style_excel_sheet(loans_ws, currency_columns=(4, 6, 7, 8, 9))

    subs_ws = workbook.create_sheet("Subcontratos")
    subs_ws.append(
        [
            "Obra",
            "Contratista",
            "Especialidad",
            "Partida",
            "Presupuesto sin IVA",
            "Avance físico",
            "Pagado sin IVA",
            "Comprometido según avance",
            "Saldo vs avance",
            "Saldo total",
            "Próximo pago sugerido",
            "Estatus",
            "Observaciones",
        ]
    )
    sub_query = Subcontract.query.options(joinedload(Subcontract.payments)).order_by(
        Subcontract.project_id, Subcontract.especialidad
    )
    if selected_project_ids:
        sub_query = sub_query.filter(Subcontract.project_id.in_(selected_project_ids))
    for sub in sub_query:
        subs_ws.append(
            [
                sub.project.nombre,
                sub.contractor.nombre,
                sub.especialidad,
                sub.budget_item.etiqueta,
                float(decimal_value(sub.presupuesto_sin_iva)),
                float(decimal_value(sub.avance_fisico)),
                float(sub.pagado_sin_iva),
                float(sub.comprometido),
                float(sub.saldo_vs_avance),
                float(sub.saldo_total),
                float(sub.proximo_pago_sugerido),
                sub.estatus_control,
                sub.observaciones,
            ]
        )
    style_excel_sheet(subs_ws, currency_columns=[5, 7, 8, 9, 10, 11], percent_columns=[6])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    audit("EXPORTAR", "REPORTE_NOMINAS", detail=f"{len(lines)} líneas")
    db.session.commit()
    filename = f"nominas_semanales_{date.today().isoformat()}.xlsx"
    # FileWrapper evita el error de uwsgi_sendfile/fileno de PythonAnywhere
    # cuando se envía un Excel creado en memoria con BytesIO.
    response = Response(
        FileWrapper(output),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        direct_passthrough=True,
    )
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@nominas_bp.route("/auditoria")
@permission_required("seguridad", "ver")
def audit_list():
    logs = BitacoraAuditoria.query.options(joinedload(BitacoraAuditoria.usuario)).order_by(BitacoraAuditoria.fecha_hora.desc()).limit(500).all()
    return render_template("audit/list.html", logs=logs)


@nominas_bp.errorhandler(400)
def bad_request(error):
    return render_template(
        "errors/error.html",
        code=400,
        title="Solicitud no válida",
        message=str(error),
    ), 400


@nominas_bp.errorhandler(404)
def not_found(_error):
    return render_template(
        "errors/error.html",
        code=404,
        title="No encontrado",
        message="El registro o página no existe.",
    ), 404


@nominas_bp.errorhandler(409)
def conflict(_error):
    return render_template(
        "errors/error.html",
        code=409,
        title="Acción no disponible",
        message="El estado actual del registro no permite esa acción.",
    ), 409
