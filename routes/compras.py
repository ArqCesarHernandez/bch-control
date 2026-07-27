"""Fase 4 final: Compras semanales, crédito, cotizaciones y SMNC."""

from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps
from io import BytesIO

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from flask_mail import Message
from email_validator import EmailNotValidError, validate_email
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload, selectinload

from forms import (
    CreditCardForm,
    CreditCardPaymentForm,
    FinanceBeneficiaryValidationForm,
    OperationalPurchaseOrderForm,
    PaymentScheduleDecisionForm,
    PurchaseOrderRevisionForm,
)

from compras_models import (
    ACTIVE_ORDER_STATES,
    BudgetExplosionItem,
    CreditCard,
    CreditCardPayment,
    ExplosionRevision,
    GoodsReceipt,
    GoodsReceiptLine,
    MaterialChangeRequest,
    MaterialChangeRequestLine,
    OPERATION_CATEGORIES,
    PaymentMethod,
    PurchaseAlertRun,
    PurchaseNotification,
    PurchaseOrder,
    PurchaseOrderLine,
    PurchaseOrderPaymentSchedule,
    PurchaseRequisition,
    PurchaseRequisitionLine,
    Quotation,
    QuotationLine,
    QuotationLineSource,
    Supplier,
    SupplierAdvanceMovement,
    SupplierSupplyItem,
    SupplyItem,
    SupplyProjectCatalog,
    quotation_requisitions,
)
from models import BitacoraAuditoria, CentroCosto, Usuario, db, usuario_centros_nomina, utc_now
from nominas_models import (
    AdditionalPayment,
    BudgetItem,
    Company,
    decimal_value,
    money,
)
from utils.decorators import any_permission_required, permission_required
from utils.access import verificar_acceso_obra, verificar_asignacion_obra
from services.fase5 import (
    actualizar_conciliacion_pagada,
    conciliacion_aprobada_para_pago,
    usuarios_con_permiso,
)
from services.actualizacion_operativa import (
    activar_revision_explosion,
    ajustar_reserva_pendiente,
    centros_operativos_accesibles,
    clasificar_y_liberar_requisicion,
    crear_o_actualizar_rfq_liberada,
    crear_programacion_pago,
    items_explosion_vigente,
    liberar_reservas_pendientes,
    registrar_revision_orden,
    revision_explosion_vigente,
    siguiente_revision_explosion,
    sincronizar_estado_requisicion,
    sincronizar_programacion_recepcion,
    snapshot_orden,
)
from services.weekly_resources import (
    RESOURCE_COMPONENTS,
    week_starts_between,
    weekly_resource_breakdown,
)
from utils.project_scope import (
    obra_activa_id,
    obras_asignadas_supervisor,
    seleccionar_obra_activa,
)


compras_bp = Blueprint("compras", __name__, url_prefix="/compras")

QUANTITY_STEP = Decimal("0.0001")
MONEY_STEP = Decimal("0.01")
SUPPLY_TYPES = {"MATERIAL", "EQUIPO", "MANO_OBRA", "SUBCONTRATO", "INDIRECTO"}
TYPE_LABELS = {
    "MATERIAL": "Material",
    "EQUIPO": "Equipo",
    "MANO_OBRA": "Mano de Obra",
    "SUBCONTRATO": "Subcontrato",
    "INDIRECTO": "Indirecto",
}
TYPE_INPUT_MAP = {normalize: key for key, normalize in TYPE_LABELS.items()}
TYPE_INPUT_MAP.update({label.upper(): key for key, label in TYPE_LABELS.items()})
EXPLOSION_HEADERS = [
    "Partida",
    "Subpartida",
    "Tipo",
    "Clave Insumo",
    "Descripción",
    "Unidad",
    "Cantidad",
    "Precio Unitario",
    "Importe",
]
EXPLOSION_OPTIONAL_HEADERS = [
    "Clasificación",
    "Requiere autorización previa",
    "Observación de clasificación",
]
EXPLOSION_CLASSIFICATIONS = {
    "NORMAL",
    "OPERATIVO",
    "EQUIPO_ESPECIAL",
    "ELECTRODOMESTICO",
}
HISTORICAL_REQUIRED_HEADERS = [
    "Proveedor",
    "Clave Insumo",
    "Descripción del insumo",
    "Unidad",
    "Precio unitario histórico",
]
HISTORICAL_OPTIONAL_HEADERS = ["RFC (opcional)", "Fecha última compra"]
HISTORICAL_TEMPLATE_HEADERS = [
    "Proveedor",
    "RFC (opcional)",
    "Clave Insumo",
    "Descripción del insumo",
    "Unidad",
    "Precio unitario histórico",
    "Fecha última compra",
]
DEFAULT_PAYMENT_METHODS = [
    ("TRANSFERENCIA", "Transferencia bancaria"),
    ("CHEQUE", "Cheque"),
    ("EFECTIVO", "Efectivo"),
    ("TARJETA DE CRÉDITO", "Tarjeta de crédito"),
    ("TARJETA DE DÉBITO", "Tarjeta de débito"),
    ("CONTADO", "Pago de contado"),
    ("CRÉDITO 30 DÍAS", "Crédito a 30 días"),
    ("CRÉDITO 60 DÍAS", "Crédito a 60 días"),
    ("CRÉDITO 90 DÍAS", "Crédito a 90 días"),
]

PURCHASE_ACCESS_MODULES = (
    "compras_dashboard",
    "explosion_insumos",
    "insumos",
    "requisiciones",
    "cotizaciones_rfq",
    "oc_compras",
    "oc_operaciones",
    "programacion_pagos",
    "pagos_proveedores",
    "proveedores",
    "reportes_compras",
    "smnc",
    "tarjetas_credito",
    "metodos_pago",
    "direcciones_entrega",
)


# ---------------------------------------------------------------------------
# Seguridad, fechas y utilidades comunes
# ---------------------------------------------------------------------------


def role_required(*roles):
    def decorator(view):
        @wraps(view)
        @login_required
        def wrapped(*args, **kwargs):
            if current_user.rol not in set(roles):
                abort(403)
            return view(*args, **kwargs)

        return wrapped

    return decorator


def require_permission(module: str, action: str = "ver") -> None:
    if not current_user.tiene_permiso(module, action):
        abort(403)


def allowed_requisition_types(action: str = "ver") -> set[str]:
    """Tipos efectivos a partir de la matriz configurable del usuario."""

    if not current_user.tiene_permiso("requisiciones", action):
        return set()
    # Una requisición normal es una solicitud, no una autorización para crear
    # una OC de Compras. Por eso el Supervisor puede enviarla a la bandeja del
    # Comprador sin recibir permiso de OC.
    allowed = {"COMPRAS"}
    # Las requisiciones operativas históricas siguen siendo consultables, pero
    # las altas nuevas se realizan exclusivamente como OC de Operaciones
    # directa, sin recrear el flujo requisición/cotización.
    if action == "ver" and current_user.tiene_permiso("oc_operaciones", "ver"):
        allowed.add("OPERACIONES")
    return allowed


def allowed_order_types(action: str = "ver") -> set[str]:
    allowed = set()
    if current_user.tiene_permiso("oc_compras", action):
        allowed.add("COMPRAS")
    if current_user.tiene_permiso("oc_operaciones", action):
        allowed.add("OPERACIONES")
    return allowed


def require_order_permission(order: PurchaseOrder, action: str = "ver") -> None:
    if order.tipo_oc not in allowed_order_types(action):
        abort(403)


def detect_operation_category(description: str | None, supply_type: str | None = None) -> str | None:
    """Sugiere una clasificación conservadora para catálogos importados.

    La bandera puede corregirse desde Catálogo de insumos; la validación del
    flujo siempre usa el campo guardado, no vuelve a decidir por texto.
    """

    text = normalize(description)
    if any(term in text for term in ("RETIRO DE ESCOMBRO", "RETIRO ESCOMBRO", "ACARREO DE ESCOMBRO")):
        return "RETIRO_ESCOMBRO"
    if any(term in text for term in ("TIERRA PARA RELLENO", "TIERRA DE RELLENO", "RELLENO CON TIERRA")):
        return "TIERRA_RELLENO"
    if "ARENA" in text:
        return "ARENA"
    if "GRAVA" in text:
        return "GRAVA"
    if "AGREGADO" in text:
        return "AGREGADOS"
    if re.search(r"\bAGUA\b", text):
        return "AGUA"
    if any(term in text for term in ("RETROEXCAVADORA", "EXCAVADORA", "BOBCAT")) and any(
        term in text for term in ("RENTA", "HORA", "HORAS")
    ):
        return "RENTA_EQUIPO"
    if normalize(supply_type) == "INDIRECTO" and any(
        term in text for term in ("OFICINA", "PAPELERIA", "PAPELERÍA")
    ):
        return "GASTO_OFICINA"
    return None


admin_required = role_required("admin")
buyer_required = role_required("admin", "comprador")
costs_required = role_required("admin", "costos")
supervisor_required = role_required("admin", "supervisor")
reports_required = role_required("admin", "costos")


@compras_bp.before_request
def enforce_purchase_module_role():
    """Exige al menos un módulo visible antes de entrar al Blueprint."""

    if current_user.is_authenticated and not any(
        current_user.tiene_permiso(module, "ver")
        for module in PURCHASE_ACCESS_MODULES
    ):
        abort(403)
    if not current_user.is_authenticated:
        return None
    endpoint = (request.endpoint or "").removeprefix("compras.")
    context = purchase_permission_context(endpoint)
    if context and not current_user.tiene_permiso(*context):
        abort(403)
    return None


def purchase_permission_context(endpoint: str) -> tuple[str, str] | None:
    """Mapea cada pantalla del sidebar a su permiso real de servidor."""

    exact = {
        "dashboard": ("compras_dashboard", "ver"),
        "suppliers_list": (
            "proveedores",
            "crear" if request.method == "POST" else "ver",
        ),
        "supplier_edit": ("proveedores", "editar"),
        "supplier_toggle": ("proveedores", "eliminar"),
        "supplier_statement": ("proveedores_sensibles", "ver"),
        "historical_import_template": ("proveedores_sensibles", "crear"),
        "historical_import": ("proveedores_sensibles", "crear"),
        "payment_methods_list": (
            "metodos_pago",
            "crear" if request.method == "POST" else "ver",
        ),
        "payment_method_toggle": ("metodos_pago", "eliminar"),
        "supplies_list": (
            "insumos",
            "crear" if request.method == "POST" else "ver",
        ),
        "supply_toggle": ("insumos", "eliminar"),
        "supply_operation_classification": ("insumos", "editar"),
        "explosion_list": ("explosion_insumos", "ver"),
        "explosion_template": ("explosion_insumos", "crear"),
        "explosion_import": ("explosion_insumos", "crear"),
        "explosion_export": ("explosion_insumos", "ver"),
        "explosion_api": ("explosion_insumos", "ver"),
        "requisitions_list": ("requisiciones", "ver"),
        "requisition_new": ("requisiciones", "crear"),
        "requisition_detail": ("requisiciones", "ver"),
        "requisition_line_add": ("requisiciones", "crear"),
        "requisition_line_delete": ("requisiciones", "editar"),
        "requisition_lines_update": ("requisiciones", "editar"),
        "requisition_cancel": ("requisiciones", "editar"),
        "requisition_delete": ("requisiciones", "editar"),
        "requisition_partidas_api": ("requisiciones", "crear"),
        "requisition_materials_api": ("requisiciones", "crear"),
        "requisition_submit": ("requisiciones", "editar"),
        "requisition_approve": ("requisiciones", "aprobar"),
        "requisition_reject": ("requisiciones", "aprobar"),
        "requisition_confirm_receipt": ("cotizaciones_rfq", "editar"),
        "quotations_list": ("cotizaciones_rfq", "ver"),
        "rfq_print": ("cotizaciones_rfq", "ver"),
        "quotations_create": ("cotizaciones_rfq", "crear"),
        "quotations_consolidate": ("cotizaciones_rfq", "crear"),
        "quotation_detail": ("cotizaciones_rfq", "ver"),
        "quotation_rfq_print": ("cotizaciones_rfq", "ver"),
        "quotation_send_email": ("cotizaciones_rfq", "editar"),
        "quotation_mark_whatsapp": ("cotizaciones_rfq", "editar"),
        "quotation_response": ("cotizaciones_rfq", "editar"),
        "quotation_compare": ("cotizaciones_rfq", "ver"),
        "quotation_select": ("cotizaciones_rfq", "aprobar"),
        "operational_order_new": ("oc_operaciones", "crear"),
        "order_validate_beneficiary": ("programacion_pagos", "aprobar"),
        "payment_schedules_list": ("programacion_pagos", "ver"),
        "payment_schedule_decide": ("programacion_pagos", "aprobar"),
        "order_authorize_advance": ("programacion_pagos", "aprobar"),
        "supplier_payments_list": ("pagos_proveedores", "ver"),
        "supplier_payment_new": ("pagos_proveedores", "pagar"),
        "advance_balances": ("programacion_pagos", "ver"),
        "advance_balance_move": ("programacion_pagos", "editar"),
        "credit_cards_list": ("tarjetas_credito", "ver"),
        "credit_card_new": ("tarjetas_credito", "crear"),
        "credit_card_edit": ("tarjetas_credito", "editar"),
        "credit_card_payment_new": ("tarjetas_credito", "pagar"),
        "credit_card_toggle": ("tarjetas_credito", "eliminar"),
        "smnc_list": ("smnc", "ver"),
        "smnc_new": ("smnc", "crear"),
        "smnc_detail": ("smnc", "ver"),
        "smnc_approve": ("smnc", "aprobar"),
        "smnc_reject": ("smnc", "aprobar"),
        "payroll_operations_report": ("reportes_compras", "ver"),
        "supplier_payments_report": ("reportes_compras", "ver"),
        "reports_index": ("reportes_compras", "ver"),
        "weekly_supplier_payments": ("reportes_compras", "ver"),
        "report_consumption": ("reportes_compras", "ver"),
        "report_smnc": ("reportes_compras", "ver"),
        "report_suppliers": ("reportes_compras", "ver"),
        "delivery_address_edit": ("direcciones_entrega", "editar"),
    }
    return exact.get(endpoint)


def today_value() -> date:
    configured = current_app.config.get("COMPRAS_TODAY")
    if isinstance(configured, date):
        return configured
    if configured:
        return datetime.strptime(str(configured), "%Y-%m-%d").date()
    return date.today()


def add_business_days(start: date, days: int) -> date:
    result = start
    remaining = days
    while remaining:
        result += timedelta(days=1)
        if result.weekday() < 5:
            remaining -= 1
    return result


def week_bounds(day: date | None = None) -> tuple[date, date]:
    current = day or today_value()
    start = current - timedelta(days=current.weekday())
    return start, start + timedelta(days=6)


def normalize(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).upper()


def normalize_email_address(value: str | None, field: str = "Correo") -> str:
    raw = (value or "").strip().lower()
    if not raw:
        raise ValueError(f"{field} es obligatorio.")
    try:
        return validate_email(raw, check_deliverability=False).normalized
    except EmailNotValidError as exc:
        raise ValueError(f"{field} no tiene un formato válido.") from exc


def slug(value: str) -> str:
    plain = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    cleaned = re.sub(r"[^A-Z0-9]+", "-", plain.upper()).strip("-")
    return cleaned[:24] or "PARTIDA"


def excel_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_decimal_value(
    value,
    field: str,
    *,
    required: bool = False,
    positive: bool = False,
    scale: Decimal = QUANTITY_STEP,
) -> Decimal:
    raw = str(value if value is not None else "").strip().replace("$", "").replace(",", "")
    if not raw:
        if required:
            raise ValueError(f"{field} es obligatorio.")
        return Decimal("0").quantize(scale)
    try:
        parsed = Decimal(raw).quantize(scale, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field} no es un número válido.") from exc
    if parsed < 0 or (positive and parsed <= 0):
        qualifier = "mayor que cero" if positive else "igual o mayor que cero"
        raise ValueError(f"{field} debe ser {qualifier}.")
    return parsed


def form_decimal(field: str, **kwargs) -> Decimal:
    return parse_decimal_value(request.form.get(field), field, **kwargs)


def form_date(field: str, *, required: bool = True) -> date | None:
    raw = (request.form.get(field) or "").strip()
    if not raw and not required:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"La fecha {field} no es válida.") from exc


def amount_without_vat(amount: Decimal, amount_type: str) -> Decimal:
    return money(amount if amount_type == "SIN_IVA" else amount / Decimal("1.16"))


def accessible_projects_query(*, include_inactive: bool = False):
    query = CentroCosto.query.filter(CentroCosto.tipo == "obra")
    if not include_inactive:
        query = query.filter(CentroCosto.estado == "activa")
    if not current_user.acceso_global_obras:
        query = query.join(usuario_centros_nomina).filter(
            usuario_centros_nomina.c.user_id == current_user.id
        )
    if current_user.rol == "supervisor":
        selected_id = obra_activa_id(
            current_user,
            incluir_inactivas=include_inactive,
        )
        query = query.filter(CentroCosto.id == (selected_id or -1))
    return query.order_by(CentroCosto.nombre)


def can_access_project(project: CentroCosto) -> bool:
    if current_user.acceso_global_obras:
        return True
    if project not in current_user.projects:
        return False
    return (
        current_user.rol != "supervisor"
        or obra_activa_id(current_user, incluir_inactivas=False) == project.id
    )


def require_project_access(project: CentroCosto) -> None:
    if project.tipo == "garantia":
        verificar_asignacion_obra(current_user, project.id)
    else:
        verificar_acceso_obra(current_user, project.id)


def project_or_403(project_id: int) -> CentroCosto:
    project = db.get_or_404(CentroCosto, project_id)
    if project.tipo != "obra":
        abort(404)
    require_project_access(project)
    return project


def requisition_project_choices() -> list[CentroCosto]:
    """Obras elegibles al crear; el Supervisor puede cambiar aquí la activa."""

    if current_user.rol == "supervisor":
        return [
            project
            for project in obras_asignadas_supervisor(
                current_user,
                incluir_inactivas=False,
            )
            if project.tipo == "obra"
        ]
    return accessible_projects_query().all()


def operational_project_or_404(project_id: int) -> CentroCosto:
    project = db.get_or_404(CentroCosto, project_id)
    if project.tipo not in {"obra", "garantia"}:
        abort(404)
    require_project_access(project)
    return project


def audit(action: str, entity: str, entity_id=None, detail: str | None = None):
    db.session.add(
        BitacoraAuditoria(
            usuario_id=current_user.id,
            accion=action,
            tabla_afectada=entity,
            registro_id=entity_id,
            detalle=detail,
        )
    )


def next_folio(model, prefix: str) -> str:
    year = today_value().year
    start = f"{prefix}-{year}-"
    last = model.query.filter(model.folio.like(f"{start}%")).order_by(model.id.desc()).first()
    sequence = 1
    if last:
        try:
            sequence = int(last.folio.rsplit("-", 1)[1]) + 1
        except (ValueError, IndexError):
            sequence = model.query.count() + 1
    return f"{start}{sequence:04d}"


def ensure_payment_methods() -> None:
    if PaymentMethod.query.first():
        return
    for name, description in DEFAULT_PAYMENT_METHODS:
        db.session.add(PaymentMethod(nombre=name, descripcion=description, activo=True))
    db.session.flush()


def resolve_payment_method() -> PaymentMethod:
    """Obtiene el método elegido o crea el capturado desde el modal de la OC."""

    raw = (request.form.get("payment_method_id") or "").strip()
    if raw.startswith("new:"):
        name = normalize(raw[4:])
        if not name:
            raise ValueError("Escribe el nombre del nuevo método de pago.")
        if len(name) > 80:
            raise ValueError("El método de pago no puede exceder 80 caracteres.")
        method = PaymentMethod.query.filter(func.upper(PaymentMethod.nombre) == name).first()
        if not method:
            method = PaymentMethod(
                nombre=name,
                descripcion="Agregado durante la captura de una OC",
                activo=True,
            )
            db.session.add(method)
            db.session.flush()
        elif not method.activo:
            method.activo = True
        return method
    method = db.session.get(PaymentMethod, int(raw)) if raw.isdigit() else None
    if not method or not method.activo:
        raise ValueError("Selecciona un método de pago activo.")
    return method


def send_purchase_email(message: Message) -> None:
    """Envía con Flask-Mail y muestra una falla de configuración comprensible."""

    if not current_app.config.get("MAIL_SUPPRESS_SEND"):
        if not current_app.config.get("MAIL_SERVER"):
            raise RuntimeError("MAIL_SERVER no está configurado")
        if not current_app.config.get("MAIL_DEFAULT_SENDER"):
            raise RuntimeError("MAIL_DEFAULT_SENDER no está configurado")
    from app import mail

    mail.send(message)


def quotation_email_message(quotation: Quotation) -> Message:
    requisition = quotation.requisition
    # La copia corresponde al comprador que preparó la cotización. Si una
    # administración hizo esa tarea de respaldo, se usa quien confirmó la
    # recepción formal de la requisición.
    buyer = (
        quotation.created_by
        if quotation.created_by and quotation.created_by.rol == "comprador"
        else requisition.buyer_received_by or quotation.created_by
    )
    cc = normalize_email_address(buyer.correo, "Correo del comprador")
    recipient = normalize_email_address(
        quotation.supplier.email, "Correo del proveedor"
    )
    project_codes = ", ".join(project.codigo for project in quotation.projects)
    subject = f"Solicitud de cotización {quotation.folio} · {project_codes}"
    body = render_template(
        "compras/emails/quotation.txt",
        quotation=quotation,
        requisition=requisition,
        buyer=buyer,
    )
    html = render_template(
        "compras/emails/quotation.html",
        quotation=quotation,
        requisition=requisition,
        buyer=buyer,
    )
    return Message(
        subject=subject,
        recipients=[recipient],
        cc=[cc],
        body=body,
        html=html,
    )


def receipt_email_message(receipt: GoodsReceipt) -> Message:
    order = receipt.order
    buyer_email = normalize_email_address(order.buyer.correo, "Correo del comprador")
    recipients = [buyer_email]
    if order.supplier and order.supplier.email:
        recipients.insert(
            0,
            normalize_email_address(order.supplier.email, "Correo del proveedor"),
        )
    subject = f"Confirmación de recepción {receipt.folio} · {order.folio}"
    body = render_template(
        "compras/emails/receipt.txt", receipt=receipt, order=order
    )
    html = render_template(
        "compras/emails/receipt.html", receipt=receipt, order=order
    )
    return Message(
        subject=subject,
        recipients=recipients,
        body=body,
        html=html,
    )


def notify(users, notification_type: str, message: str, link: str | None = None):
    seen = set()
    for user in users:
        if user and user.activo and user.id not in seen:
            seen.add(user.id)
            db.session.add(
                PurchaseNotification(
                    user_id=user.id,
                    tipo=notification_type,
                    mensaje=message[:500],
                    enlace=link,
                )
            )


def users_by_roles(*roles):
    return Usuario.query.filter(Usuario.rol.in_(roles), Usuario.activo.is_(True)).all()


def users_with_permission(module: str, action: str = "ver") -> list[Usuario]:
    return [
        user
        for user in Usuario.query.filter(Usuario.activo.is_(True)).all()
        if user.tiene_permiso(module, action)
    ]


def refresh_requisition_status(requisition: PurchaseRequisition) -> None:
    if requisition.estado not in {"APROBADA", "PARCIAL"}:
        return
    # La requisición se cierra únicamente cuando cada renglón alcanzó la
    # cantidad originalmente solicitada. Una OC parcial nunca la cierra.
    if requisition.lines and all(
        line.cantidad_pendiente_solicitada <= Decimal("0")
        for line in requisition.lines
    ):
        requisition.estado = "CERRADA"
    elif any(line.cantidad_ordenada > 0 for line in requisition.lines):
        requisition.estado = "PARCIAL"
    else:
        requisition.estado = "APROBADA"


def refresh_order_status(order: PurchaseOrder) -> None:
    if order.estado in {"BORRADOR", "CANCELADA", "PENDIENTE_ANTICIPO"}:
        return
    if order.lines and all(line.cantidad_pendiente <= 0 for line in order.lines):
        order.fecha_surtido_real = max((receipt.fecha for receipt in order.receipts), default=today_value())
        order.estado = (
            "CERRADA"
            if order.saldo_pendiente <= Decimal("0.01")
            and order.direccion_entrega_confirmada_at
            else "RECEPCION_TOTAL"
        )
        return
    if any(line.cantidad_recibida > 0 for line in order.lines):
        order.estado = "RECEPCION_PARCIAL"
        return
    if order.modalidad_pago == "CREDITO":
        order.estado = "EMITIDA"
    elif order.advance_authorizer:
        if order.monto_pagado >= decimal_value(order.anticipo_monto) > 0:
            order.estado = "ANTICIPO_PAGADO"
        elif order.monto_pagado > 0:
            order.estado = "ANTICIPO_PARCIAL"
        else:
            order.estado = "ANTICIPO_AUTORIZADO"


def run_deadline_checks() -> dict[str, int]:
    current = today_value()
    changed = False
    counts = {"requisitions": 0, "deliveries": 0}
    requisitions = PurchaseRequisition.query.filter(
        PurchaseRequisition.estado.in_({"APROBADA", "PARCIAL"}),
        PurchaseRequisition.fecha_limite_oc.isnot(None),
        PurchaseRequisition.fecha_limite_oc < current,
    ).all()
    for req in requisitions:
        if req.total_pendiente_compra <= 0:
            refresh_requisition_status(req)
            continue
        liberar_reservas_pendientes(req)
        req.estado = "VENCIDA"
        if not req.expiry_notified_at:
            req.expiry_notified_at = utc_now()
            recipients = [req.requested_by] + users_by_roles("admin")
            if req.tipo_requisicion == "COMPRAS":
                recipients.extend(users_by_roles("comprador"))
            notify(
                recipients,
                "REQUISICION_VENCIDA",
                f"La requisición {req.folio} venció sin convertirse totalmente en OC.",
                f"/compras/requisiciones/{req.id}",
            )
            counts["requisitions"] += 1
        changed = True

    orders = PurchaseOrder.query.filter(
        PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES - {"CERRADA"}),
        PurchaseOrder.fecha_entrega_estimada < current,
    ).all()
    for order in orders:
        if order.porcentaje_recepcion >= 1 or order.delivery_notified_at:
            continue
        order.delivery_notified_at = utc_now()
        recipients = (
            users_by_roles("comprador")
            if order.tipo_oc == "COMPRAS"
            else users_by_roles("admin")
        )
        recipients.extend(req.requested_by for req in order.requisitions)
        notify(
            recipients,
            "ENTREGA_VENCIDA",
            f"La fecha estimada de surtido de {order.folio} ya venció.",
            f"/compras/ordenes/{order.id}",
        )
        counts["deliveries"] += 1
        changed = True
    return counts


def run_payable_deadline_checks(current: date) -> dict[str, int]:
    """Genera avisos idempotentes de proveedores y tarjetas por vencer."""

    counts = {"payments": 0, "cards": 0}
    due_orders = PurchaseOrder.query.filter(
        PurchaseOrder.modalidad_pago == "CREDITO",
        PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
        PurchaseOrder.fecha_vencimiento.isnot(None),
        PurchaseOrder.fecha_vencimiento <= current + timedelta(days=3),
    ).all()
    for order in due_orders:
        if order.saldo_pendiente <= 0 or order.payment_due_notified_on == current:
            continue
        order.payment_due_notified_on = current
        days = (order.fecha_vencimiento - current).days
        timing = (
            f"venció hace {-days} día(s)"
            if days < 0
            else "vence hoy"
            if days == 0
            else f"vence en {days} día(s)"
        )
        module = "oc_operaciones" if order.tipo_oc == "OPERACIONES" else "compras"
        notify(
            users_with_permission(module, "ver") + [order.buyer],
            "PAGO_PROVEEDOR_POR_VENCER",
            f"{order.folio} de {order.beneficiario_nombre} {timing}; saldo {order.saldo_pendiente} MXN.",
            f"/compras/ordenes/{order.id}",
        )
        counts["payments"] += 1

    due_cards = CreditCard.query.filter(
        CreditCard.activa.is_(True),
        CreditCard.saldo_actual > 0,
        CreditCard.fecha_pago <= current + timedelta(days=3),
    ).all()
    for card in due_cards:
        if card.payment_due_notified_on == current:
            continue
        card.payment_due_notified_on = current
        days = card.dias_para_pago(current)
        timing = (
            f"venció hace {-days} día(s)"
            if days < 0
            else "vence hoy"
            if days == 0
            else f"vence en {days} día(s)"
        )
        notify(
            users_with_permission("tarjetas_credito", "ver"),
            "TARJETA_POR_PAGAR",
            f"La tarjeta {card.numero_tarjeta} de {card.empresa.nombre} {timing}; saldo {money(card.saldo_actual)} MXN.",
            "/compras/tarjetas",
        )
        counts["cards"] += 1
    return counts


def run_daily_purchase_alerts(*, force: bool = False) -> dict[str, int]:
    """Ejecuta una vez al día todos los vencimientos de Compras.

    Además de la ejecución automática al primer acceso al ERP, esta función es
    utilizada por el comando ``flask compras-alertas`` para instalaciones que
    decidan programarlo con el Programador de tareas de Windows.
    """

    current = today_value()
    previous = PurchaseAlertRun.query.filter_by(fecha=current).first()
    if previous and not force:
        # La bitácora evita repetir el barrido completo de cuentas por pagar,
        # pero durante el mismo día todavía pueden capturarse registros con una
        # fecha ya vencida (por ejemplo al importar o corregir una requisición).
        # Revalidamos esos dos vencimientos idempotentes y acumulamos cualquier
        # hallazgo nuevo en la ejecución del día.
        late_counts = run_deadline_checks()
        payable_counts = run_payable_deadline_checks(current)
        if late_counts["requisitions"] or late_counts["deliveries"]:
            previous.requisiciones_vencidas += late_counts["requisitions"]
            previous.entregas_vencidas += late_counts["deliveries"]
            previous.executed_at = utc_now()
        if payable_counts["payments"] or payable_counts["cards"]:
            previous.pagos_por_vencer += payable_counts["payments"]
            previous.tarjetas_por_vencer += payable_counts["cards"]
            previous.executed_at = utc_now()
        # También persiste el cierre de una requisición que ya quedó cubierta,
        # aunque ese cambio no genere una alerta nueva.
        db.session.commit()
        return {
            "requisitions": previous.requisiciones_vencidas,
            "deliveries": previous.entregas_vencidas,
            "payments": previous.pagos_por_vencer,
            "cards": previous.tarjetas_por_vencer,
        }

    counts = run_deadline_checks()
    counts.update(run_payable_deadline_checks(current))

    run = previous or PurchaseAlertRun(fecha=current)
    run.requisiciones_vencidas = counts["requisitions"]
    run.entregas_vencidas = counts["deliveries"]
    run.pagos_por_vencer = counts["payments"]
    run.tarjetas_por_vencer = counts["cards"]
    run.executed_at = utc_now()
    db.session.add(run)
    db.session.commit()
    return counts


@compras_bp.context_processor
def inject_helpers():
    notifications = []
    unread = 0
    if current_user.is_authenticated:
        notifications = (
            PurchaseNotification.query.filter_by(user_id=current_user.id, leida=False)
            .order_by(PurchaseNotification.created_at.desc())
            .limit(8)
            .all()
        )
        unread = PurchaseNotification.query.filter_by(user_id=current_user.id, leida=False).count()
    endpoint = request.endpoint or ""
    if endpoint == "compras.dashboard":
        back_url = url_for("auth.dashboard")
        back_label = "Volver al ERP"
    elif endpoint in {
        "compras.requisitions_list",
        "compras.orders_list",
        "compras.suppliers_list",
        "compras.supplier_payments_list",
        "compras.explosion_list",
        "compras.smnc_list",
        "compras.reports_index",
        "compras.notifications_list",
        "compras.payment_methods_list",
        "compras.supplies_list",
        "compras.credit_cards_list",
    }:
        back_url = url_for("compras.dashboard")
        back_label = "Volver al Dashboard"
    elif endpoint == "compras.advance_balances":
        back_url = url_for("compras.orders_list")
        back_label = "Volver a órdenes"
    elif endpoint.startswith("compras.requisition") or endpoint.startswith(
        "compras.quotation"
    ) or endpoint == "compras.rfq_print":
        back_url = url_for("compras.requisitions_list")
        back_label = "Volver a requisiciones"
    elif endpoint.startswith("compras.order"):
        back_url = url_for("compras.orders_list")
        back_label = "Volver a órdenes"
    elif endpoint in {
        "compras.payroll_operations_report",
        "compras.supplier_payments_report",
    }:
        back_url = url_for("compras.dashboard")
        back_label = "Volver al Dashboard"
    elif endpoint.startswith("compras.supplier_payment"):
        back_url = url_for("compras.supplier_payments_list")
        back_label = "Volver a pagos"
    elif endpoint.startswith("compras.credit_card"):
        back_url = url_for("compras.credit_cards_list")
        back_label = "Volver a tarjetas"
    elif endpoint.startswith("compras.supplier") or endpoint.startswith(
        "compras.historical"
    ):
        back_url = url_for("compras.suppliers_list")
        back_label = "Volver a proveedores"
    elif endpoint.startswith("compras.explosion"):
        back_url = url_for("compras.explosion_list")
        back_label = "Volver a explosión"
    elif endpoint.startswith("compras.smnc"):
        back_url = url_for("compras.smnc_list")
        back_label = "Volver a SMNC"
    elif endpoint.startswith("compras.report") or endpoint.startswith(
        "compras.weekly_supplier"
    ):
        back_url = url_for("compras.reports_index")
        back_label = "Volver a reportes"
    else:
        back_url = url_for("compras.dashboard")
        back_label = "Volver a Compras"
    return {
        "today": today_value(),
        "money": money,
        "purchase_notifications": notifications,
        "purchase_unread": unread,
        "type_labels": TYPE_LABELS,
        "purchase_back_url": back_url,
        "purchase_back_label": back_label,
    }


# ---------------------------------------------------------------------------
# Dashboard semanal y notificaciones
# ---------------------------------------------------------------------------


@compras_bp.get("/")
@login_required
def dashboard():
    ensure_payment_methods()
    db.session.commit()
    run_daily_purchase_alerts()
    projects = accessible_projects_query().all()
    project_ids = [project.id for project in projects]
    selected_project_id = (
        None
        if current_user.es_comprador
        else request.args.get("project_id", type=int)
    )
    if selected_project_id and selected_project_id not in project_ids:
        abort(404)
    scoped_ids = [selected_project_id] if selected_project_id else project_ids
    scoped = scoped_ids or [-1]
    current = today_value()
    week_start, week_end = week_bounds(current)
    requisition_types = allowed_requisition_types() or {"__NONE__"}
    order_types = allowed_order_types("ver") or {"__NONE__"}

    pending_requisition_query = PurchaseRequisition.query.options(
        joinedload(PurchaseRequisition.project)
    ).filter(
        PurchaseRequisition.project_id.in_(scoped),
        PurchaseRequisition.tipo_requisicion.in_(requisition_types),
    )
    if not current_user.acceso_global_obras:
        pending_requisition_query = pending_requisition_query.filter(
            PurchaseRequisition.requested_by_id == current_user.id,
            PurchaseRequisition.estado.notin_({"RECHAZADA", "CANCELADA"}),
        )
    else:
        pending_requisition_query = pending_requisition_query.filter(
            PurchaseRequisition.estado.in_({"APROBADA", "PARCIAL"})
        )
    pending_requisitions = pending_requisition_query.order_by(
        PurchaseRequisition.fecha_limite_oc, PurchaseRequisition.id
    ).all()
    for req in pending_requisitions:
        req.days_to_deadline = (req.fecha_limite_oc - current).days if req.fecha_limite_oc else None

    pending_requisition_counts = {
        project_id: count
        for project_id, count in (
            db.session.query(
                PurchaseRequisition.project_id,
                func.count(PurchaseRequisition.id),
            )
            .filter(
                PurchaseRequisition.project_id.in_(project_ids or [-1]),
                PurchaseRequisition.tipo_requisicion == "COMPRAS",
                PurchaseRequisition.estado.in_(("APROBADA", "PARCIAL")),
            )
            .group_by(PurchaseRequisition.project_id)
            .all()
        )
    }
    missing_delivery_projects = (
        [
            project
            for project in projects
            if not (project.direccion_entrega or "").strip()
        ]
        if current_user.es_comprador
        and current_user.tiene_permiso("direcciones_entrega", "editar")
        else []
    )

    weekly_orders = (
        PurchaseOrder.query.options(joinedload(PurchaseOrder.supplier), joinedload(PurchaseOrder.project))
        .filter(
            PurchaseOrder.project_id.in_(scoped),
            PurchaseOrder.tipo_oc.in_(order_types),
            PurchaseOrder.fecha_orden.between(week_start, week_end),
            PurchaseOrder.estado != "BORRADOR",
        )
        .order_by(PurchaseOrder.fecha_orden.desc(), PurchaseOrder.id.desc())
        .all()
    )
    pending_advances = (
        PurchaseOrder.query.options(joinedload(PurchaseOrder.supplier))
        .filter(
            PurchaseOrder.project_id.in_(scoped),
            PurchaseOrder.tipo_oc.in_(order_types),
            PurchaseOrder.estado == "PENDIENTE_ANTICIPO",
        )
        .order_by(PurchaseOrder.fecha_orden)
        .all()
    )
    pending_receipts = (
        PurchaseOrder.query.options(joinedload(PurchaseOrder.supplier), joinedload(PurchaseOrder.project))
        .filter(
            PurchaseOrder.project_id.in_(scoped),
            PurchaseOrder.tipo_oc.in_(order_types),
            PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES - {"CERRADA", "RECEPCION_TOTAL"}),
        )
        .order_by(PurchaseOrder.fecha_entrega_estimada)
        .all()
    )
    for order in pending_receipts:
        order.delivery_days = (order.fecha_entrega_estimada - current).days
        order.delivery_color = "danger" if order.delivery_days < 0 else "warning" if order.delivery_days <= 1 else "success"

    explosion_total = money(
        db.session.query(func.coalesce(func.sum(BudgetExplosionItem.importe_presupuestado), 0))
        .filter(BudgetExplosionItem.project_id.in_(scoped), BudgetExplosionItem.activo.is_(True))
        .scalar()
    )
    center_total = money(
        db.session.query(func.coalesce(func.sum(CentroCosto.presupuesto_total), 0))
        .filter(CentroCosto.id.in_(scoped))
        .scalar()
    )
    budget_total = center_total if center_total > 0 else explosion_total
    active_requisitions = PurchaseRequisition.query.options(
        selectinload(PurchaseRequisition.lines).joinedload(
            PurchaseRequisitionLine.explosion_item
        ),
        selectinload(PurchaseRequisition.lines)
        .selectinload(PurchaseRequisitionLine.order_lines)
        .joinedload(PurchaseOrderLine.order),
    ).filter(
        PurchaseRequisition.project_id.in_(scoped),
        PurchaseRequisition.tipo_requisicion.in_(requisition_types),
        PurchaseRequisition.estado.in_({"APROBADA", "PARCIAL"}),
    ).all()
    pending_approved_total = money(
        sum(
            (
                line.importe_pendiente_compra
                for requisition in active_requisitions
                for line in requisition.lines
                if line.estado_linea == "APROBADA"
            ),
            Decimal("0"),
        )
    )
    active_orders = PurchaseOrder.query.options(
        selectinload(PurchaseOrder.lines).selectinload(
            PurchaseOrderLine.receipt_lines
        )
    ).filter(
        PurchaseOrder.project_id.in_(scoped),
        PurchaseOrder.tipo_oc.in_(order_types),
        PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
    ).all()
    consumed_total = money(
        sum((order.monto_consumido_real for order in active_orders), Decimal("0"))
    )
    committed_total = money(
        pending_approved_total
        + sum((order.saldo_comprometido for order in active_orders), Decimal("0"))
    )
    available_total = money(budget_total - committed_total - consumed_total)
    bought_total = money(
        db.session.query(func.coalesce(func.sum(PurchaseOrderLine.importe_sin_iva), 0))
        .join(PurchaseOrder)
        .filter(
            PurchaseOrder.project_id.in_(scoped),
            PurchaseOrder.tipo_oc.in_(order_types),
            PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
        )
        .scalar()
    )
    utilization = Decimal("0") if budget_total <= 0 else (committed_total + consumed_total) / budget_total
    budget_alert = "danger" if utilization >= 1 else "warning" if utilization >= Decimal("0.80") else "success"

    weekly_lines = (
        PurchaseOrderLine.query.join(PurchaseOrder)
        .filter(
            PurchaseOrder.project_id.in_(scoped),
            PurchaseOrder.tipo_oc.in_(order_types),
            PurchaseOrder.fecha_orden.between(week_start, week_end),
            PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
        )
        .all()
    )
    quantity_week = sum((decimal_value(line.cantidad) for line in weekly_lines), Decimal("0"))
    real_week = money(sum((line.importe_sin_iva for line in weekly_lines), Decimal("0")))
    budget_week = money(
        sum(
            (
                decimal_value(line.cantidad) * decimal_value(line.explosion_item.precio_unitario_sin_iva)
                for line in weekly_lines
            ),
            Decimal("0"),
        )
    )
    selected_quote_ids = [order.quotation_id for order in weekly_orders if order.quotation_id]
    quoted_week = money(
        db.session.query(func.coalesce(func.sum(QuotationLine.importe_cotizado), 0))
        .filter(QuotationLine.quotation_id.in_(selected_quote_ids or [-1]))
        .scalar()
    )
    comparison = {
        "presupuestado": budget_week,
        "cotizado": quoted_week,
        "real": real_week,
        "precio_presupuestado": money(budget_week / quantity_week) if quantity_week else Decimal("0.00"),
        "precio_cotizado": money(quoted_week / quantity_week) if quantity_week else Decimal("0.00"),
        "precio_real": money(real_week / quantity_week) if quantity_week else Decimal("0.00"),
    }
    top_rows = (
        db.session.query(
            SupplyItem.clave,
            SupplyItem.descripcion,
            func.sum(PurchaseOrderLine.importe_sin_iva).label("importe"),
        )
        .join(BudgetExplosionItem, BudgetExplosionItem.supply_item_id == SupplyItem.id)
        .join(PurchaseOrderLine, PurchaseOrderLine.explosion_item_id == BudgetExplosionItem.id)
        .join(PurchaseOrder)
        .filter(
            PurchaseOrder.project_id.in_(scoped),
            PurchaseOrder.tipo_oc.in_(order_types),
            PurchaseOrder.fecha_orden.between(week_start, week_end),
            PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
        )
        .group_by(SupplyItem.id, SupplyItem.clave, SupplyItem.descripcion)
        .order_by(func.sum(PurchaseOrderLine.importe_sin_iva).desc())
        .limit(5)
        .all()
    )

    credit_suppliers = Supplier.query.filter(Supplier.activo.is_(True), Supplier.tiene_credito.is_(True)).order_by(Supplier.nombre).all()
    account_orders = (
        PurchaseOrder.query.options(joinedload(PurchaseOrder.supplier), joinedload(PurchaseOrder.project))
        .filter(
            PurchaseOrder.project_id.in_(scoped),
            PurchaseOrder.tipo_oc.in_(order_types),
            PurchaseOrder.modalidad_pago == "CREDITO",
            PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
        )
        .order_by(PurchaseOrder.fecha_vencimiento, PurchaseOrder.id)
        .all()
    )
    account_orders = [order for order in account_orders if order.saldo_pendiente > 0]
    for order in account_orders:
        order.due_color = order.semaforo_vencimiento(current)
        order.due_days = (order.fecha_vencimiento - current).days if order.fecha_vencimiento else None

    card_alerts = []
    if current_user.tiene_permiso("tarjetas_credito", "ver"):
        card_alerts = (
            CreditCard.query.options(joinedload(CreditCard.empresa))
            .filter(
                CreditCard.activa.is_(True),
                CreditCard.saldo_actual > 0,
                CreditCard.fecha_pago <= current + timedelta(days=3),
            )
            .order_by(CreditCard.fecha_pago, CreditCard.id)
            .all()
        )
        for card in card_alerts:
            card.due_days = card.dias_para_pago(current)

    return render_template(
        "compras/dashboard.html",
        projects=projects,
        selected_project_id=selected_project_id,
        week_start=week_start,
        week_end=week_end,
        pending_requisitions=pending_requisitions,
        weekly_orders=weekly_orders,
        pending_advances=pending_advances,
        pending_receipts=pending_receipts,
        budget_total=budget_total,
        committed_total=committed_total,
        consumed_total=consumed_total,
        available_total=available_total,
        bought_total=bought_total,
        budget_alert=budget_alert,
        utilization=utilization,
        comparison=comparison,
        top_rows=top_rows,
        credit_suppliers=credit_suppliers,
        account_orders=account_orders,
        card_alerts=card_alerts,
        operations_only=order_types == {"OPERACIONES"},
        pending_requisition_counts=pending_requisition_counts,
        missing_delivery_projects=missing_delivery_projects,
    )


@compras_bp.route(
    "/obras/<int:project_id>/direccion-entrega",
    methods=["GET", "POST"],
)
@permission_required("direcciones_entrega", "editar")
def delivery_address_edit(project_id):
    project = project_or_403(project_id)
    if request.method == "POST":
        address = " ".join(
            (request.form.get("direccion_entrega") or "").split()
        )
        if len(address) < 10:
            flash(
                "Captura una dirección de entrega completa.",
                "danger",
            )
        elif len(address) > 500:
            flash(
                "La dirección no puede superar 500 caracteres.",
                "danger",
            )
        else:
            project.direccion_entrega = address
            audit(
                "EDITAR_DIRECCION_ENTREGA",
                "CENTRO_COSTO",
                project.id,
                address,
            )
            db.session.commit()
            flash(
                f"Dirección de entrega actualizada para {project.codigo}.",
                "success",
            )
            return redirect(url_for("compras.dashboard"))
    return render_template(
        "compras/projects/delivery_address.html",
        project=project,
    )


@compras_bp.get("/notificaciones")
@login_required
def notifications_list():
    notifications = PurchaseNotification.query.filter_by(user_id=current_user.id).order_by(PurchaseNotification.created_at.desc()).all()
    return render_template("compras/notifications/list.html", notifications=notifications)


@compras_bp.post("/notificaciones/<int:notification_id>/leer")
@login_required
def notification_read(notification_id):
    notification = db.get_or_404(PurchaseNotification, notification_id)
    if notification.user_id != current_user.id:
        abort(403)
    notification.leida = True
    db.session.commit()
    return redirect(notification.enlace or url_for("compras.notifications_list"))


# ---------------------------------------------------------------------------
# Proveedores, métodos e insumos
# ---------------------------------------------------------------------------


def fill_supplier(supplier: Supplier) -> None:
    sensitive = current_user.tiene_permiso("proveedores_sensibles", "editar")
    supplier.codigo = normalize(request.form.get("codigo"))
    supplier.nombre = normalize(request.form.get("nombre"))
    supplier.contacto = (request.form.get("contacto") or "").strip() or None
    supplier.telefono = (request.form.get("telefono") or "").strip() or None
    supplier.email = normalize_email_address(
        request.form.get("email"), "Correo del proveedor"
    )
    supplier.moneda = "MXN"
    if sensitive:
        supplier.rfc = normalize(request.form.get("rfc")) or None
        supplier.company_id = request.form.get("company_id", type=int)
        supplier.tiene_credito = request.form.get("tiene_credito") == "on"
        supplier.limite_credito = form_decimal(
            "limite_credito", scale=MONEY_STEP
        )
        supplier.dias_credito = request.form.get("dias_credito", type=int) or 0
        supplier.notas = (request.form.get("notas") or "").strip() or None
    elif supplier.id is None:
        # Un alta básica nunca puede introducir datos fiscales o financieros.
        # En una edición limitada se preservan los valores sensibles existentes
        # en lugar de borrarlos por el simple hecho de no mostrarlos.
        supplier.rfc = None
        supplier.company_id = None
        supplier.tiene_credito = False
        supplier.limite_credito = Decimal("0")
        supplier.dias_credito = 0
        supplier.notas = None
    if not supplier.codigo or not supplier.nombre:
        raise ValueError("Código y nombre del proveedor son obligatorios.")
    if supplier.tiene_credito and (supplier.limite_credito <= 0 or supplier.dias_credito <= 0):
        raise ValueError("Captura un monto de línea y días de crédito mayores que cero.")
    if not supplier.tiene_credito:
        supplier.limite_credito = Decimal("0")
        supplier.dias_credito = 0


@compras_bp.route("/proveedores", methods=["GET", "POST"])
@permission_required("proveedores", "ver")
def suppliers_list():
    if request.method == "POST":
        require_permission("proveedores", "crear")
        try:
            supplier = Supplier(activo=True)
            fill_supplier(supplier)
            db.session.add(supplier)
            db.session.flush()
            audit("CREAR", "PROVEEDOR_COMPRAS", supplier.id, supplier.etiqueta)
            db.session.commit()
            flash("Proveedor registrado con sus condiciones de crédito.", "success")
            return redirect(url_for("compras.suppliers_list"))
        except IntegrityError:
            db.session.rollback()
            flash("Ya existe un proveedor con ese código o nombre.", "danger")
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    can_view_sensitive = current_user.tiene_permiso(
        "proveedores_sensibles", "ver"
    )
    company_id = (
        request.args.get("company_id", type=int) if can_view_sensitive else None
    )
    credit_state = (
        normalize(request.args.get("estado_credito"))
        if can_view_sensitive
        else ""
    )
    name_filter = (request.args.get("nombre") or "").strip()
    query = Supplier.query.options(selectinload(Supplier.orders))
    if company_id:
        query = query.filter(Supplier.company_id == company_id)
    if name_filter:
        query = query.filter(Supplier.nombre.ilike(f"%{name_filter}%"))
    suppliers = query.order_by(Supplier.activo.desc(), Supplier.nombre).all()
    current = today_value()
    for supplier in suppliers:
        supplier.credit_state = supplier.estado_credito(current)
    if credit_state in {"ACTIVO", "VENCIDO"}:
        suppliers = [
            supplier for supplier in suppliers if supplier.credit_state == credit_state
        ]
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    return render_template(
        "compras/suppliers/list.html",
        suppliers=suppliers,
        companies=companies,
        selected_company_id=company_id,
        selected_credit_state=credit_state,
        name_filter=name_filter,
        can_view_sensitive=can_view_sensitive,
    )


@compras_bp.route("/proveedores/<int:supplier_id>/editar", methods=["GET", "POST"])
@permission_required("proveedores", "editar")
def supplier_edit(supplier_id):
    supplier = db.get_or_404(Supplier, supplier_id)
    if request.method == "POST":
        try:
            fill_supplier(supplier)
            audit("EDITAR", "PROVEEDOR_COMPRAS", supplier.id, supplier.etiqueta)
            db.session.commit()
            flash("Proveedor actualizado.", "success")
            return redirect(url_for("compras.suppliers_list"))
        except (ValueError, IntegrityError) as exc:
            db.session.rollback()
            flash(str(exc) if isinstance(exc, ValueError) else "Código o nombre duplicado.", "danger")
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    return render_template(
        "compras/suppliers/form.html",
        supplier=supplier,
        companies=companies,
        can_view_sensitive=current_user.tiene_permiso(
            "proveedores_sensibles", "ver"
        ),
    )


@compras_bp.post("/proveedores/<int:supplier_id>/estado")
@permission_required("proveedores", "eliminar")
def supplier_toggle(supplier_id):
    supplier = db.get_or_404(Supplier, supplier_id)
    supplier.activo = not supplier.activo
    audit("ACTIVAR" if supplier.activo else "DESACTIVAR", "PROVEEDOR_COMPRAS", supplier.id, supplier.nombre)
    db.session.commit()
    flash("Estado del proveedor actualizado.", "success")
    return redirect(url_for("compras.suppliers_list"))


def unique_supplier_code(name: str, rfc: str | None = None) -> str:
    """Genera una clave estable para proveedores creados por importación."""

    base = f"HIST-{slug(rfc or name)}"[:40]
    candidate = base
    sequence = 2
    while Supplier.query.filter(func.upper(Supplier.codigo) == candidate).first():
        suffix = f"-{sequence}"
        candidate = f"{base[:40-len(suffix)]}{suffix}"
        sequence += 1
    return candidate


def catalog_project() -> CentroCosto:
    """Devuelve el contenedor inactivo del catálogo histórico general."""

    project = CentroCosto.query.filter(
        func.upper(CentroCosto.codigo) == "CAT-GENERAL"
    ).first()
    if project:
        return project
    project = CentroCosto(
        codigo="CAT-GENERAL",
        nombre="Catálogo general",
        tipo="obra",
        estado="cerrada",
        fecha_apertura=today_value(),
        fecha_cierre=today_value(),
        presupuesto_total=0,
        presupuesto_mano_obra=0,
        descripcion=(
            "Contenedor técnico de insumos históricos. No participa en "
            "requisiciones ni en el presupuesto de obras activas."
        ),
    )
    db.session.add(project)
    db.session.flush()
    return project


def historical_supplier_match(name: str, rfc: str | None) -> Supplier | None:
    if rfc:
        supplier = Supplier.query.filter(func.upper(Supplier.rfc) == rfc).first()
        if supplier:
            return supplier
    return Supplier.query.filter(func.upper(Supplier.nombre) == name).first()


def parse_optional_excel_date(value, field: str) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"{field} debe usar AAAA-MM-DD o DD/MM/AAAA.")


@compras_bp.get("/historico/plantilla.xlsx")
@permission_required("proveedores", "crear")
def historical_import_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Proveedores e insumos"
    sheet.append(HISTORICAL_TEMPLATE_HEADERS)
    sheet.append(
        [
            "Aceros del Norte",
            "ACN010101AA1",
            "MAT-001",
            'Acero de refuerzo 1/2"',
            "kg",
            35.50,
            today_value(),
        ]
    )
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, width in enumerate([28, 19, 18, 42, 14, 24, 20], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet["G2"].number_format = "yyyy-mm-dd"
    notes = workbook.create_sheet("Instrucciones")
    instructions = [
        ["Regla", "Detalle"],
        ["Obligatorias", ", ".join(HISTORICAL_REQUIRED_HEADERS)],
        ["RFC", "Opcional; se usa para identificar un proveedor ya existente."],
        ["Fecha", "Opcional. Formato AAAA-MM-DD o DD/MM/AAAA."],
        ["Importes", "Precios unitarios sin IVA y en MXN."],
        ["Presupuesto", "La importación histórica no suma cantidades al presupuesto de la obra."],
    ]
    for row in instructions:
        notes.append(row)
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 95
    for cell in notes[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_proveedores_insumos_historicos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@compras_bp.route("/historico/importar", methods=["GET", "POST"])
@permission_required("proveedores", "crear")
def historical_import():
    projects = accessible_projects_query().all()
    result = None
    if request.method == "POST":
        file = request.files.get("archivo")
        project_id = request.form.get("project_id", type=int)
        project = db.session.get(CentroCosto, project_id) if project_id else None
        if project and (project.tipo != "obra" or project.id not in {p.id for p in projects}):
            abort(404)
        if not file or not file.filename.lower().endswith(".xlsx"):
            flash("Selecciona un archivo Excel .xlsx.", "danger")
        else:
            try:
                workbook = load_workbook(file, data_only=True)
                sheet = workbook.active
                headers = [excel_text(cell.value) for cell in sheet[1]]
                missing = [header for header in HISTORICAL_REQUIRED_HEADERS if header not in headers]
                if missing:
                    raise ValueError(
                        "Faltan columnas obligatorias: " + ", ".join(missing) + "."
                    )
                indexes = {header: index for index, header in enumerate(headers) if header}
                destination = project or catalog_project()
                result = {
                    "processed": 0,
                    "suppliers_created": 0,
                    "supplies_created": 0,
                    "relations_created": 0,
                    "relations_updated": 0,
                    "errors": [],
                    "project": destination,
                }
                for row_number, values in enumerate(
                    sheet.iter_rows(min_row=2, values_only=True), start=2
                ):
                    if not any(value not in (None, "") for value in values):
                        continue
                    savepoint = db.session.begin_nested()
                    try:
                        supplier_created = False
                        supply_created = False
                        relation_created = False
                        def value_for(header):
                            index = indexes.get(header)
                            return values[index] if index is not None and index < len(values) else None

                        supplier_name = normalize(excel_text(value_for("Proveedor")))
                        rfc = normalize(excel_text(value_for("RFC (opcional)"))) or None
                        supply_key = normalize(excel_text(value_for("Clave Insumo")))
                        description = normalize(excel_text(value_for("Descripción del insumo")))
                        unit = normalize(excel_text(value_for("Unidad")))
                        price = parse_decimal_value(
                            value_for("Precio unitario histórico"),
                            "Precio unitario histórico",
                            required=True,
                            positive=True,
                            scale=QUANTITY_STEP,
                        )
                        last_purchase = parse_optional_excel_date(
                            value_for("Fecha última compra"), "Fecha última compra"
                        )
                        if not supplier_name or not supply_key or not description or not unit:
                            raise ValueError(
                                "Proveedor, clave, descripción y unidad son obligatorios."
                            )

                        supplier = historical_supplier_match(supplier_name, rfc)
                        if not supplier:
                            supplier = Supplier(
                                codigo=unique_supplier_code(supplier_name, rfc),
                                nombre=supplier_name,
                                rfc=rfc,
                                tiene_credito=False,
                                limite_credito=0,
                                dias_credito=0,
                                moneda="MXN",
                                activo=True,
                            )
                            db.session.add(supplier)
                            db.session.flush()
                            supplier_created = True
                        elif rfc and not supplier.rfc:
                            supplier.rfc = rfc

                        supply = SupplyItem.query.filter(
                            func.upper(SupplyItem.clave) == supply_key
                        ).first()
                        if not supply:
                            operation_category = detect_operation_category(
                                description, "MATERIAL"
                            )
                            supply = SupplyItem(
                                clave=supply_key,
                                descripcion=description,
                                tipo="MATERIAL",
                                unidad=unit,
                                clave_sat="00000000",
                                moneda="MXN",
                                precio_variable=True,
                                es_operacion=bool(operation_category),
                                categoria_operacion=operation_category,
                                activo=True,
                            )
                            db.session.add(supply)
                            db.session.flush()
                            supply_created = True
                        elif supply.unidad != unit:
                            raise ValueError(
                                f"La clave {supply_key} ya existe con unidad {supply.unidad}."
                            )

                        default_item = find_or_create_budget_path(
                            destination,
                            "Materiales importados",
                            "",
                            "ADICIONAL",
                        )
                        catalog_entry = SupplyProjectCatalog.query.filter_by(
                            project_id=destination.id,
                            supply_item_id=supply.id,
                        ).first()
                        if not catalog_entry:
                            catalog_entry = SupplyProjectCatalog(
                                project_id=destination.id,
                                budget_item_id=default_item.id,
                                supply_item_id=supply.id,
                                created_by_id=current_user.id,
                            )
                            db.session.add(catalog_entry)

                        relation = SupplierSupplyItem.query.filter_by(
                            supplier_id=supplier.id, supply_item_id=supply.id
                        ).first()
                        if relation:
                            relation.precio_historico = price
                            if last_purchase:
                                relation.fecha_ultima_compra = last_purchase
                            relation.origen = "IMPORTACION"
                        else:
                            relation = SupplierSupplyItem(
                                supplier_id=supplier.id,
                                supply_item_id=supply.id,
                                precio_historico=price,
                                fecha_ultima_compra=last_purchase,
                                origen="IMPORTACION",
                            )
                            db.session.add(relation)
                            relation_created = True
                        db.session.flush()
                        savepoint.commit()
                        result["suppliers_created"] += int(supplier_created)
                        result["supplies_created"] += int(supply_created)
                        if relation_created:
                            result["relations_created"] += 1
                        else:
                            result["relations_updated"] += 1
                        result["processed"] += 1
                    except (ValueError, IntegrityError) as exc:
                        savepoint.rollback()
                        result["errors"].append(f"Fila {row_number}: {exc}")

                audit(
                    "IMPORTAR",
                    "HISTORICO_PROVEEDOR_INSUMO",
                    destination.id,
                    f"{result['processed']} filas procesadas",
                )
                db.session.commit()
                flash(
                    f"Importación terminada: {result['suppliers_created']} proveedores nuevos, "
                    f"{result['supplies_created']} insumos nuevos, "
                    f"{result['relations_created']} relaciones creadas y "
                    f"{result['relations_updated']} actualizadas.",
                    "success" if not result["errors"] else "warning",
                )
            except (ValueError, OSError) as exc:
                db.session.rollback()
                flash(str(exc), "danger")
    return render_template(
        "compras/imports/historical.html",
        projects=projects,
        selected_project_id=request.form.get("project_id", type=int),
        result=result,
    )


@compras_bp.route("/metodos-pago", methods=["GET", "POST"])
@permission_required("compras", "editar")
def payment_methods_list():
    ensure_payment_methods()
    if request.method == "POST":
        name = normalize(request.form.get("nombre"))
        if not name:
            flash("El nombre del método es obligatorio.", "danger")
        elif PaymentMethod.query.filter(func.upper(PaymentMethod.nombre) == name).first():
            flash("Ese método ya existe.", "danger")
        else:
            method = PaymentMethod(nombre=name, descripcion=(request.form.get("descripcion") or "").strip() or None)
            db.session.add(method)
            db.session.flush()
            audit("CREAR", "METODO_PAGO", method.id, name)
            db.session.commit()
            flash("Método de pago agregado.", "success")
            return redirect(url_for("compras.payment_methods_list"))
    db.session.commit()
    methods = PaymentMethod.query.order_by(PaymentMethod.activo.desc(), PaymentMethod.nombre).all()
    return render_template("compras/settings/payment_methods.html", methods=methods)


@compras_bp.post("/metodos-pago/<int:method_id>/estado")
@permission_required("compras", "editar")
def payment_method_toggle(method_id):
    method = db.get_or_404(PaymentMethod, method_id)
    method.activo = not method.activo
    audit("ACTIVAR" if method.activo else "DESACTIVAR", "METODO_PAGO", method.id, method.nombre)
    db.session.commit()
    return redirect(url_for("compras.payment_methods_list"))


@compras_bp.route("/insumos", methods=["GET", "POST"])
@permission_required("compras", "ver")
def supplies_list():
    if request.method == "POST":
        require_permission("compras", "crear")
        try:
            operation_category = normalize(request.form.get("categoria_operacion"))
            is_operation = request.form.get("es_operacion") == "on"
            if is_operation and operation_category not in OPERATION_CATEGORIES:
                raise ValueError("Selecciona una categoría operativa válida.")
            supply = SupplyItem(
                clave=normalize(request.form.get("clave")),
                descripcion=normalize(request.form.get("descripcion")),
                tipo=normalize(request.form.get("tipo")),
                unidad=normalize(request.form.get("unidad")),
                clave_sat=normalize(request.form.get("clave_sat")) or "00000000",
                moneda="MXN",
                precio_variable=request.form.get("precio_variable") == "on",
                es_operacion=is_operation,
                categoria_operacion=operation_category if is_operation else None,
            )
            if not supply.clave or not supply.descripcion or not supply.unidad:
                raise ValueError("Clave, descripción y unidad son obligatorias.")
            if supply.tipo not in SUPPLY_TYPES:
                raise ValueError("Selecciona un tipo de insumo válido.")
            db.session.add(supply)
            db.session.flush()
            audit("CREAR", "INSUMO", supply.id, supply.etiqueta)
            db.session.commit()
            flash("Insumo registrado.", "success")
            return redirect(url_for("compras.supplies_list"))
        except IntegrityError:
            db.session.rollback()
            flash("Ya existe un insumo con esa clave.", "danger")
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    supplies = SupplyItem.query.order_by(SupplyItem.activo.desc(), SupplyItem.clave).all()
    return render_template(
        "compras/supplies/list.html",
        supplies=supplies,
        operation_categories=OPERATION_CATEGORIES,
    )


@compras_bp.post("/insumos/<int:supply_id>/estado")
@permission_required("compras", "eliminar")
def supply_toggle(supply_id):
    supply = db.get_or_404(SupplyItem, supply_id)
    supply.activo = not supply.activo
    audit("ACTIVAR" if supply.activo else "DESACTIVAR", "INSUMO", supply.id, supply.etiqueta)
    db.session.commit()
    return redirect(url_for("compras.supplies_list"))


@compras_bp.post("/insumos/<int:supply_id>/clasificacion-operativa")
@permission_required("compras", "editar")
def supply_operation_classification(supply_id):
    supply = db.get_or_404(SupplyItem, supply_id)
    category = normalize(request.form.get("categoria_operacion"))
    is_operation = request.form.get("es_operacion") == "on"
    if is_operation and category not in OPERATION_CATEGORIES:
        flash("Selecciona una categoría operativa válida.", "danger")
    else:
        supply.es_operacion = is_operation
        supply.categoria_operacion = category if is_operation else None
        audit(
            "CLASIFICAR_OPERACION",
            "INSUMO",
            supply.id,
            supply.categoria_operacion or "NO_OPERATIVO",
        )
        db.session.commit()
        flash("Clasificación operativa actualizada.", "success")
    return redirect(url_for("compras.supplies_list"))


# ---------------------------------------------------------------------------
# Explosión de insumos
# ---------------------------------------------------------------------------


def unique_budget_code(project_id: int, base: str) -> str:
    candidate = base[:40]
    sequence = 2
    while BudgetItem.query.filter_by(project_id=project_id, codigo=candidate).first():
        suffix = f"-{sequence}"
        candidate = f"{base[:40-len(suffix)]}{suffix}"
        sequence += 1
    return candidate


def find_or_create_budget_path(project: CentroCosto, partida: str, subpartida: str, category: str):
    parent = BudgetItem.query.filter(
        BudgetItem.project_id == project.id,
        BudgetItem.parent_id.is_(None),
        func.upper(BudgetItem.nombre) == normalize(partida),
    ).first()
    if not parent:
        parent = BudgetItem(
            project_id=project.id,
            codigo=unique_budget_code(project.id, f"P-{slug(partida)}"),
            nombre=normalize(partida),
            categoria=category,
            presupuesto=0,
            activa=True,
        )
        db.session.add(parent)
        db.session.flush()
    if not subpartida:
        return parent
    child = BudgetItem.query.filter(
        BudgetItem.project_id == project.id,
        BudgetItem.parent_id == parent.id,
        func.upper(BudgetItem.nombre) == normalize(subpartida),
    ).first()
    if not child:
        child = BudgetItem(
            project_id=project.id,
            parent_id=parent.id,
            codigo=unique_budget_code(project.id, f"{parent.codigo}-S-{slug(subpartida)}"),
            nombre=normalize(subpartida),
            categoria=category,
            presupuesto=0,
            activa=True,
        )
        db.session.add(child)
        db.session.flush()
    return child


@compras_bp.get("/explosion")
@permission_required("compras", "ver")
def explosion_list():
    projects = accessible_projects_query().all()
    project_ids = [p.id for p in projects]
    project_id = request.args.get("project_id", type=int)
    if project_id and project_id not in project_ids:
        abort(404)
    selected_ids = [
        entry.id
        for scoped_project_id in project_ids
        if not project_id or scoped_project_id == project_id
        for entry in items_explosion_vigente(scoped_project_id)
    ]
    query = BudgetExplosionItem.query.options(
        joinedload(BudgetExplosionItem.project),
        joinedload(BudgetExplosionItem.budget_item),
        joinedload(BudgetExplosionItem.supply_item),
        selectinload(BudgetExplosionItem.order_lines).joinedload(PurchaseOrderLine.order),
        selectinload(BudgetExplosionItem.order_lines).selectinload(PurchaseOrderLine.receipt_lines),
        selectinload(BudgetExplosionItem.requisition_lines).joinedload(PurchaseRequisitionLine.requisition),
    ).filter(BudgetExplosionItem.id.in_(selected_ids or [-1]))
    entries = query.order_by(BudgetExplosionItem.project_id, BudgetExplosionItem.budget_item_id, BudgetExplosionItem.id).all()
    totals = {
        "presupuestado": money(sum((decimal_value(e.importe_presupuestado) for e in entries), Decimal("0"))),
        "comprometido": money(sum((e.cantidad_aprobada_pendiente * decimal_value(e.precio_unitario_sin_iva) for e in entries), Decimal("0"))),
        "ordenado": money(sum((e.importe_ordenado for e in entries), Decimal("0"))),
        "pagado": money(sum((e.importe_pagado for e in entries), Decimal("0"))),
        "disponible": money(sum((e.importe_disponible for e in entries), Decimal("0"))),
    }
    revisions = {
        item.id: revision_explosion_vigente(item.id) for item in projects
    }
    return render_template(
        "compras/explosion/list.html",
        projects=projects,
        selected_project_id=project_id,
        entries=entries,
        totals=totals,
        revisions=revisions,
    )


@compras_bp.get("/explosion/plantilla.xlsx")
@permission_required("compras", "crear")
def explosion_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Explosión de insumos"
    sheet.append(EXPLOSION_HEADERS + EXPLOSION_OPTIONAL_HEADERS)
    sheet.append(
        [
            "Cimentación",
            "Concreto",
            "Material",
            "MAT-001",
            "Concreto f'c=250 kg/cm2",
            "m3",
            50,
            2500,
            125000,
            "NORMAL",
            "No",
            "",
        ]
    )
    sheet.append(
        [
            "Equipamiento",
            "Cocina",
            "Equipo",
            "EQ-001",
            "Electrodoméstico especial",
            "pza",
            1,
            30000,
            30000,
            "ELECTRODOMESTICO",
            "Sí",
            "Requiere autorización previa de Costos.",
        ]
    )
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for letter, width in zip(
        "ABCDEFGHIJKL",
        [24, 24, 18, 18, 42, 12, 14, 18, 18, 24, 22, 42],
    ):
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"
    notes = workbook.create_sheet("Instrucciones")
    for row in [
        ["Regla", "Detalle"],
        ["Encabezados", "Conserva el orden de los encabezados de la primera hoja."],
        ["Tipos", "Material, Mano de Obra, Subcontrato, Equipo o Indirecto."],
        [
            "Clasificación",
            "NORMAL, OPERATIVO, EQUIPO_ESPECIAL o ELECTRODOMESTICO.",
        ],
        [
            "Autorización",
            "Captura Sí únicamente en conceptos que requieren autorización previa.",
        ],
        ["Partidas", "Si no existen, el ERP las crea automáticamente."],
        ["Importes", "Cantidad × Precio Unitario debe coincidir con Importe."],
    ]:
        notes.append(row)
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 90
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="plantilla_explosion_insumos.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@compras_bp.route("/explosion/importar", methods=["GET", "POST"])
@permission_required("compras", "crear")
def explosion_import():
    projects = accessible_projects_query().all()
    result = None
    if request.method == "POST":
        file = request.files.get("archivo")
        project_id = request.form.get("project_id", type=int)
        project = db.session.get(CentroCosto, project_id) if project_id else None
        if not project or project.tipo != "obra":
            flash("Selecciona la obra de la explosión.", "danger")
        elif not file or not file.filename.lower().endswith(".xlsx"):
            flash("Selecciona un archivo Excel .xlsx.", "danger")
        else:
            try:
                workbook = load_workbook(file, data_only=True)
                sheet = workbook.active
                headers = [excel_text(cell.value) for cell in sheet[1]]
                missing = [header for header in EXPLOSION_HEADERS if header not in headers]
                accepted = EXPLOSION_HEADERS + EXPLOSION_OPTIONAL_HEADERS
                extras = [header for header in headers if header and header not in accepted]
                if missing:
                    raise ValueError("Faltan columnas obligatorias: " + ", ".join(missing) + ".")
                if extras or headers[: len(EXPLOSION_HEADERS)] != EXPLOSION_HEADERS:
                    raise ValueError("Los encabezados obligatorios deben conservar su orden.")
                optional_present = [
                    header for header in EXPLOSION_OPTIONAL_HEADERS if header in headers
                ]
                if optional_present and headers[
                    len(EXPLOSION_HEADERS):
                    len(EXPLOSION_HEADERS) + len(optional_present)
                ] != optional_present:
                    raise ValueError(
                        "Las columnas opcionales de clasificación deben conservar su orden."
                    )
                indexes = {header: index for index, header in enumerate(headers)}
                result = {"processed": 0, "created": 0, "updated": 0, "errors": []}
                previous_entries = {
                    (entry.budget_item_id, entry.supply_item_id)
                    for entry in items_explosion_vigente(project.id)
                }
                revision = ExplosionRevision(
                    project_id=project.id,
                    numero_revision=siguiente_revision_explosion(project.id),
                    estado="CANCELADA",
                    es_vigente=False,
                    archivo_origen=(file.filename or "")[:255],
                    loaded_by_id=current_user.id,
                )
                db.session.add(revision)
                db.session.flush()
                for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(value not in (None, "") for value in values):
                        continue
                    savepoint = db.session.begin_nested()
                    try:
                        data = {header: values[indexes[header]] for header in EXPLOSION_HEADERS}
                        partida = excel_text(data["Partida"])
                        subpartida = excel_text(data["Subpartida"])
                        type_raw = normalize(excel_text(data["Tipo"]))
                        supply_type = TYPE_INPUT_MAP.get(type_raw.replace(" ", "_")) or TYPE_INPUT_MAP.get(type_raw)
                        supply_key = normalize(excel_text(data["Clave Insumo"]))
                        description = normalize(excel_text(data["Descripción"]))
                        unit = normalize(excel_text(data["Unidad"]))
                        if not partida or not supply_type or not supply_key or not description or not unit:
                            raise ValueError("Partida, Tipo, Clave Insumo, Descripción y Unidad son obligatorios.")
                        quantity = parse_decimal_value(data["Cantidad"], "Cantidad", required=True, positive=True)
                        unit_price = parse_decimal_value(data["Precio Unitario"], "Precio Unitario", required=True, scale=QUANTITY_STEP)
                        amount = parse_decimal_value(data["Importe"], "Importe", required=True, scale=MONEY_STEP)
                        calculated = money(quantity * unit_price)
                        if abs(calculated - amount) > Decimal("0.01"):
                            raise ValueError(f"Importe no coincide: Cantidad × Precio Unitario = {calculated}.")
                        classification = normalize(
                            excel_text(
                                values[indexes["Clasificación"]]
                                if "Clasificación" in indexes
                                else "NORMAL"
                            )
                        ).replace(" ", "_")
                        if not classification:
                            classification = "NORMAL"
                        if classification not in EXPLOSION_CLASSIFICATIONS:
                            raise ValueError(
                                "Clasificación inválida. Usa NORMAL, OPERATIVO, "
                                "EQUIPO_ESPECIAL o ELECTRODOMESTICO."
                            )
                        authorization_raw = normalize(
                            excel_text(
                                values[indexes["Requiere autorización previa"]]
                                if "Requiere autorización previa" in indexes
                                else "NO"
                            )
                        )
                        if authorization_raw not in {
                            "",
                            "NO",
                            "N",
                            "0",
                            "SI",
                            "SÍ",
                            "S",
                            "1",
                        }:
                            raise ValueError(
                                "Requiere autorización previa debe indicar Sí o No."
                            )
                        requires_authorization = authorization_raw in {
                            "SI",
                            "SÍ",
                            "S",
                            "1",
                        }
                        classification_note = (
                            normalize(
                                excel_text(
                                    values[indexes["Observación de clasificación"]]
                                )
                            )
                            if "Observación de clasificación" in indexes
                            else ""
                        )
                        category = {"MANO_OBRA": "MANO_OBRA", "SUBCONTRATO": "SUBCONTRATO", "INDIRECTO": "INDIRECTO"}.get(supply_type, "ADICIONAL")
                        budget_item = find_or_create_budget_path(project, partida, subpartida, category)
                        supply = SupplyItem.query.filter(func.upper(SupplyItem.clave) == supply_key).first()
                        if not supply:
                            operation_category = detect_operation_category(
                                description, supply_type
                            )
                            supply = SupplyItem(
                                clave=supply_key,
                                descripcion=description,
                                tipo=supply_type,
                                unidad=unit,
                                clave_sat="00000000",
                                moneda="MXN",
                                es_operacion=bool(operation_category),
                                categoria_operacion=operation_category,
                                activo=True,
                            )
                            db.session.add(supply)
                            db.session.flush()
                        elif (supply.tipo, supply.unidad) != (supply_type, unit):
                            raise ValueError(f"La clave {supply_key} ya existe con tipo o unidad diferente.")
                        created = (
                            budget_item.id,
                            supply.id,
                        ) not in previous_entries
                        entry = BudgetExplosionItem(
                            revision_id=revision.id,
                            project_id=project.id,
                            budget_item_id=budget_item.id,
                            supply_item_id=supply.id,
                            created_by_id=current_user.id,
                        )
                        entry.cantidad_presupuestada = quantity
                        entry.precio_unitario_sin_iva = unit_price
                        entry.importe_presupuestado = amount
                        entry.clasificacion = classification
                        entry.requiere_autorizacion_previa = requires_authorization
                        entry.observacion_clasificacion = classification_note or None
                        entry.origen = "EXPLOSION"
                        entry.activo = False
                        db.session.add(entry)
                        db.session.flush()
                        savepoint.commit()
                        result["processed"] += 1
                        result["created" if created else "updated"] += 1
                    except Exception as exc:
                        savepoint.rollback()
                        result["errors"].append((row_number, str(exc)))
                if not result["processed"]:
                    result["errors"].append((1, "La explosión no contiene renglones válidos."))
                if not result["errors"]:
                    activar_revision_explosion(revision)
                else:
                    revision.estado = "CANCELADA"
                    revision.es_vigente = False
                    for entry in revision.items:
                        entry.activo = False
                audit("IMPORTAR", "EXPLOSION_INSUMOS", project.id, f"{result['processed']} renglones; {len(result['errors'])} errores")
                db.session.commit()
                flash(
                    (
                        f"Explosión revisión {revision.numero_revision} activada."
                        if not result["errors"]
                        else "La carga contiene errores y no sustituyó la explosión vigente."
                    ),
                    "success" if not result["errors"] else "warning",
                )
            except (ValueError, OSError) as exc:
                db.session.rollback()
                flash(str(exc), "danger")
    return render_template("compras/explosion/import.html", result=result, projects=projects)


@compras_bp.get("/explosion/<int:project_id>/exportar.xlsx")
@permission_required("compras", "ver")
def explosion_export(project_id):
    project = project_or_403(project_id)
    entry_ids = [entry.id for entry in items_explosion_vigente(project.id)]
    entries = (
        BudgetExplosionItem.query.options(
            joinedload(BudgetExplosionItem.budget_item),
            joinedload(BudgetExplosionItem.supply_item),
        )
        .filter(BudgetExplosionItem.id.in_(entry_ids or [-1]))
        .all()
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Control de insumos"
    headers = [
        "Partida",
        "Subpartida",
        "Clave",
        "Descripción",
        "Clasificación",
        "Requiere Autorización Previa",
        "Observación de Clasificación",
        "Tipo",
        "Unidad",
        "Presupuestado",
        "Requisitado",
        "Aprobado",
        "Comprado",
        "Recibido",
        "Pagado",
        "Disponible",
    ]
    sheet.append(headers)
    for entry in entries:
        parent = entry.budget_item.parent
        requested = sum((decimal_value(line.cantidad_solicitada) for line in entry.requisition_lines), Decimal("0"))
        approved = sum((decimal_value(line.cantidad_aprobada) for line in entry.requisition_lines if line.requisition.estado not in {"RECHAZADA", "VENCIDA", "CANCELADA"}), Decimal("0"))
        sheet.append([
            parent.nombre if parent else entry.budget_item.nombre,
            entry.budget_item.nombre if parent else "",
            entry.supply_item.clave,
            entry.supply_item.descripcion,
            entry.clasificacion,
            "Sí" if entry.requiere_autorizacion_previa else "No",
            entry.observacion_clasificacion or "",
            TYPE_LABELS.get(entry.supply_item.tipo, entry.supply_item.tipo),
            entry.supply_item.unidad,
            float(entry.cantidad_presupuestada),
            float(requested),
            float(approved),
            float(entry.cantidad_ordenada),
            float(entry.cantidad_recibida),
            float(entry.importe_pagado),
            float(entry.cantidad_disponible),
        ])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"control_insumos_{project.codigo}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Requisiciones y autorización parcial
# ---------------------------------------------------------------------------


def available_to_request(entry: BudgetExplosionItem, exclude_requisition_id: int | None = None) -> Decimal:
    available = entry.cantidad_disponible
    if exclude_requisition_id:
        current = next(
            (
                line
                for line in entry.requisition_lines
                if line.requisition_id == exclude_requisition_id
                and line.estado_linea == "PENDIENTE"
            ),
            None,
        )
        if current:
            available += decimal_value(current.cantidad_solicitada)
    return max(Decimal("0"), available)


def require_requisition_access(
    requisition: PurchaseRequisition, action: str = "ver"
) -> None:
    """Aplica permiso, tipo de flujo y alcance de obra/propietario."""

    require_project_access(requisition.project)
    require_permission("requisiciones", action)
    if requisition.tipo_requisicion not in allowed_requisition_types(action):
        abort(403)


@compras_bp.get("/requisiciones")
@permission_required("requisiciones", "ver")
def requisitions_list():
    run_daily_purchase_alerts()
    project_ids = [p.id for p in accessible_projects_query().all()]
    visible_types = allowed_requisition_types("ver")
    query = PurchaseRequisition.query.options(joinedload(PurchaseRequisition.project), joinedload(PurchaseRequisition.requested_by), selectinload(PurchaseRequisition.lines).joinedload(PurchaseRequisitionLine.explosion_item)).filter(
        PurchaseRequisition.project_id.in_(project_ids or [-1]),
        PurchaseRequisition.tipo_requisicion.in_(visible_types or {"__NONE__"}),
    )
    state = request.args.get("estado", "")
    project_id = request.args.get("project_id", type=int)
    requested_by_id = request.args.get("requested_by_id", type=int)
    requested_type = normalize(request.args.get("tipo_requisicion"))
    if state:
        query = query.filter(PurchaseRequisition.estado == state)
    if project_id:
        if project_id not in project_ids:
            abort(404)
        query = query.filter(PurchaseRequisition.project_id == project_id)
    if requested_by_id:
        query = query.filter(PurchaseRequisition.requested_by_id == requested_by_id)
    if requested_type:
        if requested_type not in visible_types:
            abort(403)
        query = query.filter(PurchaseRequisition.tipo_requisicion == requested_type)
    requisitions = query.order_by(PurchaseRequisition.created_at.desc()).all()
    requesters = (
        Usuario.query.join(PurchaseRequisition, PurchaseRequisition.requested_by_id == Usuario.id)
        .filter(
            PurchaseRequisition.project_id.in_(project_ids or [-1]),
            PurchaseRequisition.tipo_requisicion.in_(visible_types or {"__NONE__"}),
        )
        .distinct()
        .order_by(Usuario.nombre_completo)
        .all()
    )
    projects = accessible_projects_query().all()
    can_create = bool(allowed_requisition_types("crear"))
    can_consolidate = (
        current_user.tiene_permiso("cotizaciones_rfq", "crear")
        and "COMPRAS" in visible_types
    )
    suppliers = (
        Supplier.query.filter_by(activo=True).order_by(Supplier.nombre).all()
        if can_consolidate
        else []
    )
    return render_template(
        "compras/requisitions/list.html",
        requisitions=requisitions,
        state=state,
        can_create=can_create,
        projects=projects,
        requesters=requesters,
        selected_project_id=project_id,
        selected_requester_id=requested_by_id,
        selected_type=requested_type,
        visible_types=visible_types,
        operations_only=visible_types == {"OPERACIONES"},
        can_consolidate=can_consolidate,
        suppliers=suppliers,
    )


@compras_bp.route("/requisiciones/nueva", methods=["GET", "POST"])
@permission_required("requisiciones", "crear")
def requisition_new():
    current = today_value()
    projects = requisition_project_choices()
    visible_types = allowed_requisition_types("crear")
    if not visible_types:
        abort(404)
    selected_project_id = request.form.get("project_id", type=int)
    form_state = request.form.to_dict(flat=False) if request.method == "POST" else {}
    wants_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or request.accept_mimetypes.best == "application/json"
    )
    if request.method == "POST":
        try:
            requested_type = normalize(request.form.get("tipo_requisicion"))
            if not requested_type and len(visible_types) == 1:
                requested_type = next(iter(visible_types))
            if requested_type not in visible_types:
                raise ValueError("No tienes acceso al tipo de requisición seleccionado.")
            allowed_project_ids = {project.id for project in projects}
            if selected_project_id not in allowed_project_ids:
                raise ValueError("Selecciona una obra asignada y activa.")
            if current_user.rol == "supervisor":
                project = seleccionar_obra_activa(
                    current_user,
                    selected_project_id,
                )
                if project is None:
                    raise ValueError("La obra seleccionada no está asignada.")
            else:
                project = project_or_403(selected_project_id)
            required_date = form_date("fecha_requerida")
            if required_date < current:
                raise ValueError("La fecha requerida no puede ser anterior a la requisición.")
            requisition = PurchaseRequisition(
                folio=next_folio(PurchaseRequisition, "REQ"),
                project_id=project.id,
                fecha_solicitud=current,
                fecha_requerida=required_date,
                tipo_requisicion=requested_type,
                estado="BORRADOR",
                motivo=(request.form.get("motivo") or "").strip(),
                observaciones=(request.form.get("observaciones") or "").strip() or None,
                requested_by_id=current_user.id,
            )
            if not requisition.motivo:
                raise ValueError("El motivo o frente de trabajo es obligatorio.")
            db.session.add(requisition)
            db.session.flush()
            selected_material_ids = {
                int(value)
                for value in request.form.getlist("material_ids")
                if value.isdigit()
            }
            if selected_material_ids:
                current_entries = {
                    entry.id: entry
                    for entry in items_explosion_vigente(project.id)
                    if entry.id in selected_material_ids
                }
                if len(current_entries) != len(selected_material_ids):
                    raise ValueError(
                        "Uno de los materiales ya no pertenece a la explosión vigente."
                    )
                for entry_id in sorted(selected_material_ids):
                    entry = current_entries[entry_id]
                    quantity = parse_decimal_value(
                        request.form.get(f"cantidad_{entry.id}"),
                        f"Cantidad de {entry.supply_item.clave}",
                        required=True,
                        positive=True,
                    )
                    maximum = entry.cantidad_disponible
                    if quantity > maximum:
                        raise ValueError(
                            f"{entry.supply_item.descripcion}: la cantidad "
                            "solicitada excede la disponible. "
                            f"Cantidad máxima permitida: {maximum:.4f} "
                            f"{entry.supply_item.unidad}."
                        )
                    ajustar_reserva_pendiente(entry.id, quantity)
                    entry = db.session.get(BudgetExplosionItem, entry.id)
                    requisition.lines.append(
                        PurchaseRequisitionLine(
                            explosion_item_id=entry.id,
                            cantidad_solicitada=quantity,
                            cantidad_aprobada=0,
                            estado_linea="PENDIENTE",
                            notas=(
                                request.form.get(f"notas_{entry.id}") or ""
                            ).strip()[:240]
                            or None,
                        )
                    )
            audit("CREAR", "REQUISICION", requisition.id, requisition.folio)
            db.session.commit()
            destination = url_for(
                "compras.requisition_detail",
                requisition_id=requisition.id,
            )
            if wants_json:
                return (
                    jsonify(
                        {
                            "ok": True,
                            "message": (
                                "Borrador guardado y cantidades reservadas."
                                if selected_material_ids
                                else "Borrador creado."
                            ),
                            "redirect": destination,
                        }
                    ),
                    201,
                )
            return redirect(destination)
        except ValueError as exc:
            db.session.rollback()
            if wants_json:
                return jsonify({"ok": False, "error": str(exc)}), 422
            flash(str(exc), "danger")
    return render_template(
        "compras/requisitions/form.html",
        projects=projects,
        requisition_date=current,
        visible_types=visible_types,
        selected_project_id=selected_project_id,
        form_state=form_state,
    )


@compras_bp.get("/api/requisiciones/obras/<int:project_id>/partidas")
@permission_required("requisiciones", "crear")
def requisition_partidas_api(project_id):
    project = db.get_or_404(CentroCosto, project_id)
    if project.tipo != "obra" or project.estado != "activa":
        abort(404)
    verificar_asignacion_obra(current_user, project.id)
    entries = items_explosion_vigente(project.id)
    top_level: dict[int, dict] = {}
    for entry in entries:
        item = entry.budget_item
        parent = item.parent or item
        record = top_level.setdefault(
            parent.id,
            {
                "id": parent.id,
                "codigo": parent.codigo,
                "nombre": parent.nombre,
                "subpartidas": {},
            },
        )
        if item.parent_id:
            record["subpartidas"][item.id] = {
                "id": item.id,
                "codigo": item.codigo,
                "nombre": item.nombre,
            }
    partidas = []
    for record in sorted(
        top_level.values(),
        key=lambda value: (value["codigo"], value["nombre"]),
    ):
        record["subpartidas"] = sorted(
            record["subpartidas"].values(),
            key=lambda value: (value["codigo"], value["nombre"]),
        )
        partidas.append(record)
    return jsonify(
        {
            "ok": True,
            "project": {
                "id": project.id,
                "codigo": project.codigo,
                "nombre": project.nombre,
            },
            "partidas": partidas,
        }
    )


@compras_bp.get("/api/requisiciones/obras/<int:project_id>/materiales")
@permission_required("requisiciones", "crear")
def requisition_materials_api(project_id):
    project = db.get_or_404(CentroCosto, project_id)
    if project.tipo != "obra" or project.estado != "activa":
        abort(404)
    verificar_asignacion_obra(current_user, project.id)
    partida_id = request.args.get("partida_id", type=int)
    subpartida_id = request.args.get("subpartida_id", type=int)
    partida = db.session.get(BudgetItem, partida_id) if partida_id else None
    if (
        not partida
        or partida.project_id != project.id
        or partida.parent_id is not None
    ):
        return jsonify(
            {"ok": False, "error": "Selecciona una partida válida."}
        ), 422
    subpartida = (
        db.session.get(BudgetItem, subpartida_id)
        if subpartida_id
        else None
    )
    if subpartida and (
        subpartida.project_id != project.id
        or subpartida.parent_id != partida.id
    ):
        return jsonify(
            {"ok": False, "error": "Selecciona una subpartida válida."}
        ), 422

    materials = []
    for entry in items_explosion_vigente(project.id):
        item = entry.budget_item
        if subpartida:
            if item.id != subpartida.id:
                continue
        elif item.id != partida.id and item.parent_id != partida.id:
            continue
        available = entry.cantidad_disponible
        materials.append(
            {
                "id": entry.id,
                "clave": entry.supply_item.clave,
                "descripcion": entry.supply_item.descripcion,
                "unidad": entry.supply_item.unidad,
                "disponible": f"{available:.4f}",
                "cantidad_predeterminada": f"{available:.4f}",
                "partida": partida.nombre,
                "subpartida": (
                    item.nombre if item.parent_id else ""
                ),
            }
        )
    materials.sort(
        key=lambda item: (
            item["descripcion"].casefold(),
            item["clave"].casefold(),
        )
    )
    return jsonify({"ok": True, "materials": materials})


@compras_bp.get("/requisiciones/<int:requisition_id>")
@login_required
def requisition_detail(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition)
    require_project_access(requisition.project)
    current_entry_ids = [
        entry.id for entry in items_explosion_vigente(requisition.project_id)
    ]
    entries_query = BudgetExplosionItem.query.options(joinedload(BudgetExplosionItem.supply_item), joinedload(BudgetExplosionItem.budget_item), selectinload(BudgetExplosionItem.requisition_lines).joinedload(PurchaseRequisitionLine.requisition), selectinload(BudgetExplosionItem.order_lines).joinedload(PurchaseOrderLine.order)).filter(BudgetExplosionItem.id.in_(current_entry_ids or [-1]))
    if requisition.tipo_requisicion == "OPERACIONES":
        entries_query = entries_query.join(BudgetExplosionItem.supply_item).filter(
            SupplyItem.es_operacion.is_(True)
        )
    entries = entries_query.order_by(BudgetExplosionItem.budget_item_id, BudgetExplosionItem.id).all()
    available = {entry.id: available_to_request(entry, requisition.id) for entry in entries}
    suppliers = Supplier.query.filter_by(activo=True).order_by(Supplier.nombre).all()
    can_edit = (
        requisition.estado == "BORRADOR"
        and requisition.requested_by_id == current_user.id
        and current_user.tiene_permiso("requisiciones", "editar")
        and requisition.tipo_requisicion in allowed_requisition_types("editar")
    )
    return render_template(
        "compras/requisitions/detail.html",
        requisition=requisition,
        entries=entries,
        available=available,
        suppliers=suppliers,
        visible_lines=list(requisition.lines),
        can_edit=can_edit,
    )


@compras_bp.post("/requisiciones/<int:requisition_id>/lineas")
@permission_required("requisiciones", "crear")
def requisition_line_add(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "crear")
    require_project_access(requisition.project)
    if requisition.estado != "BORRADOR" or requisition.requested_by_id != current_user.id:
        abort(403)
    try:
        entry = db.session.get(BudgetExplosionItem, request.form.get("explosion_item_id", type=int))
        quantity = form_decimal("cantidad_solicitada", required=True, positive=True)
        if not entry or entry.project_id != requisition.project_id or not entry.activo:
            raise ValueError("Selecciona un insumo activo de la obra.")
        if requisition.tipo_requisicion == "OPERACIONES" and not entry.supply_item.es_operacion:
            raise ValueError(
                "Las requisiciones de Operaciones solo aceptan materiales marcados como operativos."
            )
        existing = PurchaseRequisitionLine.query.filter_by(requisition_id=requisition.id, explosion_item_id=entry.id).first()
        maximum = available_to_request(entry, requisition.id)
        if quantity > maximum:
            raise ValueError(
                "La cantidad solicitada excede la disponible. "
                f"Cantidad máxima permitida: {maximum:.4f} "
                f"{entry.supply_item.unidad}."
            )
        previous_quantity = (
            decimal_value(existing.cantidad_solicitada)
            if existing
            else Decimal("0")
        )
        ajustar_reserva_pendiente(
            entry.id,
            quantity - previous_quantity,
        )
        if existing:
            existing = db.session.get(PurchaseRequisitionLine, existing.id)
        line = existing or PurchaseRequisitionLine(requisition_id=requisition.id, explosion_item_id=entry.id)
        line.cantidad_solicitada = quantity
        line.cantidad_aprobada = 0
        line.estado_linea = "PENDIENTE"
        line.notas = (request.form.get("notas") or "").strip() or None
        db.session.add(line)
        db.session.commit()
        flash("Insumo agregado a la requisición.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))


@compras_bp.post("/requisiciones/<int:requisition_id>/lineas/<int:line_id>/eliminar")
@permission_required("requisiciones", "editar")
def requisition_line_delete(requisition_id, line_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "editar")
    line = db.get_or_404(PurchaseRequisitionLine, line_id)
    if line.requisition_id != requisition.id or requisition.estado != "BORRADOR" or requisition.requested_by_id != current_user.id:
        abort(403)
    if line.estado_linea == "PENDIENTE":
        ajustar_reserva_pendiente(
            line.explosion_item_id,
            -decimal_value(line.cantidad_solicitada),
        )
    db.session.delete(line)
    db.session.commit()
    return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))


@compras_bp.post("/requisiciones/<int:requisition_id>/lineas/actualizar")
@permission_required("requisiciones", "editar")
def requisition_lines_update(requisition_id):
    """Actualiza cantidades del borrador y responde sin recargar para la UI."""

    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "editar")
    if (
        requisition.estado != "BORRADOR"
        or requisition.requested_by_id != current_user.id
    ):
        abort(403)
    wants_json = (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or request.accept_mimetypes.best == "application/json"
    )
    try:
        for line in requisition.lines:
            quantity = parse_decimal_value(
                request.form.get(f"cantidad_{line.id}"),
                f"Cantidad de {line.explosion_item.supply_item.clave}",
                required=True,
                positive=True,
            )
            maximum = available_to_request(
                line.explosion_item,
                requisition.id,
            )
            if quantity > maximum:
                raise ValueError(
                    f"{line.explosion_item.supply_item.descripcion}: "
                    "la cantidad solicitada excede la disponible. "
                    f"Cantidad máxima permitida: {maximum:.4f} "
                    f"{line.explosion_item.supply_item.unidad}."
                )
            previous = decimal_value(line.cantidad_solicitada)
            ajustar_reserva_pendiente(
                line.explosion_item_id,
                quantity - previous,
            )
            line = db.session.get(PurchaseRequisitionLine, line.id)
            line.cantidad_solicitada = quantity
            line.notas = (
                request.form.get(f"notas_{line.id}") or ""
            ).strip()[:240] or None
        audit(
            "EDITAR_LINEAS",
            "REQUISICION",
            requisition.id,
            requisition.folio,
        )
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        if wants_json:
            return jsonify({"ok": False, "error": str(exc)}), 422
        flash(str(exc), "danger")
        return redirect(
            url_for(
                "compras.requisition_detail",
                requisition_id=requisition.id,
            )
        )
    if wants_json:
        return jsonify(
            {
                "ok": True,
                "message": "Cantidades y observaciones guardadas.",
            }
        )
    flash("Cantidades y observaciones guardadas.", "success")
    return redirect(
        url_for("compras.requisition_detail", requisition_id=requisition.id)
    )


@compras_bp.post("/requisiciones/<int:requisition_id>/cancelar")
@permission_required("requisiciones", "editar")
def requisition_cancel(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "editar")
    if (
        requisition.estado != "BORRADOR"
        or requisition.requested_by_id != current_user.id
    ):
        abort(403)
    liberar_reservas_pendientes(requisition)
    requisition.estado = "CANCELADA"
    audit("CANCELAR", "REQUISICION", requisition.id, requisition.folio)
    db.session.commit()
    flash(
        "Borrador cancelado; las cantidades se reintegraron a la explosión.",
        "success",
    )
    return redirect(url_for("compras.requisitions_list"))


@compras_bp.post("/requisiciones/<int:requisition_id>/eliminar")
@permission_required("requisiciones", "editar")
def requisition_delete(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "editar")
    if (
        requisition.estado != "BORRADOR"
        or requisition.requested_by_id != current_user.id
    ):
        abort(403)
    folio = requisition.folio
    liberar_reservas_pendientes(requisition)
    audit("ELIMINAR_BORRADOR", "REQUISICION", requisition.id, folio)
    db.session.delete(requisition)
    db.session.commit()
    flash(
        f"{folio} se eliminó y sus cantidades se reintegraron.",
        "success",
    )
    return redirect(url_for("compras.requisitions_list"))


@compras_bp.post("/requisiciones/<int:requisition_id>/enviar")
@permission_required("requisiciones", "editar")
def requisition_submit(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "editar")
    require_project_access(requisition.project)
    if requisition.estado != "BORRADOR" or requisition.requested_by_id != current_user.id:
        abort(403)
    if not requisition.lines:
        flash("Agrega al menos un insumo antes de enviar.", "danger")
    else:
        requisition.submitted_at = utc_now()
        clasificar_y_liberar_requisicion(
            requisition,
            usuario_id=current_user.id,
        )
        requisition.fecha_limite_oc = add_business_days(today_value(), 3)
        pending_special = sum(
            line.estado_linea == "PENDIENTE" for line in requisition.lines
        )
        if pending_special:
            notify(
                usuarios_con_permiso(
                    "requisiciones",
                    "aprobar",
                    centro_costo_id=requisition.project_id,
                ),
                "REQUISICION_PENDIENTE",
                (
                    f"{requisition.folio} tiene {pending_special} concepto(s) "
                    "especial(es) pendientes de autorización."
                ),
                url_for(
                    "compras.requisition_detail",
                    requisition_id=requisition.id,
                ),
            )
        notify(
            usuarios_con_permiso(
                "cotizaciones_rfq",
                "ver",
                centro_costo_id=requisition.project_id,
            ),
            "RFQ_AUTOMATICA",
            f"{requisition.folio} tiene conceptos liberados en la bandeja de RFQ.",
            url_for(
                "comprador_fase5.licitaciones_lista",
            ),
        )
        audit(
            "LIBERAR_POR_CLASIFICACION",
            "REQUISICION",
            requisition.id,
            f"{requisition.folio} · estado {requisition.estado}",
        )
        db.session.commit()
        flash(
            (
                "Conceptos normales liberados a Compras; los especiales "
                "permanecen pendientes de autorización."
                if pending_special
                else "Requisición liberada a Compras y RFQ generada automáticamente."
            ),
            "success",
        )
    return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))


@compras_bp.post("/requisiciones/<int:requisition_id>/confirmar-recepcion")
@permission_required("requisiciones", "editar")
def requisition_confirm_receipt(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "editar")
    if requisition.tipo_requisicion != "COMPRAS":
        abort(403)
    if requisition.estado not in {"APROBADA", "PARCIAL"}:
        flash("Solo puede confirmarse una requisición aprobada y abierta.", "danger")
    elif requisition.buyer_received_at:
        flash("La recepción por Compras ya había sido confirmada.", "info")
    else:
        requisition.buyer_received_by_id = current_user.id
        requisition.buyer_received_at = utc_now()
        audit(
            "CONFIRMAR_RECEPCION_COMPRAS",
            "REQUISICION",
            requisition.id,
            requisition.folio,
        )
        db.session.commit()
        flash(
            "Recepción confirmada. Ya puedes preparar y enviar cotizaciones a proveedores.",
            "success",
        )
    return redirect(
        url_for("compras.requisition_detail", requisition_id=requisition.id)
    )


@compras_bp.post("/requisiciones/<int:requisition_id>/aprobar")
@permission_required("requisiciones", "aprobar")
def requisition_approve(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_project_access(requisition.project)
    if requisition.estado not in {"PENDIENTE_AUTORIZACION", "PARCIAL"}:
        flash("La requisición ya no está pendiente de autorización.", "danger")
        return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))
    try:
        approved_any = False
        for line in requisition.lines:
            if line.estado_linea != "PENDIENTE":
                continue
            quantity = parse_decimal_value(request.form.get(f"aprobada_{line.id}"), f"Cantidad aprobada de {line.explosion_item.supply_item.clave}")
            if quantity > decimal_value(line.cantidad_solicitada):
                raise ValueError("Una cantidad aprobada no puede superar la solicitada.")
            ajustar_reserva_pendiente(
                line.explosion_item_id,
                -decimal_value(line.cantidad_solicitada),
            )
            line = db.session.get(PurchaseRequisitionLine, line.id)
            line.cantidad_aprobada = quantity
            line.estado_linea = "APROBADA" if quantity > 0 else "RECHAZADA"
            line.liberada_at = utc_now() if quantity > 0 else None
            approved_any = approved_any or quantity > 0
        if not approved_any:
            raise ValueError("Aprueba al menos una cantidad o rechaza la requisición completa.")
        sincronizar_estado_requisicion(requisition)
        requisition.approved_by_id = current_user.id
        requisition.approved_at = utc_now()
        requisition.fecha_limite_oc = add_business_days(today_value(), 3)
        crear_o_actualizar_rfq_liberada(
            requisition,
            usuario_id=current_user.id,
        )
        recipients = [requisition.requested_by]
        if requisition.tipo_requisicion == "COMPRAS":
            recipients.extend(users_by_roles("comprador"))
        notify(recipients, "REQUISICION_APROBADA", f"{requisition.folio} fue aprobada y debe convertirse en OC antes del {requisition.fecha_limite_oc.strftime('%d/%m/%Y')}.", url_for("compras.requisition_detail", requisition_id=requisition.id))
        audit("APROBAR", "REQUISICION", requisition.id, f"{requisition.folio} · {requisition.total_aprobado}")
        db.session.commit()
        flash("Requisición aprobada. Ya está disponible la Solicitud de Cotización genérica.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))


@compras_bp.post("/requisiciones/<int:requisition_id>/rechazar")
@permission_required("requisiciones", "aprobar")
def requisition_reject(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_project_access(requisition.project)
    reason = (request.form.get("motivo_rechazo") or "").strip()
    if requisition.estado not in {"PENDIENTE_AUTORIZACION", "PARCIAL"} or not reason:
        flash("La requisición debe estar pendiente y el motivo es obligatorio.", "danger")
    else:
        requisition.rejection_reason = reason
        for line in requisition.lines:
            if line.estado_linea != "PENDIENTE":
                continue
            ajustar_reserva_pendiente(
                line.explosion_item_id,
                -decimal_value(line.cantidad_solicitada),
            )
            line = db.session.get(PurchaseRequisitionLine, line.id)
            line.cantidad_aprobada = 0
            line.estado_linea = "RECHAZADA"
        sincronizar_estado_requisicion(requisition)
        notify([requisition.requested_by], "REQUISICION_RECHAZADA", f"{requisition.folio} fue rechazada: {reason}", url_for("compras.requisition_detail", requisition_id=requisition.id))
        audit("RECHAZAR", "REQUISICION", requisition.id, reason)
        db.session.commit()
        flash("Requisición rechazada.", "success")
    return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))


# ---------------------------------------------------------------------------
# Solicitud genérica y comparativo de cotizaciones
# ---------------------------------------------------------------------------


def require_quotation_access(quotation: Quotation) -> None:
    requisitions = quotation.requisition_set
    if not requisitions:
        abort(404)
    for requisition in requisitions:
        require_project_access(requisition.project)


def quotation_source_lines(
    quotation: Quotation,
    *,
    project_id: int | None = None,
) -> dict[int, Decimal]:
    """Mapea cada línea original al precio de su renglón agrupado."""

    prices: dict[int, Decimal] = {}
    for quote_line in quotation.lines:
        sources = quote_line.sources or []
        if not sources:
            source_line = quote_line.requisition_line
            if (
                source_line
                and (
                    project_id is None
                    or source_line.requisition.project_id == project_id
                )
            ):
                prices[source_line.id] = quote_line.precio_unitario_cotizado
            continue
        for source in sources:
            if (
                project_id is None
                or source.requisition_line.requisition.project_id == project_id
            ):
                prices[source.requisition_line_id] = (
                    quote_line.precio_unitario_cotizado
                )
    return prices


@compras_bp.get("/cotizaciones")
@permission_required("cotizaciones_rfq", "ver")
def quotations_list():
    """Bandeja única de RFQ automáticas y cotizaciones por proveedor."""

    from fase5_models import Licitacion

    project_ids = [project.id for project in accessible_projects_query().all()]
    rfqs = (
        Licitacion.query.options(
            joinedload(Licitacion.requisicion).joinedload(
                PurchaseRequisition.project
            ),
            selectinload(Licitacion.lineas),
            selectinload(Licitacion.ofertas),
        )
        .join(PurchaseRequisition)
        .filter(PurchaseRequisition.project_id.in_(project_ids or [-1]))
        .order_by(Licitacion.id.desc())
        .all()
    )
    quotations = (
        Quotation.query.options(
            joinedload(Quotation.requisition).joinedload(
                PurchaseRequisition.project
            ),
            joinedload(Quotation.supplier),
        )
        .join(PurchaseRequisition)
        .filter(PurchaseRequisition.project_id.in_(project_ids or [-1]))
        .order_by(Quotation.created_at.desc(), Quotation.id.desc())
        .all()
    )
    return render_template(
        "compras/quotations/list.html",
        rfqs=rfqs,
        quotations=quotations,
    )


@compras_bp.get("/requisiciones/<int:requisition_id>/solicitud-cotizacion")
@permission_required("requisiciones", "ver")
def rfq_print(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "ver")
    require_project_access(requisition.project)
    if requisition.tipo_requisicion != "COMPRAS" or requisition.estado not in {"APROBADA", "PARCIAL", "CERRADA"}:
        abort(404)
    return render_template(
        "compras/quotations/rfq.html",
        requisition=requisition,
        fecha_limite_respuesta=requisition.fecha_requerida,
    )


@compras_bp.post("/requisiciones/<int:requisition_id>/cotizaciones")
@permission_required("compras", "crear")
def quotations_create(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "crear")
    require_project_access(requisition.project)
    if requisition.tipo_requisicion != "COMPRAS":
        abort(403)
    if requisition.estado not in {"APROBADA", "PARCIAL"} or (requisition.fecha_limite_oc and requisition.fecha_limite_oc < today_value()):
        flash("Solo pueden cotizarse requisiciones aprobadas y vigentes.", "danger")
        return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))
    if not requisition.buyer_received_at:
        flash(
            "Primero confirma que Compras recibió y revisó la requisición.",
            "danger",
        )
        return redirect(
            url_for("compras.requisition_detail", requisition_id=requisition.id)
        )
    supplier_ids = {int(value) for value in request.form.getlist("supplier_ids") if value.isdigit()}
    if not supplier_ids:
        flash("Selecciona al menos un proveedor.", "danger")
        return redirect(url_for("compras.requisition_detail", requisition_id=requisition.id))
    requested_line_ids = {
        int(value)
        for value in request.form.getlist("line_ids")
        if value.isdigit()
    }
    eligible_lines = [
        line
        for line in requisition.lines
        if line.estado_linea == "APROBADA"
        and line.cantidad_pendiente_compra > 0
        and (not requested_line_ids or line.id in requested_line_ids)
    ]
    if not eligible_lines:
        flash("Selecciona al menos un material pendiente para cotizar.", "danger")
        return redirect(
            url_for("compras.requisition_detail", requisition_id=requisition.id)
        )
    created = 0
    for supplier in Supplier.query.filter(Supplier.id.in_(supplier_ids), Supplier.activo.is_(True)).all():
        if Quotation.query.filter_by(requisition_id=requisition.id, supplier_id=supplier.id).first():
            continue
        quotation = Quotation(folio=next_folio(Quotation, "COT"), requisition_id=requisition.id, supplier_id=supplier.id, fecha_solicitud=today_value(), estado="SOLICITADA", created_by_id=current_user.id)
        db.session.add(quotation)
        quotation.requisitions.append(requisition)
        for line in eligible_lines:
            history = SupplierSupplyItem.query.filter_by(
                supplier_id=supplier.id,
                supply_item_id=line.explosion_item.supply_item_id,
            ).first()
            historical_price = (
                decimal_value(history.precio_historico)
                if history
                else Decimal("0")
            )
            quote_line = QuotationLine(
                    requisition_line_id=line.id,
                    supply_item_id=line.explosion_item.supply_item_id,
                    cantidad=line.cantidad_pendiente_compra,
                    precio_unitario_cotizado=historical_price,
                    importe_cotizado=money(
                        line.cantidad_pendiente_compra * historical_price
                    ),
                )
            quote_line.sources.append(
                QuotationLineSource(
                    requisition_line_id=line.id,
                    cantidad=line.cantidad_pendiente_compra,
                )
            )
            quotation.lines.append(quote_line)
        created += 1
    db.session.commit()
    flash(f"Se generaron {created} solicitud(es) de cotización.", "success")
    return redirect(url_for("compras.quotation_compare", requisition_id=requisition.id))


@compras_bp.post("/cotizaciones/consolidar")
@permission_required("cotizaciones_rfq", "crear")
def quotations_consolidate():
    """Agrupa dos o más requisiciones y suma materiales repetidos."""

    requisition_ids = sorted(
        {
            int(value)
            for value in request.form.getlist("requisition_ids")
            if value.isdigit()
        }
    )
    if len(requisition_ids) < 2:
        flash(
            "Selecciona por lo menos dos requisiciones para consolidar.",
            "danger",
        )
        return redirect(url_for("compras.requisitions_list"))
    supplier = db.session.get(
        Supplier,
        request.form.get("supplier_id", type=int),
    )
    if not supplier or not supplier.activo:
        flash("Selecciona un proveedor activo.", "danger")
        return redirect(url_for("compras.requisitions_list"))

    requisitions = (
        PurchaseRequisition.query.options(
            joinedload(PurchaseRequisition.project),
            selectinload(PurchaseRequisition.lines)
            .joinedload(PurchaseRequisitionLine.explosion_item)
            .joinedload(BudgetExplosionItem.supply_item),
        )
        .filter(PurchaseRequisition.id.in_(requisition_ids))
        .order_by(PurchaseRequisition.id)
        .all()
    )
    if len(requisitions) != len(requisition_ids):
        abort(404)
    try:
        grouped: dict[int, list[tuple[PurchaseRequisitionLine, Decimal]]] = (
            defaultdict(list)
        )
        for requisition in requisitions:
            require_requisition_access(requisition, "ver")
            if (
                requisition.tipo_requisicion != "COMPRAS"
                or requisition.estado not in {"APROBADA", "PARCIAL"}
            ):
                raise ValueError(
                    f"{requisition.folio} no está aprobada y pendiente de compra."
                )
            if not requisition.buyer_received_at:
                requisition.buyer_received_by_id = current_user.id
                requisition.buyer_received_at = utc_now()
            for line in requisition.lines:
                pending = line.cantidad_pendiente_compra
                if line.estado_linea == "APROBADA" and pending > 0:
                    grouped[line.explosion_item.supply_item_id].append(
                        (line, pending)
                    )
        if not grouped:
            raise ValueError(
                "Las requisiciones seleccionadas no tienen materiales pendientes."
            )

        quotation = Quotation(
            folio=next_folio(Quotation, "COT"),
            requisition_id=requisitions[0].id,
            supplier_id=supplier.id,
            fecha_solicitud=today_value(),
            estado="SOLICITADA",
            created_by_id=current_user.id,
            notas=(
                f"Consolidada de {len(requisitions)} requisiciones."
            ),
        )
        db.session.add(quotation)
        quotation.requisitions.extend(requisitions)
        for supply_item_id, sources in sorted(grouped.items()):
            anchor, _anchor_quantity = sources[0]
            total_quantity = sum(
                (quantity for _line, quantity in sources),
                Decimal("0"),
            )
            history = SupplierSupplyItem.query.filter_by(
                supplier_id=supplier.id,
                supply_item_id=supply_item_id,
            ).first()
            historical_price = (
                decimal_value(history.precio_historico)
                if history
                else Decimal("0")
            )
            quote_line = QuotationLine(
                requisition_line_id=anchor.id,
                supply_item_id=supply_item_id,
                cantidad=total_quantity,
                precio_unitario_cotizado=historical_price,
                importe_cotizado=money(total_quantity * historical_price),
            )
            for source_line, source_quantity in sources:
                quote_line.sources.append(
                    QuotationLineSource(
                        requisition_line_id=source_line.id,
                        cantidad=source_quantity,
                    )
                )
            quotation.lines.append(quote_line)
        db.session.flush()
        audit(
            "CONSOLIDAR",
            "COTIZACION",
            quotation.id,
            ", ".join(requisition.folio for requisition in requisitions),
        )
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
        return redirect(url_for("compras.requisitions_list"))

    flash(
        f"{quotation.folio} consolidó {len(requisitions)} requisiciones y "
        f"{len(quotation.lines)} material(es) agrupado(s).",
        "success",
    )
    return redirect(
        url_for("compras.quotation_detail", quotation_id=quotation.id)
    )


@compras_bp.get("/cotizaciones/<int:quotation_id>")
@permission_required("compras", "ver")
def quotation_detail(quotation_id):
    quotation = db.get_or_404(Quotation, quotation_id)
    require_quotation_access(quotation)
    history_by_line = {
        line.id: SupplierSupplyItem.query.filter_by(
            supplier_id=quotation.supplier_id,
            supply_item_id=line.material.id,
        ).first()
        for line in quotation.lines
    }
    return render_template(
        "compras/quotations/detail.html",
        quotation=quotation,
        history_by_line=history_by_line,
    )


@compras_bp.get("/cotizaciones/<int:quotation_id>/solicitud")
@permission_required("cotizaciones_rfq", "ver")
def quotation_rfq_print(quotation_id):
    quotation = db.get_or_404(Quotation, quotation_id)
    require_quotation_access(quotation)
    return render_template(
        "compras/quotations/rfq.html",
        quotation=quotation,
        requisition=quotation.requisition,
        fecha_limite_respuesta=min(
            requisition.fecha_requerida
            for requisition in quotation.requisition_set
        ),
    )


@compras_bp.post("/cotizaciones/<int:quotation_id>/enviar-correo")
@permission_required("compras", "editar")
def quotation_send_email(quotation_id):
    quotation = db.get_or_404(Quotation, quotation_id)
    require_quotation_access(quotation)
    if any(
        not requisition.buyer_received_at
        for requisition in quotation.requisition_set
    ):
        flash("Primero confirma la recepción de la requisición por Compras.", "danger")
        return redirect(
            url_for("compras.quotation_detail", quotation_id=quotation.id)
        )
    try:
        message = quotation_email_message(quotation)
        send_purchase_email(message)
    except Exception as exc:  # Un fallo SMTP no debe romper la requisición.
        quotation.email_error = str(exc)[:500]
        db.session.commit()
        flash(
            f"No fue posible enviar la cotización: {exc}. Revisa el correo del proveedor y la configuración SMTP.",
            "danger",
        )
    else:
        quotation.email_sent_at = utc_now()
        quotation.email_sent_by_id = current_user.id
        quotation.email_to = message.recipients[0]
        quotation.email_cc = message.cc[0]
        quotation.email_error = None
        audit(
            "ENVIAR_CORREO",
            "COTIZACION",
            quotation.id,
            f"Para {quotation.email_to}; CC {quotation.email_cc}",
        )
        db.session.commit()
        flash(
            f"Cotización enviada exitosamente a {quotation.email_to}; copia enviada a {quotation.email_cc}.",
            "success",
        )
    return redirect(url_for("compras.quotation_detail", quotation_id=quotation.id))


@compras_bp.post("/cotizaciones/<int:quotation_id>/contacto-whatsapp")
@permission_required("compras", "editar")
def quotation_mark_whatsapp(quotation_id):
    quotation = db.get_or_404(Quotation, quotation_id)
    require_quotation_access(quotation)
    notes = (request.form.get("notas_whatsapp") or "").strip()
    if len(notes) > 500:
        flash("Las notas de WhatsApp no pueden exceder 500 caracteres.", "danger")
    else:
        quotation.whatsapp_contacted_at = utc_now()
        quotation.whatsapp_contacted_by_id = current_user.id
        quotation.whatsapp_notes = notes or None
        audit(
            "CONTACTAR_WHATSAPP",
            "COTIZACION",
            quotation.id,
            notes or quotation.supplier.nombre,
        )
        db.session.commit()
        flash("Contacto por WhatsApp registrado.", "success")
    return redirect(url_for("compras.quotation_detail", quotation_id=quotation.id))


@compras_bp.post("/cotizaciones/<int:quotation_id>/respuesta")
@permission_required("compras", "editar")
def quotation_response(quotation_id):
    quotation = db.get_or_404(Quotation, quotation_id)
    require_quotation_access(quotation)
    try:
        quotation.fecha_respuesta = form_date("fecha_respuesta")
        quotation.fecha_entrega_ofertada = form_date("fecha_entrega_ofertada")
        quotation.notas = (request.form.get("notas") or "").strip() or None
        for line in quotation.lines:
            price = parse_decimal_value(request.form.get(f"precio_{line.id}"), "Precio cotizado", required=True, positive=True)
            line.precio_unitario_cotizado = price
            line.importe_cotizado = money(decimal_value(line.cantidad) * price)
        quotation.estado = "RESPONDIDA"
        audit("RESPUESTA", "COTIZACION", quotation.id, quotation.folio)
        db.session.commit()
        flash("Respuesta del proveedor registrada.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("compras.quotation_detail", quotation_id=quotation.id))


@compras_bp.get("/requisiciones/<int:requisition_id>/comparativo")
@permission_required("compras", "ver")
def quotation_compare(requisition_id):
    requisition = db.get_or_404(PurchaseRequisition, requisition_id)
    require_requisition_access(requisition, "ver")
    if requisition.tipo_requisicion != "COMPRAS":
        abort(404)
    require_project_access(requisition.project)
    quotations = (
        Quotation.query.options(
            joinedload(Quotation.supplier),
            selectinload(Quotation.lines),
        )
        .join(
            quotation_requisitions,
            quotation_requisitions.c.quotation_id == Quotation.id,
        )
        .filter(
            quotation_requisitions.c.requisition_id == requisition.id
        )
        .order_by(Quotation.id)
        .all()
    )
    return render_template("compras/quotations/compare.html", requisition=requisition, quotations=quotations)


@compras_bp.post("/cotizaciones/<int:quotation_id>/seleccionar")
@permission_required("compras", "editar")
def quotation_select(quotation_id):
    quotation = db.get_or_404(Quotation, quotation_id)
    require_quotation_access(quotation)
    if quotation.estado != "RESPONDIDA":
        flash("Primero registra la respuesta completa del proveedor.", "danger")
        return redirect(
            url_for("compras.quotation_detail", quotation_id=quotation.id)
        )
    requisition_ids = {
        requisition.id for requisition in quotation.requisition_set
    }
    candidates = (
        Quotation.query.join(
            quotation_requisitions,
            quotation_requisitions.c.quotation_id == Quotation.id,
        )
        .filter(
            quotation_requisitions.c.requisition_id.in_(requisition_ids),
            Quotation.estado == "SELECCIONADA",
        )
        .all()
    )
    for candidate in candidates:
        candidate.estado = "RESPONDIDA"
    quotation.estado = "SELECCIONADA"
    db.session.commit()
    if len(quotation.projects) == 1:
        return redirect(
            url_for(
                "compras.order_new",
                quotation_id=quotation.id,
                project_id=quotation.projects[0].id,
            )
        )
    flash(
        "Cotización seleccionada. Genera una OC independiente por cada obra "
        "para conservar centros de costo y direcciones.",
        "success",
    )
    return redirect(
        url_for("compras.quotation_detail", quotation_id=quotation.id)
    )


# ---------------------------------------------------------------------------
# Órdenes consolidadas, anticipos y recepciones
# ---------------------------------------------------------------------------


def eligible_requisition_lines(
    project_id: int | None = None,
    requisition_type: str | None = None,
):
    query = PurchaseRequisitionLine.query.join(PurchaseRequisition).options(
        joinedload(PurchaseRequisitionLine.requisition).joinedload(PurchaseRequisition.project),
        joinedload(PurchaseRequisitionLine.explosion_item).joinedload(BudgetExplosionItem.supply_item),
        joinedload(PurchaseRequisitionLine.explosion_item).joinedload(BudgetExplosionItem.budget_item),
        selectinload(PurchaseRequisitionLine.order_lines).joinedload(PurchaseOrderLine.order),
    ).filter(
        PurchaseRequisition.estado.in_({"APROBADA", "PARCIAL"}),
        PurchaseRequisition.fecha_limite_oc >= today_value(),
        PurchaseRequisitionLine.estado_linea == "APROBADA",
    )
    if project_id:
        query = query.filter(PurchaseRequisition.project_id == project_id)
    if requisition_type:
        query = query.filter(
            PurchaseRequisition.tipo_requisicion == requisition_type
        )
    return [line for line in query.order_by(PurchaseRequisition.project_id, PurchaseRequisition.id, PurchaseRequisitionLine.id).all() if line.cantidad_pendiente_compra > 0]


def emitir_orden(order: PurchaseOrder, *, actor: Usuario) -> None:
    """Emite una OC una sola vez y conserva las reglas históricas aplicables."""

    if order.estado not in {"BORRADOR", "PENDIENTE_AUTORIZACION"}:
        raise ValueError("Solo puede emitirse una orden en borrador o pendiente.")
    for req in order.requisitions:
        if req.estado not in {"APROBADA", "PARCIAL"} or (
            req.fecha_limite_oc and req.fecha_limite_oc < today_value()
        ):
            raise ValueError("La OC contiene una requisición vencida o no aprobada.")
    if (
        order.modalidad_pago == "CREDITO"
        and order.supplier
        and order.supplier.tiene_credito
    ):
        projected = money(
            order.supplier.credito_utilizado + order.subtotal_sin_iva
        )
        if projected > decimal_value(order.supplier.limite_credito):
            raise ValueError(
                "La OC excede la línea de crédito disponible del proveedor."
            )

    order.issued_by_id = actor.id
    order.issued_at = utc_now()
    order.requiere_autorizacion = False
    # Las OC nuevas separan el estado documental de la autorización financiera.
    order.estado = (
        "EMITIDA"
        if order.payment_schedules
        else "PENDIENTE_ANTICIPO"
        if order.modalidad_pago == "ANTICIPO"
        else "EMITIDA"
    )
    if order.modalidad_pago == "ANTICIPO":
        notify(
            usuarios_con_permiso(
                "programacion_pagos",
                "aprobar",
                centro_costo_id=order.project_id,
            ),
            "ANTICIPO_SOLICITADO",
            (
                f"{order.folio} solicita un anticipo por "
                f"{money(order.anticipo_monto)} MXN."
            ),
            url_for("compras.order_detail", order_id=order.id),
        )
    for req in order.requisitions:
        refresh_requisition_status(req)
    if order.supplier:
        for line in order.lines:
            history = SupplierSupplyItem.query.filter_by(
                supplier_id=order.supplier_id,
                supply_item_id=line.explosion_item.supply_item_id,
            ).first()
            if not history:
                history = SupplierSupplyItem(
                    supplier_id=order.supplier_id,
                    supply_item_id=line.explosion_item.supply_item_id,
                )
                db.session.add(history)
            history.precio_historico = line.precio_unitario_sin_iva
            history.fecha_ultima_compra = order.fecha_orden
            history.origen = "ORDEN_COMPRA"
    audit("EMITIR", "ORDEN_COMPRA", order.id, order.folio)


@compras_bp.route("/ordenes-operaciones/nueva", methods=["GET", "POST"])
@permission_required("oc_operaciones", "crear")
def operational_order_new():
    """Captura directa de campo sin requisición, cotización ni proveedor forzoso."""

    centers = centros_operativos_accesibles(current_user)
    center_ids = {center.id for center in centers}
    suppliers = (
        Supplier.query.filter_by(activo=True).order_by(Supplier.nombre).all()
        if current_user.tiene_permiso("proveedores", "ver")
        else []
    )
    form = OperationalPurchaseOrderForm()
    form.project_id.choices = [
        (center.id, f"{center.codigo} · {center.nombre}") for center in centers
    ]
    form.supplier_id.choices = [(0, "Sin proveedor de catálogo")] + [
        (supplier.id, supplier.nombre) for supplier in suppliers
    ]
    if request.method == "GET":
        form.fecha_entrega_estimada.data = today_value()
        form.condicion_saldo.data = "CONTRA_RECEPCION"

    selected_project_id = (
        form.project_id.data
        if request.method == "POST"
        else request.args.get("project_id", type=int)
    )
    if selected_project_id and selected_project_id not in center_ids:
        abort(404)
    if request.method == "GET" and selected_project_id:
        form.project_id.data = selected_project_id
    entries = (
        items_explosion_vigente(selected_project_id)
        if selected_project_id
        else []
    )
    operational_entries = [
        entry
        for entry in entries
        if entry.clasificacion == "OPERATIVO"
    ]

    if form.validate_on_submit():
        try:
            project = operational_project_or_404(form.project_id.data)
            current_entries = {
                entry.id: entry for entry in items_explosion_vigente(project.id)
            }
            if not current_entries:
                raise ValueError(
                    "Sin explosión vigente. Carga una explosión antes de crear la OC."
                )
            supplier = (
                db.session.get(Supplier, form.supplier_id.data)
                if form.supplier_id.data
                else None
            )
            if supplier and not supplier.activo:
                raise ValueError("El proveedor seleccionado ya no está activo.")
            beneficiary = (form.beneficiario_libre.data or "").strip() or None
            if not supplier and not beneficiary:
                raise ValueError(
                    "Selecciona un proveedor o captura el beneficiario libre."
                )
            order = PurchaseOrder(
                folio=next_folio(PurchaseOrder, "OCO"),
                project_id=project.id,
                supplier_id=supplier.id if supplier else None,
                beneficiario_libre=beneficiary,
                beneficiario_validado=False,
                company_id=None,
                buyer_id=current_user.id,
                payment_method_id=None,
                fecha_orden=today_value(),
                fecha_entrega_estimada=form.fecha_entrega_estimada.data,
                fecha_limite=form.fecha_entrega_estimada.data,
                tipo_oc="OPERACIONES",
                categoria_pago="OPERACIONES",
                estado="BORRADOR",
                modalidad_pago=(
                    "ANTICIPO"
                    if form.requiere_anticipo.data
                    else "PAGO_CONTRA_ENTREGA"
                ),
                condicion_saldo=form.condicion_saldo.data,
                justificacion_anticipo=(
                    (form.justificacion_anticipo.data or "").strip() or None
                ),
                requiere_conciliacion=False,
                notas=(form.notas.data or "").strip() or None,
                direccion_entrega=project.direccion_entrega,
                created_by_id=current_user.id,
            )
            db.session.add(order)
            requested_by_entry: dict[int, Decimal] = {}
            for line_form in form.lineas:
                entry = current_entries.get(line_form.explosion_item_id.data)
                if not entry:
                    raise ValueError(
                        "Un concepto no pertenece a la explosión vigente."
                    )
                if entry.clasificacion != "OPERATIVO":
                    raise ValueError(
                        f"{entry.supply_item.clave} no está clasificado como operativo."
                    )
                quantity = decimal_value(line_form.cantidad.data)
                price = decimal_value(line_form.precio_unitario_sin_iva.data)
                accumulated = (
                    requested_by_entry.get(entry.id, Decimal("0")) + quantity
                ).quantize(QUANTITY_STEP, rounding=ROUND_HALF_UP)
                if accumulated > entry.cantidad_disponible:
                    raise ValueError(
                        f"{entry.supply_item.clave} supera el saldo disponible "
                        f"({entry.cantidad_disponible})."
                    )
                requested_by_entry[entry.id] = accumulated
                observation = (line_form.observacion.data or "").strip()
                if not observation:
                    raise ValueError("La observación de cada concepto es obligatoria.")
                smnc = None
                if entry.origen == "SMNC":
                    detail = (
                        MaterialChangeRequestLine.query.join(
                            MaterialChangeRequest,
                            MaterialChangeRequest.id
                            == MaterialChangeRequestLine.request_id,
                        )
                        .filter(
                            MaterialChangeRequestLine.generated_explosion_item_id
                            == entry.id,
                            MaterialChangeRequest.estado == "APROBADA",
                        )
                        .first()
                    )
                    if not detail:
                        raise ValueError(
                            f"{entry.supply_item.clave} no tiene una SMNC aprobada."
                        )
                    smnc = detail.request
                order.lines.append(
                    PurchaseOrderLine(
                        requisition_line_id=None,
                        explosion_item_id=entry.id,
                        cantidad=quantity,
                        precio_unitario_sin_iva=price,
                        importe_sin_iva=money(quantity * price),
                        clasificacion_explosion=entry.clasificacion,
                        observacion_operativa=observation,
                        notas=observation[:240],
                        smnc_id=smnc.id if smnc else None,
                    )
                )
            if not order.lines:
                raise ValueError("Agrega al menos un concepto a la OC.")
            db.session.flush()
            if order.subtotal_sin_iva <= 0:
                raise ValueError("El total de la OC debe ser mayor que cero.")
            if form.requiere_anticipo.data:
                if form.anticipo_tipo.data == "PORCENTAJE":
                    order.anticipo_porcentaje = decimal_value(
                        form.anticipo_porcentaje.data
                    )
                    order.anticipo_monto = money(
                        order.subtotal_sin_iva
                        * order.anticipo_porcentaje
                        / Decimal("100")
                    )
                else:
                    order.anticipo_monto = money(form.anticipo_monto.data)
                    order.anticipo_porcentaje = (
                        order.anticipo_monto
                        * Decimal("100")
                        / order.subtotal_sin_iva
                    )
                if order.anticipo_monto > order.subtotal_sin_iva:
                    raise ValueError(
                        "El anticipo no puede exceder el total de la OC."
                    )
                order.anticipo_pendiente = order.anticipo_monto
            crear_programacion_pago(
                order,
                solicitado_por_id=current_user.id,
            )
            if current_user.tiene_permiso("oc_operaciones", "emitir"):
                emitir_orden(order, actor=current_user)
                message = "OC de Operaciones guardada y emitida."
            else:
                order.estado = "PENDIENTE_AUTORIZACION"
                order.requiere_autorizacion = True
                order.autorizacion_solicitada_at = utc_now()
                notify(
                    usuarios_con_permiso(
                        "oc_operaciones",
                        "aprobar",
                        centro_costo_id=order.project_id,
                    ),
                    "OC_OPERACIONES_PENDIENTE",
                    f"{order.folio} espera autorización para emitirse.",
                    url_for("compras.order_detail", order_id=order.id),
                )
                message = (
                    "OC guardada y enviada a autorización; se emitirá al aprobarse."
                )
            audit("CREAR", "ORDEN_COMPRA_OPERACIONES", order.id, order.folio)
            db.session.commit()
            flash(message, "success")
            return redirect(
                url_for("compras.order_detail", order_id=order.id)
            )
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")

    return render_template(
        "compras/orders/operational_form.html",
        form=form,
        centers=centers,
        entries=operational_entries,
    )


@compras_bp.route("/ordenes/nueva", methods=["GET", "POST"])
@login_required
def order_new():
    run_daily_purchase_alerts()
    ensure_payment_methods()
    # El formulario consolidado pertenece únicamente al Comprador. Las OC de
    # Operaciones tienen su captura directa y reglas propias.
    creatable_types = (
        {"COMPRAS"}
        if current_user.tiene_permiso("oc_compras", "crear")
        else set()
    )
    if not creatable_types:
        abort(403)
    projects = accessible_projects_query().all()
    project_ids = [p.id for p in projects]
    quotation_id = request.args.get("quotation_id", type=int) or request.form.get("quotation_id", type=int)
    requisition_id = request.args.get("requisition_id", type=int)
    quotation = db.session.get(Quotation, quotation_id) if quotation_id else None
    requested_order_type = normalize(
        request.form.get("tipo_oc") or request.args.get("tipo_oc")
    )
    selected_project_id = request.args.get("project_id", type=int) or request.form.get("project_id", type=int)
    if quotation:
        require_quotation_access(quotation)
        requested_order_type = "COMPRAS"
        quotation_project_ids = {project.id for project in quotation.projects}
        if not selected_project_id and len(quotation_project_ids) == 1:
            selected_project_id = next(iter(quotation_project_ids))
        if selected_project_id not in quotation_project_ids:
            abort(404)
    elif requisition_id:
        req = db.session.get(PurchaseRequisition, requisition_id)
        if req:
            selected_project_id = req.project_id
            requested_order_type = req.tipo_requisicion
    if not requested_order_type and len(creatable_types) == 1:
        requested_order_type = next(iter(creatable_types))
    if requested_order_type and requested_order_type not in creatable_types:
        abort(403)
    if selected_project_id and selected_project_id not in project_ids:
        abort(404)
    lines = eligible_requisition_lines(selected_project_id, requested_order_type)
    if requisition_id:
        lines = [line for line in lines if line.requisition_id == requisition_id]
    quote_prices = {}
    if quotation:
        quote_prices = quotation_source_lines(
            quotation,
            project_id=selected_project_id,
        )
        lines = [line for line in lines if line.id in quote_prices]

    if request.method == "POST":
        try:
            project = project_or_403(selected_project_id)
            order_type = normalize(request.form.get("tipo_oc"))
            if not order_type and len(creatable_types) == 1:
                order_type = next(iter(creatable_types))
            if order_type not in creatable_types:
                raise ValueError("No tienes permiso para crear este tipo de OC.")
            supplier = db.session.get(Supplier, request.form.get("supplier_id", type=int))
            company = db.session.get(Company, request.form.get("company_id", type=int))
            method = resolve_payment_method()
            modality = normalize(request.form.get("modalidad_pago"))
            delivery = form_date("fecha_entrega_estimada")
            if not supplier or not supplier.activo or not company or not company.activa:
                raise ValueError("Proveedor, empresa pagadora y método de pago activos son obligatorios.")
            if modality not in {"CREDITO", "ANTICIPO"}:
                raise ValueError("Selecciona una modalidad válida.")
            order = PurchaseOrder(
                folio=next_folio(PurchaseOrder, "OC"),
                project_id=project.id,
                quotation_id=quotation.id if quotation else None,
                supplier_id=supplier.id,
                company_id=company.id,
                buyer_id=current_user.id,
                payment_method_id=method.id,
                fecha_orden=today_value(),
                fecha_entrega_estimada=delivery,
                fecha_limite=today_value(),
                tipo_oc=order_type,
                categoria_pago="NOMINA" if order_type == "OPERACIONES" else "COMPRAS",
                estado="BORRADOR",
                modalidad_pago=modality,
                notas=(request.form.get("notas") or "").strip() or None,
                direccion_entrega=project.direccion_entrega,
                created_by_id=current_user.id,
            )
            selected = 0
            deadline_dates = []
            for line in eligible_requisition_lines(project.id, order_type):
                if order_type == "OPERACIONES" and not line.explosion_item.supply_item.es_operacion:
                    raise ValueError(
                        "La OC de Operaciones contiene un material fuera del catálogo permitido."
                    )
                reject = request.form.get(f"rechazar_{line.id}") == "on"
                if reject:
                    reason = (request.form.get(f"motivo_rechazo_{line.id}") or "").strip()
                    if not reason:
                        raise ValueError("El motivo es obligatorio para rechazar un ítem.")
                    line.estado_linea = "RECHAZADA_COMPRAS"
                    line.motivo_rechazo_compras = reason
                    continue
                quantity = parse_decimal_value(request.form.get(f"cantidad_{line.id}"), "Cantidad a ordenar")
                if quantity <= 0:
                    continue
                if quantity > line.cantidad_pendiente_compra:
                    raise ValueError("Una cantidad a ordenar supera la aprobada pendiente.")
                price = parse_decimal_value(request.form.get(f"precio_{line.id}"), "Precio unitario real", required=True, positive=True)
                order.lines.append(PurchaseOrderLine(requisition_line_id=line.id, explosion_item_id=line.explosion_item_id, cantidad=quantity, precio_unitario_sin_iva=price, importe_sin_iva=money(quantity * price)))
                selected += 1
                if line.requisition.fecha_limite_oc:
                    deadline_dates.append(line.requisition.fecha_limite_oc)
            if not selected:
                raise ValueError("Selecciona al menos un ítem aprobado para la OC.")
            order.fecha_limite = min(deadline_dates) if deadline_dates else today_value()
            if modality == "ANTICIPO":
                order.anticipo_monto = form_decimal("anticipo_monto", required=True, positive=True, scale=MONEY_STEP)
                order.anticipo_pendiente = order.anticipo_monto
                order.justificacion_anticipo = (request.form.get("justificacion_anticipo") or "").strip()
                if not order.justificacion_anticipo:
                    raise ValueError("La justificación del anticipo es obligatoria.")
            db.session.add(order)
            db.session.flush()
            if order.anticipo_monto > order.subtotal_sin_iva:
                raise ValueError("El anticipo no puede exceder el total de la OC.")
            audit("CREAR", "ORDEN_COMPRA", order.id, order.folio)
            if request.form.get("guardar_emitir") == "1":
                permission_module = (
                    "oc_operaciones"
                    if order.tipo_oc == "OPERACIONES"
                    else "oc_compras"
                )
                if current_user.tiene_permiso(permission_module, "emitir"):
                    emitir_orden(order, actor=current_user)
                    success_message = "Orden guardada y emitida."
                else:
                    order.estado = "PENDIENTE_AUTORIZACION"
                    order.requiere_autorizacion = True
                    order.autorizacion_solicitada_at = utc_now()
                    notify(
                        usuarios_con_permiso(
                            permission_module,
                            "aprobar",
                            centro_costo_id=order.project_id,
                        ),
                        "OC_PENDIENTE_AUTORIZACION",
                        f"{order.folio} espera autorización para emitirse.",
                        url_for("compras.order_detail", order_id=order.id),
                    )
                    success_message = (
                        "Orden guardada y enviada a autorización; se emitirá al aprobarse."
                    )
            else:
                # Compatibilidad de API y pruebas históricas que guardan
                # explícitamente antes de emitir.
                success_message = (
                    "Orden creada en borrador. Revísala y emítela para reconocer la compra."
                )
            db.session.commit()
            flash(success_message, "success")
            return redirect(url_for("compras.order_detail", order_id=order.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")

    suppliers = Supplier.query.filter_by(activo=True).order_by(Supplier.nombre).all()
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    methods = PaymentMethod.query.filter_by(activo=True).order_by(PaymentMethod.nombre).all()
    db.session.commit()
    return render_template(
        "compras/orders/form.html",
        projects=projects,
        selected_project_id=selected_project_id,
        lines=lines,
        quotation=quotation,
        quote_prices=quote_prices,
        suppliers=suppliers,
        companies=companies,
        methods=methods,
        order=None,
        order_type=requested_order_type,
        creatable_types=creatable_types,
        editing=False,
        existing_order_lines={},
    )


@compras_bp.route("/ordenes/<int:order_id>/editar", methods=["GET", "POST"])
@login_required
def order_edit(order_id):
    """Permite modificar sin restricciones operativas una OC aún no emitida."""

    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    require_order_permission(order, "editar")
    if order.estado != "BORRADOR":
        flash("Solo pueden editarse órdenes en estado borrador.", "danger")
        return redirect(url_for("compras.order_detail", order_id=order.id))
    if order.requiere_autorizacion:
        flash(
            "La OC ya fue enviada a autorización; cancela la solicitud antes de editarla.",
            "danger",
        )
        return redirect(url_for("compras.order_detail", order_id=order.id))

    ensure_payment_methods()
    lines = eligible_requisition_lines(order.project_id, order.tipo_oc)
    existing_by_requisition = {
        line.requisition_line_id: line
        for line in order.lines
        if line.requisition_line_id is not None
    }
    # Un borrador histórico puede contener un renglón que dejó de ser elegible;
    # se conserva visible para poder quitarlo o cancelar la orden.
    candidates = {line.id: line for line in lines}
    for order_line in order.lines:
        if order_line.requisition_line:
            candidates.setdefault(
                order_line.requisition_line_id, order_line.requisition_line
            )
    lines = sorted(
        candidates.values(),
        key=lambda line: (
            line.requisition_id,
            line.id,
        ),
    )

    if request.method == "POST":
        try:
            supplier = db.session.get(
                Supplier, request.form.get("supplier_id", type=int)
            )
            company = db.session.get(
                Company, request.form.get("company_id", type=int)
            )
            method = resolve_payment_method()
            modality = normalize(request.form.get("modalidad_pago"))
            delivery = form_date("fecha_entrega_estimada")
            if not supplier or not supplier.activo or not company or not company.activa:
                raise ValueError(
                    "Proveedor, empresa pagadora y método de pago activos son obligatorios."
                )
            if modality not in {"CREDITO", "ANTICIPO"}:
                raise ValueError("Selecciona una modalidad válida.")

            selected: dict[int, tuple[PurchaseRequisitionLine, Decimal, Decimal]] = {}
            deadline_dates = []
            for line in lines:
                quantity = parse_decimal_value(
                    request.form.get(f"cantidad_{line.id}"),
                    "Cantidad a ordenar",
                )
                if quantity <= 0:
                    continue
                if line.requisition.estado not in {"APROBADA", "PARCIAL"}:
                    raise ValueError(
                        f"{line.requisition.folio} ya no está abierta para compra."
                    )
                existing_quantity = decimal_value(
                    existing_by_requisition.get(line.id).cantidad
                    if existing_by_requisition.get(line.id)
                    else 0
                )
                editable_available = line.cantidad_pendiente_compra + existing_quantity
                if quantity > editable_available:
                    raise ValueError(
                        "Una cantidad a ordenar supera la aprobada pendiente."
                    )
                if order.tipo_oc == "OPERACIONES" and not line.explosion_item.supply_item.es_operacion:
                    raise ValueError(
                        "Las OC de Operaciones solo aceptan materiales del catálogo operativo."
                    )
                price = parse_decimal_value(
                    request.form.get(f"precio_{line.id}"),
                    "Precio unitario real",
                    required=True,
                    positive=True,
                )
                selected[line.id] = (line, quantity, price)
                if line.requisition.fecha_limite_oc:
                    deadline_dates.append(line.requisition.fecha_limite_oc)
            if not selected:
                raise ValueError(
                    "La OC debe conservar al menos un ítem; para eliminarla usa Cancelar OC."
                )

            for requisition_line_id, order_line in list(
                existing_by_requisition.items()
            ):
                if requisition_line_id not in selected:
                    # ``delete-orphan`` elimina el renglón al retirarlo de la
                    # colección. Marcarlo además con session.delete() lo
                    # registraría dos veces en el mismo flush.
                    order.lines.remove(order_line)

            for requisition_line_id, (line, quantity, price) in selected.items():
                order_line = existing_by_requisition.get(requisition_line_id)
                if not order_line:
                    order_line = PurchaseOrderLine(
                        requisition_line_id=line.id,
                        explosion_item_id=line.explosion_item_id,
                    )
                    order.lines.append(order_line)
                order_line.cantidad = quantity
                order_line.precio_unitario_sin_iva = price
                order_line.importe_sin_iva = money(quantity * price)

            order.supplier_id = supplier.id
            order.company_id = company.id
            order.payment_method_id = method.id
            order.modalidad_pago = modality
            order.fecha_entrega_estimada = delivery
            order.fecha_limite = (
                min(deadline_dates) if deadline_dates else today_value()
            )
            order.notas = (request.form.get("notas") or "").strip() or None
            if order.quotation and order.quotation.supplier_id != supplier.id:
                order.quotation_id = None
            if modality == "ANTICIPO":
                order.anticipo_monto = form_decimal(
                    "anticipo_monto", required=True, positive=True, scale=MONEY_STEP
                )
                order.anticipo_pendiente = order.anticipo_monto
                order.justificacion_anticipo = (
                    request.form.get("justificacion_anticipo") or ""
                ).strip()
                if not order.justificacion_anticipo:
                    raise ValueError("La justificación del anticipo es obligatoria.")
            else:
                order.anticipo_monto = Decimal("0")
                order.anticipo_pendiente = Decimal("0")
                order.justificacion_anticipo = None

            db.session.flush()
            if order.anticipo_monto > order.subtotal_sin_iva:
                raise ValueError("El anticipo no puede exceder el total de la OC.")
            audit("EDITAR", "ORDEN_COMPRA", order.id, order.folio)
            if request.form.get("guardar_emitir") == "1":
                module = "oc_compras"
                if current_user.tiene_permiso(module, "emitir"):
                    emitir_orden(order, actor=current_user)
                    message = "Borrador actualizado y OC emitida."
                else:
                    order.estado = "PENDIENTE_AUTORIZACION"
                    order.requiere_autorizacion = True
                    order.autorizacion_solicitada_at = utc_now()
                    notify(
                        usuarios_con_permiso(
                            module,
                            "aprobar",
                            centro_costo_id=order.project_id,
                        ),
                        "OC_PENDIENTE_AUTORIZACION",
                        f"{order.folio} espera autorización para emitirse.",
                        url_for("compras.order_detail", order_id=order.id),
                    )
                    message = (
                        "Borrador actualizado y enviado a autorización; "
                        "se emitirá automáticamente al aprobarse."
                    )
            else:
                message = "Borrador de OC actualizado."
            db.session.commit()
            flash(message, "success")
            return redirect(url_for("compras.order_detail", order_id=order.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")

    suppliers = Supplier.query.filter_by(activo=True).order_by(Supplier.nombre).all()
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    methods = PaymentMethod.query.filter_by(activo=True).order_by(PaymentMethod.nombre).all()
    return render_template(
        "compras/orders/form.html",
        projects=[order.project],
        selected_project_id=order.project_id,
        lines=lines,
        quotation=order.quotation,
        quote_prices={},
        suppliers=suppliers,
        companies=companies,
        methods=methods,
        order=order,
        order_type=order.tipo_oc,
        creatable_types={order.tipo_oc},
        editing=True,
        existing_order_lines={
            line.requisition_line_id: line
            for line in order.lines
            if line.requisition_line_id is not None
        },
    )


@compras_bp.get("/ordenes")
@any_permission_required(("compras", "ver"), ("oc_operaciones", "ver"))
def orders_list():
    run_daily_purchase_alerts()
    projects = centros_operativos_accesibles(current_user)
    project_ids = [p.id for p in projects]
    query = PurchaseOrder.query.options(
        joinedload(PurchaseOrder.project),
        joinedload(PurchaseOrder.supplier),
        joinedload(PurchaseOrder.company),
        selectinload(PurchaseOrder.lines).selectinload(
            PurchaseOrderLine.receipt_lines
        ),
    ).filter(
        PurchaseOrder.project_id.in_(project_ids or [-1]),
        PurchaseOrder.tipo_oc.in_(allowed_order_types("ver") or {"__NONE__"}),
    )

    project_id = request.args.get("project_id", type=int)
    supplier_id = request.args.get("supplier_id", type=int)
    order_type = normalize(request.args.get("tipo_oc"))
    state = normalize(request.args.get("estado"))
    start_raw = (request.args.get("fecha_desde") or "").strip()
    end_raw = (request.args.get("fecha_hasta") or "").strip()
    start = None
    end = None
    try:
        start = datetime.strptime(start_raw, "%Y-%m-%d").date() if start_raw else None
        end = datetime.strptime(end_raw, "%Y-%m-%d").date() if end_raw else None
        if start and end and start > end:
            raise ValueError("La fecha inicial no puede ser posterior a la final.")
    except ValueError as exc:
        flash(
            str(exc) if "posterior" in str(exc) else "El rango de fechas no es válido.",
            "danger",
        )
        start = end = None

    if project_id:
        if project_id not in project_ids:
            abort(404)
        query = query.filter(PurchaseOrder.project_id == project_id)
    if supplier_id:
        query = query.filter(PurchaseOrder.supplier_id == supplier_id)
    visible_order_types = allowed_order_types("ver")
    if order_type:
        if order_type not in visible_order_types:
            abort(403)
        query = query.filter(PurchaseOrder.tipo_oc == order_type)
    allowed_states = {
        "BORRADOR",
        "EMITIDA",
        "PENDIENTE_ANTICIPO",
        "ANTICIPO_AUTORIZADO",
        "ANTICIPO_PARCIAL",
        "ANTICIPO_PAGADO",
        "RECEPCION_PARCIAL",
        "RECEPCION_TOTAL",
        "CERRADA",
        "CANCELADA",
        "PENDIENTE_AUTORIZACION",
    }
    if state:
        if state not in allowed_states:
            abort(400)
        if state == "PENDIENTE_AUTORIZACION":
            query = query.filter(
                PurchaseOrder.estado == "PENDIENTE_AUTORIZACION",
                PurchaseOrder.requiere_autorizacion.is_(True),
            )
        else:
            query = query.filter(PurchaseOrder.estado == state)
    if start:
        query = query.filter(PurchaseOrder.fecha_orden >= start)
    if end:
        query = query.filter(PurchaseOrder.fecha_orden <= end)
    orders = query.order_by(
        PurchaseOrder.fecha_orden.desc(), PurchaseOrder.id.desc()
    ).all()
    suppliers = (
        Supplier.query.join(PurchaseOrder)
        .filter(
            PurchaseOrder.project_id.in_(project_ids or [-1]),
            PurchaseOrder.tipo_oc.in_(visible_order_types or {"__NONE__"}),
        )
        .distinct()
        .order_by(Supplier.nombre)
        .all()
    )
    return render_template(
        "compras/orders/list.html",
        orders=orders,
        projects=projects,
        suppliers=suppliers,
        selected_project_id=project_id,
        selected_supplier_id=supplier_id,
        selected_order_type=order_type,
        visible_order_types=visible_order_types,
        selected_state=state,
        fecha_desde=start_raw,
        fecha_hasta=end_raw,
        order_states=sorted(allowed_states),
        operations_only=visible_order_types == {"OPERACIONES"},
    )


@compras_bp.get("/ordenes/<int:order_id>")
@any_permission_required(("compras", "ver"), ("oc_operaciones", "ver"))
def order_detail(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    require_order_permission(order, "ver")
    beneficiary_form = FinanceBeneficiaryValidationForm()
    beneficiary_form.company_id.choices = [
        (company.id, company.nombre)
        for company in Company.query.filter_by(activa=True).order_by(Company.nombre)
    ]
    beneficiary_form.payment_method_id.choices = [
        (method.id, method.nombre)
        for method in PaymentMethod.query.filter_by(activo=True).order_by(
            PaymentMethod.nombre
        )
    ]
    beneficiary_form.beneficiario_confirmado.data = order.beneficiario_nombre
    return render_template(
        "compras/orders/detail.html",
        order=order,
        beneficiary_form=beneficiary_form,
    )


@compras_bp.get("/ordenes/<int:order_id>/imprimir")
@login_required
def order_print(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    require_order_permission(order, "ver")
    partidas = sorted(
        {
            line.explosion_item.budget_item.codigo
            for line in order.lines
            if line.explosion_item and line.explosion_item.budget_item
        }
    )
    return render_template(
        "compras/orders/print.html",
        order=order,
        partidas=partidas,
    )


@compras_bp.post("/ordenes/<int:order_id>/confirmar-direccion")
@login_required
def order_delivery_address_confirm(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    module = "oc_operaciones" if order.es_operaciones else "oc_compras"
    if not current_user.tiene_permiso(module, "editar"):
        abort(403)
    if order.estado == "CANCELADA":
        abort(403)
    address = " ".join(
        (request.form.get("direccion_entrega") or "").split()
    )
    if len(address) < 10:
        flash("Confirma una dirección de entrega completa.", "danger")
    elif len(address) > 500:
        flash(
            "La dirección de entrega no puede superar 500 caracteres.",
            "danger",
        )
    else:
        order.direccion_entrega = address
        order.direccion_entrega_confirmada_por_id = current_user.id
        order.direccion_entrega_confirmada_at = utc_now()
        order.project.direccion_entrega = address
        refresh_order_status(order)
        audit(
            "CONFIRMAR_DIRECCION_ENTREGA",
            "ORDEN_COMPRA",
            order.id,
            address,
        )
        db.session.commit()
        flash(
            "Dirección de entrega confirmada para la OC y la obra.",
            "success",
        )
    return redirect(url_for("compras.order_detail", order_id=order.id))


@compras_bp.route(
    "/ordenes/<int:order_id>/revision",
    methods=["GET", "POST"],
)
@login_required
def order_revision(order_id):
    order = (
        PurchaseOrder.query.options(
            joinedload(PurchaseOrder.project),
            selectinload(PurchaseOrder.lines).selectinload(
                PurchaseOrderLine.receipt_lines
            ),
            selectinload(PurchaseOrder.payment_schedules),
        )
        .filter_by(id=order_id)
        .first_or_404()
    )
    require_project_access(order.project)
    module = "oc_operaciones" if order.es_operaciones else "oc_compras"
    require_permission(module, "editar")
    if not order.issued_at or order.estado in {
        "BORRADOR",
        "PENDIENTE_AUTORIZACION",
        "CANCELADA",
        "CERRADA",
    }:
        flash("La OC no está disponible para crear una revisión.", "danger")
        return redirect(url_for("compras.order_detail", order_id=order.id))

    if request.method == "GET":
        form = PurchaseOrderRevisionForm(
            data={
                "version_actual": str(order.version_actual),
                "beneficiario_libre": order.beneficiario_libre,
                "fecha_entrega_estimada": order.fecha_entrega_estimada,
                "notas": order.notas,
                "lineas": [
                    {
                        "explosion_item_id": line.explosion_item_id,
                        "cantidad": line.cantidad,
                        "precio_unitario_sin_iva": line.precio_unitario_sin_iva,
                        "observacion": line.observacion_operativa or line.notas,
                    }
                    for line in order.lines
                ],
            }
        )
    else:
        form = PurchaseOrderRevisionForm()

    entry_ids = {
        item.id for item in items_explosion_vigente(order.project_id)
    } | {line.explosion_item_id for line in order.lines}
    entries = (
        BudgetExplosionItem.query.options(
            joinedload(BudgetExplosionItem.supply_item),
            joinedload(BudgetExplosionItem.budget_item),
        )
        .filter(BudgetExplosionItem.id.in_(entry_ids or [-1]))
        .all()
    )
    if form.validate_on_submit():
        try:
            if int(form.version_actual.data) != int(order.version_actual):
                abort(
                    409,
                    description=(
                        "La OC cambió mientras la editabas. Recarga antes de continuar."
                    ),
                )
            if len(form.lineas) != len(order.lines):
                raise ValueError(
                    "Una revisión no puede eliminar silenciosamente conceptos."
                )
            before = snapshot_orden(order)
            previous_total = order.subtotal_sin_iva
            valid_entries = {entry.id: entry for entry in entries}
            for existing, line_form in zip(order.lines, form.lineas):
                entry = valid_entries.get(line_form.explosion_item_id.data)
                if not entry or entry.project_id != order.project_id:
                    raise ValueError(
                        "Un concepto no pertenece a la explosión de la obra."
                    )
                if entry.id != existing.explosion_item_id:
                    raise ValueError(
                        "Para sustituir un concepto cancela la OC y genera otra; "
                        "la revisión solo ajusta cantidades, precio y observación."
                    )
                quantity = decimal_value(line_form.cantidad.data)
                received = existing.cantidad_recibida
                if quantity < received:
                    raise ValueError(
                        "La cantidad revisada no puede ser menor que lo ya recibido."
                    )
                observation = (line_form.observacion.data or "").strip()
                if order.es_operaciones and not observation:
                    raise ValueError(
                        "La observación sigue siendo obligatoria en cada concepto operativo."
                    )
                existing.cantidad = quantity
                existing.precio_unitario_sin_iva = decimal_value(
                    line_form.precio_unitario_sin_iva.data
                )
                existing.importe_sin_iva = money(
                    existing.cantidad * existing.precio_unitario_sin_iva
                )
                existing.observacion_operativa = observation or None
                existing.notas = observation[:240] if observation else None
            order.fecha_entrega_estimada = form.fecha_entrega_estimada.data
            order.notas = (form.notas.data or "").strip() or None
            if order.es_operaciones and not order.supplier:
                beneficiary = (form.beneficiario_libre.data or "").strip()
                if not beneficiary:
                    raise ValueError("El beneficiario libre es obligatorio.")
                if beneficiary != (order.beneficiario_libre or ""):
                    order.beneficiario_validado = False
                    order.beneficiario_validado_por_id = None
                    order.beneficiario_validado_at = None
                    order.company_id = None
                    order.payment_method_id = None
                order.beneficiario_libre = beneficiary
            db.session.flush()
            if order.monto_pagado > 0 and order.subtotal_sin_iva != previous_total:
                raise ValueError(
                    "No puede cambiarse el importe de una OC con pagos registrados."
                )
            if order.payment_schedules:
                advance = next(
                    (
                        item
                        for item in order.payment_schedules
                        if item.tipo == "ANTICIPO"
                    ),
                    None,
                )
                balance = next(
                    (
                        item
                        for item in order.payment_schedules
                        if item.tipo == "SALDO"
                    ),
                    None,
                )
                advance_amount = (
                    decimal_value(advance.monto_programado)
                    if advance
                    else Decimal("0")
                )
                if advance_amount > order.subtotal_sin_iva:
                    raise ValueError(
                        "El nuevo total no puede ser menor que el anticipo programado."
                    )
                if balance:
                    balance.monto_programado = money(
                        order.subtotal_sin_iva - advance_amount
                    )
            revision = registrar_revision_orden(
                order,
                valores_anteriores=before,
                motivo=form.motivo.data,
                usuario_id=current_user.id,
            )
            audit(
                "REVISAR",
                "ORDEN_COMPRA",
                order.id,
                f"{order.folio} · revisión {revision.version}",
            )
            db.session.commit()
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        else:
            flash(
                f"Revisión {order.version_actual} guardada sin borrar la versión anterior.",
                "success",
            )
            return redirect(url_for("compras.order_detail", order_id=order.id))
    return render_template(
        "compras/orders/revision_form.html",
        order=order,
        form=form,
        entries=entries,
    )


@compras_bp.post("/ordenes/<int:order_id>/validar-beneficiario")
@permission_required("programacion_pagos", "aprobar")
def order_validate_beneficiary(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    if not order.es_operaciones:
        abort(404)
    form = FinanceBeneficiaryValidationForm()
    form.company_id.choices = [
        (company.id, company.nombre)
        for company in Company.query.filter_by(activa=True).order_by(Company.nombre)
    ]
    form.payment_method_id.choices = [
        (method.id, method.nombre)
        for method in PaymentMethod.query.filter_by(activo=True).order_by(
            PaymentMethod.nombre
        )
    ]
    if not form.validate_on_submit():
        flash("Completa la empresa, método y beneficiario validados.", "danger")
    else:
        company = db.session.get(Company, form.company_id.data)
        method = db.session.get(PaymentMethod, form.payment_method_id.data)
        if not company or not company.activa or not method or not method.activo:
            abort(404)
        confirmed = (form.beneficiario_confirmado.data or "").strip()
        order.company_id = company.id
        order.payment_method_id = method.id
        if not order.supplier:
            order.beneficiario_libre = confirmed
        order.beneficiario_validado = True
        order.beneficiario_validado_por_id = current_user.id
        order.beneficiario_validado_at = utc_now()
        audit(
            "VALIDAR_BENEFICIARIO",
            "ORDEN_COMPRA",
            order.id,
            (
                f"{confirmed} · {company.nombre} · {method.nombre} · "
                f"{(form.comentario.data or '').strip()}"
            ),
        )
        db.session.commit()
        flash(
            "Beneficiario validado por Finanzas; esto no registra un pago.",
            "success",
        )
    return redirect(url_for("compras.order_detail", order_id=order.id))


@compras_bp.get("/programacion-pagos")
@permission_required("programacion_pagos", "ver")
def payment_schedules_list():
    centers = centros_operativos_accesibles(
        current_user, incluir_obras_inactivas=True
    )
    center_ids = [center.id for center in centers]
    schedules = (
        PurchaseOrderPaymentSchedule.query.options(
            joinedload(PurchaseOrderPaymentSchedule.order).joinedload(
                PurchaseOrder.project
            )
        )
        .join(PurchaseOrder)
        .filter(PurchaseOrder.project_id.in_(center_ids or [-1]))
        .order_by(
            PurchaseOrderPaymentSchedule.fecha_solicitud.desc(),
            PurchaseOrderPaymentSchedule.id.desc(),
        )
        .all()
    )
    return render_template(
        "compras/payments/schedules.html",
        schedules=schedules,
        decision_form=PaymentScheduleDecisionForm(),
    )


@compras_bp.post("/programacion-pagos/<int:schedule_id>/resolver")
@permission_required("programacion_pagos", "aprobar")
def payment_schedule_decide(schedule_id):
    schedule = db.get_or_404(PurchaseOrderPaymentSchedule, schedule_id)
    require_project_access(schedule.order.project)
    form = PaymentScheduleDecisionForm()
    if not form.validate_on_submit():
        abort(400)
    if schedule.tipo != "ANTICIPO" or schedule.estado != "SOLICITADO":
        flash("La programación ya no está pendiente de decisión.", "danger")
    elif form.accion.data == "AUTORIZAR":
        schedule.estado = "AUTORIZADO"
        schedule.autorizado_por_id = current_user.id
        schedule.fecha_autorizacion = utc_now()
        schedule.order.autorizado_anticipo_por_id = current_user.id
        schedule.order.fecha_autorizacion_anticipo = utc_now()
        audit(
            "AUTORIZAR_ANTICIPO",
            "PROGRAMACION_PAGO",
            schedule.id,
            form.comentario.data,
        )
        db.session.commit()
        flash("Anticipo autorizado; todavía no está pagado.", "success")
    else:
        schedule.estado = "CANCELADO"
        order = schedule.order
        order.modalidad_pago = "PAGO_CONTRA_ENTREGA"
        order.anticipo_monto = Decimal("0")
        order.anticipo_porcentaje = Decimal("0")
        order.anticipo_pendiente = Decimal("0")
        balance = next(
            (
                item
                for item in order.payment_schedules
                if item.tipo == "SALDO"
            ),
            None,
        )
        if balance:
            balance.monto_programado = order.subtotal_sin_iva
            balance.porcentaje = Decimal("100")
        audit(
            "RECHAZAR_ANTICIPO",
            "PROGRAMACION_PAGO",
            schedule.id,
            form.comentario.data,
        )
        db.session.commit()
        flash(
            "Anticipo rechazado; la OC quedó íntegramente contra recepción.",
            "success",
        )
    return redirect(url_for("compras.payment_schedules_list"))


@compras_bp.post("/ordenes/<int:order_id>/emitir")
@login_required
def order_issue(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    permission_module = (
        "oc_operaciones" if order.tipo_oc == "OPERACIONES" else "oc_compras"
    )
    if not (
        current_user.tiene_permiso(permission_module, "emitir")
        or current_user.tiene_permiso(permission_module, "crear")
    ):
        abort(403)
    if order.estado != "BORRADOR":
        flash("Solo puede emitirse una orden en borrador.", "danger")
        return redirect(url_for("compras.order_detail", order_id=order.id))
    if not current_user.tiene_permiso(permission_module, "emitir"):
        if order.requiere_autorizacion:
            flash("La OC ya está pendiente de autorización administrativa.", "info")
        else:
            order.requiere_autorizacion = True
            order.estado = "PENDIENTE_AUTORIZACION"
            order.autorizacion_solicitada_at = utc_now()
            notify(
                usuarios_con_permiso(
                    permission_module,
                    "aprobar",
                    centro_costo_id=order.project_id,
                ),
                "OC_PENDIENTE_AUTORIZACION",
                f"{order.folio} requiere autorización para emitirse.",
                url_for("compras.order_detail", order_id=order.id),
            )
            audit("SOLICITAR_AUTORIZACION", "ORDEN_COMPRA", order.id, order.folio)
            db.session.commit()
            flash("OC enviada a autorización administrativa.", "success")
        return redirect(url_for("compras.order_detail", order_id=order.id))
    try:
        emitir_orden(order, actor=current_user)
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("compras.order_detail", order_id=order.id))
    db.session.commit()
    flash("Orden emitida; la compra ya fue reconocida.", "success")
    return redirect(url_for("compras.order_detail", order_id=order.id))


@compras_bp.post("/ordenes/<int:order_id>/autorizar-operacion")
@permission_required("oc_operaciones", "aprobar")
def order_authorize_operation(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    if (
        order.tipo_oc != "OPERACIONES"
        or order.estado not in {"BORRADOR", "PENDIENTE_AUTORIZACION"}
        or not order.requiere_autorizacion
    ):
        flash("La OC no está pendiente de autorización operativa.", "danger")
        return redirect(url_for("compras.order_detail", order_id=order.id))
    try:
        emitir_orden(order, actor=current_user)
        audit("APROBAR_EMISION", "ORDEN_COMPRA", order.id, order.folio)
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    else:
        flash("OC aprobada y emitida automáticamente.", "success")
    return redirect(url_for("compras.order_detail", order_id=order.id))


@compras_bp.post("/ordenes/<int:order_id>/aprobar-emision")
@login_required
def order_approve_issue(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    module = "oc_operaciones" if order.es_operaciones else "oc_compras"
    require_permission(module, "aprobar")
    if order.estado != "PENDIENTE_AUTORIZACION":
        flash("La OC ya no está pendiente de autorización.", "danger")
    else:
        try:
            emitir_orden(order, actor=current_user)
            audit("APROBAR_EMISION", "ORDEN_COMPRA", order.id, order.folio)
            db.session.commit()
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        else:
            flash("OC aprobada y emitida automáticamente.", "success")
    return redirect(url_for("compras.order_detail", order_id=order.id))


@compras_bp.post("/ordenes/<int:order_id>/autorizar-anticipo")
@permission_required("programacion_pagos", "aprobar")
def order_authorize_advance(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    advance_schedule = next(
        (
            schedule
            for schedule in order.payment_schedules
            if schedule.tipo == "ANTICIPO"
        ),
        None,
    )
    if advance_schedule:
        if advance_schedule.estado != "SOLICITADO":
            flash("El anticipo ya no está pendiente de autorización.", "danger")
        else:
            advance_schedule.estado = "AUTORIZADO"
            advance_schedule.autorizado_por_id = current_user.id
            advance_schedule.fecha_autorizacion = utc_now()
            order.autorizado_anticipo_por_id = current_user.id
            order.fecha_autorizacion_anticipo = utc_now()
            notify([order.buyer], "ANTICIPO_AUTORIZADO", f"El anticipo de {order.folio} fue autorizado.", url_for("compras.order_detail", order_id=order.id))
            audit("AUTORIZAR_ANTICIPO", "PROGRAMACION_PAGO", advance_schedule.id, str(advance_schedule.monto_programado))
            db.session.commit()
            flash("Anticipo autorizado; todavía no se ha registrado ningún pago.", "success")
    elif order.estado != "PENDIENTE_ANTICIPO":
        flash("La orden no está pendiente de autorización de anticipo.", "danger")
    else:
        order.autorizado_anticipo_por_id = current_user.id
        order.fecha_autorizacion_anticipo = utc_now()
        order.estado = "ANTICIPO_AUTORIZADO"
        notify([order.buyer], "ANTICIPO_AUTORIZADO", f"El anticipo de {order.folio} fue autorizado.", url_for("compras.order_detail", order_id=order.id))
        audit("AUTORIZAR_ANTICIPO", "ORDEN_COMPRA", order.id, str(order.anticipo_monto))
        db.session.commit()
        flash("Anticipo autorizado.", "success")
    return redirect(url_for("compras.order_detail", order_id=order.id))


@compras_bp.post("/ordenes/<int:order_id>/cancelar")
@login_required
def order_cancel(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    require_order_permission(
        order, "editar" if order.estado == "BORRADOR" else "cancelar"
    )
    if order.estado == "CANCELADA":
        flash("La OC ya está cancelada.", "info")
    elif order.receipts or order.monto_pagado > 0:
        flash("No puede cancelarse una OC con recepciones o pagos.", "danger")
    else:
        order.estado = "CANCELADA"
        for req in order.requisitions:
            refresh_requisition_status(req)
        audit("CANCELAR", "ORDEN_COMPRA", order.id, order.folio)
        db.session.commit()
        flash("Orden cancelada; sus cantidades volvieron a estar disponibles.", "success")
    return redirect(url_for("compras.order_detail", order_id=order.id))


@compras_bp.post("/ordenes/<int:order_id>/recibir")
@login_required
def order_receive(order_id):
    order = db.get_or_404(PurchaseOrder, order_id)
    require_project_access(order.project)
    require_permission("recepcion_materiales", "crear")
    if order.estado not in ACTIVE_ORDER_STATES - {"CERRADA", "RECEPCION_TOTAL", "PENDIENTE_ANTICIPO"}:
        flash("La OC no está habilitada para recepción.", "danger")
        return redirect(url_for("compras.order_detail", order_id=order.id))
    try:
        receipt_date = form_date("fecha_recepcion")
        invoice_number = (request.form.get("documento_proveedor") or "").strip()
        invoice_date = form_date("fecha_factura", required=False)
        if order.modalidad_pago == "CREDITO" and not order.fecha_factura:
            if not invoice_number or not invoice_date:
                raise ValueError("En compras a crédito, número y fecha de factura son obligatorios al recibir.")
            order.numero_factura = invoice_number
            order.fecha_factura = invoice_date
            order.fecha_vencimiento = invoice_date + timedelta(days=order.supplier.dias_credito if order.supplier.tiene_credito else 0)
        receipt = GoodsReceipt(folio=next_folio(GoodsReceipt, "REC"), order_id=order.id, fecha=receipt_date, tipo="PARCIAL", documento_proveedor=invoice_number or order.numero_factura, fecha_factura=invoice_date or order.fecha_factura, notas=(request.form.get("notas_recepcion") or "").strip() or None, received_by_id=current_user.id)
        for line in order.lines:
            quantity = parse_decimal_value(request.form.get(f"recibir_{line.id}"), "Cantidad recibida")
            if quantity <= 0:
                continue
            if quantity > line.cantidad_pendiente:
                raise ValueError("Una recepción supera la cantidad pendiente de la OC.")
            # Vincular mediante la relación mantiene actualizado el acumulado de
            # la línea dentro de esta misma transacción. Asignar solo el FK deja
            # en caché la colección anterior y retrasa el cambio de estado hasta
            # la siguiente petición.
            receipt.lines.append(
                GoodsReceiptLine(order_line=line, cantidad_recibida=quantity)
            )
        if not receipt.lines:
            raise ValueError("Captura al menos una cantidad recibida.")
        db.session.add(receipt)
        db.session.flush()
        receipt.tipo = "TOTAL" if all(line.cantidad_pendiente <= 0 for line in order.lines) else "PARCIAL"
        refresh_order_status(order)
        sincronizar_programacion_recepcion(order)
        notify([order.buyer], "MATERIAL_RECIBIDO", f"Se registró una recepción {receipt.tipo.lower()} de {order.folio}.", url_for("compras.order_detail", order_id=order.id))
        audit("RECIBIR", "ORDEN_COMPRA", order.id, receipt.folio)
        db.session.commit()
        try:
            message = receipt_email_message(receipt)
            send_purchase_email(message)
        except Exception as exc:  # La recepción guardada debe conservarse.
            receipt.notification_email_error = str(exc)[:500]
            db.session.commit()
            flash(
                "Recepción registrada, pero no se pudo enviar el correo de confirmación: "
                f"{exc}. Revisa el correo del proveedor y la configuración SMTP.",
                "warning",
            )
        else:
            receipt.notification_email_sent_at = utc_now()
            receipt.notification_email_error = None
            db.session.commit()
            flash(
                "Recepción registrada y correo de confirmación enviado al comprador y al proveedor.",
                "success",
            )
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("compras.order_detail", order_id=order.id))


# ---------------------------------------------------------------------------
# Pagos dirigidos y estados de cuenta
# ---------------------------------------------------------------------------


def line_paid_amount(line: PurchaseOrderLine) -> Decimal:
    """Pago efectivo de un renglón, con compatibilidad histórica controlada."""

    direct = money(
        db.session.query(func.coalesce(func.sum(AdditionalPayment.monto_sin_iva), 0))
        .filter(AdditionalPayment.purchase_order_line_id == line.id)
        .scalar()
    )
    legacy = money(
        db.session.query(func.coalesce(func.sum(AdditionalPayment.monto_sin_iva), 0))
        .filter(
            AdditionalPayment.purchase_order_id == line.order_id,
            AdditionalPayment.explosion_item_id == line.explosion_item_id,
            AdditionalPayment.purchase_order_line_id.is_(None),
        )
        .scalar()
    )
    incoming = money(
        sum(
            (
                decimal_value(movement.monto)
                for movement in line.advance_movements_in
                if movement.tipo == "APLICACION"
            ),
            Decimal("0"),
        )
    )
    outgoing = money(
        sum(
            (decimal_value(movement.monto) for movement in line.advance_movements_out),
            Decimal("0"),
        )
    )
    return max(Decimal("0.00"), money(direct + legacy + incoming - outgoing))


def monto_pagable_programado(order: PurchaseOrder) -> Decimal:
    if order.payment_schedules:
        return money(
            sum(
                (schedule.monto_liberado for schedule in order.payment_schedules),
                Decimal("0"),
            )
        )
    return order.saldo_pagable


@compras_bp.get("/pagos")
@permission_required("pagos_proveedores", "ver")
def supplier_payments_list():
    project_ids = [
        project.id
        for project in centros_operativos_accesibles(
            current_user, incluir_obras_inactivas=True
        )
    ]
    visible_order_types = allowed_order_types("ver")
    query = AdditionalPayment.query.options(joinedload(AdditionalPayment.project), joinedload(AdditionalPayment.budget_item), joinedload(AdditionalPayment.explosion_item).joinedload(BudgetExplosionItem.supply_item), joinedload(AdditionalPayment.supplier), joinedload(AdditionalPayment.purchase_order), joinedload(AdditionalPayment.company)).join(
        PurchaseOrder, PurchaseOrder.id == AdditionalPayment.purchase_order_id
    ).filter(
        AdditionalPayment.project_id.in_(project_ids or [-1]),
        PurchaseOrder.tipo_oc.in_(visible_order_types or {"__NONE__"}),
    )
    supplier_id = request.args.get("supplier_id", type=int)
    payment_type = normalize(request.args.get("tipo_pago"))
    start_raw = (request.args.get("fecha_desde") or "").strip()
    end_raw = (request.args.get("fecha_hasta") or "").strip()
    start = end = None
    try:
        start = datetime.strptime(start_raw, "%Y-%m-%d").date() if start_raw else None
        end = datetime.strptime(end_raw, "%Y-%m-%d").date() if end_raw else None
        if start and end and start > end:
            raise ValueError("La fecha inicial no puede ser posterior a la final.")
    except ValueError as exc:
        flash(str(exc) if "posterior" in str(exc) else "El rango de fechas no es válido.", "danger")
        start = end = None
    if supplier_id:
        query = query.filter(AdditionalPayment.supplier_id == supplier_id)
    if start:
        query = query.filter(AdditionalPayment.fecha >= start)
    if end:
        query = query.filter(AdditionalPayment.fecha <= end)
    if payment_type == "NOMINA":
        query = query.filter(PurchaseOrder.categoria_pago == "NOMINA")
    elif payment_type == "COMPRAS":
        query = query.filter(PurchaseOrder.categoria_pago == "COMPRAS")
    elif payment_type == "OPERACIONES":
        query = query.filter(PurchaseOrder.categoria_pago == "OPERACIONES")
    elif payment_type == "CREDITO":
        query = query.filter(PurchaseOrder.modalidad_pago == "CREDITO")
    elif payment_type:
        abort(400)
    payments = query.order_by(AdditionalPayment.fecha.desc(), AdditionalPayment.id.desc()).all()
    eligible = PurchaseOrder.query.filter(
        PurchaseOrder.project_id.in_(project_ids or [-1]),
        PurchaseOrder.tipo_oc.in_(allowed_order_types("crear") or {"__NONE__"}),
        PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES - {"CERRADA"}),
    ).all()
    has_payable = any(monto_pagable_programado(order) > 0 for order in eligible)
    suppliers = Supplier.query.filter_by(activo=True).order_by(Supplier.nombre).all()
    return render_template(
        "compras/payments/list.html",
        payments=payments,
        has_payable=has_payable,
        suppliers=suppliers,
        selected_supplier_id=supplier_id,
        selected_payment_type=payment_type,
        fecha_desde=start_raw,
        fecha_hasta=end_raw,
    )


@compras_bp.route("/pagos/nuevo", methods=["GET", "POST"])
@permission_required("pagos_proveedores", "pagar")
def supplier_payment_new():
    ensure_payment_methods()
    payable_types = allowed_order_types("ver")
    if not payable_types:
        abort(403)
    project_ids = [
        project.id
        for project in centros_operativos_accesibles(
            current_user, incluir_obras_inactivas=True
        )
    ]
    order_id = request.args.get("order_id", type=int) or request.form.get("purchase_order_id", type=int)
    order = db.session.get(PurchaseOrder, order_id) if order_id else None
    if order_id and order is None:
        abort(404)
    if order:
        require_project_access(order.project)
    if order and order.tipo_oc not in payable_types:
        abort(403)
    eligible_orders = PurchaseOrder.query.options(joinedload(PurchaseOrder.supplier), joinedload(PurchaseOrder.project), selectinload(PurchaseOrder.lines).joinedload(PurchaseOrderLine.explosion_item)).filter(
        PurchaseOrder.project_id.in_(project_ids or [-1]),
        PurchaseOrder.tipo_oc.in_(payable_types),
        PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES - {"CERRADA"}),
    ).order_by(PurchaseOrder.fecha_orden.desc()).all()
    eligible_orders = [
        candidate
        for candidate in eligible_orders
        if monto_pagable_programado(candidate) > 0
    ]
    if request.method == "POST":
        try:
            if not order or order.estado not in ACTIVE_ORDER_STATES or order.estado == "CERRADA":
                raise ValueError("Selecciona una OC emitida y vigente.")
            if (
                current_app.config.get("REQUIRE_THREE_WAY_MATCH", True)
                and not conciliacion_aprobada_para_pago(order)
            ):
                raise ValueError(
                    "El pago está bloqueado hasta que pedido, recepción y "
                    "factura coincidan en la conciliación de tres vías."
                )
            if order.es_operaciones and not order.beneficiario_validado:
                raise ValueError(
                    "Finanzas debe validar al beneficiario y sus datos antes de pagar."
                )
            line = db.session.get(PurchaseOrderLine, request.form.get("order_line_id", type=int))
            method = db.session.get(PaymentMethod, request.form.get("payment_method_id", type=int))
            company = db.session.get(Company, request.form.get("company_id", type=int))
            if not line or line.order_id != order.id:
                raise ValueError("Selecciona un renglón de la OC.")
            if not method or not method.activo or not company or not company.activa:
                raise ValueError("Método y empresa de pago activos son obligatorios.")
            if order.es_operaciones and (
                order.company_id != company.id
                or order.payment_method_id != method.id
            ):
                raise ValueError(
                    "La empresa y el método deben coincidir con la validación "
                    "financiera del beneficiario."
                )
            amount_type = normalize(request.form.get("tipo_monto"))
            captured = form_decimal("monto_capturado", required=True, positive=True, scale=MONEY_STEP)
            base_amount = amount_without_vat(captured, amount_type)
            paid_line = line_paid_amount(line)
            schedule = None
            if order.payment_schedules:
                schedule_id = request.form.get("payment_schedule_id", type=int)
                schedule = (
                    db.session.get(PurchaseOrderPaymentSchedule, schedule_id)
                    if schedule_id
                    else next(
                        (
                            item
                            for item in order.payment_schedules
                            if item.monto_liberado > 0
                        ),
                        None,
                    )
                )
                if not schedule or schedule.order_id != order.id:
                    raise ValueError("Selecciona una programación liberada.")
                if base_amount > schedule.monto_liberado:
                    raise ValueError(
                        f"El pago excede el monto liberado de la programación "
                        f"({schedule.monto_liberado})."
                    )
            if schedule and schedule.tipo == "ANTICIPO":
                line_limit = decimal_value(line.importe_sin_iva)
            elif order.modalidad_pago == "ANTICIPO" and order.estado in {"ANTICIPO_AUTORIZADO", "ANTICIPO_PARCIAL", "ANTICIPO_PAGADO"} and order.monto_recibido <= 0:
                line_limit = min(decimal_value(line.importe_sin_iva), decimal_value(order.anticipo_monto))
            else:
                line_limit = money(line.cantidad_recibida * decimal_value(line.precio_unitario_sin_iva))
            available_line = max(Decimal("0.00"), money(line_limit - paid_line))
            payable_order = monto_pagable_programado(order)
            if base_amount > available_line or base_amount > payable_order:
                raise ValueError(f"El pago excede el saldo habilitado ({min(available_line, payable_order)}).")
            payment = AdditionalPayment(
                fecha=form_date("fecha"),
                project_id=order.project_id,
                budget_item_id=line.explosion_item.budget_item_id,
                explosion_item_id=line.explosion_item_id,
                supplier_id=order.supplier_id,
                purchase_order_id=order.id,
                purchase_order_line_id=line.id,
                payment_method_id=method.id,
                payment_schedule_id=schedule.id if schedule else None,
                beneficiario=order.beneficiario_nombre,
                concepto=(request.form.get("concepto") or f"PAGO {order.folio}").strip(),
                monto_capturado=captured,
                tipo_monto=amount_type,
                monto_sin_iva=base_amount,
                metodo_pago=method.nombre,
                company_id=company.id,
                notas=(request.form.get("notas") or "").strip() or None,
                created_by_id=current_user.id,
            )
            db.session.add(payment)
            db.session.flush()
            if schedule:
                schedule.monto_pagado = money(
                    decimal_value(schedule.monto_pagado) + base_amount
                )
                schedule.estado = (
                    "PAGADO" if schedule.pendiente <= 0 else "PARCIAL"
                )
                if schedule.estado == "PAGADO":
                    schedule.fecha_pago = utc_now()
            order.anticipo_pendiente = max(Decimal("0"), money(decimal_value(order.anticipo_monto) - order.monto_pagado))
            refresh_order_status(order)
            actualizar_conciliacion_pagada(order)
            audit("PAGAR", "ORDEN_COMPRA", order.id, f"{payment.id} · {base_amount}")
            db.session.commit()
            flash("Pago registrado y aplicado al estado de cuenta del proveedor.", "success")
            return redirect(url_for("compras.order_detail", order_id=order.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    methods = PaymentMethod.query.filter_by(activo=True).order_by(PaymentMethod.nombre).all()
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    for candidate in eligible_orders:
        candidate.monto_pagable_actual = monto_pagable_programado(candidate)
    if order:
        order.monto_pagable_actual = monto_pagable_programado(order)
    db.session.commit()
    return render_template("compras/payments/form.html", orders=eligible_orders, order=order, methods=methods, companies=companies)


@compras_bp.get("/proveedores/<int:supplier_id>/estado-cuenta")
@permission_required("proveedores", "ver")
def supplier_statement(supplier_id):
    supplier = db.get_or_404(Supplier, supplier_id)
    project_ids = [p.id for p in accessible_projects_query().all()]
    orders = PurchaseOrder.query.options(joinedload(PurchaseOrder.project), selectinload(PurchaseOrder.lines)).filter(PurchaseOrder.supplier_id == supplier.id, PurchaseOrder.project_id.in_(project_ids or [-1]), PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES)).order_by(PurchaseOrder.fecha_vencimiento, PurchaseOrder.fecha_orden).all()
    current = today_value()
    for order in orders:
        order.due_color = order.semaforo_vencimiento(current)
        order.due_days = (order.fecha_vencimiento - current).days if order.fecha_vencimiento else None
    return render_template("compras/suppliers/statement.html", supplier=supplier, orders=orders)


# ---------------------------------------------------------------------------
# Saldos a favor derivados de anticipos
# ---------------------------------------------------------------------------


def line_payment_limit(order: PurchaseOrder, line: PurchaseOrderLine) -> Decimal:
    if (
        order.modalidad_pago == "ANTICIPO"
        and order.estado
        in {"ANTICIPO_AUTORIZADO", "ANTICIPO_PARCIAL", "ANTICIPO_PAGADO"}
        and order.monto_recibido <= 0
    ):
        return money(
            min(decimal_value(line.importe_sin_iva), decimal_value(order.anticipo_monto))
        )
    return line.monto_recibido


@compras_bp.get("/anticipos/saldos")
@any_permission_required(("compras", "ver"), ("oc_operaciones", "ver"))
def advance_balances():
    project_ids = [project.id for project in accessible_projects_query().all()]
    visible_types = allowed_order_types("ver")
    orders = (
        PurchaseOrder.query.options(
            joinedload(PurchaseOrder.supplier),
            joinedload(PurchaseOrder.company),
            selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.payments),
            selectinload(PurchaseOrder.lines).selectinload(
                PurchaseOrderLine.advance_movements_out
            ),
            selectinload(PurchaseOrder.advance_movements_out),
        )
        .filter(
            PurchaseOrder.project_id.in_(project_ids or [-1]),
            PurchaseOrder.tipo_oc.in_(visible_types or {"__NONE__"}),
            PurchaseOrder.modalidad_pago == "ANTICIPO",
            PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
        )
        .order_by(PurchaseOrder.fecha_orden.desc(), PurchaseOrder.id.desc())
        .all()
    )
    source_orders = [order for order in orders if order.saldo_favor_disponible > 0]
    target_orders = (
        PurchaseOrder.query.options(
            joinedload(PurchaseOrder.supplier),
            joinedload(PurchaseOrder.project),
            selectinload(PurchaseOrder.lines),
        )
        .filter(
            PurchaseOrder.project_id.in_(project_ids or [-1]),
            PurchaseOrder.tipo_oc.in_(visible_types or {"__NONE__"}),
            PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES - {"CERRADA"}),
        )
        .order_by(PurchaseOrder.fecha_orden.desc())
        .all()
    )
    target_orders = [order for order in target_orders if order.saldo_pagable > 0]
    movements = (
        SupplierAdvanceMovement.query.options(
            joinedload(SupplierAdvanceMovement.source_order),
            joinedload(SupplierAdvanceMovement.target_order),
            joinedload(SupplierAdvanceMovement.supplier),
            joinedload(SupplierAdvanceMovement.company),
            joinedload(SupplierAdvanceMovement.payment_method),
        )
        .join(PurchaseOrder, PurchaseOrder.id == SupplierAdvanceMovement.source_order_id)
        .filter(
            PurchaseOrder.project_id.in_(project_ids or [-1]),
            PurchaseOrder.tipo_oc.in_(visible_types or {"__NONE__"}),
        )
        .order_by(
            SupplierAdvanceMovement.fecha.desc(),
            SupplierAdvanceMovement.id.desc(),
        )
        .all()
    )
    methods = PaymentMethod.query.filter_by(activo=True).order_by(PaymentMethod.nombre).all()
    return render_template(
        "compras/advances/balances.html",
        source_orders=source_orders,
        target_orders=target_orders,
        movements=movements,
        methods=methods,
    )


@compras_bp.post("/anticipos/saldos/movimiento")
@any_permission_required(("compras", "editar"), ("oc_operaciones", "editar"))
def advance_balance_move():
    try:
        movement_type = normalize(request.form.get("tipo"))
        source = db.get_or_404(
            PurchaseOrder, request.form.get("source_order_id", type=int)
        )
        source_line = db.get_or_404(
            PurchaseOrderLine, request.form.get("source_order_line_id", type=int)
        )
        method = db.get_or_404(
            PaymentMethod, request.form.get("payment_method_id", type=int)
        )
        amount = form_decimal(
            "monto", required=True, positive=True, scale=MONEY_STEP
        )
        if source_line.order_id != source.id or source.modalidad_pago != "ANTICIPO":
            raise ValueError("Selecciona un renglón válido de una OC con anticipo.")
        require_order_permission(source, "editar")
        require_project_access(source.project)
        if not source.receipts:
            raise ValueError(
                "El saldo a favor solo puede conciliarse después de registrar una recepción."
            )
        if amount > source_line.saldo_favor_disponible:
            raise ValueError(
                f"El movimiento excede el saldo a favor del renglón ({source_line.saldo_favor_disponible})."
            )
        if not method.activo or movement_type not in {"APLICACION", "REEMBOLSO"}:
            raise ValueError("Selecciona tipo y método activos.")

        target = None
        target_line = None
        if movement_type == "APLICACION":
            target = db.get_or_404(
                PurchaseOrder, request.form.get("target_order_id", type=int)
            )
            target_line = db.get_or_404(
                PurchaseOrderLine,
                request.form.get("target_order_line_id", type=int),
            )
            require_order_permission(target, "editar")
            require_project_access(target.project)
            if target.id == source.id or target_line.order_id != target.id:
                raise ValueError("Selecciona otra OC y uno de sus renglones.")
            if (
                target.supplier_id != source.supplier_id
                or target.company_id != source.company_id
            ):
                raise ValueError(
                    "El saldo solo puede aplicarse al mismo proveedor y empresa pagadora."
                )
            if target.estado not in ACTIVE_ORDER_STATES - {"CERRADA"}:
                raise ValueError("La OC destino no está vigente.")
            paid = line_paid_amount(target_line)
            available_line = max(
                Decimal("0.00"),
                money(line_payment_limit(target, target_line) - paid),
            )
            if amount > available_line or amount > target.saldo_pagable:
                raise ValueError(
                    f"La aplicación excede el saldo habilitado de la OC destino ({min(available_line, target.saldo_pagable)})."
                )

        movement = SupplierAdvanceMovement(
            source_order=source,
            source_order_line=source_line,
            target_order=target,
            target_order_line=target_line,
            supplier_id=source.supplier_id,
            company_id=source.company_id,
            payment_method_id=method.id,
            tipo=movement_type,
            fecha=form_date("fecha"),
            monto=amount,
            referencia=(request.form.get("referencia") or "").strip() or None,
            notas=(request.form.get("notas") or "").strip() or None,
            created_by_id=current_user.id,
        )
        db.session.add(movement)
        db.session.flush()
        refresh_order_status(source)
        if target:
            refresh_order_status(target)
            notify(
                [target.buyer],
                "SALDO_ANTICIPO_APLICADO",
                f"Se aplicaron {amount} MXN de {source.folio} a {target.folio}.",
                f"/compras/ordenes/{target.id}",
            )
        audit(
            "APLICAR" if movement_type == "APLICACION" else "REEMBOLSAR",
            "SALDO_ANTICIPO",
            movement.id,
            f"{source.folio} · {amount}",
        )
        db.session.commit()
        flash(
            "Saldo aplicado a la OC destino."
            if movement_type == "APLICACION"
            else "Reembolso del saldo a favor registrado.",
            "success",
        )
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("compras.advance_balances"))


# ---------------------------------------------------------------------------
# Tarjetas de crédito de empresas pagadoras
# ---------------------------------------------------------------------------


def masked_card_number(value: str | None) -> str:
    raw = re.sub(r"\s+", " ", (value or "").strip())
    digits = re.sub(r"\D", "", raw)
    if len(digits) != 4:
        raise ValueError(
            "Captura únicamente los últimos cuatro dígitos de la tarjeta."
        )
    if "*" not in raw and len(raw) > 4:
        raise ValueError("El número de tarjeta debe estar enmascarado.")
    return f"**** **** **** {digits}"


def fill_credit_card(card: CreditCard, form: CreditCardForm) -> None:
    company = db.session.get(Company, form.empresa_id.data)
    if not company or not company.activa:
        raise ValueError("Selecciona una empresa pagadora activa.")
    if form.fecha_pago.data < form.fecha_corte.data:
        raise ValueError("La fecha de pago no puede ser anterior al corte.")
    if decimal_value(form.saldo_actual.data) > decimal_value(
        form.limite_credito.data
    ):
        raise ValueError("El saldo actual no puede superar el límite de crédito.")
    card.empresa_id = company.id
    card.numero_tarjeta = masked_card_number(form.numero_tarjeta.data)
    card.fecha_corte = form.fecha_corte.data
    card.fecha_pago = form.fecha_pago.data
    card.limite_credito = money(form.limite_credito.data)
    card.saldo_actual = money(form.saldo_actual.data)
    card.payment_due_notified_on = None


@compras_bp.get("/tarjetas")
@permission_required("tarjetas_credito", "ver")
def credit_cards_list():
    current = today_value()
    cards = (
        CreditCard.query.options(
            joinedload(CreditCard.empresa), selectinload(CreditCard.pagos)
        )
        .order_by(CreditCard.activa.desc(), CreditCard.fecha_pago, CreditCard.id)
        .all()
    )
    for card in cards:
        card.due_days = card.dias_para_pago(current)
        card.alert_due = card.saldo_actual > 0 and card.due_days <= 3
    return render_template("compras/cards/list.html", cards=cards)


@compras_bp.route("/tarjetas/nueva", methods=["GET", "POST"])
@permission_required("tarjetas_credito", "crear")
def credit_card_new():
    form = CreditCardForm()
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    if form.validate_on_submit():
        try:
            card = CreditCard(activa=True)
            fill_credit_card(card, form)
            db.session.add(card)
            db.session.flush()
            audit("CREAR", "TARJETA_CREDITO", card.id, card.numero_tarjeta)
            db.session.commit()
            flash("Tarjeta registrada correctamente.", "success")
            return redirect(url_for("compras.credit_cards_list"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "compras/cards/form.html", form=form, companies=companies, card=None
    )


@compras_bp.route("/tarjetas/<int:card_id>/editar", methods=["GET", "POST"])
@permission_required("tarjetas_credito", "editar")
def credit_card_edit(card_id):
    card = db.get_or_404(CreditCard, card_id)
    form = CreditCardForm(obj=card)
    if request.method == "GET":
        form.empresa_id.data = card.empresa_id
    companies = Company.query.filter_by(activa=True).order_by(Company.nombre).all()
    if form.validate_on_submit():
        try:
            fill_credit_card(card, form)
            audit("EDITAR", "TARJETA_CREDITO", card.id, card.numero_tarjeta)
            db.session.commit()
            flash("Tarjeta actualizada correctamente.", "success")
            return redirect(url_for("compras.credit_cards_list"))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "compras/cards/form.html", form=form, companies=companies, card=card
    )


@compras_bp.route("/tarjetas/<int:card_id>/pagos/nuevo", methods=["GET", "POST"])
@permission_required("tarjetas_credito", "crear")
def credit_card_payment_new(card_id):
    card = db.get_or_404(CreditCard, card_id)
    form = CreditCardPaymentForm()
    if request.method == "GET":
        form.fecha.data = today_value()
    if form.validate_on_submit():
        amount = money(form.monto.data)
        previous_balance = money(card.saldo_actual)
        if amount > previous_balance:
            flash(
                f"El pago supera el saldo actual ({previous_balance:,.2f}).",
                "danger",
            )
        else:
            new_balance = money(previous_balance - amount)
            payment = CreditCardPayment(
                tarjeta_id=card.id,
                fecha=form.fecha.data,
                monto=amount,
                saldo_anterior=previous_balance,
                saldo_nuevo=new_balance,
                referencia=(form.referencia.data or "").strip() or None,
                notas=(form.notas.data or "").strip() or None,
                created_by_id=current_user.id,
            )
            card.saldo_actual = new_balance
            db.session.add(payment)
            db.session.flush()
            audit(
                "PAGAR",
                "TARJETA_CREDITO",
                card.id,
                f"{payment.id} · {amount}",
            )
            db.session.commit()
            flash("Pago de tarjeta registrado y saldo actualizado.", "success")
            return redirect(url_for("compras.credit_cards_list"))
    return render_template(
        "compras/cards/payment_form.html", form=form, card=card
    )


@compras_bp.post("/tarjetas/<int:card_id>/estado")
@permission_required("tarjetas_credito", "eliminar")
def credit_card_toggle(card_id):
    card = db.get_or_404(CreditCard, card_id)
    card.activa = not card.activa
    audit(
        "ACTIVAR" if card.activa else "DESACTIVAR",
        "TARJETA_CREDITO",
        card.id,
        card.numero_tarjeta,
    )
    db.session.commit()
    flash("Estado de la tarjeta actualizado.", "success")
    return redirect(url_for("compras.credit_cards_list"))


# ---------------------------------------------------------------------------
# SMNC
# ---------------------------------------------------------------------------


@compras_bp.get("/smnc")
@permission_required("smnc", "ver")
def smnc_list():
    project_ids = [
        p.id
        for p in centros_operativos_accesibles(
            current_user, incluir_obras_inactivas=True
        )
    ]
    requests = MaterialChangeRequest.query.options(joinedload(MaterialChangeRequest.project), joinedload(MaterialChangeRequest.requested_by), selectinload(MaterialChangeRequest.details)).filter(MaterialChangeRequest.project_id.in_(project_ids or [-1])).order_by(MaterialChangeRequest.created_at.desc()).all()
    return render_template("compras/smnc/list.html", requests=requests)


@compras_bp.route("/smnc/nueva", methods=["GET", "POST"])
@permission_required("smnc", "crear")
def smnc_new():
    projects = centros_operativos_accesibles(current_user)
    garantia_id = request.args.get("garantia_id", type=int) or request.form.get(
        "garantia_id", type=int
    )
    warranty = None
    if garantia_id:
        from fase5_models import GarantiaObra

        warranty = db.get_or_404(GarantiaObra, garantia_id)
        require_project_access(warranty.centro_garantia)
        if warranty.estado in {"cerrada", "rechazada"}:
            abort(404)
    selected_project_id = (
        warranty.centro_garantia_id
        if warranty
        else request.args.get("project_id", type=int)
        or request.form.get("project_id", type=int)
    )
    items = BudgetItem.query.filter_by(project_id=selected_project_id, activa=True).order_by(BudgetItem.codigo).all() if selected_project_id else []
    entries = items_explosion_vigente(selected_project_id) if selected_project_id else []
    if request.method == "POST":
        try:
            project = operational_project_or_404(selected_project_id)
            if project.tipo == "garantia":
                if not warranty or warranty.centro_garantia_id != project.id:
                    raise ValueError(
                        "Selecciona la garantía a la que pertenece la SMNC."
                    )
            item = db.session.get(BudgetItem, request.form.get("budget_item_id", type=int))
            action = normalize(request.form.get("action_type"))
            if not item or item.project_id != project.id or action not in {"NUEVO", "AUMENTO"}:
                raise ValueError("Selecciona una partida y un tipo de solicitud válidos.")
            existing = db.session.get(BudgetExplosionItem, request.form.get("existing_explosion_item_id", type=int)) if action == "AUMENTO" else None
            if action == "AUMENTO" and (not existing or existing.project_id != project.id):
                raise ValueError("Selecciona el insumo existente que aumentará.")
            classification = (
                existing.clasificacion
                if existing
                else (
                    normalize(request.form.get("clasificacion")).replace(" ", "_")
                    or "NORMAL"
                )
            )
            if classification not in EXPLOSION_CLASSIFICATIONS:
                raise ValueError(
                    "Selecciona una clasificación válida para la explosión."
                )
            description = existing.supply_item.descripcion if existing else normalize(request.form.get("descripcion"))
            unit = existing.supply_item.unidad if existing else normalize(request.form.get("unidad"))
            supply_type = existing.supply_item.tipo if existing else normalize(request.form.get("supply_type"))
            if not description or not unit or supply_type not in SUPPLY_TYPES:
                raise ValueError("Descripción, unidad y tipo de insumo son obligatorios.")
            smnc = MaterialChangeRequest(
                folio=next_folio(MaterialChangeRequest, "SMNC"),
                project_id=project.id,
                garantia_id=warranty.id if warranty else None,
                estado="PENDIENTE_AUTORIZACION",
                requested_by_id=current_user.id,
            )
            smnc.details.append(MaterialChangeRequestLine(
                budget_item_id=item.id,
                existing_explosion_item_id=existing.id if existing else None,
                action_type=action,
                supply_key=existing.supply_item.clave if existing else normalize(request.form.get("supply_key")),
                supply_type=supply_type,
                clasificacion=classification,
                descripcion=description,
                unidad=unit,
                cantidad=form_decimal("cantidad", required=True, positive=True),
                precio_estimado=form_decimal("precio_estimado", required=True, positive=True),
                justificacion_tipo=normalize(request.form.get("justificacion_tipo")),
                justificacion=(request.form.get("justificacion") or "").strip(),
            ))
            if not smnc.details[0].supply_key or not smnc.details[0].justificacion:
                raise ValueError("Clave de insumo y justificación son obligatorias.")
            db.session.add(smnc)
            db.session.flush()
            notify(
                usuarios_con_permiso(
                    "smnc", "aprobar", centro_costo_id=project.id
                ),
                "SMNC_PENDIENTE",
                f"{smnc.folio} requiere autorización.",
                url_for("compras.smnc_detail", smnc_id=smnc.id),
            )
            audit("CREAR", "SMNC", smnc.id, smnc.folio)
            db.session.commit()
            flash("SMNC enviada a autorización.", "success")
            return redirect(url_for("compras.smnc_detail", smnc_id=smnc.id))
        except ValueError as exc:
            db.session.rollback()
            flash(str(exc), "danger")
    return render_template(
        "compras/smnc/form.html",
        projects=projects,
        selected_project_id=selected_project_id,
        items=items,
        entries=entries,
        garantia=warranty,
    )


@compras_bp.get("/smnc/<int:smnc_id>")
@permission_required("smnc", "ver")
def smnc_detail(smnc_id):
    smnc = db.get_or_404(MaterialChangeRequest, smnc_id)
    require_project_access(smnc.project)
    return render_template("compras/smnc/detail.html", smnc=smnc)


@compras_bp.post("/smnc/<int:smnc_id>/aprobar")
@permission_required("smnc", "aprobar")
def smnc_approve(smnc_id):
    smnc = db.get_or_404(MaterialChangeRequest, smnc_id)
    require_project_access(smnc.project)
    if smnc.estado != "PENDIENTE_AUTORIZACION":
        flash("La SMNC ya fue atendida.", "danger")
        return redirect(url_for("compras.smnc_detail", smnc_id=smnc.id))
    try:
        for line in smnc.details:
            if line.action_type == "AUMENTO":
                entry = line.existing_explosion_item
                entry.cantidad_presupuestada = decimal_value(entry.cantidad_presupuestada) + decimal_value(line.cantidad)
                entry.importe_presupuestado = money(entry.cantidad_presupuestada * decimal_value(entry.precio_unitario_sin_iva))
            else:
                operation_category = (
                    detect_operation_category(line.descripcion, line.supply_type)
                    if line.clasificacion == "OPERATIVO"
                    else None
                )
                supply = SupplyItem.query.filter(
                    func.upper(SupplyItem.clave) == line.supply_key
                ).first()
                if supply and (
                    supply.tipo != line.supply_type
                    or supply.unidad != line.unidad
                ):
                    raise ValueError(
                        f"La clave {line.supply_key} existe con tipo o unidad diferente."
                    )
                if not supply:
                    supply = SupplyItem(
                        clave=line.supply_key,
                        descripcion=line.descripcion,
                        tipo=line.supply_type,
                        unidad=line.unidad,
                        clave_sat="00000000",
                        moneda="MXN",
                        es_operacion=line.clasificacion == "OPERATIVO",
                        categoria_operacion=operation_category,
                        activo=True,
                    )
                    db.session.add(supply)
                    db.session.flush()
                revision = revision_explosion_vigente(smnc.project_id)
                if not revision:
                    revision = ExplosionRevision(
                        project_id=smnc.project_id,
                        numero_revision=siguiente_revision_explosion(smnc.project_id),
                        estado="VIGENTE",
                        es_vigente=True,
                        archivo_origen="SMNC",
                        observaciones="Revisión creada por SMNC aprobada.",
                        loaded_by_id=current_user.id,
                    )
                    db.session.add(revision)
                    db.session.flush()
                existing_in_revision = BudgetExplosionItem.query.filter_by(
                    revision_id=revision.id,
                    budget_item_id=line.budget_item_id,
                    supply_item_id=supply.id,
                ).first()
                if existing_in_revision:
                    raise ValueError(
                        f"{line.supply_key} ya está contemplado; usa Aumento."
                    )
                entry = BudgetExplosionItem(
                    revision_id=revision.id,
                    project_id=smnc.project_id,
                    budget_item_id=line.budget_item_id,
                    supply_item_id=supply.id,
                    cantidad_presupuestada=line.cantidad,
                    precio_unitario_sin_iva=line.precio_estimado,
                    importe_presupuestado=line.importe_estimado,
                    clasificacion=line.clasificacion,
                    requiere_autorizacion_previa=line.clasificacion
                    in {"EQUIPO_ESPECIAL", "ELECTRODOMESTICO"},
                    observacion_clasificacion=line.justificacion,
                    origen="SMNC",
                    activo=True,
                    created_by_id=current_user.id,
                )
                db.session.add(entry)
                db.session.flush()
            line.generated_explosion_item_id = entry.id
        smnc.estado = "APROBADA"
        smnc.approved_by_id = current_user.id
        smnc.approved_at = utc_now()
        notify([smnc.requested_by], "SMNC_APROBADA", f"{smnc.folio} fue aprobada; el insumo ya está disponible.", url_for("compras.smnc_detail", smnc_id=smnc.id))
        audit("APROBAR", "SMNC", smnc.id, smnc.folio)
        db.session.commit()
        flash("SMNC aprobada y explosión actualizada.", "success")
    except ValueError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
    return redirect(url_for("compras.smnc_detail", smnc_id=smnc.id))


@compras_bp.post("/smnc/<int:smnc_id>/rechazar")
@permission_required("smnc", "aprobar")
def smnc_reject(smnc_id):
    smnc = db.get_or_404(MaterialChangeRequest, smnc_id)
    require_project_access(smnc.project)
    reason = (request.form.get("motivo_rechazo") or "").strip()
    if smnc.estado != "PENDIENTE_AUTORIZACION" or not reason:
        flash("El motivo de rechazo es obligatorio.", "danger")
    else:
        smnc.estado = "RECHAZADA"
        smnc.rejection_reason = reason
        notify([smnc.requested_by], "SMNC_RECHAZADA", f"{smnc.folio} fue rechazada: {reason}", url_for("compras.smnc_detail", smnc_id=smnc.id))
        db.session.commit()
    return redirect(url_for("compras.smnc_detail", smnc_id=smnc.id))


# ---------------------------------------------------------------------------
# Reportes
# ---------------------------------------------------------------------------


def filtered_date_range() -> tuple[date, date]:
    default_start, default_end = week_bounds(today_value())
    start_raw = (request.args.get("fecha_desde") or default_start.isoformat()).strip()
    end_raw = (request.args.get("fecha_hasta") or default_end.isoformat()).strip()
    try:
        start = datetime.strptime(start_raw, "%Y-%m-%d").date()
        end = datetime.strptime(end_raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("El rango de fechas no es válido.") from exc
    if start > end:
        raise ValueError("La fecha inicial no puede ser posterior a la final.")
    return start, end


@compras_bp.get("/reportes/nominas-operaciones")
@permission_required("reportes_compras", "ver")
def payroll_operations_report():
    """Recurso real por semana, obra y método de pago."""

    if not current_user.tiene_permiso("reportes_nomina", "ver"):
        abort(403)
    projects = accessible_projects_query().all()
    project_ids = [project.id for project in projects]
    selected_project_id = request.args.get("project_id", type=int)
    if selected_project_id:
        if selected_project_id not in project_ids:
            abort(404)
        scoped_ids = [selected_project_id]
    else:
        scoped_ids = project_ids
    try:
        start, end = filtered_date_range()
    except ValueError as exc:
        flash(str(exc), "danger")
        start, end = week_bounds(today_value())
    weeks = week_starts_between(start, end)
    summaries = []
    totals = {
        **{component: Decimal("0") for component in RESOURCE_COMPONENTS},
        "efectivo": Decimal("0"),
        "transferencia": Decimal("0"),
        "total": Decimal("0"),
    }
    for week in weeks:
        for project in projects:
            if project.id not in scoped_ids:
                continue
            resource = weekly_resource_breakdown(week, [project.id])
            if resource["requerido_total"] <= 0:
                continue
            summary = {
                "semana": resource["week_start"],
                "project": project,
                "methods": resource["methods"],
                "total": resource["requerido_total"],
            }
            summaries.append(summary)
            cash = resource["methods"]["EFECTIVO"]
            bank = resource["methods"]["TRANSFERENCIA"]
            for component in RESOURCE_COMPONENTS:
                totals[component] += cash[component] + bank[component]
            totals["efectivo"] += cash["requerido"]
            totals["transferencia"] += bank["requerido"]
            totals["total"] += resource["requerido_total"]
    for key in totals:
        totals[key] = money(totals[key])

    # Las OC emitidas se conservan como trazabilidad documental. Su valor
    # nominal no se suma al requerido: el egreso real entra cuando existe un
    # AdditionalPayment y la fuente única lo clasifica como operación.
    operation_orders = (
        PurchaseOrder.query.options(
            joinedload(PurchaseOrder.project),
            joinedload(PurchaseOrder.supplier),
            selectinload(PurchaseOrder.lines),
        )
        .filter(
            PurchaseOrder.project_id.in_(scoped_ids or [-1]),
            PurchaseOrder.tipo_oc == "OPERACIONES",
            PurchaseOrder.fecha_orden.between(start, end),
            PurchaseOrder.estado.notin_({"BORRADOR", "CANCELADA"}),
        )
        .all()
    )
    summaries.sort(key=lambda row: (row["semana"], row["project"].codigo))

    if request.args.get("export") == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Recurso semanal"
        sheet.append(
            [
                "Semana",
                "Obra",
                "Método",
                "Nómina",
                "Préstamos nuevos",
                "Gastos operativos",
                "Pagos adicionales",
                "Subcontratos",
                "Requerido",
            ]
        )
        for summary in summaries:
            for method in ("EFECTIVO", "TRANSFERENCIA"):
                row = summary["methods"][method]
                sheet.append(
                    [
                        summary["semana"],
                        f"{summary['project'].codigo} · {summary['project'].nombre}",
                        method,
                        float(row["nomina"]),
                        float(row["prestamos"]),
                        float(row["gastos_operativos"]),
                        float(row["pagos_adicionales"]),
                        float(row["subcontratos"]),
                        float(row["requerido"]),
                    ]
                )
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="17365D")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row_index, 1).number_format = "dd/mm/yyyy"
            for column in range(4, 10):
                sheet.cell(row_index, column).number_format = (
                    '$#,##0.00;[Red]-$#,##0.00'
                )
        for index, width in enumerate((14, 34, 18, 18, 20, 20, 20, 18, 18), 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"recurso_semanal_{start}_{end}.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    return render_template(
        "compras/reports/payroll_operations.html",
        projects=projects,
        selected_project_id=selected_project_id,
        fecha_desde=start,
        fecha_hasta=end,
        summaries=summaries,
        operation_orders=operation_orders,
        totals=totals,
    )


@compras_bp.get("/reportes/pagos-proveedores")
@permission_required("reportes_compras", "ver")
def supplier_payments_report():
    """OC de Compras, pagos por categoría, crédito y tarjetas empresariales."""

    try:
        start, end = filtered_date_range()
    except ValueError as exc:
        flash(str(exc), "danger")
        start, end = week_bounds(today_value())
    supplier_id = request.args.get("supplier_id", type=int)
    payment_type = normalize(request.args.get("tipo_pago"))
    visible_types = allowed_order_types("ver")
    project_ids = [
        project.id
        for project in centros_operativos_accesibles(
            current_user, incluir_obras_inactivas=True
        )
    ]

    payments_query = (
        AdditionalPayment.query.options(
            joinedload(AdditionalPayment.project),
            joinedload(AdditionalPayment.supplier),
            joinedload(AdditionalPayment.purchase_order),
            joinedload(AdditionalPayment.company),
        )
        .join(PurchaseOrder, PurchaseOrder.id == AdditionalPayment.purchase_order_id)
        .filter(
            PurchaseOrder.tipo_oc.in_(visible_types or {"__NONE__"}),
            AdditionalPayment.project_id.in_(project_ids or [-1]),
            AdditionalPayment.fecha.between(start, end),
        )
    )
    if supplier_id:
        payments_query = payments_query.filter(
            AdditionalPayment.supplier_id == supplier_id
        )
    if payment_type == "NOMINA":
        payments_query = payments_query.filter(PurchaseOrder.categoria_pago == "NOMINA")
    elif payment_type == "COMPRAS":
        payments_query = payments_query.filter(PurchaseOrder.categoria_pago == "COMPRAS")
    elif payment_type == "OPERACIONES":
        payments_query = payments_query.filter(PurchaseOrder.categoria_pago == "OPERACIONES")
    elif payment_type == "CREDITO":
        payments_query = payments_query.filter(PurchaseOrder.modalidad_pago == "CREDITO")
    elif payment_type:
        abort(400)
    payments = payments_query.order_by(
        AdditionalPayment.fecha.desc(), AdditionalPayment.id.desc()
    ).all()

    orders_query = PurchaseOrder.query.options(
        joinedload(PurchaseOrder.project), joinedload(PurchaseOrder.supplier)
    ).filter(
        PurchaseOrder.project_id.in_(project_ids or [-1]),
        PurchaseOrder.tipo_oc == "COMPRAS",
        PurchaseOrder.fecha_orden.between(start, end),
        PurchaseOrder.estado.notin_({"BORRADOR", "CANCELADA"}),
    )
    if supplier_id:
        orders_query = orders_query.filter(PurchaseOrder.supplier_id == supplier_id)
    purchase_orders = orders_query.order_by(PurchaseOrder.fecha_orden.desc()).all()

    card_payments = []
    if current_user.tiene_permiso("tarjetas_credito", "ver") and not supplier_id:
        card_payments = (
            CreditCardPayment.query.options(
                joinedload(CreditCardPayment.tarjeta).joinedload(CreditCard.empresa)
            )
            .filter(CreditCardPayment.fecha.between(start, end))
            .order_by(CreditCardPayment.fecha.desc(), CreditCardPayment.id.desc())
            .all()
        )
    suppliers = Supplier.query.order_by(Supplier.nombre).all()
    totals = {
        "nomina": money(
            sum(
                (
                    decimal_value(payment.monto_sin_iva)
                    for payment in payments
                    if payment.purchase_order.categoria_pago == "NOMINA"
                ),
                Decimal("0"),
            )
        ),
        "compras": money(
            sum(
                (
                    decimal_value(payment.monto_sin_iva)
                    for payment in payments
                    if payment.purchase_order.categoria_pago == "COMPRAS"
                ),
                Decimal("0"),
            )
        ),
        "operaciones": money(
            sum(
                (
                    decimal_value(payment.monto_sin_iva)
                    for payment in payments
                    if payment.purchase_order.categoria_pago == "OPERACIONES"
                ),
                Decimal("0"),
            )
        ),
        "credito": money(
            sum(
                (
                    decimal_value(payment.monto_sin_iva)
                    for payment in payments
                    if payment.purchase_order.modalidad_pago == "CREDITO"
                ),
                Decimal("0"),
            )
        ),
        "tarjetas": money(
            sum((decimal_value(payment.monto) for payment in card_payments), Decimal("0"))
        ),
    }
    return render_template(
        "compras/reports/supplier_payments_admin.html",
        payments=payments,
        purchase_orders=purchase_orders,
        card_payments=card_payments,
        suppliers=suppliers,
        selected_supplier_id=supplier_id,
        selected_payment_type=payment_type,
        fecha_desde=start,
        fecha_hasta=end,
        totals=totals,
    )


REPORT_STATES = {
    "": "Todos los estados",
    "REQUISITADO": "Requisitado",
    "APROBADO": "Aprobado",
    "EN_OC": "En OC",
    "RECEPCIONADO": "Recepcionado",
    "PAGADO": "Pagado",
}


def report_column(key, label, kind="text", *, total=False):
    return {"key": key, "label": label, "kind": kind, "total": total}


REPORT_DEFINITIONS = {
    "cantidades": {
        "title": "Cantidades por insumo y obra",
        "description": "Presupuestado, requisitado, aprobado, comprado, recibido y pagado.",
        "columns": [
            report_column("obra", "Obra"),
            report_column("partida", "Partida"),
            report_column("subpartida", "Subpartida"),
            report_column("clave", "Clave insumo"),
            report_column("descripcion", "Descripción"),
            report_column("unidad", "Unidad"),
            report_column("cantidad_presupuestada", "Cant. presupuestada", "quantity", total=True),
            report_column("importe_presupuestado", "Imp. presupuestado", "money", total=True),
            report_column("cantidad_requisitada", "Cant. requisitada", "quantity", total=True),
            report_column("importe_requisitado", "Imp. requisitado", "money", total=True),
            report_column("cantidad_aprobada", "Cant. aprobada", "quantity", total=True),
            report_column("importe_aprobado", "Imp. aprobado", "money", total=True),
            report_column("cantidad_comprada", "Cant. comprada", "quantity", total=True),
            report_column("importe_comprado", "Imp. comprado", "money", total=True),
            report_column("cantidad_recibida", "Cant. recibida", "quantity", total=True),
            report_column("importe_recibido", "Imp. recibido", "money", total=True),
            report_column("cantidad_pagada", "Cant. pagada", "quantity", total=True),
            report_column("importe_pagado", "Imp. pagado", "money", total=True),
        ],
    },
    "comprados": {
        "title": "Materiales requisitados y comprados",
        "description": "Renglones que ya se convirtieron en Orden de Compra.",
        "columns": [
            report_column("fecha", "Fecha compra", "date"),
            report_column("obra", "Obra"),
            report_column("oc", "OC"),
            report_column("requisicion", "Requisición"),
            report_column("proveedor", "Proveedor"),
            report_column("partida", "Partida"),
            report_column("clave", "Clave insumo"),
            report_column("descripcion", "Descripción"),
            report_column("unidad", "Unidad"),
            report_column("cantidad", "Cantidad", "quantity", total=True),
            report_column("precio_unitario", "P.U. real", "money"),
            report_column("importe", "Importe", "money", total=True),
            report_column("estado", "Estado OC"),
        ],
    },
    "pendientes": {
        "title": "Materiales requisitados sin compra",
        "description": "Cantidades aprobadas aún no cubiertas totalmente por una OC.",
        "columns": [
            report_column("fecha", "Fecha aprobación", "date"),
            report_column("obra", "Obra"),
            report_column("requisicion", "Requisición"),
            report_column("limite", "Límite OC", "date"),
            report_column("partida", "Partida"),
            report_column("clave", "Clave insumo"),
            report_column("descripcion", "Descripción"),
            report_column("unidad", "Unidad"),
            report_column("cantidad_aprobada", "Aprobada", "quantity", total=True),
            report_column("cantidad_comprada", "Comprada", "quantity", total=True),
            report_column("cantidad_pendiente", "Pendiente", "quantity", total=True),
            report_column("importe_pendiente", "Importe pendiente", "money", total=True),
            report_column("estado", "Estado"),
        ],
    },
    "proveedores": {
        "title": "Compras por proveedor",
        "description": "Volumen comprado por proveedor, desglosado por obra.",
        "columns": [
            report_column("proveedor", "Proveedor"),
            report_column("rfc", "RFC"),
            report_column("obra", "Obra"),
            report_column("ordenes", "OC", "integer", total=True),
            report_column("comprado", "Comprado", "money", total=True),
            report_column("recibido", "Recibido", "money", total=True),
            report_column("pagado", "Pagado", "money", total=True),
            report_column("saldo", "Saldo", "money", total=True),
        ],
    },
    "obras": {
        "title": "Compras por obra",
        "description": "Detalle de partidas e insumos comprados en cada obra.",
        "columns": [
            report_column("obra", "Obra"),
            report_column("partida", "Partida"),
            report_column("subpartida", "Subpartida"),
            report_column("clave", "Clave insumo"),
            report_column("descripcion", "Descripción"),
            report_column("unidad", "Unidad"),
            report_column("cantidad", "Cantidad comprada", "quantity", total=True),
            report_column("comprado", "Comprado", "money", total=True),
            report_column("recibido", "Recibido", "money", total=True),
            report_column("pagado", "Pagado", "money", total=True),
            report_column("saldo", "Saldo", "money", total=True),
        ],
    },
}


def in_report_period(value, start: date, end: date) -> bool:
    if value is None:
        return False
    candidate = value.date() if isinstance(value, datetime) else value
    return start <= candidate <= end


def report_date_range() -> tuple[date, date, str]:
    period = request.args.get("period", "month")
    current = today_value()
    if period == "custom":
        try:
            start = datetime.strptime(request.args.get("date_from"), "%Y-%m-%d").date()
            end = datetime.strptime(request.args.get("date_to"), "%Y-%m-%d").date()
        except (TypeError, ValueError) as exc:
            raise ValueError("Captura un rango de fechas válido.") from exc
        period = "custom"
    elif period == "week":
        start, end = week_bounds(current)
    else:
        start = current.replace(day=1)
        next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = next_month - timedelta(days=1)
        period = "month"
    if start > end:
        raise ValueError("La fecha inicial no puede ser posterior a la final.")
    return start, end, period


def budget_path_labels(item: BudgetItem) -> tuple[str, str]:
    if item.parent:
        return item.parent.nombre, item.nombre
    return item.nombre, ""


def report_state_matches(state: str, values: dict[str, Decimal]) -> bool:
    if not state:
        return True
    key = {
        "REQUISITADO": "requested",
        "APROBADO": "approved",
        "EN_OC": "ordered",
        "RECEPCIONADO": "received",
        "PAGADO": "paid",
    }[state]
    return decimal_value(values.get(key)) > 0


def scoped_report_entries(project_ids: list[int]):
    return (
        BudgetExplosionItem.query.options(
            joinedload(BudgetExplosionItem.project),
            joinedload(BudgetExplosionItem.budget_item).joinedload(BudgetItem.parent),
            joinedload(BudgetExplosionItem.supply_item),
            selectinload(BudgetExplosionItem.requisition_lines).joinedload(
                PurchaseRequisitionLine.requisition
            ),
            selectinload(BudgetExplosionItem.order_lines).joinedload(
                PurchaseOrderLine.order
            ).joinedload(PurchaseOrder.supplier),
            selectinload(BudgetExplosionItem.order_lines)
            .selectinload(PurchaseOrderLine.receipt_lines)
            .joinedload(GoodsReceiptLine.receipt),
        )
        .filter(
            BudgetExplosionItem.project_id.in_(project_ids or [-1]),
            BudgetExplosionItem.origen != "HISTORICO",
        )
        .order_by(
            BudgetExplosionItem.project_id,
            BudgetExplosionItem.budget_item_id,
            BudgetExplosionItem.id,
        )
        .all()
    )


def build_quantities_report(start, end, project_ids, supplier_id, state):
    payments = (
        AdditionalPayment.query.options(joinedload(AdditionalPayment.purchase_order_line))
        .join(PurchaseOrder, PurchaseOrder.id == AdditionalPayment.purchase_order_id)
        .filter(
            AdditionalPayment.project_id.in_(project_ids or [-1]),
            AdditionalPayment.fecha.between(start, end),
            AdditionalPayment.purchase_order_id.isnot(None),
            PurchaseOrder.tipo_oc == "COMPRAS",
        )
        .all()
    )
    if supplier_id:
        payments = [p for p in payments if p.supplier_id == supplier_id]
    payment_map = defaultdict(list)
    for payment in payments:
        payment_map[payment.explosion_item_id].append(payment)
    movements = (
        SupplierAdvanceMovement.query.options(
            joinedload(SupplierAdvanceMovement.source_order),
            joinedload(SupplierAdvanceMovement.target_order),
        )
        .filter(SupplierAdvanceMovement.fecha.between(start, end))
        .all()
    )
    incoming_map = defaultdict(list)
    outgoing_map = defaultdict(list)
    for movement in movements:
        if (
            movement.target_order_line
            and movement.target_order
            and movement.target_order.tipo_oc == "COMPRAS"
        ):
            incoming_map[movement.target_order_line.explosion_item_id].append(movement)
        if movement.source_order and movement.source_order.tipo_oc == "COMPRAS":
            outgoing_map[movement.source_order_line.explosion_item_id].append(movement)

    rows = []
    for entry in scoped_report_entries(project_ids):
        requisition_lines = [
            line
            for line in entry.requisition_lines
            if in_report_period(line.requisition.fecha_solicitud, start, end)
            and line.requisition.tipo_requisicion == "COMPRAS"
            and line.requisition.estado not in {"RECHAZADA", "CANCELADA"}
        ]
        requested = sum(
            (decimal_value(line.cantidad_solicitada) for line in requisition_lines),
            Decimal("0"),
        )
        approved = sum(
            (
                decimal_value(line.cantidad_aprobada)
                for line in requisition_lines
                if line.estado_linea in {"APROBADA", "RECHAZADA_COMPRAS"}
            ),
            Decimal("0"),
        )
        order_lines = [
            line
            for line in entry.order_lines
            if line.order
            and line.order.tipo_oc == "COMPRAS"
            and line.order.estado in ACTIVE_ORDER_STATES
            and in_report_period(line.order.fecha_orden, start, end)
            and (not supplier_id or line.order.supplier_id == supplier_id)
        ]
        ordered = sum((decimal_value(line.cantidad) for line in order_lines), Decimal("0"))
        ordered_amount = money(
            sum((decimal_value(line.importe_sin_iva) for line in order_lines), Decimal("0"))
        )
        received = Decimal("0")
        received_amount = Decimal("0")
        for line in entry.order_lines:
            if (
                not line.order
                or line.order.tipo_oc != "COMPRAS"
                or line.order.estado not in ACTIVE_ORDER_STATES
            ):
                continue
            if supplier_id and line.order.supplier_id != supplier_id:
                continue
            for receipt_line in line.receipt_lines:
                if in_report_period(receipt_line.receipt.fecha, start, end):
                    quantity = decimal_value(receipt_line.cantidad_recibida)
                    received += quantity
                    received_amount += quantity * decimal_value(
                        line.precio_unitario_sin_iva
                    )

        paid_amount = Decimal("0")
        paid_quantity = Decimal("0")
        for payment in payment_map.get(entry.id, []):
            amount = decimal_value(payment.monto_sin_iva)
            price = (
                decimal_value(payment.purchase_order_line.precio_unitario_sin_iva)
                if payment.purchase_order_line
                else decimal_value(entry.precio_unitario_sin_iva)
            )
            paid_amount += amount
            if price > 0:
                paid_quantity += amount / price
        for movement in incoming_map.get(entry.id, []):
            amount = decimal_value(movement.monto)
            price = decimal_value(movement.target_order_line.precio_unitario_sin_iva)
            paid_amount += amount
            if price > 0:
                paid_quantity += amount / price
        for movement in outgoing_map.get(entry.id, []):
            amount = decimal_value(movement.monto)
            price = decimal_value(movement.source_order_line.precio_unitario_sin_iva)
            paid_amount -= amount
            if price > 0:
                paid_quantity -= amount / price
        paid_amount = max(Decimal("0"), money(paid_amount))
        paid_quantity = max(Decimal("0"), paid_quantity)
        status_values = {
            "requested": requested,
            "approved": approved,
            "ordered": ordered,
            "received": received,
            "paid": paid_amount,
        }
        if not report_state_matches(state, status_values):
            continue
        partida, subpartida = budget_path_labels(entry.budget_item)
        budget_price = decimal_value(entry.precio_unitario_sin_iva)
        rows.append(
            {
                "obra": f"{entry.project.codigo} · {entry.project.nombre}",
                "partida": partida,
                "subpartida": subpartida,
                "clave": entry.supply_item.clave,
                "descripcion": entry.supply_item.descripcion,
                "unidad": entry.supply_item.unidad,
                "cantidad_presupuestada": decimal_value(entry.cantidad_presupuestada),
                "importe_presupuestado": money(entry.importe_presupuestado),
                "cantidad_requisitada": requested,
                "importe_requisitado": money(requested * budget_price),
                "cantidad_aprobada": approved,
                "importe_aprobado": money(approved * budget_price),
                "cantidad_comprada": ordered,
                "importe_comprado": ordered_amount,
                "cantidad_recibida": received,
                "importe_recibido": money(received_amount),
                "cantidad_pagada": paid_quantity,
                "importe_pagado": paid_amount,
            }
        )
    return rows


def report_orders(start, end, project_ids, supplier_id):
    query = PurchaseOrder.query.options(
        joinedload(PurchaseOrder.project),
        joinedload(PurchaseOrder.supplier),
        selectinload(PurchaseOrder.lines)
        .joinedload(PurchaseOrderLine.explosion_item)
        .joinedload(BudgetExplosionItem.budget_item)
        .joinedload(BudgetItem.parent),
        selectinload(PurchaseOrder.lines)
        .joinedload(PurchaseOrderLine.explosion_item)
        .joinedload(BudgetExplosionItem.supply_item),
        selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.receipt_lines),
        selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.payments),
        selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.advance_movements_in),
        selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.advance_movements_out),
    ).filter(
        PurchaseOrder.project_id.in_(project_ids or [-1]),
        PurchaseOrder.fecha_orden.between(start, end),
        PurchaseOrder.estado.in_(ACTIVE_ORDER_STATES),
        PurchaseOrder.tipo_oc == "COMPRAS",
    )
    if supplier_id:
        query = query.filter(PurchaseOrder.supplier_id == supplier_id)
    return query.order_by(PurchaseOrder.fecha_orden, PurchaseOrder.id).all()


def build_purchased_report(start, end, project_ids, supplier_id, state):
    rows = []
    for order in report_orders(start, end, project_ids, supplier_id):
        for line in order.lines:
            paid = line_paid_amount(line)
            values = {
                "requested": line.requisition_line.cantidad_solicitada if line.requisition_line else line.cantidad,
                "approved": line.requisition_line.cantidad_aprobada if line.requisition_line else line.cantidad,
                "ordered": line.cantidad,
                "received": line.cantidad_recibida,
                "paid": paid,
            }
            if not report_state_matches(state, values):
                continue
            partida, _ = budget_path_labels(line.explosion_item.budget_item)
            rows.append(
                {
                    "fecha": order.fecha_orden,
                    "obra": f"{order.project.codigo} · {order.project.nombre}",
                    "oc": order.folio,
                    "requisicion": (
                        line.requisition_line.requisition.folio
                        if line.requisition_line
                        else "Histórica"
                    ),
                    "proveedor": order.supplier.nombre,
                    "partida": partida,
                    "clave": line.explosion_item.supply_item.clave,
                    "descripcion": line.explosion_item.supply_item.descripcion,
                    "unidad": line.explosion_item.supply_item.unidad,
                    "cantidad": decimal_value(line.cantidad),
                    "precio_unitario": money(line.precio_unitario_sin_iva),
                    "importe": money(line.importe_sin_iva),
                    "estado": order.estado,
                }
            )
    return rows


def build_pending_report(start, end, project_ids, supplier_id, state):
    requisitions = (
        PurchaseRequisition.query.options(
            joinedload(PurchaseRequisition.project),
            selectinload(PurchaseRequisition.quotations),
            selectinload(PurchaseRequisition.lines)
            .joinedload(PurchaseRequisitionLine.explosion_item)
            .joinedload(BudgetExplosionItem.budget_item)
            .joinedload(BudgetItem.parent),
            selectinload(PurchaseRequisition.lines)
            .joinedload(PurchaseRequisitionLine.explosion_item)
            .joinedload(BudgetExplosionItem.supply_item),
            selectinload(PurchaseRequisition.lines)
            .selectinload(PurchaseRequisitionLine.order_lines)
            .joinedload(PurchaseOrderLine.order),
        )
        .filter(
            PurchaseRequisition.project_id.in_(project_ids or [-1]),
            PurchaseRequisition.estado.in_({"APROBADA", "PARCIAL"}),
            PurchaseRequisition.tipo_requisicion == "COMPRAS",
        )
        .all()
    )
    rows = []
    for requisition in requisitions:
        approval_date = (
            requisition.approved_at.date()
            if isinstance(requisition.approved_at, datetime)
            else requisition.fecha_solicitud
        )
        if not in_report_period(approval_date, start, end):
            continue
        if supplier_id and not any(
            quote.supplier_id == supplier_id for quote in requisition.quotations
        ):
            continue
        for line in requisition.lines:
            pending = line.cantidad_pendiente_compra
            if pending <= 0:
                continue
            values = {
                "requested": line.cantidad_solicitada,
                "approved": line.cantidad_aprobada,
                "ordered": line.cantidad_ordenada,
                "received": 0,
                "paid": 0,
            }
            if not report_state_matches(state, values):
                continue
            partida, _ = budget_path_labels(line.explosion_item.budget_item)
            rows.append(
                {
                    "fecha": approval_date,
                    "obra": f"{requisition.project.codigo} · {requisition.project.nombre}",
                    "requisicion": requisition.folio,
                    "limite": requisition.fecha_limite_oc,
                    "partida": partida,
                    "clave": line.explosion_item.supply_item.clave,
                    "descripcion": line.explosion_item.supply_item.descripcion,
                    "unidad": line.explosion_item.supply_item.unidad,
                    "cantidad_aprobada": decimal_value(line.cantidad_aprobada),
                    "cantidad_comprada": line.cantidad_ordenada,
                    "cantidad_pendiente": pending,
                    "importe_pendiente": line.importe_pendiente_compra,
                    "estado": requisition.estado,
                }
            )
    return rows


def build_supplier_report(start, end, project_ids, supplier_id, state):
    grouped = {}
    for order in report_orders(start, end, project_ids, supplier_id):
        key = (order.supplier_id, order.project_id)
        row = grouped.setdefault(
            key,
            {
                "proveedor": order.supplier.nombre,
                "rfc": order.supplier.rfc or "—",
                "obra": f"{order.project.codigo} · {order.project.nombre}",
                "order_ids": set(),
                "comprado": Decimal("0"),
                "recibido": Decimal("0"),
                "pagado": Decimal("0"),
            },
        )
        row["order_ids"].add(order.id)
        row["comprado"] += order.subtotal_sin_iva
        row["recibido"] += order.monto_recibido
        row["pagado"] += order.monto_pagado
    rows = []
    for row in grouped.values():
        values = {
            "requested": row["comprado"],
            "approved": row["comprado"],
            "ordered": row["comprado"],
            "received": row["recibido"],
            "paid": row["pagado"],
        }
        if not report_state_matches(state, values):
            continue
        row["ordenes"] = len(row.pop("order_ids"))
        row["comprado"] = money(row["comprado"])
        row["recibido"] = money(row["recibido"])
        row["pagado"] = money(row["pagado"])
        row["saldo"] = max(Decimal("0"), money(row["comprado"] - row["pagado"]))
        rows.append(row)
    return sorted(rows, key=lambda row: (row["proveedor"], row["obra"]))


def build_project_report(start, end, project_ids, supplier_id, state):
    grouped = {}
    for order in report_orders(start, end, project_ids, supplier_id):
        for line in order.lines:
            entry = line.explosion_item
            key = (order.project_id, entry.budget_item_id, entry.supply_item_id)
            partida, subpartida = budget_path_labels(entry.budget_item)
            row = grouped.setdefault(
                key,
                {
                    "obra": f"{order.project.codigo} · {order.project.nombre}",
                    "partida": partida,
                    "subpartida": subpartida,
                    "clave": entry.supply_item.clave,
                    "descripcion": entry.supply_item.descripcion,
                    "unidad": entry.supply_item.unidad,
                    "cantidad": Decimal("0"),
                    "comprado": Decimal("0"),
                    "recibido": Decimal("0"),
                    "pagado": Decimal("0"),
                },
            )
            row["cantidad"] += decimal_value(line.cantidad)
            row["comprado"] += decimal_value(line.importe_sin_iva)
            row["recibido"] += line.monto_recibido
            row["pagado"] += line_paid_amount(line)
    rows = []
    for row in grouped.values():
        values = {
            "requested": row["comprado"],
            "approved": row["comprado"],
            "ordered": row["comprado"],
            "received": row["recibido"],
            "paid": row["pagado"],
        }
        if not report_state_matches(state, values):
            continue
        for key in ("comprado", "recibido", "pagado"):
            row[key] = money(row[key])
        row["saldo"] = max(Decimal("0"), money(row["comprado"] - row["pagado"]))
        rows.append(row)
    return sorted(rows, key=lambda row: (row["obra"], row["partida"], row["clave"]))


REPORT_BUILDERS = {
    "cantidades": build_quantities_report,
    "comprados": build_purchased_report,
    "pendientes": build_pending_report,
    "proveedores": build_supplier_report,
    "obras": build_project_report,
}


def style_report_sheet(sheet, columns):
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(
            14, min(42, len(column["label"]) + 5)
        )
        if column["kind"] == "money":
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for item in cell:
                    item.number_format = '$#,##0.00;[Red]-$#,##0.00'
        elif column["kind"] == "quantity":
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for item in cell:
                    item.number_format = '#,##0.0000'


def excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def report_excel(report, rows, columns, start, end):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte"
    sheet.append([column["label"] for column in columns])
    for row in rows:
        sheet.append([excel_value(row.get(column["key"], "")) for column in columns])
    if rows and any(column["total"] for column in columns):
        total_row = sheet.max_row + 1
        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(total_row, index)
            if index == 1:
                cell.value = "TOTAL"
                cell.font = Font(bold=True)
            elif column["total"]:
                cell.value = f"=SUM({get_column_letter(index)}2:{get_column_letter(index)}{total_row-1})"
                cell.font = Font(bold=True)
    style_report_sheet(sheet, columns)
    info = workbook.create_sheet("Filtros")
    for row in [
        ["Reporte", report["title"]],
        ["Desde", start],
        ["Hasta", end],
        ["Generado", datetime.now()],
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 60
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@compras_bp.get("/reportes")
@permission_required("reportes_compras", "ver")
def reports_index():
    report_type = request.args.get("report", "cantidades")
    if report_type not in REPORT_DEFINITIONS:
        abort(404)
    try:
        start, end, period = report_date_range()
    except ValueError as exc:
        flash(str(exc), "danger")
        start = today_value().replace(day=1)
        end = today_value()
        period = "custom"
    projects = accessible_projects_query().all()
    allowed_ids = [project.id for project in projects]
    project_id = request.args.get("project_id", type=int)
    if project_id and project_id not in allowed_ids:
        abort(404)
    project_ids = [project_id] if project_id else allowed_ids
    supplier_id = request.args.get("supplier_id", type=int)
    if supplier_id and not db.session.get(Supplier, supplier_id):
        abort(404)
    state = normalize(request.args.get("state"))
    if state not in REPORT_STATES:
        state = ""
    report = REPORT_DEFINITIONS[report_type]
    rows = REPORT_BUILDERS[report_type](
        start, end, project_ids, supplier_id, state
    )
    selected_keys = request.args.getlist("columns")
    available_keys = {column["key"] for column in report["columns"]}
    selected_keys = [key for key in selected_keys if key in available_keys]
    if not selected_keys:
        selected_keys = [column["key"] for column in report["columns"]]
    selected_columns = [
        column for column in report["columns"] if column["key"] in selected_keys
    ]

    if request.args.get("export") == "xlsx":
        output = report_excel(report, rows, selected_columns, start, end)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"compras_{report_type}_{start}_{end}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    per_page = min(100, max(10, request.args.get("per_page", 25, type=int)))
    total_rows = len(rows)
    total_pages = max(1, (total_rows + per_page - 1) // per_page)
    page = min(total_pages, max(1, request.args.get("page", 1, type=int)))
    page_rows = rows[(page - 1) * per_page : page * per_page]
    totals = {
        column["key"]: sum(
            (decimal_value(row.get(column["key"])) for row in rows), Decimal("0")
        )
        for column in selected_columns
        if column["total"]
    }
    suppliers = Supplier.query.order_by(Supplier.nombre).all()
    return render_template(
        "compras/reports/index.html",
        definitions=REPORT_DEFINITIONS,
        report_type=report_type,
        report=report,
        rows=page_rows,
        columns=selected_columns,
        selected_keys=selected_keys,
        totals=totals,
        projects=projects,
        suppliers=suppliers,
        project_id=project_id,
        supplier_id=supplier_id,
        state=state,
        states=REPORT_STATES,
        start=start,
        end=end,
        period=period,
        page=page,
        total_pages=total_pages,
        total_rows=total_rows,
        per_page=per_page,
    )


def weekly_payment_rows(start: date, end: date, project_ids: list[int]):
    grouped = {}
    payments = (
        AdditionalPayment.query.options(
            joinedload(AdditionalPayment.purchase_order).joinedload(PurchaseOrder.supplier),
            joinedload(AdditionalPayment.purchase_order).joinedload(PurchaseOrder.project),
            joinedload(AdditionalPayment.purchase_order_line)
            .joinedload(PurchaseOrderLine.explosion_item)
            .joinedload(BudgetExplosionItem.supply_item),
            joinedload(AdditionalPayment.payment_method),
            joinedload(AdditionalPayment.company),
        )
        .join(PurchaseOrder, PurchaseOrder.id == AdditionalPayment.purchase_order_id)
        .filter(
            AdditionalPayment.project_id.in_(project_ids or [-1]),
            AdditionalPayment.purchase_order_id.isnot(None),
            AdditionalPayment.fecha.between(start, end),
            PurchaseOrder.tipo_oc == "COMPRAS",
        )
        .order_by(AdditionalPayment.fecha, AdditionalPayment.id)
        .all()
    )
    for payment in payments:
        order = payment.purchase_order
        key = (
            order.id,
            payment.fecha,
            payment.payment_method_id,
            payment.company_id,
            "DIRECTO",
        )
        row = grouped.setdefault(
            key,
            {
                "supplier": order.supplier,
                "project": order.project,
                "order": order,
                "concepts": set(),
                "method": payment.payment_method.nombre if payment.payment_method else payment.metodo_pago,
                "company": payment.company,
                "modalidad": order.modalidad_pago,
                "amount": Decimal("0"),
                "date": payment.fecha,
            },
        )
        concept = (
            payment.purchase_order_line.explosion_item.supply_item.descripcion
            if payment.purchase_order_line
            else payment.concepto
        )
        row["concepts"].add(concept)
        row["amount"] += decimal_value(payment.monto_sin_iva)

    movements = (
        SupplierAdvanceMovement.query.options(
            joinedload(SupplierAdvanceMovement.target_order).joinedload(PurchaseOrder.supplier),
            joinedload(SupplierAdvanceMovement.target_order).joinedload(PurchaseOrder.project),
            joinedload(SupplierAdvanceMovement.target_order_line)
            .joinedload(PurchaseOrderLine.explosion_item)
            .joinedload(BudgetExplosionItem.supply_item),
            joinedload(SupplierAdvanceMovement.company),
        )
        .join(
            PurchaseOrder,
            PurchaseOrder.id == SupplierAdvanceMovement.target_order_id,
        )
        .filter(
            SupplierAdvanceMovement.tipo == "APLICACION",
            SupplierAdvanceMovement.fecha.between(start, end),
            SupplierAdvanceMovement.target_order_id.isnot(None),
            PurchaseOrder.tipo_oc == "COMPRAS",
        )
        .all()
    )
    for movement in movements:
        order = movement.target_order
        if order.project_id not in project_ids:
            continue
        key = (order.id, movement.fecha, None, movement.company_id, "APLICACION")
        row = grouped.setdefault(
            key,
            {
                "supplier": order.supplier,
                "project": order.project,
                "order": order,
                "concepts": set(),
                "method": "APLICACIÓN DE SALDO",
                "company": movement.company,
                "modalidad": order.modalidad_pago,
                "amount": Decimal("0"),
                "date": movement.fecha,
            },
        )
        row["concepts"].add(
            movement.target_order_line.explosion_item.supply_item.descripcion
        )
        row["amount"] += decimal_value(movement.monto)

    rows = []
    for row in grouped.values():
        order = row["order"]
        rows.append(
            {
                "proveedor": row["supplier"].nombre,
                "rfc": row["supplier"].rfc or "—",
                "obra": f"{row['project'].codigo} · {row['project'].nombre}",
                "project_id": row["project"].id,
                "oc": order.folio,
                "concepto": ", ".join(sorted(row["concepts"])),
                "metodo": row["method"],
                "empresa": f"{row['company'].codigo} · {row['company'].nombre}",
                "modalidad": row["modalidad"],
                "total_oc": order.subtotal_sin_iva,
                "pagado_semana": money(row["amount"]),
                "saldo": order.saldo_pendiente,
                "fecha": row["date"],
            }
        )
    return sorted(rows, key=lambda row: (row["obra"], row["fecha"], row["proveedor"]))


WEEKLY_PAYMENT_COLUMNS = [
    report_column("proveedor", "Proveedor"),
    report_column("rfc", "RFC"),
    report_column("obra", "Obra"),
    report_column("oc", "OC"),
    report_column("concepto", "Concepto"),
    report_column("metodo", "Método de pago"),
    report_column("empresa", "Empresa que paga"),
    report_column("modalidad", "Modalidad"),
    report_column("total_oc", "Monto total OC", "money"),
    report_column("pagado_semana", "Pagado en la semana", "money", total=True),
    report_column("saldo", "Saldo pendiente", "money"),
    report_column("fecha", "Fecha del pago", "date"),
]


def weekly_payments_excel(rows, start, end, projects):
    workbook = Workbook()
    workbook.remove(workbook.active)
    groups = [("Resumen general", rows)]
    for project in projects:
        project_rows = [row for row in rows if row["project_id"] == project.id]
        if project_rows:
            groups.append((project.codigo[:31], project_rows))
    for title, group_rows in groups:
        sheet = workbook.create_sheet(title)
        sheet.append([column["label"] for column in WEEKLY_PAYMENT_COLUMNS])
        for row in group_rows:
            sheet.append(
                [excel_value(row.get(column["key"], "")) for column in WEEKLY_PAYMENT_COLUMNS]
            )
        total_row = sheet.max_row + 1
        sheet.cell(total_row, 1, "TOTAL")
        sheet.cell(total_row, 10, f"=SUM(J2:J{total_row-1})")
        sheet.cell(total_row, 1).font = Font(bold=True)
        sheet.cell(total_row, 10).font = Font(bold=True)
        style_report_sheet(sheet, WEEKLY_PAYMENT_COLUMNS)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@compras_bp.get("/reportes/pagos-semanales")
@permission_required("reportes_compras", "ver")
def weekly_supplier_payments():
    raw_week = request.args.get("week")
    try:
        selected = (
            datetime.strptime(raw_week, "%Y-%m-%d").date()
            if raw_week
            else today_value()
        )
    except ValueError:
        selected = today_value()
        flash("La semana seleccionada no es válida.", "danger")
    start, end = week_bounds(selected)
    projects = accessible_projects_query().all()
    allowed_ids = [project.id for project in projects]
    project_id = request.args.get("project_id", type=int)
    if project_id and project_id not in allowed_ids:
        abort(404)
    project_ids = [project_id] if project_id else allowed_ids
    rows = weekly_payment_rows(start, end, project_ids)
    totals_by_project = defaultdict(Decimal)
    for row in rows:
        totals_by_project[row["obra"]] += row["pagado_semana"]
    general_total = money(sum(totals_by_project.values(), Decimal("0")))
    if request.args.get("export") == "xlsx":
        output = weekly_payments_excel(rows, start, end, projects)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"pagos_proveedores_{start}_{end}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return render_template(
        "compras/reports/weekly_payments.html",
        rows=rows,
        columns=WEEKLY_PAYMENT_COLUMNS,
        start=start,
        end=end,
        projects=projects,
        project_id=project_id,
        totals_by_project=dict(totals_by_project),
        general_total=general_total,
    )


@compras_bp.get("/reportes/consumo-insumos")
@permission_required("reportes_compras", "ver")
def report_consumption():
    projects = accessible_projects_query().all()
    project_ids = [p.id for p in projects]
    project_id = request.args.get("project_id", type=int)
    if project_id and project_id not in project_ids:
        abort(404)
    scoped = [project_id] if project_id else project_ids
    entries = BudgetExplosionItem.query.options(joinedload(BudgetExplosionItem.project), joinedload(BudgetExplosionItem.budget_item), joinedload(BudgetExplosionItem.supply_item), selectinload(BudgetExplosionItem.requisition_lines).joinedload(PurchaseRequisitionLine.requisition), selectinload(BudgetExplosionItem.order_lines).joinedload(PurchaseOrderLine.order), selectinload(BudgetExplosionItem.order_lines).selectinload(PurchaseOrderLine.receipt_lines)).filter(BudgetExplosionItem.project_id.in_(scoped or [-1])).order_by(BudgetExplosionItem.project_id, BudgetExplosionItem.budget_item_id).all()
    rows = []
    for entry in entries:
        purchase_lines = [
            line
            for line in entry.requisition_lines
            if line.requisition.tipo_requisicion == "COMPRAS"
        ]
        requested = sum(
            (decimal_value(line.cantidad_solicitada) for line in purchase_lines),
            Decimal("0"),
        )
        approved = sum(
            (
                decimal_value(line.cantidad_aprobada)
                for line in purchase_lines
                if line.requisition.estado
                not in {"RECHAZADA", "VENCIDA", "CANCELADA"}
            ),
            Decimal("0"),
        )
        rows.append({"entry": entry, "requested": requested, "approved": approved})
    return render_template("compras/reports/consumption.html", projects=projects, selected_project_id=project_id, rows=rows)


@compras_bp.get("/reportes/smnc")
@permission_required("reportes_compras", "ver")
def report_smnc():
    project_ids = [p.id for p in accessible_projects_query().all()]
    requests = MaterialChangeRequest.query.options(joinedload(MaterialChangeRequest.project), joinedload(MaterialChangeRequest.requested_by), selectinload(MaterialChangeRequest.details)).filter(MaterialChangeRequest.project_id.in_(project_ids or [-1])).order_by(MaterialChangeRequest.created_at.desc()).all()
    return render_template("compras/reports/smnc.html", requests=requests)


@compras_bp.get("/reportes/proveedores")
@permission_required("reportes_compras", "ver")
def report_suppliers():
    project_ids = [p.id for p in accessible_projects_query().all()]
    suppliers = Supplier.query.options(selectinload(Supplier.orders).selectinload(PurchaseOrder.lines), selectinload(Supplier.orders).selectinload(PurchaseOrder.receipts)).order_by(Supplier.nombre).all()
    rows = []
    for supplier in suppliers:
        orders = [
            order
            for order in supplier.orders
            if order.project_id in project_ids
            and order.estado in ACTIVE_ORDER_STATES
            and order.tipo_oc == "COMPRAS"
        ]
        estimated_days = []
        real_days = []
        for order in orders:
            estimated_days.append((order.fecha_entrega_estimada - order.fecha_orden).days)
            if order.fecha_surtido_real:
                real_days.append((order.fecha_surtido_real - order.fecha_orden).days)
        rows.append({"supplier": supplier, "orders": orders, "total": money(sum((order.subtotal_sin_iva for order in orders), Decimal("0"))), "estimated_avg": sum(estimated_days) / len(estimated_days) if estimated_days else None, "real_avg": sum(real_days) / len(real_days) if real_days else None})
    return render_template("compras/reports/suppliers.html", rows=rows)


@compras_bp.get("/api/obras/<int:project_id>/insumos")
@login_required
def explosion_api(project_id):
    project = project_or_403(project_id)
    entry_ids = [entry.id for entry in items_explosion_vigente(project.id)]
    entries = (
        BudgetExplosionItem.query.options(
            joinedload(BudgetExplosionItem.supply_item),
            joinedload(BudgetExplosionItem.budget_item),
        )
        .filter(BudgetExplosionItem.id.in_(entry_ids or [-1]))
        .all()
    )
    return jsonify(
        [
            {
                "id": entry.id,
                "label": entry.etiqueta,
                "budget_item_id": entry.budget_item_id,
                "available": str(entry.cantidad_disponible),
                "classification": entry.clasificacion,
                "requires_authorization": entry.requiere_autorizacion_previa,
            }
            for entry in entries
        ]
    )
