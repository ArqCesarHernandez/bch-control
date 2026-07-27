"""Pruebas integrales de la Fase 4 final de Compras."""

import os
import unittest
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook

os.environ["FLASK_ENV"] = "testing"
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["SECRET_KEY"] = "test-compras-secret"

from app import create_app, mail  # noqa: E402
from compras_models import (  # noqa: E402
    BudgetExplosionItem,
    ExplosionRevision,
    GoodsReceipt,
    MaterialChangeRequest,
    PaymentMethod,
    PurchaseAlertRun,
    CreditCard,
    CreditCardPayment,
    PurchaseNotification,
    PurchaseOrder,
    PurchaseRequisition,
    PurchaseRequisitionLine,
    Quotation,
    QuotationLineSource,
    SupplierAdvanceMovement,
    SupplierSupplyItem,
    Supplier,
    SupplyItem,
    SupplyProjectCatalog,
)
from models import (  # noqa: E402
    ACCIONES_PERMISO,
    MODULOS_PERMISOS,
    CentroCosto,
    Permiso,
    Usuario,
    db,
)
from nominas_models import (  # noqa: E402
    AdditionalPayment,
    BudgetItem,
    Company,
    Payroll,
)
from routes.compras import (  # noqa: E402
    EXPLOSION_HEADERS,
    HISTORICAL_TEMPLATE_HEADERS,
)


app = create_app()


class PurchaseFlowTest(unittest.TestCase):
    def setUp(self):
        app.config.update(
            TESTING=True,
            WTF_CSRF_ENABLED=False,
            COMPRAS_TODAY=date(2026, 7, 21),  # martes
        )
        self.client = app.test_client()
        with app.app_context():
            db.drop_all()
            db.create_all()
            company = Company(codigo="BCH", nombre="Baja Custom Homes")
            project = CentroCosto(
                nombre="Casa Compras",
                codigo="L-COMPRA",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2026, 7, 1),
                presupuesto_total=500000,
                presupuesto_mano_obra=100000,
            )
            db.session.add_all([company, project])
            db.session.flush()
            parent = BudgetItem(
                project_id=project.id,
                codigo="MAT-01",
                nombre="Cimentación",
                categoria="ADICIONAL",
                presupuesto=150000,
            )
            child = BudgetItem(
                project_id=project.id,
                parent=parent,
                codigo="MAT-01.01",
                nombre="Concreto",
                categoria="ADICIONAL",
                presupuesto=150000,
            )
            db.session.add_all([parent, child])
            users = []
            for role, email, name in (
                ("admin", "admin@example.com", "Grecia Administradora"),
                ("capturista", "capturista@example.com", "Capturista Nómina"),
                ("supervisor", "supervisor@example.com", "Amir Supervisor"),
                ("comprador", "comprador@example.com", "Comprador Compras"),
                ("costos", "costos@example.com", "Costos Presupuestos"),
            ):
                user = Usuario(
                    nombre_completo=name,
                    correo=email,
                    rol=role,
                    centro_costo_id=(
                        project.id if role in {"capturista", "supervisor"} else None
                    ),
                    activo=True,
                )
                user.set_password("Password123!")
                if role in {"capturista", "supervisor", "comprador"}:
                    user.projects = [project]
                user.asignar_permisos_predeterminados()
                users.append(user)
            method = PaymentMethod(nombre="TRANSFERENCIA", descripcion="Transferencia")
            db.session.add_all(users + [method])
            db.session.commit()
            self.company_id = company.id
            self.project_id = project.id
            self.child_id = child.id
            self.method_id = method.id
        self.login("admin@example.com")

    def login(self, email):
        self.client.post("/logout", follow_redirects=True)
        response = self.client.post(
            "/login",
            data={"correo": email, "contrasena": "Password123!"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        return response

    @staticmethod
    def explosion_file(headers=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers or EXPLOSION_HEADERS)
        values = {
            "Partida": "Cimentación",
            "Subpartida": "Concreto",
            "Tipo": "Material",
            "Clave Insumo": "MAT-001",
            "Descripción": "Concreto f'c=250 kg/cm2",
            "Unidad": "m3",
            "Cantidad": 100,
            "Precio Unitario": 2500,
            "Importe": 250000,
        }
        sheet.append([values.get(header, "") for header in (headers or EXPLOSION_HEADERS)])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def historical_file():
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(HISTORICAL_TEMPLATE_HEADERS)
        sheet.append(
            [
                "Aceros Históricos",
                "AHI010101AA1",
                "MAT-001",
                "Concreto f'c=250 kg/cm2",
                "m3",
                2350,
                date(2026, 6, 30),
            ]
        )
        # Esta fila comprueba que un error aislado no revierte la fila válida.
        sheet.append(
            [
                "Aceros Históricos",
                "AHI010101AA1",
                "MAT-001",
                "Concreto f'c=250 kg/cm2",
                "kg",
                2400,
                date(2026, 7, 1),
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def create_explosion(self):
        response = self.client.post(
            "/compras/explosion/importar",
            data={
                "project_id": str(self.project_id),
                "archivo": (self.explosion_file(), "explosion.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("Todo el archivo", response.get_data(as_text=True))
        with app.app_context():
            return BudgetExplosionItem.query.filter_by(
                project_id=self.project_id, origen="EXPLOSION", activo=True
            ).one().id

    def create_supplier(
        self,
        credit=True,
        *,
        code="PROV-001",
        name="Materiales del Pacifico",
        email="cotizaciones@materiales.example.com",
    ):
        response = self.client.post(
            "/compras/proveedores",
            data={
                "codigo": code,
                "nombre": name,
                "rfc": "MAP010101AA1",
                "email": email,
                "tiene_credito": "on" if credit else "",
                "limite_credito": "100000",
                "dias_credito": "30",
                "company_id": str(self.company_id),
            },
            follow_redirects=True,
        )
        self.assertIn("Proveedor registrado", response.get_data(as_text=True))
        with app.app_context():
            return Supplier.query.filter_by(codigo=code).one().id

    def create_approved_requisition(
        self,
        entry_id,
        requested="40",
        approved="30",
        *,
        notes=None,
        suggested_supplier=None,
    ):
        # Estas regresiones históricas ejercitan deliberadamente el flujo de
        # autorización parcial. Desde la actualización, los conceptos normales
        # se liberan automáticamente; por eso el dato de prueba se clasifica
        # como especial antes de solicitarlo.
        with app.app_context():
            entry = db.session.get(BudgetExplosionItem, entry_id)
            entry.requiere_autorizacion_previa = True
            db.session.commit()
        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/requisiciones/nueva",
            data={
                "project_id": str(self.project_id),
                "fecha_requerida": "2026-07-23",
                "motivo": "Colado de cimentación",
                "tipo_requisicion": "COMPRAS",
            },
            follow_redirects=True,
        )
        self.assertIn("REQ-2026-", response.get_data(as_text=True))
        with app.app_context():
            req = PurchaseRequisition.query.order_by(PurchaseRequisition.id.desc()).first()
            req_id = req.id
        self.client.post(
            f"/compras/requisiciones/{req_id}/lineas",
            data={
                "explosion_item_id": str(entry_id),
                "cantidad_solicitada": requested,
                "notas": notes or "",
                "proveedor_sugerido": suggested_supplier or "",
            },
            follow_redirects=True,
        )
        self.client.post(
            f"/compras/requisiciones/{req_id}/enviar", follow_redirects=True
        )
        with app.app_context():
            line_id = db.session.get(PurchaseRequisition, req_id).lines[0].id
        self.login("admin@example.com")
        response = self.client.post(
            f"/compras/requisiciones/{req_id}/aprobar",
            data={f"aprobada_{line_id}": approved},
            follow_redirects=True,
        )
        self.assertIn("Solicitud de Cotización genérica", response.get_data(as_text=True))
        return req_id, line_id

    def create_extra_explosion(
        self,
        *,
        key="MAT-002",
        description="Block hueco 15x20x40",
        unit="PZA",
        quantity="1000",
        price="25",
    ):
        """Agrega un segundo material sin depender de otra importación Excel."""

        with app.app_context():
            supply = SupplyItem(
                clave=key,
                descripcion=description.upper(),
                tipo="MATERIAL",
                unidad=unit,
                clave_sat="00000000",
                moneda="MXN",
                activo=True,
            )
            admin = Usuario.query.filter_by(rol="admin").one()
            entry = BudgetExplosionItem(
                project_id=self.project_id,
                budget_item_id=self.child_id,
                supply_item=supply,
                cantidad_presupuestada=quantity,
                precio_unitario_sin_iva=price,
                importe_presupuestado=Decimal(quantity) * Decimal(price),
                origen="EXPLOSION",
                activo=True,
                created_by_id=admin.id,
            )
            db.session.add(entry)
            db.session.commit()
            return entry.id

    def create_bulk_draft(
        self,
        entry_id,
        *,
        quantity="25",
        notes="Entregar en acceso norte",
    ):
        """Crea el borrador con el mismo POST AJAX usado por la pantalla."""

        response = self.client.post(
            "/compras/requisiciones/nueva",
            data={
                "project_id": str(self.project_id),
                "fecha_requerida": "2026-07-23",
                "motivo": "Materiales para frente de cimentación",
                "tipo_requisicion": "COMPRAS",
                "material_ids": str(entry_id),
                f"cantidad_{entry_id}": quantity,
                f"notas_{entry_id}": notes,
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        return response

    def create_paid_credit_order(self, quantity="10", price="2500", paid="10000"):
        entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=True)
        req_id, requisition_line_id = self.create_approved_requisition(
            entry_id, requested=quantity, approved=quantity
        )
        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-22",
                f"cantidad_{requisition_line_id}": quantity,
                f"precio_{requisition_line_id}": price,
            },
            follow_redirects=True,
        )
        self.assertIn("Orden creada en borrador", response.get_data(as_text=True))
        with app.app_context():
            order = PurchaseOrder.query.order_by(PurchaseOrder.id.desc()).first()
            order_id = order.id
            order_line_id = order.lines[0].id
        self.client.post(f"/compras/ordenes/{order_id}/emitir", follow_redirects=True)
        self.login("comprador@example.com")
        response = self.client.post(
            f"/compras/ordenes/{order_id}/recibir",
            data={
                "fecha_recepcion": "2026-07-22",
                "documento_proveedor": f"FAC-{order_id:03d}",
                "fecha_factura": "2026-07-22",
                f"recibir_{order_line_id}": quantity,
            },
            follow_redirects=True,
        )
        self.assertIn("Recepción registrada", response.get_data(as_text=True))
        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/pagos/nuevo",
            data={
                "purchase_order_id": str(order_id),
                "order_line_id": str(order_line_id),
                "fecha": "2026-07-22",
                "monto_capturado": paid,
                "tipo_monto": "SIN_IVA",
                "payment_method_id": str(self.method_id),
                "company_id": str(self.company_id),
                "concepto": "Pago para pruebas de reportes",
            },
            follow_redirects=True,
        )
        self.assertIn("Pago registrado", response.get_data(as_text=True))
        return entry_id, supplier_id, req_id, order_id, order_line_id

    def test_excel_exact_and_auto_budget_path(self):
        entry_id = self.create_explosion()
        with app.app_context():
            entry = db.session.get(BudgetExplosionItem, entry_id)
            self.assertEqual(entry.supply_item.clave, "MAT-001")
            self.assertEqual(entry.budget_item.nombre, "CONCRETO")
            self.assertEqual(entry.budget_item.parent.nombre, "CIMENTACIÓN")
            self.assertEqual(str(entry.importe_presupuestado), "250000.00")

        bad_headers = EXPLOSION_HEADERS[:-1]
        response = self.client.post(
            "/compras/explosion/importar",
            data={
                "project_id": str(self.project_id),
                "archivo": (self.explosion_file(bad_headers), "mal.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertIn("Faltan columnas obligatorias: Importe", response.get_data(as_text=True))

    def test_dynamic_partida_material_api_and_bulk_draft_reservation(self):
        entry_id = self.create_explosion()
        self.login("supervisor@example.com")

        partidas_response = self.client.get(
            f"/compras/api/requisiciones/obras/{self.project_id}/partidas"
        )
        self.assertEqual(partidas_response.status_code, 200)
        partidas = partidas_response.get_json()["partidas"]
        self.assertEqual(len(partidas), 1)
        self.assertEqual(partidas[0]["nombre"], "CIMENTACIÓN")
        self.assertEqual(partidas[0]["subpartidas"][0]["nombre"], "CONCRETO")

        materials_response = self.client.get(
            f"/compras/api/requisiciones/obras/{self.project_id}/materiales",
            query_string={
                "partida_id": partidas[0]["id"],
                "subpartida_id": partidas[0]["subpartidas"][0]["id"],
            },
        )
        self.assertEqual(materials_response.status_code, 200)
        materials = materials_response.get_json()["materials"]
        self.assertEqual(len(materials), 1)
        self.assertEqual(materials[0]["clave"], "MAT-001")
        self.assertEqual(materials[0]["disponible"], "100.0000")

        created = self.create_bulk_draft(entry_id, quantity="25")
        self.assertEqual(created.status_code, 201)
        self.assertTrue(created.get_json()["ok"])
        with app.app_context():
            requisition = PurchaseRequisition.query.one()
            entry = db.session.get(BudgetExplosionItem, entry_id)
            self.assertEqual(requisition.estado, "BORRADOR")
            self.assertEqual(len(requisition.lines), 1)
            self.assertEqual(
                requisition.lines[0].cantidad_solicitada,
                Decimal("25.0000"),
            )
            self.assertEqual(
                entry.cantidad_reservada_borrador,
                Decimal("25.0000"),
            )
            self.assertEqual(entry.cantidad_disponible, Decimal("75.0000"))
            requisition_id = requisition.id

        cancelled = self.client.post(
            f"/compras/requisiciones/{requisition_id}/cancelar",
            follow_redirects=True,
        )
        self.assertIn(
            "cantidades se reintegraron",
            cancelled.get_data(as_text=True),
        )
        with app.app_context():
            entry = db.session.get(BudgetExplosionItem, entry_id)
            self.assertEqual(
                entry.cantidad_reservada_borrador,
                Decimal("0.0000"),
            )
            self.assertEqual(entry.cantidad_disponible, Decimal("100.0000"))

    def test_ajax_over_limit_keeps_existing_draft_data(self):
        entry_id = self.create_explosion()
        self.login("supervisor@example.com")
        created = self.create_bulk_draft(
            entry_id,
            quantity="30",
            notes="Conservar esta observación",
        )
        self.assertEqual(created.status_code, 201)
        with app.app_context():
            requisition = PurchaseRequisition.query.one()
            requisition_id = requisition.id
            line_id = requisition.lines[0].id

        second = self.create_bulk_draft(
            entry_id,
            quantity="80",
            notes="Este intento no debe persistir",
        )
        self.assertEqual(second.status_code, 422)
        payload = second.get_json()
        self.assertFalse(payload["ok"])
        self.assertIn("Cantidad máxima permitida: 70.0000", payload["error"])

        invalid_update = self.client.post(
            f"/compras/requisiciones/{requisition_id}/lineas/actualizar",
            data={
                f"cantidad_{line_id}": "101",
                f"notas_{line_id}": "Tampoco debe sustituir el dato",
            },
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        self.assertEqual(invalid_update.status_code, 422)
        self.assertIn(
            "Cantidad máxima permitida: 100.0000",
            invalid_update.get_json()["error"],
        )
        with app.app_context():
            line = db.session.get(PurchaseRequisitionLine, line_id)
            entry = db.session.get(BudgetExplosionItem, entry_id)
            self.assertEqual(PurchaseRequisition.query.count(), 1)
            self.assertEqual(line.cantidad_solicitada, Decimal("30.0000"))
            self.assertEqual(line.notas, "Conservar esta observación")
            self.assertEqual(
                entry.cantidad_reservada_borrador,
                Decimal("30.0000"),
            )

    def test_admin_assigns_multiple_projects_and_supervisor_switches_scope(self):
        with app.app_context():
            second_project = CentroCosto(
                nombre="Casa Compras Dos",
                codigo="L-COMPRA-2",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2026, 7, 2),
                presupuesto_total=300000,
                presupuesto_mano_obra=80000,
            )
            db.session.add(second_project)
            db.session.flush()
            supervisor = Usuario.query.filter_by(rol="supervisor").one()
            supervisor_id = supervisor.id
            second_project_id = second_project.id
            first_request = PurchaseRequisition(
                folio="REQ-SCOPE-01",
                project_id=self.project_id,
                fecha_solicitud=date(2026, 7, 21),
                fecha_requerida=date(2026, 7, 23),
                tipo_requisicion="COMPRAS",
                estado="BORRADOR",
                motivo="Alcance obra uno",
                requested_by_id=supervisor.id,
            )
            second_request = PurchaseRequisition(
                folio="REQ-SCOPE-02",
                project_id=second_project_id,
                fecha_solicitud=date(2026, 7, 21),
                fecha_requerida=date(2026, 7, 23),
                tipo_requisicion="COMPRAS",
                estado="BORRADOR",
                motivo="Alcance obra dos",
                requested_by_id=supervisor.id,
            )
            first_payroll = Payroll(
                project_id=self.project_id,
                semana_inicio=date(2026, 7, 13),
                semana_fin=date(2026, 7, 17),
                estado="borrador",
                created_by_id=supervisor.id,
            )
            second_payroll = Payroll(
                project_id=second_project_id,
                semana_inicio=date(2026, 7, 13),
                semana_fin=date(2026, 7, 17),
                estado="borrador",
                created_by_id=supervisor.id,
            )
            first_order = PurchaseOrder(
                folio="OCO-SCOPE-01",
                project_id=self.project_id,
                buyer_id=supervisor.id,
                beneficiario_libre="Operador Obra Uno",
                fecha_orden=date(2026, 7, 21),
                fecha_entrega_estimada=date(2026, 7, 23),
                fecha_limite=date(2026, 7, 23),
                tipo_oc="OPERACIONES",
                categoria_pago="OPERACIONES",
                estado="BORRADOR",
                modalidad_pago="PAGO_CONTRA_ENTREGA",
                created_by_id=supervisor.id,
            )
            second_order = PurchaseOrder(
                folio="OCO-SCOPE-02",
                project_id=second_project_id,
                buyer_id=supervisor.id,
                beneficiario_libre="Operador Obra Dos",
                fecha_orden=date(2026, 7, 21),
                fecha_entrega_estimada=date(2026, 7, 23),
                fecha_limite=date(2026, 7, 23),
                tipo_oc="OPERACIONES",
                categoria_pago="OPERACIONES",
                estado="BORRADOR",
                modalidad_pago="PAGO_CONTRA_ENTREGA",
                created_by_id=supervisor.id,
            )
            db.session.add_all(
                [
                    first_request,
                    second_request,
                    first_payroll,
                    second_payroll,
                    first_order,
                    second_order,
                ]
            )
            db.session.flush()
            first_request_id = first_request.id
            first_payroll_id = first_payroll.id
            second_payroll_id = second_payroll.id
            data = {
                "nombre_completo": supervisor.nombre_completo,
                "correo": supervisor.correo,
                "contrasena": "",
                "rol": "supervisor",
                "centro_costo_id": str(self.project_id),
                "project_ids": [
                    str(self.project_id),
                    str(second_project_id),
                ],
            }
            for permission in supervisor.permisos:
                for action in ACCIONES_PERMISO:
                    if getattr(permission, f"puede_{action}"):
                        data[
                            f"perm_{permission.modulo}_{action}"
                        ] = "on"
            db.session.commit()

        updated = self.client.post(
            f"/admin/usuarios/{supervisor_id}/editar",
            data=data,
            follow_redirects=True,
        )
        self.assertIn(
            "Usuario y permisos actualizados correctamente",
            updated.get_data(as_text=True),
        )
        with app.app_context():
            supervisor = db.session.get(Usuario, supervisor_id)
            self.assertEqual(
                {project.id for project in supervisor.projects},
                {self.project_id, second_project_id},
            )

        self.login("supervisor@example.com")
        topbar = self.client.get(
            "/campo/dashboard-supervisor"
        ).get_data(as_text=True)
        self.assertIn("L-COMPRA", topbar)
        self.assertIn("L-COMPRA-2", topbar)
        first_requisitions = self.client.get(
            "/compras/requisiciones"
        ).get_data(as_text=True)
        self.assertIn("REQ-SCOPE-01", first_requisitions)
        self.assertNotIn("REQ-SCOPE-02", first_requisitions)
        first_payrolls = self.client.get("/nominas").get_data(as_text=True)
        self.assertIn(f"/nominas/{first_payroll_id}", first_payrolls)
        self.assertNotIn(f"/nominas/{second_payroll_id}", first_payrolls)
        first_orders = self.client.get(
            "/compras/ordenes"
        ).get_data(as_text=True)
        self.assertIn("OCO-SCOPE-01", first_orders)
        self.assertNotIn("OCO-SCOPE-02", first_orders)
        switched = self.client.post(
            "/obra-activa",
            data={
                "project_id": str(second_project_id),
                "return_to": "/campo/dashboard-supervisor",
            },
            follow_redirects=True,
        )
        self.assertIn("Obra activa: L-COMPRA-2", switched.get_data(as_text=True))
        second_requisitions = self.client.get(
            "/compras/requisiciones"
        ).get_data(as_text=True)
        self.assertIn("REQ-SCOPE-02", second_requisitions)
        self.assertNotIn("REQ-SCOPE-01", second_requisitions)
        second_payrolls = self.client.get("/nominas").get_data(as_text=True)
        self.assertIn(f"/nominas/{second_payroll_id}", second_payrolls)
        self.assertNotIn(f"/nominas/{first_payroll_id}", second_payrolls)
        second_orders = self.client.get(
            "/compras/ordenes"
        ).get_data(as_text=True)
        self.assertIn("OCO-SCOPE-02", second_orders)
        self.assertNotIn("OCO-SCOPE-01", second_orders)
        self.assertEqual(
            self.client.get(
                f"/compras/requisiciones/{first_request_id}"
            ).status_code,
            404,
        )
        self.assertEqual(
            self.client.get(f"/nominas/{first_payroll_id}").status_code,
            404,
        )

    def test_buyer_gets_new_project_and_dashboard_address_alert(self):
        new_page = self.client.get("/obras/nueva").get_data(as_text=True)
        self.assertIn("Dirección de entrega", new_page)
        created = self.client.post(
            "/obras/nueva",
            data={
                "nombre": "Casa Entrega",
                "codigo": "L-ENTREGA",
                "tipo": "obra",
                "presupuesto_total": "250000",
                "presupuesto_mano_obra": "60000",
                "direccion_entrega": (
                    "Calle del Mar 120, Colonia Centro, La Paz, BCS"
                ),
                "activa": "on",
            },
            follow_redirects=True,
        )
        self.assertIn("Obra/oficina guardada", created.get_data(as_text=True))
        with app.app_context():
            project = CentroCosto.query.filter_by(codigo="L-ENTREGA").one()
            buyer = Usuario.query.filter_by(rol="comprador").one()
            admin = Usuario.query.filter_by(rol="admin").one()
            self.assertIn(project, buyer.projects)
            self.assertEqual(
                project.direccion_entrega,
                "Calle del Mar 120, Colonia Centro, La Paz, BCS",
            )
            # Simula una requisición histórica creada antes de que una
            # integración externa materialice user_projects. El Comprador
            # debe verla de todos modos por su alcance global.
            late_project = CentroCosto(
                nombre="Casa Asignación Tardía",
                codigo="L-LATE",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2026, 7, 3),
                presupuesto_total=180000,
                presupuesto_mano_obra=45000,
            )
            late_request = PurchaseRequisition(
                folio="REQ-LATE-01",
                project=late_project,
                fecha_solicitud=date(2026, 7, 21),
                fecha_requerida=date(2026, 7, 24),
                tipo_requisicion="COMPRAS",
                estado="APROBADA",
                motivo="Requisición anterior a la asignación",
                requested_by_id=admin.id,
                fecha_limite_oc=date(2026, 7, 24),
            )
            db.session.add_all([late_project, late_request])
            db.session.commit()
            self.assertNotIn(late_project, buyer.projects)

        self.login("comprador@example.com")
        dashboard = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("L-ENTREGA", dashboard)
        self.assertIn("L-LATE", dashboard)
        self.assertIn("REQ-LATE-01", dashboard)
        self.assertIn("Requisiciones pendientes por obra", dashboard)
        self.assertIn("L-COMPRA · Casa Compras", dashboard)
        self.assertIn("no tiene dirección de entrega", dashboard)
        self.assertIn("Haz clic aquí para asignarla", dashboard)

    def test_consolidation_groups_repeated_material_across_projects(self):
        first_entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=False)
        with app.app_context():
            admin = Usuario.query.filter_by(rol="admin").one()
            buyer = Usuario.query.filter_by(rol="comprador").one()
            first_entry = db.session.get(
                BudgetExplosionItem,
                first_entry_id,
            )
            second_project = CentroCosto(
                nombre="Casa Consolidada",
                codigo="L-CONS",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2026, 7, 2),
                presupuesto_total=350000,
                presupuesto_mano_obra=75000,
                direccion_entrega=(
                    "Avenida Consolidación 450, Cabo San Lucas, BCS"
                ),
            )
            second_item = BudgetItem(
                project=second_project,
                codigo="MAT-02",
                nombre="Estructura",
                categoria="ADICIONAL",
                presupuesto=100000,
            )
            second_revision = ExplosionRevision(
                project=second_project,
                numero_revision=1,
                estado="VIGENTE",
                es_vigente=True,
                loaded_by_id=admin.id,
            )
            second_entry = BudgetExplosionItem(
                revision=second_revision,
                project=second_project,
                budget_item=second_item,
                supply_item_id=first_entry.supply_item_id,
                cantidad_presupuestada=20,
                precio_unitario_sin_iva=2450,
                importe_presupuestado=49000,
                clasificacion="NORMAL",
                origen="EXPLOSION",
                activo=True,
                created_by_id=admin.id,
            )
            first_request = PurchaseRequisition(
                folio="REQ-CONS-01",
                project_id=self.project_id,
                fecha_solicitud=date(2026, 7, 21),
                fecha_requerida=date(2026, 7, 24),
                tipo_requisicion="COMPRAS",
                estado="APROBADA",
                motivo="Consolidación uno",
                requested_by_id=buyer.id,
                fecha_limite_oc=date(2026, 7, 24),
            )
            second_request = PurchaseRequisition(
                folio="REQ-CONS-02",
                project=second_project,
                fecha_solicitud=date(2026, 7, 21),
                fecha_requerida=date(2026, 7, 25),
                tipo_requisicion="COMPRAS",
                estado="APROBADA",
                motivo="Consolidación dos",
                requested_by_id=buyer.id,
                fecha_limite_oc=date(2026, 7, 24),
            )
            db.session.add_all(
                [
                    second_project,
                    second_item,
                    second_revision,
                    second_entry,
                    first_request,
                    second_request,
                ]
            )
            db.session.flush()
            first_line = PurchaseRequisitionLine(
                requisition=first_request,
                explosion_item_id=first_entry.id,
                cantidad_solicitada=4,
                cantidad_aprobada=4,
                estado_linea="APROBADA",
                notas="Descarga por acceso principal",
            )
            second_line = PurchaseRequisitionLine(
                requisition=second_request,
                explosion_item=second_entry,
                cantidad_solicitada=6,
                cantidad_aprobada=6,
                estado_linea="APROBADA",
                notas="Descarga por acceso secundario",
            )
            db.session.add_all([first_line, second_line])
            db.session.commit()
            first_request_id = first_request.id
            second_request_id = second_request.id

        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/cotizaciones/consolidar",
            data={
                "supplier_id": str(supplier_id),
                "requisition_ids": [
                    str(first_request_id),
                    str(second_request_id),
                ],
            },
            follow_redirects=True,
        )
        self.assertIn(
            "consolidó 2 requisiciones y 1 material",
            response.get_data(as_text=True),
        )
        with app.app_context():
            quotation = Quotation.query.one()
            quotation_id = quotation.id
            self.assertTrue(quotation.es_consolidada)
            self.assertEqual(len(quotation.requisition_set), 2)
            self.assertEqual(len(quotation.projects), 2)
            self.assertEqual(len(quotation.lines), 1)
            self.assertEqual(
                quotation.lines[0].cantidad,
                Decimal("10.0000"),
            )
            self.assertEqual(
                QuotationLineSource.query.count(),
                2,
            )
            self.assertTrue(
                all(
                    requisition.buyer_received_at
                    for requisition in quotation.requisition_set
                )
            )

        rfq = self.client.get(
            f"/compras/cotizaciones/{quotation_id}/solicitud"
        )
        rfq_page = rfq.get_data(as_text=True)
        self.assertEqual(rfq.status_code, 200)
        self.assertIn("REQ-CONS-01", rfq_page)
        self.assertIn("REQ-CONS-02", rfq_page)
        self.assertIn("Avenida Consolidación 450", rfq_page)
        self.assertEqual(rfq_page.count("MAT-001"), 1)

    def test_order_closes_only_after_delivery_address_confirmation(self):
        _entry_id, _supplier_id, _req_id, order_id, _line_id = (
            self.create_paid_credit_order(
                quantity="1",
                price="100",
                paid="100",
            )
        )
        with app.app_context():
            order = db.session.get(PurchaseOrder, order_id)
            self.assertEqual(order.estado, "RECEPCION_TOTAL")
            self.assertIsNone(order.direccion_entrega_confirmada_at)

        confirmed = self.client.post(
            f"/compras/ordenes/{order_id}/confirmar-direccion",
            data={
                "direccion_entrega": (
                    "Calle Obra 88, Colonia Centro, La Paz, BCS"
                )
            },
            follow_redirects=True,
        )
        self.assertIn(
            "Dirección de entrega confirmada",
            confirmed.get_data(as_text=True),
        )
        with app.app_context():
            order = db.session.get(PurchaseOrder, order_id)
            self.assertEqual(order.estado, "CERRADA")
            self.assertEqual(
                order.direccion_entrega,
                "Calle Obra 88, Colonia Centro, La Paz, BCS",
            )
            self.assertIsNotNone(order.direccion_entrega_confirmada_at)
            self.assertEqual(
                order.project.direccion_entrega,
                order.direccion_entrega,
            )

        printed = self.client.get(
            f"/compras/ordenes/{order_id}/imprimir"
        ).get_data(as_text=True)
        self.assertIn("Orden de Compra", printed)
        self.assertIn("Calle Obra 88", printed)
        self.assertNotIn(">Origen<", printed)

    def test_partial_approval_commits_and_creates_generic_rfq(self):
        entry_id = self.create_explosion()
        req_id, line_id = self.create_approved_requisition(entry_id)
        with app.app_context():
            req = db.session.get(PurchaseRequisition, req_id)
            line = db.session.get(PurchaseRequisitionLine, line_id)
            self.assertEqual(req.estado, "APROBADA")
            self.assertEqual(req.fecha_limite_oc, date(2026, 7, 24))
            self.assertEqual(str(line.cantidad_solicitada), "40.0000")
            self.assertEqual(str(line.cantidad_aprobada), "30.0000")
            self.assertEqual(str(req.total_aprobado), "75000.00")
        dashboard = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("$75,000.00", dashboard)
        self.assertIn("$425,000.00", dashboard)
        response = self.client.get(
            f"/compras/requisiciones/{req_id}/solicitud-cotizacion"
        )
        page = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("No representa una compra", page)
        self.assertIn("MAT-001", page)

    def test_requisition_closes_only_after_every_requested_unit_is_ordered(self):
        entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=True)
        req_id, line_id = self.create_approved_requisition(
            entry_id, requested="10", approved="10"
        )
        self.login("comprador@example.com")

        first = self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-23",
                f"cantidad_{line_id}": "4",
                f"precio_{line_id}": "2500",
            },
            follow_redirects=True,
        )
        self.assertIn("Orden creada en borrador", first.get_data(as_text=True))
        with app.app_context():
            first_order_id = PurchaseOrder.query.one().id
        self.client.post(
            f"/compras/ordenes/{first_order_id}/emitir", follow_redirects=True
        )

        with app.app_context():
            requisition = db.session.get(PurchaseRequisition, req_id)
            line = db.session.get(PurchaseRequisitionLine, line_id)
            self.assertEqual(requisition.estado, "PARCIAL")
            self.assertEqual(line.cantidad_ordenada, Decimal("4.0000"))
            self.assertEqual(line.cantidad_pendiente_solicitada, Decimal("6.0000"))
            self.assertEqual(line.porcentaje_compra, Decimal("40"))
        list_page = self.client.get("/compras/requisiciones").get_data(as_text=True)
        self.assertNotIn("Avance comprado", list_page)
        self.assertNotIn("4.0000 de 10.0000", list_page)

        second = self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-24",
                f"cantidad_{line_id}": "6",
                f"precio_{line_id}": "2550",
            },
            follow_redirects=True,
        )
        self.assertIn("Orden creada en borrador", second.get_data(as_text=True))
        with app.app_context():
            second_order_id = PurchaseOrder.query.order_by(
                PurchaseOrder.id.desc()
            ).first().id
        self.client.post(
            f"/compras/ordenes/{second_order_id}/emitir", follow_redirects=True
        )
        with app.app_context():
            requisition = db.session.get(PurchaseRequisition, req_id)
            line = db.session.get(PurchaseRequisitionLine, line_id)
            self.assertEqual(requisition.estado, "CERRADA")
            self.assertEqual(line.cantidad_ordenada, Decimal("10.0000"))
            self.assertEqual(line.porcentaje_compra, Decimal("100"))

    def test_supervisor_comments_provider_filter_and_quote_contact_channels(self):
        entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=False)
        req_id, line_id = self.create_approved_requisition(
            entry_id,
            requested="8",
            approved="8",
            notes="Urgente: concreto bombeable, entregar antes de las 07:00",
            suggested_supplier="Materiales del Pacifico",
        )
        self.login("comprador@example.com")
        detail = self.client.get(f"/compras/requisiciones/{req_id}")
        page = detail.get_data(as_text=True)
        self.assertIn("Urgente: concreto bombeable", page)
        self.assertNotIn("Materiales del Pacifico", page)
        self.assertNotIn("Filtrar materiales por proveedor sugerido", page)
        self.assertIn("Confirmar recepción de requisición", page)

        without_suggestion = self.client.get(
            f"/compras/requisiciones/{req_id}",
            query_string={"proveedor_sugerido": "__SIN_SUGERENCIA__"},
        ).get_data(as_text=True)
        self.assertIn("Urgente: concreto bombeable", without_suggestion)
        suggested = self.client.get(
            f"/compras/requisiciones/{req_id}",
            query_string={"proveedor_sugerido": "Materiales del Pacifico"},
        ).get_data(as_text=True)
        self.assertIn("Urgente: concreto bombeable", suggested)

        confirmation = self.client.post(
            f"/compras/requisiciones/{req_id}/confirmar-recepcion",
            follow_redirects=True,
        )
        self.assertIn("Recepción confirmada", confirmation.get_data(as_text=True))
        quote_response = self.client.post(
            f"/compras/requisiciones/{req_id}/cotizaciones",
            data={"supplier_ids": str(supplier_id), "line_ids": str(line_id)},
            follow_redirects=True,
        )
        self.assertIn(
            "Se generaron 1 solicitud(es) de cotización",
            quote_response.get_data(as_text=True),
        )
        with app.app_context():
            quote_id = Quotation.query.one().id

        with mail.record_messages() as outbox:
            sent = self.client.post(
                f"/compras/cotizaciones/{quote_id}/enviar-correo",
                follow_redirects=True,
            )
        sent_page = sent.get_data(as_text=True)
        self.assertIn("Cotización enviada exitosamente", sent_page)
        self.assertEqual(len(outbox), 1)
        self.assertEqual(outbox[0].recipients, ["cotizaciones@materiales.example.com"])
        self.assertEqual(outbox[0].cc, ["comprador@example.com"])
        self.assertIn("Urgente: concreto bombeable", outbox[0].body)
        self.assertNotIn("supervisor@example.com", outbox[0].recipients + outbox[0].cc)

        whatsapp = self.client.post(
            f"/compras/cotizaciones/{quote_id}/contacto-whatsapp",
            data={"notas_whatsapp": "Respondió Laura; enviará precio hoy."},
            follow_redirects=True,
        )
        self.assertIn("Contacto por WhatsApp registrado", whatsapp.get_data(as_text=True))
        with app.app_context():
            quote = db.session.get(Quotation, quote_id)
            self.assertIsNotNone(quote.email_sent_at)
            self.assertEqual(quote.email_cc, "comprador@example.com")
            self.assertIsNotNone(quote.whatsapp_contacted_at)
            self.assertIn("Laura", quote.whatsapp_notes)

    def test_draft_order_can_change_supplier_method_and_lines_then_cancel(self):
        first_entry_id = self.create_explosion()
        second_entry_id = self.create_extra_explosion()
        first_supplier_id = self.create_supplier(credit=True)
        second_supplier_id = self.create_supplier(
            credit=False,
            code="PROV-002",
            name="Blocks y Prefabricados del Sur",
            email="ventas@blocks.example.com",
        )
        _, first_line_id = self.create_approved_requisition(
            first_entry_id, requested="5", approved="5"
        )
        _, second_line_id = self.create_approved_requisition(
            second_entry_id, requested="20", approved="20"
        )

        self.login("comprador@example.com")
        created = self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(first_supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": "new:SPEI URGENTE",
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-23",
                f"cantidad_{first_line_id}": "2",
                f"precio_{first_line_id}": "2500",
                f"cantidad_{second_line_id}": "0",
            },
            follow_redirects=True,
        )
        self.assertIn("Orden creada en borrador", created.get_data(as_text=True))
        with app.app_context():
            order = PurchaseOrder.query.one()
            order_id = order.id
            self.assertEqual(order.payment_method.nombre, "SPEI URGENTE")
            self.assertEqual(order.lines[0].requisition_line_id, first_line_id)

        edit_page = self.client.get(
            f"/compras/ordenes/{order_id}/editar"
        ).get_data(as_text=True)
        self.assertIn('list="supplier-options"', edit_page)
        self.assertIn("Agregar método de pago", edit_page)
        self.assertIn("Quitar del borrador", edit_page)

        edited = self.client.post(
            f"/compras/ordenes/{order_id}/editar",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(second_supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": "new:TARJETA OBRA",
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-24",
                "notas": "Borrador sustituido por block",
                f"cantidad_{first_line_id}": "0",
                f"precio_{first_line_id}": "2500",
                f"cantidad_{second_line_id}": "3",
                f"precio_{second_line_id}": "30",
            },
            follow_redirects=True,
        )
        self.assertIn("Borrador de OC actualizado", edited.get_data(as_text=True))
        with app.app_context():
            order = db.session.get(PurchaseOrder, order_id)
            self.assertEqual(order.supplier_id, second_supplier_id)
            self.assertEqual(order.payment_method.nombre, "TARJETA OBRA")
            self.assertEqual(len(order.lines), 1)
            self.assertEqual(order.lines[0].requisition_line_id, second_line_id)
            self.assertEqual(order.lines[0].cantidad, Decimal("3.0000"))
            self.assertEqual(order.lines[0].precio_unitario_sin_iva, Decimal("30.0000"))

        cancelled = self.client.post(
            f"/compras/ordenes/{order_id}/cancelar", follow_redirects=True
        )
        self.assertIn("Orden cancelada", cancelled.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(db.session.get(PurchaseOrder, order_id).estado, "CANCELADA")

    def test_cumulative_order_filters_and_partial_receipt_email(self):
        entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=True)
        _, line_id = self.create_approved_requisition(
            entry_id, requested="5", approved="5"
        )
        self.login("comprador@example.com")
        self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-23",
                f"cantidad_{line_id}": "5",
                f"precio_{line_id}": "2500",
            },
            follow_redirects=True,
        )
        with app.app_context():
            order = PurchaseOrder.query.one()
            order_id = order.id
            order_line_id = order.lines[0].id
            folio = order.folio
        self.client.post(f"/compras/ordenes/{order_id}/emitir", follow_redirects=True)

        filtered = self.client.get(
            "/compras/ordenes",
            query_string={
                "project_id": self.project_id,
                "supplier_id": supplier_id,
                "estado": "EMITIDA",
                "fecha_desde": "2026-07-21",
                "fecha_hasta": "2026-07-21",
            },
        ).get_data(as_text=True)
        self.assertIn(folio, filtered)
        excluded = self.client.get(
            "/compras/ordenes",
            query_string={
                "project_id": self.project_id,
                "supplier_id": supplier_id,
                "estado": "EMITIDA",
                "fecha_desde": "2026-07-22",
                "fecha_hasta": "2026-07-30",
            },
        ).get_data(as_text=True)
        self.assertNotIn(folio, excluded)

        self.login("comprador@example.com")
        with mail.record_messages() as outbox:
            received = self.client.post(
                f"/compras/ordenes/{order_id}/recibir",
                data={
                    "fecha_recepcion": "2026-07-22",
                    "documento_proveedor": "FAC-PARCIAL-01",
                    "fecha_factura": "2026-07-22",
                    "notas_recepcion": "Dos unidades recibidas sin daño",
                    f"recibir_{order_line_id}": "2",
                },
                follow_redirects=True,
            )
        self.assertIn(
            "correo de confirmación enviado al comprador y al proveedor",
            received.get_data(as_text=True),
        )
        self.assertEqual(len(outbox), 1)
        self.assertEqual(
            set(outbox[0].recipients),
            {"cotizaciones@materiales.example.com", "comprador@example.com"},
        )
        self.assertIn("FAC-PARCIAL-01", outbox[0].body)
        self.assertIn("2.0000", outbox[0].body)
        with app.app_context():
            order = db.session.get(PurchaseOrder, order_id)
            receipt = GoodsReceipt.query.one()
            self.assertEqual(order.estado, "RECEPCION_PARCIAL")
            self.assertEqual(receipt.tipo, "PARCIAL")
            self.assertIsNotNone(receipt.notification_email_sent_at)
            self.assertIsNone(receipt.notification_email_error)

        self.login("comprador@example.com")
        partial_page = self.client.get(
            "/compras/ordenes",
            query_string={
                "project_id": self.project_id,
                "supplier_id": supplier_id,
                "estado": "RECEPCION_PARCIAL",
                "fecha_desde": "2026-07-01",
                "fecha_hasta": "2026-07-31",
            },
        ).get_data(as_text=True)
        self.assertIn(folio, partial_page)

    def test_quote_credit_order_reception_payment_and_due_date(self):
        entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=True)
        req_id, line_id = self.create_approved_requisition(entry_id, approved="30")

        self.login("comprador@example.com")
        self.client.post(
            f"/compras/requisiciones/{req_id}/confirmar-recepcion",
            follow_redirects=True,
        )
        self.client.post(
            f"/compras/requisiciones/{req_id}/cotizaciones",
            data={"supplier_ids": str(supplier_id)},
            follow_redirects=True,
        )
        with app.app_context():
            quote = Quotation.query.one()
            quote_id = quote.id
            quote_line_id = quote.lines[0].id
        self.client.post(
            f"/compras/cotizaciones/{quote_id}/respuesta",
            data={
                "fecha_respuesta": "2026-07-21",
                "fecha_entrega_ofertada": "2026-07-22",
                f"precio_{quote_line_id}": "2600",
            },
            follow_redirects=True,
        )
        response = self.client.post(
            f"/compras/cotizaciones/{quote_id}/seleccionar",
            follow_redirects=False,
        )
        self.assertIn("/compras/ordenes/nueva", response.headers["Location"])
        response = self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "quotation_id": str(quote_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-22",
                f"cantidad_{line_id}": "30",
                f"precio_{line_id}": "2600",
            },
            follow_redirects=True,
        )
        self.assertIn("Orden creada en borrador", response.get_data(as_text=True))
        with app.app_context():
            order = PurchaseOrder.query.one()
            order_id = order.id
            order_line_id = order.lines[0].id
            self.assertEqual(str(order.subtotal_sin_iva), "78000.00")

        self.client.post(f"/compras/ordenes/{order_id}/emitir", follow_redirects=True)
        with app.app_context():
            self.assertEqual(db.session.get(PurchaseOrder, order_id).estado, "EMITIDA")
            # Se compraron 30 de las 40 unidades solicitadas: aunque toda la
            # cantidad autorizada esté en OC, la requisición sigue abierta.
            self.assertEqual(db.session.get(PurchaseRequisition, req_id).estado, "PARCIAL")

        # Regla de oro: crédito no puede pagarse antes de recibir.
        response = self.client.post(
            "/compras/pagos/nuevo",
            data={
                "purchase_order_id": str(order_id),
                "order_line_id": str(order_line_id),
                "fecha": "2026-07-21",
                "monto_capturado": "1000",
                "tipo_monto": "SIN_IVA",
                "payment_method_id": str(self.method_id),
                "company_id": str(self.company_id),
                "concepto": "Pago indebido",
            },
            follow_redirects=True,
        )
        self.assertIn("saldo habilitado", response.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(AdditionalPayment.query.count(), 0)

        self.login("comprador@example.com")
        response = self.client.post(
            f"/compras/ordenes/{order_id}/recibir",
            data={
                "fecha_recepcion": "2026-07-22",
                "documento_proveedor": "FAC-100",
                "fecha_factura": "2026-07-22",
                f"recibir_{order_line_id}": "30",
            },
            follow_redirects=True,
        )
        self.assertIn("Recepción registrada", response.get_data(as_text=True))
        with app.app_context():
            order = db.session.get(PurchaseOrder, order_id)
            self.assertEqual(order.fecha_vencimiento, date(2026, 8, 21))
            self.assertEqual(order.estado, "RECEPCION_TOTAL")
            self.assertEqual(GoodsReceipt.query.one().tipo, "TOTAL")

        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/pagos/nuevo",
            data={
                "purchase_order_id": str(order_id),
                "order_line_id": str(order_line_id),
                "fecha": "2026-07-22",
                "monto_capturado": "50000",
                "tipo_monto": "SIN_IVA",
                "payment_method_id": str(self.method_id),
                "company_id": str(self.company_id),
                "concepto": "Pago parcial factura FAC-100",
            },
            follow_redirects=True,
        )
        self.assertIn("estado de cuenta", response.get_data(as_text=True))
        with app.app_context():
            order = db.session.get(PurchaseOrder, order_id)
            supplier = db.session.get(Supplier, supplier_id)
            self.assertEqual(str(order.monto_pagado), "50000.00")
            self.assertEqual(str(order.saldo_pendiente), "28000.00")
            self.assertEqual(str(order.monto_consumido_real), "50000.00")
            self.assertEqual(str(order.saldo_comprometido), "28000.00")
            self.assertEqual(str(supplier.credito_utilizado), "28000.00")

        # Rojo desde tres días antes del vencimiento.
        app.config["COMPRAS_TODAY"] = date(2026, 8, 19)
        page = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("due-ROJO", page)
        self.assertIn("FAC-100", page)
        self.assertIn("$28,000.00", page)
        self.assertIn("$50,000.00", page)
        self.assertIn("$422,000.00", page)

    def test_authorized_advance_is_only_pre_receipt_exception(self):
        entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=False)
        req_id, line_id = self.create_approved_requisition(entry_id, requested="10", approved="10")
        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "ANTICIPO",
                "fecha_entrega_estimada": "2026-07-23",
                "anticipo_monto": "10000",
                "justificacion_anticipo": "Proveedor exige anticipo",
                f"cantidad_{line_id}": "10",
                f"precio_{line_id}": "2500",
            },
            follow_redirects=True,
        )
        self.assertIn("Orden creada", response.get_data(as_text=True))
        with app.app_context():
            order = PurchaseOrder.query.one()
            order_id, order_line_id = order.id, order.lines[0].id
        self.client.post(f"/compras/ordenes/{order_id}/emitir", follow_redirects=True)
        with app.app_context():
            self.assertEqual(db.session.get(PurchaseOrder, order_id).estado, "PENDIENTE_ANTICIPO")
        self.login("admin@example.com")
        self.client.post(
            f"/compras/ordenes/{order_id}/autorizar-anticipo", follow_redirects=True
        )
        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/pagos/nuevo",
            data={
                "purchase_order_id": str(order_id),
                "order_line_id": str(order_line_id),
                "fecha": "2026-07-21",
                "monto_capturado": "10000",
                "tipo_monto": "SIN_IVA",
                "payment_method_id": str(self.method_id),
                "company_id": str(self.company_id),
                "concepto": "Anticipo autorizado",
            },
            follow_redirects=True,
        )
        self.assertIn("Pago registrado", response.get_data(as_text=True))
        with app.app_context():
            order = db.session.get(PurchaseOrder, order_id)
            self.assertEqual(order.estado, "ANTICIPO_PAGADO")
            self.assertEqual(str(order.anticipo_pendiente), "0.00")
            self.assertEqual(str(order.monto_consumido_real), "0.00")
            self.assertEqual(str(order.saldo_comprometido), "25000.00")

    def test_smnc_approval_adds_budget_input(self):
        self.create_explosion()
        # SMNC pertenece a Administración/Costos; Supervisión conserva solo
        # requisiciones propias y recepción de materiales.
        self.login("admin@example.com")
        response = self.client.post(
            "/compras/smnc/nueva",
            data={
                "project_id": str(self.project_id),
                "budget_item_id": str(self.child_id),
                "action_type": "NUEVO",
                "supply_key": "MAT-NC-01",
                "supply_type": "MATERIAL",
                "descripcion": "Aditivo no contemplado",
                "unidad": "L",
                "cantidad": "20",
                "precio_estimado": "100",
                "justificacion_tipo": "MATERIAL_NO_CONTEMPLADO",
                "justificacion": "Requerido por especificación del supervisor",
            },
            follow_redirects=True,
        )
        self.assertIn("SMNC enviada", response.get_data(as_text=True))
        with app.app_context():
            smnc_id = MaterialChangeRequest.query.one().id
        self.login("admin@example.com")
        response = self.client.post(
            f"/compras/smnc/{smnc_id}/aprobar", follow_redirects=True
        )
        self.assertIn("explosión actualizada", response.get_data(as_text=True))
        with app.app_context():
            entry = (
                BudgetExplosionItem.query.join(SupplyItem)
                .filter(SupplyItem.clave == "MAT-NC-01")
                .one()
            )
            self.assertEqual(entry.origen, "SMNC")
            self.assertEqual(str(entry.importe_presupuestado), "2000.00")

    def test_expired_requisition_releases_pending_and_notifies(self):
        entry_id = self.create_explosion()
        with app.app_context():
            supervisor = Usuario.query.filter_by(rol="supervisor").one()
            costs = Usuario.query.filter_by(rol="costos").one()
            req = PurchaseRequisition(
                folio="REQ-2026-9999",
                project_id=self.project_id,
                fecha_solicitud=date(2026, 7, 14),
                fecha_requerida=date(2026, 7, 15),
                estado="APROBADA",
                motivo="Prueba vencimiento",
                requested_by_id=supervisor.id,
                approved_by_id=costs.id,
                approved_at=datetime(2026, 7, 14, tzinfo=timezone.utc),
                fecha_limite_oc=date(2026, 7, 17),
            )
            req.lines.append(
                PurchaseRequisitionLine(
                    explosion_item_id=entry_id,
                    cantidad_solicitada=10,
                    cantidad_aprobada=10,
                    estado_linea="APROBADA",
                )
            )
            db.session.add(req)
            db.session.commit()
            req_id = req.id
        self.client.get("/compras/requisiciones")
        with app.app_context():
            req = db.session.get(PurchaseRequisition, req_id)
            self.assertEqual(req.estado, "VENCIDA")
            self.assertGreaterEqual(PurchaseNotification.query.filter_by(tipo="REQUISICION_VENCIDA").count(), 2)
            self.assertEqual(str(db.session.get(BudgetExplosionItem, entry_id).cantidad_disponible), "100.0000")

    def test_historical_import_creates_suggestions_without_touching_budget(self):
        template = self.client.get("/compras/historico/plantilla.xlsx")
        self.assertEqual(template.status_code, 200)
        workbook = load_workbook(BytesIO(template.data), data_only=True)
        self.assertEqual(workbook["Proveedores e insumos"]["A1"].value, "Proveedor")

        response = self.client.post(
            "/compras/historico/importar",
            data={
                "project_id": str(self.project_id),
                "archivo": (self.historical_file(), "historico.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        page = response.get_data(as_text=True)
        self.assertIn("1 proveedores nuevos", page)
        self.assertIn("Fila 3", page)
        with app.app_context():
            supplier = Supplier.query.filter_by(rfc="AHI010101AA1").one()
            relation = SupplierSupplyItem.query.one()
            catalog_entry = SupplyProjectCatalog.query.one()
            self.assertEqual(relation.supplier_id, supplier.id)
            self.assertEqual(str(relation.precio_historico), "2350.0000")
            self.assertEqual(relation.fecha_ultima_compra, date(2026, 6, 30))
            self.assertEqual(catalog_entry.project_id, self.project_id)
            self.assertEqual(catalog_entry.budget_item.nombre, "MATERIALES IMPORTADOS")
            self.assertEqual(BudgetExplosionItem.query.count(), 0)
            supplier_id = supplier.id

        entry_id = self.create_explosion()
        req_id, _ = self.create_approved_requisition(
            entry_id, requested="5", approved="5"
        )
        self.login("comprador@example.com")
        self.client.post(
            f"/compras/requisiciones/{req_id}/confirmar-recepcion",
            follow_redirects=True,
        )
        detail = self.client.get(f"/compras/requisiciones/{req_id}")
        detail_page = detail.get_data(as_text=True)
        self.assertNotIn("Proveedores sugeridos por precio histórico", detail_page)
        self.assertNotIn("$2,350.00", detail_page)

        response = self.client.post(
            f"/compras/requisiciones/{req_id}/cotizaciones",
            data={"supplier_ids": str(supplier_id)},
            follow_redirects=True,
        )
        self.assertIn("Se generaron 1 solicitud(es) de cotización", response.get_data(as_text=True))
        with app.app_context():
            quote = Quotation.query.one()
            self.assertEqual(str(quote.lines[0].precio_unitario_cotizado), "2350.0000")

    def test_report_engine_pagination_columns_and_excel_exports(self):
        self.create_paid_credit_order()
        self.login("costos@example.com")
        base = (
            "/compras/reportes?period=custom&date_from=2026-07-01"
            "&date_to=2026-07-31"
        )
        for report_type, title in (
            ("cantidades", "Cantidades por insumo y obra"),
            ("comprados", "Materiales requisitados y comprados"),
            ("pendientes", "Materiales requisitados sin compra"),
            ("proveedores", "Compras por proveedor"),
            ("obras", "Compras por obra"),
        ):
            response = self.client.get(f"{base}&report={report_type}&per_page=10")
            self.assertEqual(response.status_code, 200, report_type)
            self.assertIn(title, response.get_data(as_text=True), report_type)

        export = self.client.get(
            f"{base}&report=cantidades&columns=obra"
            "&columns=importe_comprado&export=xlsx"
        )
        self.assertEqual(export.status_code, 200)
        workbook = load_workbook(BytesIO(export.data), data_only=False)
        self.assertEqual(
            [cell.value for cell in workbook["Reporte"][1]],
            ["Obra", "Imp. comprado"],
        )
        self.assertEqual(workbook["Filtros"]["B1"].value, "Cantidades por insumo y obra")

        weekly = self.client.get("/compras/reportes/pagos-semanales?week=2026-07-20")
        weekly_page = weekly.get_data(as_text=True)
        self.assertEqual(weekly.status_code, 200)
        self.assertIn("MATERIALES DEL PACIFICO", weekly_page)
        self.assertIn("$10,000.00", weekly_page)
        weekly_export = self.client.get(
            "/compras/reportes/pagos-semanales?week=2026-07-20&export=xlsx"
        )
        workbook = load_workbook(BytesIO(weekly_export.data), data_only=False)
        self.assertIn("Resumen general", workbook.sheetnames)
        self.assertIn("L-COMPRA", workbook.sheetnames)
        self.assertEqual(workbook["Resumen general"]["J2"].value, 10000)

    def test_advance_balance_can_be_applied_and_refunded(self):
        entry_id = self.create_explosion()
        supplier_id = self.create_supplier(credit=True)
        _, source_req_line_id = self.create_approved_requisition(
            entry_id, requested="10", approved="10"
        )
        self.login("comprador@example.com")
        self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "ANTICIPO",
                "fecha_entrega_estimada": "2026-07-23",
                "anticipo_monto": "10000",
                "justificacion_anticipo": "Anticipo con sobrante",
                f"cantidad_{source_req_line_id}": "10",
                f"precio_{source_req_line_id}": "2500",
            },
            follow_redirects=True,
        )
        with app.app_context():
            source = PurchaseOrder.query.one()
            source_id = source.id
            source_line_id = source.lines[0].id
        self.client.post(f"/compras/ordenes/{source_id}/emitir", follow_redirects=True)
        self.login("admin@example.com")
        self.client.post(
            f"/compras/ordenes/{source_id}/autorizar-anticipo",
            follow_redirects=True,
        )
        self.login("comprador@example.com")
        payment_response = self.client.post(
            "/compras/pagos/nuevo",
            data={
                "purchase_order_id": str(source_id),
                "order_line_id": str(source_line_id),
                "fecha": "2026-07-21",
                "monto_capturado": "10000",
                "tipo_monto": "SIN_IVA",
                "payment_method_id": str(self.method_id),
                "company_id": str(self.company_id),
                "concepto": "Anticipo origen",
            },
            follow_redirects=True,
        )
        self.assertIn("Pago registrado", payment_response.get_data(as_text=True))
        self.login("comprador@example.com")
        receipt_response = self.client.post(
            f"/compras/ordenes/{source_id}/recibir",
            data={
                "fecha_recepcion": "2026-07-22",
                "documento_proveedor": "REC-ANT-1",
                f"recibir_{source_line_id}": "2",
            },
            follow_redirects=True,
        )
        self.assertIn("Recepción registrada", receipt_response.get_data(as_text=True))
        with app.app_context():
            source = db.session.get(PurchaseOrder, source_id)
            self.assertEqual(str(source.lines[0].saldo_favor_disponible), "5000.00")

        _, target_req_line_id = self.create_approved_requisition(
            entry_id, requested="2", approved="2"
        )
        self.login("comprador@example.com")
        self.client.post(
            "/compras/ordenes/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": str(supplier_id),
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "modalidad_pago": "CREDITO",
                "fecha_entrega_estimada": "2026-07-23",
                f"cantidad_{target_req_line_id}": "2",
                f"precio_{target_req_line_id}": "2500",
            },
            follow_redirects=True,
        )
        with app.app_context():
            target = PurchaseOrder.query.order_by(PurchaseOrder.id.desc()).first()
            target_id = target.id
            target_line_id = target.lines[0].id
        self.client.post(f"/compras/ordenes/{target_id}/emitir", follow_redirects=True)
        self.login("comprador@example.com")
        self.client.post(
            f"/compras/ordenes/{target_id}/recibir",
            data={
                "fecha_recepcion": "2026-07-23",
                "documento_proveedor": "REC-DST-1",
                "fecha_factura": "2026-07-23",
                f"recibir_{target_line_id}": "2",
            },
            follow_redirects=True,
        )

        self.login("comprador@example.com")
        response = self.client.post(
            "/compras/anticipos/saldos/movimiento",
            data={
                "tipo": "APLICACION",
                "source_order_id": str(source_id),
                "source_order_line_id": str(source_line_id),
                "target_order_id": str(target_id),
                "target_order_line_id": str(target_line_id),
                "payment_method_id": str(self.method_id),
                "fecha": "2026-07-23",
                "monto": "2000",
                "referencia": "NC-APLICADA",
            },
            follow_redirects=True,
        )
        self.assertIn("Saldo aplicado", response.get_data(as_text=True))
        response = self.client.post(
            "/compras/anticipos/saldos/movimiento",
            data={
                "tipo": "REEMBOLSO",
                "source_order_id": str(source_id),
                "source_order_line_id": str(source_line_id),
                "payment_method_id": str(self.method_id),
                "fecha": "2026-07-23",
                "monto": "1000",
                "referencia": "DEV-001",
            },
            follow_redirects=True,
        )
        self.assertIn("Reembolso", response.get_data(as_text=True))
        with app.app_context():
            source = db.session.get(PurchaseOrder, source_id)
            target = db.session.get(PurchaseOrder, target_id)
            self.assertEqual(SupplierAdvanceMovement.query.count(), 2)
            self.assertEqual(str(source.saldo_favor_disponible), "2000.00")
            self.assertEqual(str(source.monto_pagado), "7000.00")
            self.assertEqual(str(target.monto_pagado), "2000.00")
            self.assertEqual(str(target.saldo_pendiente), "3000.00")

    def test_daily_run_shared_catalog_and_role_labels(self):
        with app.app_context():
            custom = PaymentMethod(
                nombre="TARJETA EMPRESARIAL", descripcion="Tarjeta corporativa"
            )
            office = CentroCosto(
                nombre="Oficina Central",
                codigo="OF-CENTRAL",
                tipo="oficina",
                estado="activa",
                fecha_apertura=date(2026, 7, 1),
                presupuesto_total=50000,
                presupuesto_mano_obra=0,
            )
            db.session.add_all([custom, office])
            db.session.flush()
            item = BudgetItem(
                project_id=office.id,
                codigo="OF-01",
                nombre="Papelería",
                categoria="INDIRECTO",
                presupuesto=10000,
            )
            db.session.add(item)
            db.session.commit()
            custom_id, office_id, item_id = custom.id, office.id, item.id

        response = self.client.post(
            "/gastos-oficina/nuevo",
            data={
                "project_id": str(office_id),
                "budget_item_id": str(item_id),
                "fecha": "2026-07-21",
                "proveedor": "Papelería Uno",
                "concepto": "Insumos administrativos",
                "monto_capturado": "500",
                "tipo_monto": "SIN_IVA",
                "payment_method_id": str(custom_id),
                "company_id": str(self.company_id),
            },
            follow_redirects=True,
        )
        self.assertIn("Gasto de oficina registrado", response.get_data(as_text=True))
        loan_form = self.client.get("/prestamos/nuevo").get_data(as_text=True)
        self.assertNotIn("TARJETA EMPRESARIAL", loan_form)
        self.assertIn(">Efectivo<", loan_form)
        self.assertIn(">Transferencia<", loan_form)
        users_page = self.client.get("/admin/usuarios").get_data(as_text=True)
        self.assertIn("Comprador", users_page)
        self.assertIn("Costos", users_page)
        self.assertIn("Capturista", users_page)
        self.assertIn("Supervisor de obra", users_page)

        runner = app.test_cli_runner()
        result = runner.invoke(args=["compras-alertas", "--force"])
        self.assertEqual(result.exit_code, 0, result.output)
        with app.app_context():
            self.assertEqual(PurchaseAlertRun.query.count(), 1)

    def test_operational_orders_are_isolated_authorized_and_reported(self):
        operation_entry_id = self.create_explosion()
        with app.app_context():
            operation_entry = db.session.get(
                BudgetExplosionItem, operation_entry_id
            )
            operation_entry.supply_item.es_operacion = True
            operation_entry.supply_item.categoria_operacion = "AGUA"
            operation_entry.clasificacion = "OPERATIVO"
            db.session.commit()

        self.login("supervisor@example.com")
        order_response = self.client.post(
            "/compras/ordenes-operaciones/nueva",
            data={
                "project_id": str(self.project_id),
                "supplier_id": "0",
                "beneficiario_libre": "Pipa de Agua del Cabo",
                "fecha_entrega_estimada": "2026-07-22",
                "condicion_saldo": "CONTRA_RECEPCION",
                "lineas-0-explosion_item_id": str(operation_entry_id),
                "lineas-0-cantidad": "10",
                "lineas-0-precio_unitario_sin_iva": "190",
                "lineas-0-observacion": "Entregar agua antes de las 08:00.",
            },
            follow_redirects=True,
        )
        self.assertIn(
            "guardada y enviada a autorización",
            order_response.get_data(as_text=True),
        )
        with app.app_context():
            order = PurchaseOrder.query.one()
            order_id = order.id
            order_line_id = order.lines[0].id
            order_folio = order.folio
            self.assertEqual(order.tipo_oc, "OPERACIONES")
            self.assertEqual(order.categoria_pago, "OPERACIONES")
            self.assertEqual(order.estado, "PENDIENTE_AUTORIZACION")
            self.assertIsNone(order.supplier_id)
            self.assertEqual(order.beneficiario_libre, "Pipa de Agua del Cabo")
            self.assertTrue(order.requiere_autorizacion)
            self.assertEqual(PurchaseRequisition.query.count(), 0)
            self.assertEqual(len(order.payment_schedules), 1)
            self.assertEqual(
                order.payment_schedules[0].estado,
                "PENDIENTE_RECEPCION",
            )

        self.login("comprador@example.com")
        self.assertEqual(
            self.client.get(f"/compras/ordenes/{order_id}").status_code, 403
        )
        self.assertNotIn(
            order_folio, self.client.get("/compras/ordenes").get_data(as_text=True)
        )

        self.login("admin@example.com")
        authorized = self.client.post(
            f"/compras/ordenes/{order_id}/autorizar-operacion",
            follow_redirects=True,
        )
        self.assertIn(
            "OC aprobada y emitida automáticamente",
            authorized.get_data(as_text=True),
        )
        with app.app_context():
            self.assertEqual(db.session.get(PurchaseOrder, order_id).estado, "EMITIDA")

        validated = self.client.post(
            f"/compras/ordenes/{order_id}/validar-beneficiario",
            data={
                "company_id": str(self.company_id),
                "payment_method_id": str(self.method_id),
                "beneficiario_confirmado": "Pipa de Agua del Cabo",
                "comentario": "Identidad y cuenta verificadas por Finanzas.",
            },
            follow_redirects=True,
        )
        self.assertIn("Beneficiario validado por Finanzas", validated.get_data(as_text=True))
        received = self.client.post(
            f"/compras/ordenes/{order_id}/recibir",
            data={
                "fecha_recepcion": "2026-07-22",
                "documento_proveedor": "FAC-OPS-001",
                "fecha_factura": "2026-07-22",
                f"recibir_{order_line_id}": "10",
            },
            follow_redirects=True,
        )
        self.assertIn("Recepción registrada", received.get_data(as_text=True))
        paid = self.client.post(
            "/compras/pagos/nuevo",
            data={
                "purchase_order_id": str(order_id),
                "order_line_id": str(order_line_id),
                "fecha": "2026-07-22",
                "monto_capturado": "1900",
                "tipo_monto": "SIN_IVA",
                "payment_method_id": str(self.method_id),
                "company_id": str(self.company_id),
                "concepto": "Pago de agua operativa",
            },
            follow_redirects=True,
        )
        self.assertIn("Pago registrado", paid.get_data(as_text=True))
        payment_filter = self.client.get("/compras/pagos").get_data(as_text=True)
        self.assertIn("Pago de agua operativa", payment_filter)

        operations_report = self.client.get(
            "/compras/reportes/nominas-operaciones",
            query_string={
                "fecha_desde": "2026-07-01",
                "fecha_hasta": "2026-07-31",
            },
        ).get_data(as_text=True)
        self.assertIn(order_folio, operations_report)
        self.assertIn("Gasto operativo", operations_report)
        supplier_report = self.client.get(
            "/compras/reportes/pagos-proveedores",
            query_string={
                "fecha_desde": "2026-07-01",
                "fecha_hasta": "2026-07-31",
            },
        ).get_data(as_text=True)
        self.assertIn("Pago de agua operativa", supplier_report)
        purchase_report = self.client.get(
            "/compras/reportes",
            query_string={
                "report": "comprados",
                "period": "custom",
                "date_from": "2026-07-01",
                "date_to": "2026-07-31",
            },
        ).get_data(as_text=True)
        self.assertIn("0 registro(s)", purchase_report)
        report_table = purchase_report.split('<div class="table-responsive">', 1)[1]
        report_table = report_table.split("</div>", 1)[0]
        self.assertNotIn(order_folio, report_table)

    def test_user_permissions_can_be_overridden_individually(self):
        with app.app_context():
            buyer = Usuario.query.filter_by(rol="comprador").one()
            buyer_id = buyer.id
            self.assertEqual(len(buyer.permisos), len(MODULOS_PERMISOS))
            data = {
                "nombre_completo": buyer.nombre_completo,
                "correo": buyer.correo,
                "contrasena": "",
                "rol": buyer.rol,
                "centro_costo_id": str(self.project_id),
            }
            for permission in buyer.permisos:
                for action in ACCIONES_PERMISO:
                    if permission.modulo != "proveedores" and getattr(
                        permission, f"puede_{action}"
                    ):
                        data[f"perm_{permission.modulo}_{action}"] = "on"

        updated = self.client.post(
            f"/admin/usuarios/{buyer_id}/editar",
            data=data,
            follow_redirects=True,
        )
        self.assertIn(
            "Usuario y permisos actualizados correctamente",
            updated.get_data(as_text=True),
        )
        with app.app_context():
            buyer = db.session.get(Usuario, buyer_id)
            self.assertFalse(buyer.tiene_permiso("proveedores", "ver"))
            self.assertTrue(buyer.tiene_permiso("compras", "ver"))
            provider_permission = Permiso.query.filter_by(
                usuario_id=buyer_id, modulo="proveedores"
            ).one()
            self.assertFalse(provider_permission.puede_crear)

        self.login("comprador@example.com")
        self.assertEqual(self.client.get("/compras/proveedores").status_code, 403)
        self.assertEqual(self.client.get("/compras/ordenes").status_code, 200)
        nav = self.client.get("/compras/").get_data(as_text=True)
        self.assertNotIn(">Proveedores</a>", nav)

    def test_buyer_can_cancel_order_from_any_project_with_permission(self):
        supplier_id = self.create_supplier()
        with app.app_context():
            foreign_project = CentroCosto(
                nombre="Casa Ajena",
                codigo="L-AJENA",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2026, 7, 1),
                presupuesto_total=200000,
                presupuesto_mano_obra=50000,
            )
            db.session.add(foreign_project)
            db.session.flush()
            buyer = Usuario.query.filter_by(rol="comprador").one()
            order = PurchaseOrder(
                folio="OC-2026-AJENA",
                project_id=foreign_project.id,
                supplier_id=supplier_id,
                company_id=self.company_id,
                buyer_id=buyer.id,
                payment_method_id=self.method_id,
                fecha_orden=date(2026, 7, 21),
                fecha_entrega_estimada=date(2026, 7, 25),
                fecha_limite=date(2026, 7, 28),
                tipo_oc="COMPRAS",
                categoria_pago="COMPRAS",
                estado="BORRADOR",
                modalidad_pago="CREDITO",
                created_by_id=buyer.id,
            )
            db.session.add(order)
            db.session.commit()
            order_id = order.id

        self.login("comprador@example.com")
        # Estar asignado a una obra no sustituye el permiso del módulo:
        # Comprador no puede abrir ni mutar Nóminas mediante una URL directa.
        self.assertEqual(self.client.get("/nominas").status_code, 403)
        response = self.client.post(
            f"/compras/ordenes/{order_id}/cancelar",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            self.assertEqual(
                db.session.get(PurchaseOrder, order_id).estado,
                "CANCELADA",
            )

    def test_user_cannot_self_assign_admin_or_grant_missing_permission(self):
        with app.app_context():
            buyer = Usuario.query.filter_by(rol="comprador").one()
            target = Usuario.query.filter_by(rol="supervisor").one()
            buyer_id = buyer.id
            target_id = target.id
            users_permission = Permiso.query.filter_by(
                usuario_id=buyer.id,
                modulo="usuarios",
            ).one()
            users_permission.puede_ver = True
            users_permission.puede_editar = True
            db.session.commit()

            own_data = {
                "nombre_completo": buyer.nombre_completo,
                "correo": buyer.correo,
                "contrasena": "",
                "rol": "admin",
                "centro_costo_id": str(self.project_id),
            }
            for permission in buyer.permisos:
                for action in ("ver", "crear", "editar", "eliminar"):
                    if getattr(permission, f"puede_{action}"):
                        own_data[f"perm_{permission.modulo}_{action}"] = "on"

        self.login("comprador@example.com")
        admin_route = self.client.post(
            f"/admin/usuarios/{buyer_id}/editar",
            data=own_data,
            follow_redirects=False,
        )
        self.assertEqual(admin_route.status_code, 403)

        legacy_route = self.client.post(
            f"/usuarios/{buyer_id}/editar",
            data={
                "nombre_completo": "Comprador Compras",
                "email": "comprador@example.com",
                "role": "administrador",
                "project_ids": str(self.project_id),
                "password": "",
            },
            follow_redirects=False,
        )
        self.assertEqual(legacy_route.status_code, 403)

        with app.app_context():
            target = db.session.get(Usuario, target_id)
            target_data = {
                "nombre_completo": target.nombre_completo,
                "correo": target.correo,
                "contrasena": "",
                "rol": target.rol,
                "centro_costo_id": str(self.project_id),
            }
            for permission in target.permisos:
                for action in ("ver", "crear", "editar", "eliminar"):
                    if getattr(permission, f"puede_{action}"):
                        target_data[f"perm_{permission.modulo}_{action}"] = "on"
            target_data["perm_seguridad_editar"] = "on"

        excessive_grant = self.client.post(
            f"/admin/usuarios/{target_id}/editar",
            data=target_data,
            follow_redirects=False,
        )
        self.assertEqual(excessive_grant.status_code, 403)
        with app.app_context():
            buyer = db.session.get(Usuario, buyer_id)
            target = db.session.get(Usuario, target_id)
            self.assertEqual(buyer.rol, "comprador")
            self.assertFalse(target.tiene_permiso("seguridad", "editar"))

    def test_credit_cards_payments_and_same_day_alerts(self):
        def create_card(last_four, cut_date, due_date, balance):
            return self.client.post(
                "/compras/tarjetas/nueva",
                data={
                    "empresa_id": str(self.company_id),
                    "numero_tarjeta": last_four,
                    "fecha_corte": cut_date,
                    "fecha_pago": due_date,
                    "limite_credito": "50000",
                    "saldo_actual": balance,
                },
                follow_redirects=True,
            )

        first = create_card("1234", "2026-07-10", "2026-07-23", "10000")
        self.assertIn("Tarjeta registrada correctamente", first.get_data(as_text=True))
        dashboard = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("Tarjetas próximas a vencer", dashboard)
        self.assertIn("**** **** **** 1234", dashboard)

        second = create_card("5678", "2026-07-15", "2026-07-24", "5000")
        self.assertIn("Tarjeta registrada correctamente", second.get_data(as_text=True))
        same_day = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("**** **** **** 5678", same_day)
        with app.app_context():
            run = PurchaseAlertRun.query.one()
            self.assertEqual(run.tarjetas_por_vencer, 2)
            self.assertEqual(
                PurchaseNotification.query.filter_by(tipo="TARJETA_POR_PAGAR").count(),
                2,
            )
            first_card = CreditCard.query.filter_by(
                numero_tarjeta="**** **** **** 1234"
            ).one()
            first_card_id = first_card.id

        payment = self.client.post(
            f"/compras/tarjetas/{first_card_id}/pagos/nuevo",
            data={
                "fecha": "2026-07-21",
                "monto": "2500",
                "referencia": "SPEI-TDC-001",
                "notas": "Pago parcial antes del vencimiento",
            },
            follow_redirects=True,
        )
        self.assertIn("Pago de tarjeta registrado", payment.get_data(as_text=True))
        with app.app_context():
            card = db.session.get(CreditCard, first_card_id)
            self.assertEqual(card.saldo_actual, Decimal("7500.00"))
            self.assertEqual(CreditCardPayment.query.count(), 1)

        report = self.client.get(
            "/compras/reportes/pagos-proveedores",
            query_string={
                "fecha_desde": "2026-07-01",
                "fecha_hasta": "2026-07-31",
            },
        ).get_data(as_text=True)
        self.assertIn("SPEI-TDC-001", report)
        self.assertIn("$2,500.00", report)

    def test_roles_all_day_requisitions_and_old_payment_redirect(self):
        self.login("supervisor@example.com")
        self.assertEqual(self.client.get("/compras/proveedores").status_code, 200)
        self.assertEqual(self.client.get("/compras/ordenes/nueva").status_code, 403)
        self.assertEqual(
            self.client.get("/compras/ordenes-operaciones/nueva").status_code,
            200,
        )
        app.config["COMPRAS_TODAY"] = date(2026, 7, 22)  # miércoles
        response = self.client.get("/compras/requisiciones/nueva", follow_redirects=True)
        self.assertIn("Disponible todos los días", response.get_data(as_text=True))
        response = self.client.post(
            "/compras/requisiciones/nueva",
            data={
                "project_id": str(self.project_id),
                "fecha_requerida": "2026-07-23",
                "motivo": "Requisición capturada en miércoles",
            },
            follow_redirects=True,
        )
        self.assertIn("REQ-2026-0001", response.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(PurchaseRequisition.query.count(), 1)
            self.assertEqual(PurchaseRequisition.query.one().tipo_requisicion, "COMPRAS")
        response = self.client.get("/pagos-adicionales/nuevo", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/compras/pagos/nuevo", response.headers["Location"])
        self.login("admin@example.com")
        response = self.client.get("/pagos-adicionales/nuevo", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/compras/pagos/nuevo", response.headers["Location"])

    def test_role_menus_and_server_permissions_are_separated(self):
        self.login("capturista@example.com")
        payroll_home = self.client.get("/dashboard").get_data(as_text=True)
        self.assertIn("Panel de capturista de nómina", payroll_home)
        self.assertIn("Abrir Nóminas", payroll_home)
        self.assertNotIn("Abrir Compras", payroll_home)
        self.assertEqual(self.client.get("/compras/").status_code, 403)
        self.assertEqual(self.client.get("/admin/usuarios").status_code, 403)

        self.login("comprador@example.com")
        buyer_dashboard = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("Dashboard semanal de Compras", buyer_dashboard)
        self.assertIn('href="/compras/proveedores"', buyer_dashboard)
        self.assertIn(">Proveedores</span>", buyer_dashboard)
        self.assertIn('href="/compras/explosion"', buyer_dashboard)
        self.assertIn(">Explosión de insumos</span>", buyer_dashboard)
        self.assertIn('href="/compras/reportes"', buyer_dashboard)
        self.assertIn(">Reportes de compras</span>", buyer_dashboard)
        self.assertEqual(self.client.get("/compras/explosion").status_code, 200)
        self.assertEqual(self.client.get("/compras/reportes").status_code, 200)
        self.assertEqual(self.client.get("/panel").status_code, 403)

        self.login("supervisor@example.com")
        supervisor_dashboard = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("Panel de Operaciones", supervisor_dashboard)
        self.assertIn(">Proveedores</span>", supervisor_dashboard)
        self.assertEqual(self.client.get("/compras/proveedores").status_code, 200)
        self.assertEqual(self.client.get("/compras/smnc").status_code, 200)
        self.assertEqual(self.client.get("/panel").status_code, 200)

        self.login("costos@example.com")
        costs_dashboard = self.client.get("/compras/").get_data(as_text=True)
        self.assertIn("Ver reportes", costs_dashboard)
        self.assertNotIn("Líneas de crédito de proveedores", costs_dashboard)
        self.assertNotIn("Estados de cuenta y vencimientos", costs_dashboard)
        self.assertEqual(self.client.get("/compras/proveedores").status_code, 403)
        self.assertEqual(self.client.get("/compras/ordenes").status_code, 200)
        self.assertEqual(self.client.get("/compras/reportes").status_code, 200)
        self.assertEqual(self.client.get("/obras").status_code, 200)

        self.login("admin@example.com")
        self.assertEqual(self.client.get("/admin/usuarios").status_code, 200)
        self.assertEqual(self.client.get("/compras/proveedores").status_code, 200)
        self.assertEqual(self.client.get("/compras/reportes").status_code, 200)

    def test_bch_control_branding_login_and_responsive_shell(self):
        self.client.post("/logout", follow_redirects=True)
        login_response = self.client.get("/login")
        self.assertEqual(login_response.status_code, 200)
        login_html = login_response.get_data(as_text=True)
        self.assertIn("<title>Iniciar sesión · BCH Control</title>", login_html)
        self.assertIn('class="login-shell"', login_html)
        self.assertIn('src="/static/img/logo.png"', login_html)
        self.assertIn(">BCH Control</h1>", login_html)
        self.assertIn("btn-accent", login_html)

        style_response = self.client.get("/static/css/style.css")
        try:
            self.assertEqual(style_response.status_code, 200)
            style = style_response.get_data(as_text=True)
            self.assertIn("--color-primary: #1a3a5c", style)
            self.assertIn("--color-accent: #e87e2f", style)
            self.assertIn("--bs-offcanvas-bg: var(--color-sidebar)", style)
            self.assertIn(
                "background-color: var(--color-sidebar) !important", style
            )
            self.assertIn("color: rgba(255, 255, 255, .88)", style)
        finally:
            style_response.close()
        for logo_path in ("/static/img/logo.png", "/static/img/logo-sidebar.png"):
            logo_response = self.client.get(logo_path)
            try:
                self.assertEqual(logo_response.status_code, 200)
            finally:
                logo_response.close()

        self.login("admin@example.com")
        dashboard_html = self.client.get("/dashboard").get_data(as_text=True)
        self.assertIn('id="appSidebar"', dashboard_html)
        self.assertIn('data-bs-target="#appSidebar"', dashboard_html)
        self.assertIn(">BCH Control</span>", dashboard_html)
        self.assertIn("© 2026 Baja Custom Homes – BCH Control", dashboard_html)


if __name__ == "__main__":
    unittest.main()
