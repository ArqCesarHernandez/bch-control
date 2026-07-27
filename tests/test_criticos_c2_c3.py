"""Regresiones de ciclo financiero y endurecimiento de producción."""

import os
import subprocess
import sys
import unittest
from datetime import date, timedelta
from io import BytesIO

import pyotp
from openpyxl import load_workbook

os.environ["FLASK_ENV"] = "testing"
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["SECRET_KEY"] = "test-critical-security-secret"

from app import create_app  # noqa: E402
from compras_models import PaymentMethod, PurchaseNotification  # noqa: E402
from config import CRITICAL_SECRET_ERROR, validate_secret_key  # noqa: E402
from models import BitacoraAuditoria, CentroCosto, Permiso, Usuario, db, utc_now  # noqa: E402
from nominas_models import (  # noqa: E402
    BudgetItem,
    Company,
    Employee,
    Loan,
    LoanPayment,
    Payroll,
)
from routes.nominas import scheduled_loan_deduction  # noqa: E402


app = create_app()


class CriticalC2C3Test(unittest.TestCase):
    def setUp(self):
        app.config.update(
            TESTING=True,
            WTF_CSRF_ENABLED=False,
            MFA_REQUIRED_FOR_ADMINS=False,
            IS_PRODUCTION=False,
        )
        self.client = app.test_client()
        with app.app_context():
            db.drop_all()
            db.create_all()

            project = CentroCosto(
                nombre="Obra Seguridad",
                codigo="SEG-01",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2025, 1, 1),
            )
            db.session.add(project)
            db.session.flush()

            company = Company(nombre="Baja Custom Homes", codigo="BCH")
            method = PaymentMethod(nombre="EFECTIVO", descripcion="Efectivo", activo=True)
            item = BudgetItem(
                project_id=project.id,
                codigo="MO-01",
                nombre="Mano de obra",
                categoria="MANO_OBRA",
                presupuesto=100000,
            )
            db.session.add_all([company, method, item])
            db.session.flush()

            admin = self._user("Administradora General", "admin@example.com", "admin")
            supervisor = self._user(
                "Supervisor de Obra",
                "supervisor@example.com",
                "supervisor",
                project,
            )
            financial = self._user(
                "Administrador Financiero",
                "finanzas@example.com",
                "admin_financiero",
            )
            costs = self._user("Analista de Costos", "costos@example.com", "costos")
            db.session.flush()

            employee = Employee(
                nombre_completo="TRABAJADOR DE PRUEBA",
                fecha_ingreso=date(2025, 1, 1),
                puesto="OFICIAL",
                project_id=project.id,
                budget_item_id=item.id,
                salario_semanal=5000,
                descuento_infonavit=100,
                registrado_imss=True,
                nss="12345678901",
                empresa_imss_id=company.id,
                transferencia_predeterminada=2000,
                empresa_transferencia_id=company.id,
                empresa_efectivo_id=company.id,
            )
            db.session.add(employee)
            db.session.commit()

            self.project_id = project.id
            self.company_id = company.id
            self.method_id = method.id
            self.employee_id = employee.id
            self.admin_id = admin.id
            self.supervisor_id = supervisor.id
            self.financial_id = financial.id
            self.costs_id = costs.id

    @staticmethod
    def _user(name, email, role, project=None):
        user = Usuario(
            nombre_completo=name,
            correo=email,
            rol=role,
            centro_costo_id=project.id if project and role in {"supervisor", "capturista"} else None,
            activo=True,
        )
        user.set_password("Password123!")
        user.asignar_permisos_predeterminados()
        if project:
            user.projects = [project]
        db.session.add(user)
        return user

    def login(self, email, password="Password123!", *, follow=True):
        return self.client.post(
            "/login",
            data={"correo": email, "contrasena": password},
            follow_redirects=follow,
        )

    def logout(self):
        return self.client.post("/logout", follow_redirects=True)

    def request_loan(self, employee_id=None, amount="1000", loan_date="2026-07-10"):
        return self.client.post(
            "/prestamos/nuevo",
            data={
                "employee_id": str(employee_id or self.employee_id),
                "fecha_prestamo": loan_date,
                "monto": amount,
                "retencion_semanal": "200",
                "payment_method_id": str(self.method_id),
                "company_id": str(self.company_id),
                "concepto": "Préstamo de prueba",
            },
            follow_redirects=True,
        )

    def test_loan_interest_approval_and_active_only_withholding(self):
        self.login("supervisor@example.com")
        response = self.request_loan()
        self.assertIn("Solicitud enviada a los administradores", response.get_data(as_text=True))
        with app.app_context():
            loan = Loan.query.one()
            loan_id = loan.id
            self.assertEqual(loan.estado, "pendiente")
            self.assertEqual(float(loan.tasa_interes), 5.0)
            self.assertEqual(float(loan.total_pagar), 1050.0)
            self.assertEqual(float(loan.saldo_pendiente), 1050.0)
            self.assertEqual(loan.solicitante_id, self.supervisor_id)
            self.assertIsNone(loan.aprobador_id)
            self.assertEqual(PurchaseNotification.query.count(), 1)
            self.assertEqual(
                float(scheduled_loan_deduction(self.employee_id, date(2026, 7, 13))),
                0.0,
            )

        self.logout()
        self.login("admin@example.com")
        self.client.post(f"/prestamos/{loan_id}/aprobar", follow_redirects=True)
        with app.app_context():
            loan = db.session.get(Loan, loan_id)
            self.assertEqual(loan.estado, "activo")
            self.assertEqual(loan.aprobador_id, self.admin_id)
            self.assertIsNotNone(loan.fecha_aprobacion)
            self.assertEqual(
                float(scheduled_loan_deduction(self.employee_id, date(2026, 7, 6))),
                0.0,
            )
            self.assertEqual(
                float(scheduled_loan_deduction(self.employee_id, date(2026, 7, 13))),
                200.0,
            )

    def test_loan_rejects_insufficient_seniority_and_excess_total(self):
        with app.app_context():
            new_employee = Employee(
                nombre_completo="TRABAJADOR NUEVO",
                fecha_ingreso=date(2026, 3, 15),
                puesto="AYUDANTE",
                project_id=self.project_id,
                budget_item_id=BudgetItem.query.one().id,
                salario_semanal=5000,
            )
            db.session.add(new_employee)
            db.session.commit()
            new_employee_id = new_employee.id

        self.login("supervisor@example.com")
        seniority = self.request_loan(employee_id=new_employee_id)
        self.assertIn(
            "El trabajador no cumple con la antigüedad mínima de 6 meses.",
            seniority.get_data(as_text=True),
        )
        excessive = self.request_loan(amount="4700")
        self.assertIn(
            "El préstamo no puede exceder el salario semanal del trabajador.",
            excessive.get_data(as_text=True),
        )
        with app.app_context():
            self.assertEqual(Loan.query.count(), 0)

        self.request_loan(amount="500")
        with app.app_context():
            rejected_id = Loan.query.one().id
        self.logout()
        self.login("admin@example.com")
        rejected = self.client.post(
            f"/prestamos/{rejected_id}/rechazar",
            data={"motivo": "Solicitud no autorizada por Dirección"},
            follow_redirects=True,
        )
        self.assertIn("Solicitud de préstamo rechazada", rejected.get_data(as_text=True))
        with app.app_context():
            loan = db.session.get(Loan, rejected_id)
            self.assertEqual(loan.estado, "rechazado")
            self.assertEqual(loan.aprobador_id, self.admin_id)
            self.assertIsNotNone(loan.fecha_aprobacion)
            self.assertEqual(loan.motivo_rechazo, "Solicitud no autorizada por Dirección")

        self.logout()
        self.login("costos@example.com")
        self.assertEqual(self.client.get("/prestamos/nuevo").status_code, 403)

    def test_payroll_lifecycle_reversal_and_paid_lock(self):
        with app.app_context():
            loan = Loan(
                employee_id=self.employee_id,
                fecha_prestamo=date(2026, 7, 10),
                monto=1000,
                tasa_interes=5,
                total_pagar=1050,
                retencion_semanal=200,
                metodo_entrega="EFECTIVO",
                payment_method_id=self.method_id,
                company_id=self.company_id,
                estado="activo",
                solicitante_id=self.supervisor_id,
                aprobador_id=self.admin_id,
                fecha_aprobacion=utc_now(),
                created_by_id=self.supervisor_id,
            )
            db.session.add(loan)
            db.session.commit()

        self.login("supervisor@example.com")
        self.client.post(
            "/nominas/nueva",
            data={"project_id": str(self.project_id), "semana_inicio": "2026-07-13"},
            follow_redirects=True,
        )
        with app.app_context():
            payroll = Payroll.query.one()
            payroll_id = payroll.id
            line_id = payroll.lines[0].id
            item_id = BudgetItem.query.one().id

        self.client.post(
            f"/nominas/{payroll_id}/guardar",
            data={
                f"line_{line_id}_partida_id": str(item_id),
                f"line_{line_id}_subpartida_id": "",
                f"line_{line_id}_puesto": "OFICIAL",
                f"line_{line_id}_salario_semanal": "5000",
                f"line_{line_id}_lunes": "on",
                f"line_{line_id}_martes": "on",
                f"line_{line_id}_miercoles": "on",
                f"line_{line_id}_jueves": "on",
                f"line_{line_id}_viernes": "on",
                f"line_{line_id}_descuento_infonavit": "100",
                f"line_{line_id}_pago_transferencia": "2000",
                f"line_{line_id}_empresa_transferencia_id": str(self.company_id),
                f"line_{line_id}_empresa_efectivo_id": str(self.company_id),
            },
            follow_redirects=True,
        )

        self.client.post(f"/nominas/{payroll_id}/cerrar", follow_redirects=True)
        with app.app_context():
            self.assertEqual(db.session.get(Payroll, payroll_id).estado, "enviada")
            self.assertEqual(LoanPayment.query.count(), 0)

        self.logout()
        self.login("admin@example.com")
        self.client.post(f"/nominas/{payroll_id}/aprobar", follow_redirects=True)
        with app.app_context():
            self.assertEqual(db.session.get(Payroll, payroll_id).estado, "aprobada")
            self.assertEqual(LoanPayment.query.count(), 1)

        no_reason = self.client.post(f"/nominas/{payroll_id}/reabrir", follow_redirects=True)
        self.assertIn("Indica el motivo", no_reason.get_data(as_text=True))
        self.client.post(
            f"/nominas/{payroll_id}/reabrir",
            data={"motivo": "Corregir una asistencia"},
            follow_redirects=True,
        )
        with app.app_context():
            self.assertEqual(db.session.get(Payroll, payroll_id).estado, "borrador")
            self.assertEqual(LoanPayment.query.count(), 0)
            event = BitacoraAuditoria.query.filter_by(accion="REABRIR").one()
            self.assertIn("Corregir una asistencia", event.detalle)

        self.client.post(f"/nominas/{payroll_id}/aprobar", follow_redirects=True)
        self.logout()
        self.login("finanzas@example.com")
        self.client.post(f"/nominas/{payroll_id}/pagar", follow_redirects=True)
        blocked = self.client.post(
            f"/nominas/{payroll_id}/reabrir",
            data={"motivo": "Intento posterior al pago"},
            follow_redirects=True,
        )
        self.assertIn("Una nómina pagada no puede reabrirse", blocked.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(db.session.get(Payroll, payroll_id).estado, "pagada")
            self.assertEqual(LoanPayment.query.count(), 1)

        self.client.post(f"/nominas/{payroll_id}/conciliar", follow_redirects=True)
        with app.app_context():
            self.assertEqual(db.session.get(Payroll, payroll_id).estado, "conciliada")

    def test_login_rate_limit_logs_every_failure(self):
        unknown = self.login("cuenta-inexistente@example.com", "Incorrecta!", follow=True)
        self.assertIn("Correo, contraseña", unknown.get_data(as_text=True))
        for attempt in range(5):
            response = self.login("supervisor@example.com", "Incorrecta!", follow=True)
            if attempt < 4:
                self.assertIn("Correo, contraseña", response.get_data(as_text=True))
        self.assertIn(
            "Demasiados intentos. Intenta de nuevo en 15 minutos.",
            response.get_data(as_text=True),
        )
        with app.app_context():
            user = db.session.get(Usuario, self.supervisor_id)
            self.assertEqual(user.intentos_fallidos, 5)
            self.assertIsNotNone(user.bloqueado_hasta)
            self.assertEqual(
                BitacoraAuditoria.query.filter_by(
                    usuario_id=self.supervisor_id, accion="LOGIN_FALLIDO"
                ).count(),
                5,
            )
            self.assertEqual(
                BitacoraAuditoria.query.filter_by(accion="LOGIN_FALLIDO").count(),
                6,
            )

        blocked = self.login("supervisor@example.com", follow=True)
        self.assertIn("Demasiados intentos", blocked.get_data(as_text=True))
        with app.app_context():
            user = db.session.get(Usuario, self.supervisor_id)
            user.bloqueado_hasta = utc_now() - timedelta(minutes=1)
            user.ventana_intentos_inicio = utc_now() - timedelta(minutes=16)
            db.session.commit()
        allowed = self.login("supervisor@example.com", follow=False)
        self.assertEqual(allowed.status_code, 302)
        self.assertTrue(allowed.headers["Location"].endswith("/dashboard"))

    def test_admin_mfa_and_security_headers(self):
        secret = pyotp.random_base32()
        with app.app_context():
            admin = db.session.get(Usuario, self.admin_id)
            admin.mfa_secret = secret
            admin.mfa_confirmado_en = utc_now()
            db.session.commit()

        app.config["MFA_REQUIRED_FOR_ADMINS"] = True
        first = self.login("admin@example.com", follow=False)
        self.assertEqual(first.status_code, 302)
        self.assertTrue(first.headers["Location"].endswith("/mfa/verificar"))
        verified = self.client.post(
            "/mfa/verificar",
            data={"codigo": pyotp.TOTP(secret).now()},
            follow_redirects=False,
        )
        self.assertEqual(verified.status_code, 302)
        self.assertTrue(verified.headers["Location"].endswith("/dashboard"))

        principal = self.client.post(
            f"/admin/usuarios/{self.admin_id}/desactivar",
            follow_redirects=True,
        )
        self.assertIn(
            "El administrador principal (id=1) debe permanecer activo.",
            principal.get_data(as_text=True),
        )
        with app.app_context():
            self.assertTrue(db.session.get(Usuario, self.admin_id).activo)

        app.config["IS_PRODUCTION"] = True
        response = self.client.get("/dashboard")
        self.assertEqual(
            response.headers["Strict-Transport-Security"],
            "max-age=31536000; includeSubDomains",
        )
        self.assertEqual(response.headers["X-Content-Type-Options"], "nosniff")
        self.assertEqual(response.headers["X-Frame-Options"], "DENY")
        self.assertEqual(response.headers["Referrer-Policy"], "same-origin")

    def test_nss_is_masked_in_pages_and_excel_without_permission(self):
        with app.app_context():
            report_permission = Permiso.query.filter_by(
                usuario_id=self.supervisor_id, modulo="reportes_nomina"
            ).one()
            report_permission.puede_ver = True
            db.session.commit()

        self.login("supervisor@example.com")
        self.client.post(
            "/nominas/nueva",
            data={"project_id": str(self.project_id), "semana_inicio": "2026-07-13"},
            follow_redirects=True,
        )
        page = self.client.get("/trabajadores").get_data(as_text=True)
        self.assertIn("****8901", page)
        self.assertNotIn("12345678901", page)
        export = self.client.get("/reportes/exportar.xlsx")
        workbook = load_workbook(BytesIO(export.data), read_only=True, data_only=True)
        sheet = workbook["Nóminas"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        nss_column = headers.index("NSS") + 1
        self.assertEqual(sheet.cell(row=2, column=nss_column).value, "****8901")

        self.logout()
        self.login("admin@example.com")
        admin_page = self.client.get(f"/trabajadores/{self.employee_id}").get_data(as_text=True)
        self.assertIn("12345678901", admin_page)

    def test_production_rejects_insecure_secret_and_sqlite(self):
        with self.assertRaisesRegex(RuntimeError, "ERROR CRÍTICO"):
            validate_secret_key("REEMPLAZAR_CON_UNA_CLAVE")
        with self.assertRaisesRegex(RuntimeError, "ERROR CRÍTICO"):
            validate_secret_key("GENERA_UNA_CLAVE_ALEATORIA_DE_32_CARACTERES_O_MAS")
        self.assertEqual(
            validate_secret_key("A9!" + "clave-aleatoria-segura-" * 2),
            "A9!" + "clave-aleatoria-segura-" * 2,
        )

        env = os.environ.copy()
        env.update(
            FLASK_ENV="production",
            DATABASE_URL="sqlite:///produccion.db",
            SECRET_KEY="X" * 48,
        )
        result = subprocess.run(
            [sys.executable, "-c", "import config"],
            cwd=os.path.dirname(os.path.dirname(__file__)),
            env=env,
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("producción requiere PostgreSQL", result.stderr)
        self.assertIn("ERROR CRÍTICO", CRITICAL_SECRET_ERROR)

        missing_secret_env = env.copy()
        missing_secret_env.pop("SECRET_KEY", None)
        missing_secret_env.update(
            FLASK_ENV="development",
            DATABASE_URL="postgresql://usuario:contrasena@localhost/bch_control",
        )
        wsgi_result = subprocess.run(
            [sys.executable, "-c", "import wsgi_pythonanywhere"],
            cwd=os.path.dirname(os.path.dirname(__file__)),
            env=missing_secret_env,
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertNotEqual(wsgi_result.returncode, 0)
        self.assertIn(CRITICAL_SECRET_ERROR, wsgi_result.stderr)

        env.update(
            DATABASE_URL="postgresql://usuario:contrasena@localhost/bch_control",
            SECRET_KEY="Y" * 48,
        )
        secure = subprocess.run(
            [
                sys.executable,
                "-c",
                (
                    "import config; c=config.get_config(); "
                    "print(c.SESSION_COOKIE_SECURE, c.REMEMBER_COOKIE_SECURE, "
                    "c.PREFERRED_URL_SCHEME, c.IS_PRODUCTION)"
                ),
            ],
            cwd=os.path.dirname(os.path.dirname(__file__)),
            env=env,
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(secure.returncode, 0, secure.stderr)
        self.assertEqual(secure.stdout.strip(), "True True https True")


if __name__ == "__main__":
    unittest.main()
