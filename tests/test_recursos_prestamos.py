"""Regresiones del recurso semanal por entrega inicial de préstamos."""

import os
import unittest
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

os.environ["FLASK_ENV"] = "testing"
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["SECRET_KEY"] = "test-recursos-prestamos-secret"

from app import create_app  # noqa: E402
from compras_models import PaymentMethod, PurchaseOrder  # noqa: E402
from models import CentroCosto, Usuario, db, utc_now  # noqa: E402
from nominas_models import (  # noqa: E402
    AdditionalPayment,
    BudgetItem,
    Company,
    Contractor,
    Employee,
    Loan,
    LoanPayment,
    OfficeExpense,
    Payroll,
    PayrollLine,
    Subcontract,
    SubcontractPayment,
    WeeklyResourceAvailability,
)
from routes.nominas import build_weekly_closing_report  # noqa: E402
from services.weekly_resources import weekly_resource_breakdown  # noqa: E402


app = create_app()


class WeeklyLoanResourceTest(unittest.TestCase):
    """Cruza la fórmula única contra rutas, cierres y exportaciones."""

    WEEK_START = date(2026, 7, 20)
    WEEK_END = date(2026, 7, 24)

    def setUp(self):
        app.config.update(
            TESTING=True,
            WTF_CSRF_ENABLED=False,
            MFA_REQUIRED_FOR_ADMINS=False,
            COMPRAS_TODAY=self.WEEK_END,
            FASE5_TODAY=self.WEEK_END,
        )
        self.client = app.test_client()
        with app.app_context():
            db.drop_all()
            db.create_all()

            company = Company(
                codigo="BCH",
                nombre="Baja Custom Homes",
                activa=True,
            )
            project_one = CentroCosto(
                nombre="Casa Norte",
                codigo="OBRA-NORTE",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2026, 1, 1),
                presupuesto_total=500000,
                presupuesto_mano_obra=150000,
            )
            project_two = CentroCosto(
                nombre="Casa Sur",
                codigo="OBRA-SUR",
                tipo="obra",
                estado="activa",
                fecha_apertura=date(2026, 1, 1),
                presupuesto_total=500000,
                presupuesto_mano_obra=150000,
            )
            db.session.add_all([company, project_one, project_two])
            db.session.flush()

            labor_item = BudgetItem(
                project_id=project_one.id,
                codigo="MO-01",
                nombre="Mano de obra",
                categoria="MANO_OBRA",
                presupuesto=100000,
            )
            subcontract_item = BudgetItem(
                project_id=project_one.id,
                codigo="SC-01",
                nombre="Subcontratos",
                categoria="SUBCONTRATO",
                presupuesto=100000,
            )
            south_item = BudgetItem(
                project_id=project_two.id,
                codigo="MO-02",
                nombre="Mano de obra sur",
                categoria="MANO_OBRA",
                presupuesto=100000,
            )
            methods = [
                PaymentMethod(nombre="EFECTIVO", descripcion="Caja", activo=True),
                PaymentMethod(
                    nombre="TRANSFERENCIA",
                    descripcion="Banco",
                    activo=True,
                ),
                PaymentMethod(
                    nombre="TARJETA EMPRESARIAL",
                    descripcion="Método no válido para préstamos",
                    activo=True,
                ),
            ]
            db.session.add_all([labor_item, subcontract_item, south_item, *methods])
            db.session.flush()
            cash_method, transfer_method, card_method = methods

            admin = self._user(
                "Administradora General",
                "admin.recursos@example.com",
                "admin",
            )
            supervisor = self._user(
                "Supervisor Multiobra",
                "supervisor.recursos@example.com",
                "supervisor",
                project_one,
                [project_one, project_two],
            )
            ceo = self._user(
                "Dirección General",
                "ceo.recursos@example.com",
                "ceo",
            )
            db.session.flush()

            payroll_employee = self._employee(
                "TRABAJADOR CON RETENCIÓN",
                project_one,
                labor_item,
                company,
            )
            cash_employee = self._employee(
                "TRABAJADOR PRÉSTAMO EFECTIVO",
                project_one,
                labor_item,
                company,
            )
            transfer_employee = self._employee(
                "TRABAJADOR PRÉSTAMO TRANSFERENCIA",
                project_two,
                south_item,
                company,
            )
            default_employee = self._employee(
                "TRABAJADOR MÉTODO PREDETERMINADO",
                project_two,
                south_item,
                company,
            )
            db.session.flush()

            payroll = Payroll(
                project_id=project_one.id,
                semana_inicio=self.WEEK_START,
                semana_fin=self.WEEK_END,
                estado="aprobada",
                created_by_id=admin.id,
                closed_by_id=admin.id,
                closed_at=utc_now(),
            )
            line = PayrollLine(
                payroll=payroll,
                employee_id=payroll_employee.id,
                budget_item_id=labor_item.id,
                nombre_trabajador=payroll_employee.nombre_completo,
                puesto=payroll_employee.puesto,
                salario_semanal=3500,
                lunes=True,
                martes=True,
                miercoles=True,
                jueves=True,
                viernes=True,
                dias_trabajados=5,
                numero_faltas=0,
                sueldo_diario=700,
                descuento_faltas=0,
                monto_devengado=3500,
                pago_extra=0,
                descuento_infonavit=0,
                descuento_imss=0,
                descuento_prestamo=500,
                otro_descuento=0,
                vales_gasolina=0,
                pago_transferencia=2000,
                empresa_transferencia_id=company.id,
                pago_efectivo=1000,
                empresa_efectivo_id=company.id,
                neto_pagar=3000,
            )
            db.session.add(payroll)
            db.session.flush()

            old_loan = Loan(
                employee_id=payroll_employee.id,
                project_id=project_one.id,
                fecha_prestamo=self.WEEK_START - timedelta(days=7),
                monto=1000,
                tasa_interes=5,
                total_pagar=1050,
                retencion_semanal=500,
                metodo_entrega="EFECTIVO",
                payment_method_id=cash_method.id,
                company_id=company.id,
                concepto="Préstamo anterior con abono actual",
                estado="activo",
                solicitante_id=supervisor.id,
                aprobador_id=admin.id,
                fecha_aprobacion=utc_now(),
                created_by_id=supervisor.id,
            )
            cash_loan = Loan(
                employee_id=cash_employee.id,
                project_id=project_one.id,
                fecha_prestamo=self.WEEK_START + timedelta(days=2),
                monto=500,
                tasa_interes=5,
                total_pagar=525,
                retencion_semanal=100,
                metodo_entrega="EFECTIVO",
                payment_method_id=cash_method.id,
                company_id=company.id,
                concepto="Capital efectivo nuevo",
                estado="activo",
                solicitante_id=supervisor.id,
                aprobador_id=admin.id,
                fecha_aprobacion=utc_now(),
                created_by_id=supervisor.id,
            )
            transfer_loan = Loan(
                employee_id=transfer_employee.id,
                project_id=project_two.id,
                fecha_prestamo=self.WEEK_START + timedelta(days=3),
                monto=700,
                tasa_interes=5,
                total_pagar=735,
                retencion_semanal=100,
                metodo_entrega="TRANSFERENCIA",
                payment_method_id=transfer_method.id,
                company_id=company.id,
                concepto="Capital transferencia nuevo",
                estado="activo",
                solicitante_id=supervisor.id,
                aprobador_id=admin.id,
                fecha_aprobacion=utc_now(),
                created_by_id=supervisor.id,
            )
            pending_loan = Loan(
                employee_id=cash_employee.id,
                project_id=project_one.id,
                fecha_prestamo=self.WEEK_START + timedelta(days=1),
                monto=900,
                tasa_interes=5,
                total_pagar=945,
                retencion_semanal=100,
                metodo_entrega="EFECTIVO",
                payment_method_id=cash_method.id,
                company_id=company.id,
                concepto="Pendiente sin entrega",
                estado="pendiente",
                solicitante_id=supervisor.id,
                created_by_id=supervisor.id,
            )
            rejected_loan = Loan(
                employee_id=transfer_employee.id,
                project_id=project_two.id,
                fecha_prestamo=self.WEEK_START + timedelta(days=1),
                monto=800,
                tasa_interes=5,
                total_pagar=840,
                retencion_semanal=100,
                metodo_entrega="TRANSFERENCIA",
                payment_method_id=transfer_method.id,
                company_id=company.id,
                concepto="Rechazado sin entrega",
                estado="rechazado",
                solicitante_id=supervisor.id,
                aprobador_id=admin.id,
                fecha_aprobacion=utc_now(),
                motivo_rechazo="No autorizado",
                created_by_id=supervisor.id,
            )
            db.session.add_all(
                [
                    old_loan,
                    cash_loan,
                    transfer_loan,
                    pending_loan,
                    rejected_loan,
                ]
            )
            db.session.flush()
            db.session.add(
                LoanPayment(
                    loan_id=old_loan.id,
                    payroll_line_id=line.id,
                    monto=500,
                )
            )

            operation_order = PurchaseOrder(
                folio="OC-OP-RECURSO-001",
                project_id=project_one.id,
                company_id=company.id,
                buyer_id=admin.id,
                payment_method_id=cash_method.id,
                fecha_orden=self.WEEK_START,
                fecha_entrega_estimada=self.WEEK_END,
                fecha_limite=self.WEEK_END,
                tipo_oc="OPERACIONES",
                categoria_pago="OPERACIONES",
                estado="EMITIDA",
                modalidad_pago="PAGO_CONTRA_ENTREGA",
                condicion_saldo="CONTRA_ENTREGA_TOTAL",
                beneficiario_libre="Proveedor operativo",
                beneficiario_validado=True,
                created_by_id=admin.id,
                issued_by_id=admin.id,
                issued_at=utc_now(),
            )
            contractor = Contractor(
                nombre="Contratista Recurso",
                especialidad="Estructura",
                activo=True,
            )
            db.session.add_all([operation_order, contractor])
            db.session.flush()
            subcontract = Subcontract(
                project_id=project_one.id,
                budget_item_id=subcontract_item.id,
                contractor_id=contractor.id,
                especialidad="Estructura",
                presupuesto_sin_iva=10000,
                avance_fisico=Decimal("0.50"),
            )
            db.session.add(subcontract)
            db.session.flush()

            db.session.add_all(
                [
                    OfficeExpense(
                        fecha=self.WEEK_START,
                        project_id=project_one.id,
                        budget_item_id=labor_item.id,
                        proveedor="Caja chica",
                        concepto="Gasto operativo efectivo",
                        monto_capturado=100,
                        tipo_monto="SIN_IVA",
                        monto_sin_iva=100,
                        metodo_pago="EFECTIVO",
                        payment_method_id=cash_method.id,
                        company_id=company.id,
                        created_by_id=admin.id,
                    ),
                    AdditionalPayment(
                        fecha=self.WEEK_START + timedelta(days=1),
                        project_id=project_one.id,
                        budget_item_id=labor_item.id,
                        purchase_order_id=operation_order.id,
                        payment_method_id=cash_method.id,
                        beneficiario="Proveedor operativo",
                        concepto="Pago real de OC operativa",
                        monto_capturado=150,
                        tipo_monto="SIN_IVA",
                        monto_sin_iva=150,
                        metodo_pago="EFECTIVO",
                        company_id=company.id,
                        created_by_id=admin.id,
                    ),
                    AdditionalPayment(
                        fecha=self.WEEK_START + timedelta(days=2),
                        project_id=project_one.id,
                        budget_item_id=labor_item.id,
                        payment_method_id=transfer_method.id,
                        beneficiario="Apoyo de campo",
                        concepto="Pago adicional independiente",
                        monto_capturado=200,
                        tipo_monto="SIN_IVA",
                        monto_sin_iva=200,
                        metodo_pago="TRANSFERENCIA",
                        company_id=company.id,
                        created_by_id=admin.id,
                    ),
                    SubcontractPayment(
                        subcontract_id=subcontract.id,
                        fecha=self.WEEK_START + timedelta(days=3),
                        concepto="Estimación semanal",
                        monto_capturado=300,
                        tipo_monto="SIN_IVA",
                        monto_sin_iva=300,
                        metodo_pago="EFECTIVO",
                        payment_method_id=cash_method.id,
                        company_id=company.id,
                        created_by_id=admin.id,
                    ),
                    WeeklyResourceAvailability(
                        semana_inicio=self.WEEK_START,
                        metodo="EFECTIVO",
                        monto_disponible=10000,
                        updated_by_id=admin.id,
                    ),
                    WeeklyResourceAvailability(
                        semana_inicio=self.WEEK_START,
                        metodo="TRANSFERENCIA",
                        monto_disponible=10000,
                        updated_by_id=admin.id,
                    ),
                ]
            )
            db.session.commit()

            self.admin_email = admin.correo
            self.supervisor_email = supervisor.correo
            self.ceo_email = ceo.correo
            self.admin_id = admin.id
            self.project_one_id = project_one.id
            self.project_two_id = project_two.id
            self.payroll_id = payroll.id
            self.cash_employee_id = cash_employee.id
            self.default_employee_id = default_employee.id
            self.company_id = company.id
            self.cash_method_id = cash_method.id
            self.card_method_id = card_method.id

    @staticmethod
    def _user(name, email, role, primary=None, projects=None):
        user = Usuario(
            nombre_completo=name,
            correo=email,
            rol=role,
            centro_costo_id=(
                primary.id
                if primary is not None and role in {"capturista", "supervisor"}
                else None
            ),
            activo=True,
        )
        user.set_password("Password123!")
        user.asignar_permisos_predeterminados()
        if projects:
            user.projects = projects
        db.session.add(user)
        return user

    @staticmethod
    def _employee(name, project, item, company):
        employee = Employee(
            nombre_completo=name,
            fecha_ingreso=date(2025, 1, 1),
            activo=True,
            puesto="OFICIAL",
            project_id=project.id,
            budget_item_id=item.id,
            salario_semanal=3500,
            registrado_imss=False,
            descuento_infonavit=0,
            transferencia_predeterminada=2000,
            empresa_transferencia_id=company.id,
            empresa_efectivo_id=company.id,
        )
        db.session.add(employee)
        return employee

    def login(self, email):
        self.client.post("/logout", follow_redirects=True)
        response = self.client.post(
            "/login",
            data={"correo": email, "contrasena": "Password123!"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        return response

    @staticmethod
    def assert_money(testcase, actual, expected):
        testcase.assertEqual(Decimal(str(actual)), Decimal(str(expected)))

    def test_formula_unica_excluye_abonos_y_conserva_obra_historica(self):
        with app.app_context():
            resource = weekly_resource_breakdown(self.WEEK_START)
            cash = resource["methods"]["EFECTIVO"]
            bank = resource["methods"]["TRANSFERENCIA"]

            self.assert_money(self, cash["nomina"], "1000.00")
            self.assert_money(self, cash["prestamos"], "500.00")
            self.assert_money(self, cash["gastos_operativos"], "250.00")
            self.assert_money(self, cash["pagos_adicionales"], "0.00")
            self.assert_money(self, cash["subcontratos"], "300.00")
            self.assert_money(self, cash["requerido"], "2050.00")
            self.assert_money(self, cash["diferencia"], "7950.00")

            self.assert_money(self, bank["nomina"], "2000.00")
            self.assert_money(self, bank["prestamos"], "700.00")
            self.assert_money(self, bank["gastos_operativos"], "0.00")
            self.assert_money(self, bank["pagos_adicionales"], "200.00")
            self.assert_money(self, bank["subcontratos"], "0.00")
            self.assert_money(self, bank["requerido"], "2900.00")
            self.assert_money(self, bank["diferencia"], "7100.00")
            self.assert_money(self, resource["requerido_total"], "4950.00")

            # El abono de 500 ya redujo la nómina a 3,000. No se agrega como
            # una sexta categoría ni se usa el total con 5% de interés.
            self.assertEqual(LoanPayment.query.count(), 1)
            self.assertEqual(Loan.query.filter_by(estado="pendiente").count(), 1)
            self.assertEqual(Loan.query.filter_by(estado="rechazado").count(), 1)

            north = weekly_resource_breakdown(
                self.WEEK_START, [self.project_one_id]
            )
            south = weekly_resource_breakdown(
                self.WEEK_START, [self.project_two_id]
            )
            self.assert_money(self, north["requerido_total"], "4250.00")
            self.assert_money(self, south["requerido_total"], "700.00")

            report = build_weekly_closing_report(self.WEEK_START)
            self.assert_money(self, report["totals"]["nomina_neto"], "3000.00")
            self.assert_money(
                self, report["totals"]["prestamos_entregados"], "1200.00"
            )
            self.assert_money(
                self, report["totals"]["prestamos_efectivo"], "500.00"
            )
            self.assert_money(
                self, report["totals"]["prestamos_transferencia"], "700.00"
            )
            self.assert_money(
                self, report["totals"]["retencion_prestamos"], "500.00"
            )
            self.assert_money(self, report["totals"]["recurso_total"], "4950.00")
            self.assert_money(
                self, report["resource_summary"]["requerido_total"], "4950.00"
            )
            loan_funding = [
                row
                for row in report["funding_rows"]
                if row["categoria"] == "PRÉSTAMOS"
            ]
            self.assertEqual(
                {(row["metodo"], row["monto"]) for row in loan_funding},
                {
                    ("EFECTIVO", Decimal("500.00")),
                    ("TRANSFERENCIA", Decimal("700.00")),
                },
            )

            # Una reasignación posterior no mueve la salida histórica.
            cash_employee = db.session.get(Employee, self.cash_employee_id)
            cash_employee.project_id = self.project_two_id
            db.session.commit()
            north_after_move = weekly_resource_breakdown(
                self.WEEK_START, [self.project_one_id]
            )
            self.assert_money(
                self,
                north_after_move["methods"]["EFECTIVO"]["prestamos"],
                "500.00",
            )
            next_week = weekly_resource_breakdown(
                self.WEEK_START + timedelta(days=7)
            )
            self.assert_money(self, next_week["methods"]["EFECTIVO"]["prestamos"], "0.00")
            self.assert_money(
                self,
                next_week["methods"]["TRANSFERENCIA"]["prestamos"],
                "0.00",
            )

    def test_cierre_y_reapertura_no_cambian_recurso_ni_crean_adicionales(self):
        self.login(self.admin_email)
        with app.app_context():
            before = weekly_resource_breakdown(self.WEEK_START)["requerido_total"]
            self.assertEqual(LoanPayment.query.count(), 1)
            self.assertEqual(AdditionalPayment.query.count(), 2)

        reopened = self.client.post(
            f"/nominas/{self.payroll_id}/reabrir",
            data={"motivo": "Comprobar invariancia del recurso"},
            follow_redirects=True,
        )
        self.assertIn(
            "Las retenciones de préstamos quedaron revertidas",
            reopened.get_data(as_text=True),
        )
        with app.app_context():
            self.assertEqual(LoanPayment.query.count(), 0)
            self.assertEqual(AdditionalPayment.query.count(), 2)
            after_reopen = weekly_resource_breakdown(self.WEEK_START)[
                "requerido_total"
            ]
            self.assertEqual(before, after_reopen)

        approved = self.client.post(
            f"/nominas/{self.payroll_id}/aprobar",
            follow_redirects=True,
        )
        self.assertIn("Nómina aprobada", approved.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(LoanPayment.query.count(), 1)
            self.assertEqual(AdditionalPayment.query.count(), 2)
            after_close = weekly_resource_breakdown(self.WEEK_START)[
                "requerido_total"
            ]
            self.assertEqual(before, after_close)

    def test_dashboards_panel_operativo_y_multiobra_concilian(self):
        self.login(self.admin_email)
        admin_panel = self.client.get("/panel").get_data(as_text=True)
        self.assertIn("$2,050.00", admin_panel)
        self.assertIn("$2,900.00", admin_panel)
        self.assertIn("$4,950.00", admin_panel)

        operations = self.client.get(
            "/compras/reportes/nominas-operaciones"
            f"?fecha_desde={self.WEEK_START.isoformat()}"
            f"&fecha_hasta={self.WEEK_END.isoformat()}"
        ).get_data(as_text=True)
        self.assertIn("$2,050.00", operations)
        self.assertIn("$2,900.00", operations)
        self.assertIn("$1,200.00", operations)
        self.assertIn("$4,950.00", operations)

        supplier_report = self.client.get(
            "/compras/reportes/pagos-proveedores"
            f"?fecha_desde={self.WEEK_START.isoformat()}"
            f"&fecha_hasta={self.WEEK_END.isoformat()}"
        ).get_data(as_text=True)
        self.assertNotIn("Capital efectivo nuevo", supplier_report)
        self.assertNotIn("Capital transferencia nuevo", supplier_report)
        self.assertNotIn("TRABAJADOR CON RETENCIÓN", supplier_report)

        self.login(self.supervisor_email)
        north_dashboard = self.client.get(
            "/campo/dashboard-supervisor"
        ).get_data(as_text=True)
        self.assertIn("OBRA-NORTE", north_dashboard)
        self.assertIn("$4,250.00", north_dashboard)
        self.assertNotIn("$4,950.00", north_dashboard)

        switched = self.client.post(
            "/obra-activa",
            data={
                "project_id": str(self.project_two_id),
                "return_to": "/campo/dashboard-supervisor",
            },
            follow_redirects=True,
        )
        south_dashboard = switched.get_data(as_text=True)
        self.assertIn("Obra activa: OBRA-SUR", south_dashboard)
        self.assertIn("$700.00", south_dashboard)
        self.assertNotIn("$4,250.00", south_dashboard)

        self.login(self.ceo_email)
        executive = self.client.get("/direccion/").get_data(as_text=True)
        self.assertIn("$2,050.00", executive)
        self.assertIn("$2,900.00", executive)
        self.assertIn("$4,950.00", executive)
        self.assertIn("Los abonos de préstamos están excluidos", executive)

    def test_excel_repite_la_formula_por_metodo_y_obra(self):
        self.login(self.admin_email)
        export = self.client.get(
            "/reportes/exportar.xlsx"
            f"?desde={self.WEEK_START.isoformat()}"
            f"&hasta={self.WEEK_END.isoformat()}"
        )
        self.assertEqual(export.status_code, 200)
        workbook = load_workbook(
            BytesIO(export.data),
            read_only=True,
            data_only=True,
        )
        self.assertIn("Recurso semanal", workbook.sheetnames)
        sheet = workbook["Recurso semanal"]
        headers = [cell.value for cell in sheet[1]]
        values = {
            header: sheet.cell(2, index + 1).value
            for index, header in enumerate(headers)
        }
        self.assertEqual(values["Préstamos nuevos efectivo"], 500)
        self.assertEqual(values["Efectivo requerido"], 2050)
        self.assertEqual(values["Préstamos nuevos transferencia"], 700)
        self.assertEqual(values["Transferencias requeridas"], 2900)
        self.assertEqual(values["Recurso total requerido"], 4950)

        project_export = self.client.get(
            "/reportes/exportar.xlsx"
            f"?desde={self.WEEK_START.isoformat()}"
            f"&hasta={self.WEEK_END.isoformat()}"
            f"&project_id={self.project_one_id}"
        )
        project_book = load_workbook(
            BytesIO(project_export.data),
            read_only=True,
            data_only=True,
        )
        project_sheet = project_book["Recurso semanal"]
        project_headers = [cell.value for cell in project_sheet[1]]
        project_values = {
            header: project_sheet.cell(2, index + 1).value
            for index, header in enumerate(project_headers)
        }
        self.assertEqual(project_values["Préstamos nuevos efectivo"], 500)
        self.assertEqual(project_values["Préstamos nuevos transferencia"], 0)
        self.assertEqual(project_values["Recurso total requerido"], 4250)

        operations_export = self.client.get(
            "/compras/reportes/nominas-operaciones"
            f"?fecha_desde={self.WEEK_START.isoformat()}"
            f"&fecha_hasta={self.WEEK_END.isoformat()}&export=xlsx"
        )
        operations_book = load_workbook(
            BytesIO(operations_export.data),
            read_only=True,
            data_only=True,
        )
        operation_rows = list(
            operations_book["Recurso semanal"].iter_rows(values_only=True)
        )
        operation_headers = operation_rows[0]
        by_scope = {
            (row[1], row[2]): dict(zip(operation_headers, row))
            for row in operation_rows[1:]
        }
        north_cash = next(
            value
            for (project, method), value in by_scope.items()
            if project.startswith("OBRA-NORTE") and method == "EFECTIVO"
        )
        south_bank = next(
            value
            for (project, method), value in by_scope.items()
            if project.startswith("OBRA-SUR") and method == "TRANSFERENCIA"
        )
        self.assertEqual(north_cash["Préstamos nuevos"], 500)
        self.assertEqual(north_cash["Requerido"], 2050)
        self.assertEqual(south_bank["Préstamos nuevos"], 700)
        self.assertEqual(south_bank["Requerido"], 700)

    def test_metodo_de_entrega_es_obligatorio_con_fallback_seguro(self):
        self.login(self.admin_email)
        form = self.client.get("/prestamos/nuevo").get_data(as_text=True)
        self.assertIn(">Efectivo<", form)
        self.assertIn(">Transferencia<", form)
        self.assertNotIn("TARJETA EMPRESARIAL", form)

        created = self.client.post(
            "/prestamos/nuevo",
            data={
                "employee_id": str(self.default_employee_id),
                "fecha_prestamo": self.WEEK_END.isoformat(),
                "monto": "100",
                "retencion_semanal": "50",
                # Compatibilidad solicitada: una integración anterior sin
                # campo explícito queda en EFECTIVO.
                "company_id": str(self.company_id),
                "concepto": "Método por defecto",
            },
            follow_redirects=True,
        )
        self.assertIn("Solicitud enviada", created.get_data(as_text=True))
        with app.app_context():
            default_loan = Loan.query.filter_by(
                concepto="Método por defecto"
            ).one()
            self.assertEqual(default_loan.metodo_entrega, "EFECTIVO")
            self.assertEqual(default_loan.empresa_entrega_id, self.company_id)
            self.assertEqual(default_loan.project_id, self.project_two_id)
            loan_count = Loan.query.count()

        invalid = self.client.post(
            "/prestamos/nuevo",
            data={
                "employee_id": str(self.default_employee_id),
                "fecha_prestamo": self.WEEK_END.isoformat(),
                "monto": "100",
                "retencion_semanal": "50",
                "payment_method_id": str(self.card_method_id),
                "company_id": str(self.company_id),
                "concepto": "Método inválido",
            },
            follow_redirects=True,
        )
        self.assertIn(
            "debe ser Efectivo o Transferencia",
            invalid.get_data(as_text=True),
        )
        with app.app_context():
            self.assertEqual(Loan.query.count(), loan_count)
            self.assertIsNone(
                Loan.query.filter_by(concepto="Método inválido").first()
            )


if __name__ == "__main__":
    unittest.main()
