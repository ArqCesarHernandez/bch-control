"""Prueba integral del flujo más delicado: préstamo, nómina, cierre y Excel."""

import os
import re
import unittest
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

os.environ["FLASK_ENV"] = "testing"
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["SECRET_KEY"] = "test-secret-key"

from app import create_app  # noqa: E402
from models import CentroCosto as Project, Usuario as User, db  # noqa: E402
from compras_models import (  # noqa: E402
    BudgetExplosionItem,
    Supplier,
    SupplyItem,
)
from nominas_models import (  # noqa: E402
    AdditionalPayment,
    BudgetItem,
    Company,
    Contractor,
    Employee,
    Loan,
    LoanPayment,
    OfficeExpense,
    Payroll,
    PayrollLine,
    Subcontract,
    SubcontractPayment,
    WeeklyResourceAvailability,
)
from routes.nominas import (  # noqa: E402
    IMPORT_DEFINITIONS,
    build_weekly_closing_report,
    item_consumption,
    project_consumption,
    weekly_resource_breakdown,
)


app = create_app()


class PayrollFlowTest(unittest.TestCase):
    def setUp(self):
        app.config.update(TESTING=True)
        self.client = app.test_client()
        self.usernames = {"grecia.admin": "grecia@example.com"}
        with app.app_context():
            db.drop_all()
            db.create_all()
            for code in ("BCH", "RGOVC", "CA", "CN"):
                db.session.add(Company(codigo=code, nombre=code))
            db.session.commit()

    def post(self, url, data):
        payload = dict(data)
        if url == "/registro":
            username = payload.get("username", "")
            email = payload.get("email", "").lower()
            self.usernames[username] = email
            role = "admin" if payload.get("role") == "administrador" else "capturista"
            project_ids = str(payload.get("project_ids", "")).split(",")
            center_id = next((value for value in project_ids if value.isdigit()), "0")
            return self.client.post(
                "/admin/usuarios/nuevo",
                data={
                    "nombre_completo": payload.get("nombre_completo", ""),
                    "correo": email,
                    "contrasena": payload.get("password", ""),
                    "rol": role,
                    "centro_costo_id": center_id if role == "capturista" else "0",
                },
                follow_redirects=True,
            )
        return self.client.post(url, data=payload, follow_redirects=True)

    def register_admin(self):
        self.client.get("/register")
        response = self.post(
            "/register",
            {
                "nombre_completo": "Grecia Martinez",
                "correo": "grecia@example.com",
                "contrasena": "Password123!",
                "confirmar_contrasena": "Password123!",
            },
        )
        if response.status_code == 200:
            return self.login("grecia.admin", "Password123!")
        return response

    def login(self, username, password):
        self.client.get("/login")
        return self.post(
            "/login",
            {
                "correo": self.usernames.get(username, username),
                "contrasena": password,
            },
        )

    def logout(self):
        return self.post("/logout", {})

    @staticmethod
    def excel_file(kind, rows):
        """Crea en memoria un archivo igual a la plantilla de importación."""

        workbook = __import__("openpyxl").Workbook()
        worksheet = workbook.active
        headers = [column[0] for column in IMPORT_DEFINITIONS[kind]["columns"]]
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def test_complete_weekly_payroll_flow(self):
        self.assertEqual(self.client.get("/register").status_code, 200)
        response = self.register_admin()
        self.assertEqual(response.status_code, 200)

        self.post(
            "/obras/nueva",
            {
                "nombre": "Casa Prueba",
                "codigo": "LTEST",
                "tipo": "OBRA",
                "presupuesto_total": "1000000",
                "presupuesto_mano_obra": "300000",
                "activa": "on",
            },
        )
        with app.app_context():
            project_id = Project.query.filter_by(codigo="LTEST").one().id

        self.post(
            f"/obras/{project_id}/partidas/nueva",
            {
                "codigo": "MO-01",
                "nombre": "Albañilería",
                "categoria": "MANO_OBRA",
                "presupuesto": "300000",
                "parent_id": "",
                "activa": "on",
            },
        )
        with app.app_context():
            item_id = BudgetItem.query.filter_by(project_id=project_id).one().id
            bch_id = Company.query.filter_by(codigo="BCH").one().id
            rgovc_id = Company.query.filter_by(codigo="RGOVC").one().id

        self.post(
            "/trabajadores/nuevo",
            {
                "nombre_completo": "Juan Perez Lopez",
                "fecha_ingreso": "2025-07-01",
                "fecha_baja": "",
                "activo": "on",
                "puesto": "Oficial",
                "cuadrilla": "Alex",
                "supervisor": "Amir",
                "empresa_operativa": "RGOVC",
                "project_id": str(project_id),
                "budget_item_id": str(item_id),
                "salario_semanal": "5000",
                "registrado_imss": "on",
                "nss": "12345678901",
                "empresa_imss_id": str(rgovc_id),
                "descuento_infonavit": "100",
                "imss_tipo": "PORCENTAJE",
                "descuento_imss": "5",
                "transferencia_predeterminada": "2500",
                "empresa_transferencia_id": str(rgovc_id),
                "empresa_efectivo_id": str(bch_id),
            },
        )
        with app.app_context():
            employee = Employee.query.one()
            employee_id = employee.id
            self.assertIsNone(employee.budget_item_id)

        # Un capturista puede operar su obra, pero ni siquiera enviando campos
        # manipulados puede cambiar la configuración IMSS del administrador.
        self.post(
            "/registro",
            {
                "nombre_completo": "Capturista Prueba",
                "username": "captura.uno",
                "email": "captura@example.com",
                "password": "Password123!",
                "password_confirm": "Password123!",
                "role": "capturista",
                "project_ids": str(project_id),
            },
        )
        self.logout()
        self.login("captura.uno", "Password123!")
        capture_dashboard = self.client.get("/panel").get_data(as_text=True)
        self.assertIn("Panel de captura", capture_dashboard)
        self.assertIn("CASA PRUEBA", capture_dashboard)
        self.post(
            f"/trabajadores/{employee_id}/editar",
            {
                "nombre_completo": "Juan Perez Lopez",
                "fecha_ingreso": "2025-07-01",
                "fecha_baja": "",
                "activo": "on",
                "puesto": "Oficial",
                "cuadrilla": "Alex",
                "supervisor": "Amir",
                "empresa_operativa": "RGOVC",
                "project_id": str(project_id),
                "budget_item_id": str(item_id),
                "salario_semanal": "5000",
                "registrado_imss": "on",
                "imss_tipo": "FIJO",
                "descuento_imss": "999",
                "descuento_infonavit": "100",
                "transferencia_predeterminada": "2500",
                "empresa_transferencia_id": str(rgovc_id),
                "empresa_efectivo_id": str(bch_id),
            },
        )
        with app.app_context():
            protected_employee = db.session.get(Employee, employee_id)
            self.assertEqual(protected_employee.imss_tipo, "PORCENTAJE")
            self.assertEqual(float(protected_employee.descuento_imss), 5.0)
            self.assertIsNone(protected_employee.budget_item_id)

        # Primero se crea la nómina y después el préstamo. Este es el orden real
        # que provocaba que la pantalla conservara un neto anterior al préstamo.
        self.post(
            "/nominas/nueva", {"project_id": str(project_id), "semana_inicio": "2026-07-13"}
        )
        with app.app_context():
            payroll = Payroll.query.filter_by(semana_inicio=date(2026, 7, 13)).one()
            payroll_id = payroll.id
            line_id = payroll.lines[0].id
            self.assertEqual(float(payroll.lines[0].descuento_prestamo), 0.0)
        payroll_capture = self.client.get(
            f"/nominas/{payroll_id}"
        ).get_data(as_text=True)
        self.assertIn(
            f'name="line_{line_id}_partida_id"',
            payroll_capture,
        )
        self.assertIn(f'name="line_{line_id}_subpartida_id"', payroll_capture)
        self.assertIn("MO-01", payroll_capture)

        loan_response = self.post(
            "/prestamos/nuevo",
            {
                "employee_id": str(employee_id),
                "fecha_prestamo": "2026-07-10",
                "monto": "2000",
                "retencion_semanal": "500",
                "metodo_entrega": "EFECTIVO",
                "company_id": str(bch_id),
                "concepto": "Prueba",
            },
        )
        self.assertIn(
            "Solicitud enviada a los administradores.",
            loan_response.get_data(as_text=True),
        )
        with app.app_context():
            loan = Loan.query.one()
            loan_id = loan.id
            self.assertEqual(loan.estado, "pendiente")
            self.assertEqual(float(loan.total_pagar), 2100.0)
            refreshed_line = db.session.get(PayrollLine, line_id)
            self.assertEqual(float(refreshed_line.descuento_prestamo), 0.0)

        self.logout()
        self.login("grecia.admin", "Password123!")
        self.post(f"/prestamos/{loan_id}/aprobar", {})
        self.logout()
        self.login("captura.uno", "Password123!")
        with app.app_context():
            refreshed_line = db.session.get(PayrollLine, line_id)
            self.assertEqual(float(refreshed_line.descuento_prestamo), 500.0)
            self.assertEqual(float(refreshed_line.neto_pagar), 4400.0)

        save_response = self.post(
            f"/nominas/{payroll_id}/guardar",
            {
                f"line_{line_id}_partida_id": str(item_id),
                f"line_{line_id}_subpartida_id": "",
                f"line_{line_id}_puesto": "OFICIAL",
                f"line_{line_id}_cuadrilla": "ALEX",
                f"line_{line_id}_supervisor": "AMIR",
                f"line_{line_id}_empresa_operativa": "RGOVC",
                # Importe con centavos para verificar que navegador y servidor
                # redondeen el sueldo diario de la misma forma.
                f"line_{line_id}_salario_semanal": "5000.03",
                f"line_{line_id}_lunes": "on",
                f"line_{line_id}_martes": "on",
                f"line_{line_id}_miercoles": "on",
                f"line_{line_id}_jueves": "on",
                f"line_{line_id}_pago_extra": "200",
                f"line_{line_id}_descuento_infonavit": "100",
                f"line_{line_id}_otro_descuento": "0",
                f"line_{line_id}_vales_gasolina": "1",
                # La transferencia es exactamente igual al neto recalculado.
                f"line_{line_id}_pago_transferencia": "3600.04",
                f"line_{line_id}_empresa_transferencia_id": str(rgovc_id),
                f"line_{line_id}_empresa_efectivo_id": str(bch_id),
            },
        )
        page = save_response.get_data(as_text=True)
        self.assertIn("Nómina guardada y cálculos actualizados.", page)
        self.assertNotIn("supera su neto", page)
        with app.app_context():
            line = db.session.get(PayrollLine, line_id)
            self.assertEqual(line.partida_id, item_id)
            self.assertEqual(line.partida.id, item_id)
            self.assertEqual(float(line.numero_faltas), 1.0)
            self.assertEqual(float(line.descuento_prestamo), 500.0)
            self.assertEqual(float(line.descuento_imss), 200.0)
            self.assertEqual(float(line.neto_pagar), 3600.04)
            self.assertEqual(float(line.pago_transferencia), 3600.04)
            self.assertEqual(float(line.pago_efectivo), 0.0)

        # El administrador puede presentar el consolidado al CEO antes de cerrar.
        self.logout()
        self.login("grecia.admin", "Password123!")
        preclose = self.client.get("/reportes/cierre-semanal?semana=2026-07-13")
        preclose_text = preclose.get_data(as_text=True)
        self.assertEqual(preclose.status_code, 200)
        self.assertIn("Reporte conjunto previo al cierre", preclose_text)
        self.assertIn("CASA PRUEBA", preclose_text)
        self.assertIn("$3,600.04", preclose_text)
        self.assertIn("RETENCIÓN PREVISTA", preclose_text)
        with app.app_context():
            report = build_weekly_closing_report(date(2026, 7, 13))
            self.assertEqual(float(report["totals"]["nomina_neto"]), 3600.04)
            self.assertEqual(float(report["totals"]["retencion_prestamos"]), 500.0)
            self.assertEqual(report["totals"]["borradores"], 1)

        self.post(f"/nominas/{payroll_id}/cerrar", {})
        with app.app_context():
            self.assertEqual(db.session.get(Payroll, payroll_id).estado, "aprobada")
            self.assertEqual(LoanPayment.query.count(), 1)
            next_payroll = Payroll.query.filter_by(semana_inicio=date(2026, 7, 20)).one()
            self.assertEqual(next_payroll.lines[0].empresa_operativa, "RGOVC")
            self.assertEqual(next_payroll.lines[0].partida_id, item_id)
            self.assertIsNone(next_payroll.lines[0].subpartida_id)
            closed_payroll = db.session.get(Payroll, payroll_id)
            self.assertEqual(float(closed_payroll.total_costo_mano_obra), 4400.04)
            self.assertEqual(
                float(project_consumption(closed_payroll.project)["mano_obra"]), 4400.04
            )

            # Regresión principal: el abono registrado en la semana anterior no
            # vuelve a sumarse al recurso de la semana nueva. Solo se informa la
            # retención elegible actual, que ya está restada del neto.
            next_report = build_weekly_closing_report(date(2026, 7, 20))
            next_breakdown = weekly_resource_breakdown(date(2026, 7, 20))
            self.assertEqual(float(next_report["totals"]["retencion_prestamos"]), 500.0)
            self.assertEqual(
                next_report["totals"]["recurso_total"],
                next_report["totals"]["nomina_neto"],
            )
            self.assertEqual(
                next_report["totals"]["recurso_total"],
                next_breakdown["requerido_total"],
            )
            self.assertEqual(LoanPayment.query.count(), 1)

        self.logout()
        self.login("grecia.admin", "Password123!")
        export = self.client.get("/reportes/exportar.xlsx?estado=CERRADA")
        self.assertEqual(export.status_code, 200)
        self.assertTrue(export.data.startswith(b"PK"))
        workbook = load_workbook(BytesIO(export.data), read_only=True, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Nóminas",
                "Resumen por obra",
                "Recurso semanal",
                "Presupuesto por partida",
                "Control presupuestal",
                "Préstamos",
                "Subcontratos",
            ],
        )
        payroll_sheet = workbook["Nóminas"]
        headers = [cell.value for cell in next(payroll_sheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(
            payroll_sheet.cell(
                row=2,
                column=headers.index("Partida") + 1,
            ).value,
            "MO-01 · ALBAÑILERÍA",
        )
        self.assertIn("Subpartida", headers)
        self.assertIsNone(
            payroll_sheet.cell(
                row=2,
                column=headers.index("Subpartida") + 1,
            ).value
        )
        self.assertIn("Empresa operativa", headers)
        self.assertEqual(payroll_sheet.cell(row=2, column=headers.index("Empresa operativa") + 1).value, "RGOVC")
        imss_header = "Costo IMSS patronal (no se descuenta)"
        self.assertEqual(payroll_sheet.cell(row=2, column=headers.index(imss_header) + 1).value, 200)

    def test_weekly_partida_subpartida_is_required_preloaded_and_independent(self):
        """La asignación vive en la línea y concilia filtros, Excel y presupuesto."""

        self.register_admin()
        self.post(
            "/obras/nueva",
            {
                "nombre": "Obra Semanal",
                "codigo": "SEM01",
                "tipo": "OBRA",
                "presupuesto_total": "500000",
                "presupuesto_mano_obra": "200000",
                "activa": "on",
            },
        )
        self.post(
            "/obras/nueva",
            {
                "nombre": "Obra Ajena",
                "codigo": "AJN01",
                "tipo": "OBRA",
                "presupuesto_total": "500000",
                "presupuesto_mano_obra": "200000",
                "activa": "on",
            },
        )
        with app.app_context():
            project_id = Project.query.filter_by(codigo="SEM01").one().id
            foreign_project_id = Project.query.filter_by(codigo="AJN01").one().id

        def create_item(project, code, name, parent=""):
            self.post(
                f"/obras/{project}/partidas/nueva",
                {
                    "codigo": code,
                    "nombre": name,
                    "categoria": "MANO_OBRA",
                    "presupuesto": "100000",
                    "parent_id": str(parent),
                    "activa": "on",
                },
            )
            with app.app_context():
                return BudgetItem.query.filter_by(
                    project_id=project, codigo=code
                ).one().id

        partida_a_id = create_item(project_id, "MO-A", "Estructura")
        subpartida_a_id = create_item(
            project_id, "MO-A.1", "Cimentación", partida_a_id
        )
        partida_b_id = create_item(project_id, "MO-B", "Acabados")
        foreign_partida_id = create_item(
            foreign_project_id, "MO-AJENA", "No disponible"
        )

        with app.app_context():
            company_id = Company.query.filter_by(codigo="BCH").one().id
        employee_form = self.client.get("/trabajadores/nuevo").get_data(
            as_text=True
        )
        self.assertNotIn('name="budget_item_id"', employee_form)
        self.assertNotIn('name="cuadrilla"', employee_form)
        self.assertNotIn('name="transferencia_predeterminada"', employee_form)
        self.post(
            "/trabajadores/nuevo",
            {
                "nombre_completo": "Persona Semanal Prueba",
                "fecha_ingreso": "2026-07-01",
                "activo": "on",
                "puesto": "Oficial",
                "project_id": str(project_id),
                "salario_semanal": "5000",
                "descuento_infonavit": "0",
                "descuento_imss": "0",
                "transferencia_predeterminada": "0",
                "empresa_efectivo_id": str(company_id),
            },
        )
        with app.app_context():
            employee = Employee.query.one()
            self.assertIsNone(employee.budget_item_id)

        self.post(
            "/nominas/nueva",
            {"project_id": str(project_id), "semana_inicio": "2026-07-13"},
        )
        with app.app_context():
            first_payroll = Payroll.query.filter_by(
                project_id=project_id, semana_inicio=date(2026, 7, 13)
            ).one()
            first_payroll_id = first_payroll.id
            first_line_id = first_payroll.lines[0].id
            self.assertIsNone(first_payroll.lines[0].partida_id)
            self.assertIsNone(first_payroll.lines[0].subpartida_id)

        detail = self.client.get(f"/nominas/{first_payroll_id}").get_data(
            as_text=True
        )
        self.assertIn("MO-A", detail)
        self.assertIn("MO-A.1", detail)
        self.assertIn("MO-B", detail)
        self.assertNotIn("MO-AJENA", detail)

        def payroll_payload(line_id, partida="", subpartida=""):
            return {
                f"line_{line_id}_partida_id": str(partida),
                f"line_{line_id}_subpartida_id": str(subpartida),
                f"line_{line_id}_puesto": "OFICIAL",
                f"line_{line_id}_salario_semanal": "5000",
                f"line_{line_id}_lunes": "on",
                f"line_{line_id}_martes": "on",
                f"line_{line_id}_miercoles": "on",
                f"line_{line_id}_jueves": "on",
                f"line_{line_id}_viernes": "on",
                f"line_{line_id}_pago_extra": "0",
                f"line_{line_id}_descuento_infonavit": "0",
                f"line_{line_id}_otro_descuento": "0",
                f"line_{line_id}_vales_gasolina": "0",
                f"line_{line_id}_pago_transferencia": "0",
                f"line_{line_id}_empresa_efectivo_id": str(company_id),
            }

        required_message = (
            "Debe asignar una partida a cada trabajador antes de guardar."
        )
        missing = self.post(
            f"/nominas/{first_payroll_id}/guardar",
            payroll_payload(first_line_id),
        )
        self.assertIn(required_message, missing.get_data(as_text=True))
        missing_child = self.post(
            f"/nominas/{first_payroll_id}/guardar",
            payroll_payload(first_line_id, partida_a_id),
        )
        self.assertIn(required_message, missing_child.get_data(as_text=True))
        foreign_item = self.post(
            f"/nominas/{first_payroll_id}/guardar",
            payroll_payload(first_line_id, foreign_partida_id),
        )
        self.assertIn(
            "Selecciona una partida activa de la obra.",
            foreign_item.get_data(as_text=True),
        )
        with app.app_context():
            line = db.session.get(PayrollLine, first_line_id)
            self.assertIsNone(line.partida_id)
            self.assertIsNone(line.subpartida_id)
        blocked_close = self.post(f"/nominas/{first_payroll_id}/cerrar", {})
        self.assertIn(
            required_message,
            blocked_close.get_data(as_text=True),
        )
        with app.app_context():
            self.assertEqual(
                db.session.get(Payroll, first_payroll_id).estado, "borrador"
            )
            self.assertIsNone(
                Payroll.query.filter_by(
                    project_id=project_id,
                    semana_inicio=date(2026, 7, 20),
                ).first()
            )

        saved = self.post(
            f"/nominas/{first_payroll_id}/guardar",
            payroll_payload(first_line_id, partida_a_id, subpartida_a_id),
        )
        self.assertIn(
            "Nómina guardada y cálculos actualizados.",
            saved.get_data(as_text=True),
        )
        self.post(f"/nominas/{first_payroll_id}/cerrar", {})
        with app.app_context():
            first_line = db.session.get(PayrollLine, first_line_id)
            self.assertEqual(first_line.partida_id, partida_a_id)
            self.assertEqual(first_line.subpartida_id, subpartida_a_id)
            self.assertEqual(first_line.budget_item_id, subpartida_a_id)
            second_payroll = Payroll.query.filter_by(
                project_id=project_id, semana_inicio=date(2026, 7, 20)
            ).one()
            second_payroll_id = second_payroll.id
            second_line_id = second_payroll.lines[0].id
            self.assertEqual(second_payroll.lines[0].partida_id, partida_a_id)
            self.assertEqual(
                second_payroll.lines[0].subpartida_id, subpartida_a_id
            )

        second_saved = self.post(
            f"/nominas/{second_payroll_id}/guardar",
            payroll_payload(second_line_id, partida_b_id),
        )
        self.assertIn(
            "Nómina guardada y cálculos actualizados.",
            second_saved.get_data(as_text=True),
        )
        self.post(f"/nominas/{second_payroll_id}/cerrar", {})
        with app.app_context():
            first_line = db.session.get(PayrollLine, first_line_id)
            second_line = db.session.get(PayrollLine, second_line_id)
            self.assertEqual(first_line.partida_id, partida_a_id)
            self.assertEqual(first_line.subpartida_id, subpartida_a_id)
            self.assertEqual(second_line.partida_id, partida_b_id)
            self.assertIsNone(second_line.subpartida_id)
            partida_a = db.session.get(BudgetItem, partida_a_id)
            subpartida_a = db.session.get(BudgetItem, subpartida_a_id)
            partida_b = db.session.get(BudgetItem, partida_b_id)
            self.assertEqual(float(item_consumption(partida_a)["nomina"]), 5000.0)
            self.assertEqual(
                float(item_consumption(subpartida_a)["nomina"]), 5000.0
            )
            self.assertEqual(float(item_consumption(partida_b)["nomina"]), 5000.0)
            auto_third = Payroll.query.filter_by(
                project_id=project_id, semana_inicio=date(2026, 7, 27)
            ).one()
            db.session.delete(auto_third)
            db.session.commit()

        # Sin la semana inmediata anterior no se reutiliza una asignación más
        # antigua ni se elige silenciosamente la primera partida de la obra.
        self.post(
            "/nominas/nueva",
            {"project_id": str(project_id), "semana_inicio": "2026-08-03"},
        )
        with app.app_context():
            gap_payroll = Payroll.query.filter_by(
                project_id=project_id, semana_inicio=date(2026, 8, 3)
            ).one()
            self.assertIsNone(gap_payroll.lines[0].partida_id)
            self.assertIsNone(gap_payroll.lines[0].subpartida_id)

        parent_report = self.client.get(
            "/reportes",
            query_string={
                "project_id": project_id,
                "budget_item_id": partida_a_id,
            },
        ).get_data(as_text=True)
        self.assertIn(f'href="/nominas/{first_payroll_id}"', parent_report)
        self.assertNotIn(f'href="/nominas/{second_payroll_id}"', parent_report)
        second_report = self.client.get(
            "/reportes",
            query_string={
                "project_id": project_id,
                "budget_item_id": partida_b_id,
            },
        ).get_data(as_text=True)
        self.assertIn(f'href="/nominas/{second_payroll_id}"', second_report)
        self.assertNotIn(f'href="/nominas/{first_payroll_id}"', second_report)

        export = self.client.get(
            "/reportes/exportar.xlsx",
            query_string={
                "project_id": project_id,
                "budget_item_id": partida_a_id,
            },
        )
        workbook = load_workbook(
            BytesIO(export.data), read_only=True, data_only=True
        )
        payroll_sheet = workbook["Nóminas"]
        headers = [
            cell.value
            for cell in next(payroll_sheet.iter_rows(min_row=1, max_row=1))
        ]
        self.assertEqual(payroll_sheet.max_row, 2)
        self.assertEqual(
            payroll_sheet.cell(
                row=2, column=headers.index("Partida") + 1
            ).value,
            "MO-A · ESTRUCTURA",
        )
        self.assertEqual(
            payroll_sheet.cell(
                row=2, column=headers.index("Subpartida") + 1
            ).value,
            "MO-A.1 · CIMENTACIÓN",
        )

    def test_cost_controls_with_vat_and_subcontracts(self):
        """Comprueba adicionales, indirectos y balance contra avance del contratista."""

        self.register_admin()
        for name, code, project_type, total in (
            ("Obra Costos", "OC01", "OBRA", "120000"),
            ("Oficina Central", "OF01", "OFICINA", "100000"),
        ):
            self.post(
                "/obras/nueva",
                {
                    "nombre": name,
                    "codigo": code,
                    "tipo": project_type,
                    "presupuesto_total": total,
                    "presupuesto_mano_obra": "0",
                    "activa": "on",
                },
            )
        with app.app_context():
            obra_id = Project.query.filter_by(codigo="OC01").one().id
            office_id = Project.query.filter_by(codigo="OF01").one().id
            company_id = Company.query.filter_by(codigo="BCH").one().id

        self.post(
            f"/obras/{obra_id}/partidas/nueva",
            {
                "codigo": "SUB-01",
                "nombre": "Instalaciones",
                "categoria": "SUBCONTRATO",
                "presupuesto": "100000",
                "parent_id": "",
                "activa": "on",
            },
        )
        self.post(
            f"/obras/{obra_id}/partidas/nueva",
            {
                "codigo": "ADI-01",
                "nombre": "Adicionales",
                "categoria": "ADICIONAL",
                "presupuesto": "20000",
                "parent_id": "",
                "activa": "on",
            },
        )
        self.post(
            f"/obras/{office_id}/partidas/nueva",
            {
                "codigo": "IND-01",
                "nombre": "Indirectos de oficina",
                "categoria": "INDIRECTO",
                "presupuesto": "100000",
                "parent_id": "",
                "activa": "on",
            },
        )
        with app.app_context():
            subcontract_item_id = BudgetItem.query.filter_by(
                project_id=obra_id, codigo="SUB-01"
            ).one().id
            additional_item_id = BudgetItem.query.filter_by(
                project_id=obra_id, codigo="ADI-01"
            ).one().id
            office_item_id = BudgetItem.query.filter_by(
                project_id=office_id, codigo="IND-01"
            ).one().id
            supply = SupplyItem(
                clave="ADI-SERV-01",
                descripcion="SERVICIO ADICIONAL DE OBRA",
                tipo="MATERIAL",
                unidad="SER",
                clave_sat="00000000",
                moneda="MXN",
            )
            supplier = Supplier(
                codigo="PROV-TEST",
                nombre="PROVEEDOR EJEMPLO",
                moneda="MXN",
            )
            db.session.add_all([supply, supplier])
            db.session.flush()
            explosion = BudgetExplosionItem(
                project_id=obra_id,
                budget_item_id=additional_item_id,
                supply_item_id=supply.id,
                cantidad_presupuestada=10,
                precio_unitario_sin_iva=2500,
                importe_presupuestado=25000,
                created_by_id=User.query.filter_by(correo="grecia@example.com").one().id,
            )
            db.session.add(explosion)
            db.session.commit()
            explosion_id = explosion.id
            supplier_id = supplier.id

        self.post(
            "/contratistas",
            {
                "nombre": "Constructora Ejemplo",
                "especialidad": "Instalaciones",
                "telefono": "",
                "email": "",
            },
        )
        with app.app_context():
            contractor_id = Contractor.query.one().id

        self.post(
            "/subcontratos/nuevo",
            {
                "project_id": str(obra_id),
                "budget_item_id": str(subcontract_item_id),
                "contractor_id": str(contractor_id),
                "especialidad": "Instalaciones eléctricas",
                "presupuesto_sin_iva": "100000",
                "avance_fisico": "50",
                "umbral_alerta": "15",
                "activo": "on",
            },
        )
        with app.app_context():
            subcontract_id = Subcontract.query.one().id

        self.post(
            f"/subcontratos/{subcontract_id}/pagos/nuevo",
            {
                "fecha": "2026-07-15",
                "concepto": "ESTIMACIÓN",
                "monto_capturado": "11600",
                "tipo_monto": "CON_IVA",
                "metodo_pago": "TRANSFERENCIA",
                "company_id": str(company_id),
                "notas": "Prueba IVA",
            },
        )
        # Los pagos históricos sin OC se conservan en reportes, pero el nuevo
        # flujo ya no permite capturarlos. Se inserta uno como dato heredado para
        # comprobar compatibilidad sin reabrir la ruta de compra directa.
        with app.app_context():
            historical = AdditionalPayment(
                fecha=date(2026, 7, 15),
                project_id=obra_id,
                budget_item_id=additional_item_id,
                explosion_item_id=explosion_id,
                supplier_id=supplier_id,
                purchase_order_id=None,
                beneficiario="PROVEEDOR EJEMPLO",
                concepto="Trabajo adicional histórico",
                monto_capturado=2500,
                tipo_monto="SIN_IVA",
                monto_sin_iva=2500,
                metodo_pago="EFECTIVO",
                company_id=company_id,
                created_by_id=User.query.filter_by(correo="grecia@example.com").one().id,
            )
            db.session.add(historical)
            db.session.commit()
        self.post(
            "/gastos-oficina/nuevo",
            {
                "project_id": str(office_id),
                "budget_item_id": str(office_item_id),
                "fecha": "2026-07-15",
                "proveedor": "Papelería Ejemplo",
                "concepto": "Papelería",
                "monto_capturado": "1160",
                "tipo_monto": "CON_IVA",
                "metodo_pago": "TRANSFERENCIA",
                "company_id": str(company_id),
                "notas": "",
            },
        )

        with app.app_context():
            subcontract = Subcontract.query.one()
            self.assertEqual(float(SubcontractPayment.query.one().monto_sin_iva), 10000.0)
            self.assertEqual(float(AdditionalPayment.query.one().monto_sin_iva), 2500.0)
            self.assertEqual(float(OfficeExpense.query.one().monto_sin_iva), 1000.0)
            self.assertEqual(float(subcontract.comprometido), 50000.0)
            self.assertEqual(float(subcontract.saldo_vs_avance), 40000.0)
            self.assertEqual(subcontract.estatus_control, "FALTA DE PAGO")
            obra = db.session.get(Project, obra_id)
            office = db.session.get(Project, office_id)
            # Los cinco indicadores del proyecto son exclusivos de presupuesto,
            # subpartidas y nóminas. Estos costos conservan sus reportes propios.
            self.assertEqual(float(project_consumption(obra)["total"]), 0.0)
            self.assertEqual(float(project_consumption(office)["total"]), 0.0)
            self.assertEqual(float(project_consumption(obra)["total_comprometido"]), 0.0)
            self.assertFalse(project_consumption(obra)["alerta_comprometido"])
            self.assertFalse(project_consumption(obra)["alerta_consumido"])

            report = build_weekly_closing_report(date(2026, 7, 13))
            self.assertEqual(float(report["totals"]["proveedores_recurso"]), 3660.0)
            self.assertEqual(float(report["totals"]["proveedores_costo"]), 3500.0)
            self.assertEqual(float(report["totals"]["subcontratos_recurso"]), 11600.0)
            self.assertEqual(float(report["totals"]["subcontratos_costo"]), 10000.0)
            self.assertEqual(float(report["totals"]["recurso_total"]), 15260.0)
            self.assertEqual(float(report["totals"]["costo_total"]), 13500.0)
            resource = weekly_resource_breakdown(date(2026, 7, 13))
            self.assertEqual(float(resource["methods"]["EFECTIVO"]["adicionales"]), 2500.0)
            self.assertEqual(float(resource["methods"]["TRANSFERENCIA"]["adicionales"]), 1160.0)
            self.assertEqual(float(resource["methods"]["TRANSFERENCIA"]["subcontratos"]), 11600.0)

        self.post(
            "/recursos-semanales/guardar",
            {
                "semana_inicio": "2026-07-13",
                "disponible_efectivo": "3000",
                "disponible_transferencia": "15000",
                "return_to": "/panel",
            },
        )
        with app.app_context():
            self.assertEqual(WeeklyResourceAvailability.query.count(), 2)
            resource = weekly_resource_breakdown(date(2026, 7, 13))
            self.assertEqual(float(resource["methods"]["EFECTIVO"]["diferencia"]), 500.0)
            self.assertEqual(
                float(resource["methods"]["TRANSFERENCIA"]["diferencia"]), 2240.0
            )

        closing_report = self.client.get("/reportes/cierre-semanal?semana=2026-07-13")
        closing_text = closing_report.get_data(as_text=True)
        self.assertEqual(closing_report.status_code, 200)
        self.assertIn("CONSTRUCTORA EJEMPLO", closing_text)
        self.assertIn("PAPELERÍA EJEMPLO", closing_text)
        self.assertIn("$15,260.00", closing_text)
        self.assertIn('class="executive-kpi-grid"', closing_text)
        self.assertEqual(closing_text.count('class="executive-kpi-card"'), 6)
        self.assertIn('class="table-responsive loan-status-wrapper"', closing_text)
        self.assertIn('loan-status-table mb-0', closing_text)
        self.assertIn(".loan-status-wrapper { overflow: visible !important;", closing_text)
        self.assertIn("table-layout: fixed", closing_text)
        self.assertIn("<colgroup><col><col><col><col><col><col><col><col><col></colgroup>", closing_text)
        indicator_labels = (
            "Recurso total requerido",
            "Total efectivo",
            "Total transferencias",
            "Total nóminas",
            "Total subcontratos",
            "Total proveedores",
        )
        indicator_positions = [closing_text.index(label) for label in indicator_labels]
        self.assertEqual(indicator_positions, sorted(indicator_positions))

        dashboard = self.client.get("/panel")
        dashboard_text = dashboard.get_data(as_text=True)
        self.assertIn("Dashboard ejecutivo", dashboard_text)
        self.assertIn("Recurso semanal por método de pago", dashboard_text)
        self.assertIn("Disponible comprometido", dashboard_text)

        budget_export = self.client.get("/reportes/exportar.xlsx")
        budget_workbook = load_workbook(
            BytesIO(budget_export.data), read_only=True, data_only=True
        )
        control_rows = list(
            budget_workbook["Control presupuestal"].iter_rows(min_row=2, values_only=True)
        )
        obra_control = next(row for row in control_rows if row[0] == "OBRA COSTOS")
        self.assertEqual(obra_control[4], 0)
        self.assertEqual(obra_control[8], 120000)
        self.assertEqual(obra_control[9], 0)
        self.assertEqual(obra_control[10], "EN CONTROL")

        for route in (
            "/panel",
            "/obras",
            "/trabajadores",
            "/prestamos",
            "/nominas",
            "/pagos-adicionales",
            "/gastos-oficina",
            "/contratistas",
            "/subcontratos",
            f"/subcontratos/{subcontract_id}",
            "/reportes",
            "/auditoria",
        ):
            self.assertEqual(self.client.get(route).status_code, 200, route)

    def test_project_dashboard_uses_subpartidas_and_closed_payrolls(self):
        """Concilia los cinco indicadores sin mezclar costos ni borradores."""

        self.register_admin()
        with app.app_context():
            admin = User.query.filter_by(correo="grecia@example.com").one()
            project = Project(
                nombre="OBRA INDICADORES",
                codigo="KPI01",
                tipo="obra",
                estado="activa",
                presupuesto_total=100000,
                presupuesto_mano_obra=100000,
            )
            db.session.add(project)
            db.session.flush()

            parent_one = BudgetItem(
                project_id=project.id,
                codigo="P-01",
                nombre="PARTIDA UNO",
                categoria="MANO_OBRA",
                presupuesto=60000,
            )
            parent_two = BudgetItem(
                project_id=project.id,
                codigo="P-02",
                nombre="PARTIDA DOS",
                categoria="MANO_OBRA",
                presupuesto=20000,
            )
            db.session.add_all([parent_one, parent_two])
            db.session.flush()
            sub_one = BudgetItem(
                project_id=project.id,
                parent_id=parent_one.id,
                codigo="SP-01",
                nombre="SUBPARTIDA UNO",
                categoria="MANO_OBRA",
                presupuesto=30000,
            )
            sub_two = BudgetItem(
                project_id=project.id,
                parent_id=parent_one.id,
                codigo="SP-02",
                nombre="SUBPARTIDA DOS",
                categoria="MANO_OBRA",
                presupuesto=10000,
            )
            sub_three = BudgetItem(
                project_id=project.id,
                parent_id=parent_two.id,
                codigo="SP-03",
                nombre="SUBPARTIDA TRES",
                categoria="MANO_OBRA",
                presupuesto=5000,
            )
            db.session.add_all([sub_one, sub_two, sub_three])
            db.session.flush()

            employee = Employee(
                nombre_completo="TRABAJADOR INDICADORES",
                fecha_ingreso=date(2026, 7, 1),
                puesto="OFICIAL",
                project_id=project.id,
                budget_item_id=sub_one.id,
                salario_semanal=10000,
            )
            db.session.add(employee)
            db.session.flush()

            closed_payroll = Payroll(
                project_id=project.id,
                semana_inicio=date(2026, 7, 13),
                semana_fin=date(2026, 7, 17),
                estado="aprobada",
                created_by_id=admin.id,
                closed_by_id=admin.id,
            )
            draft_payroll = Payroll(
                project_id=project.id,
                semana_inicio=date(2026, 7, 20),
                semana_fin=date(2026, 7, 24),
                estado="borrador",
                created_by_id=admin.id,
            )
            db.session.add_all([closed_payroll, draft_payroll])
            db.session.flush()
            db.session.add_all(
                [
                    PayrollLine(
                        payroll_id=closed_payroll.id,
                        employee_id=employee.id,
                        budget_item_id=sub_one.id,
                        nombre_trabajador=employee.nombre_completo,
                        puesto=employee.puesto,
                        salario_semanal=10000,
                        monto_devengado=10000,
                        pago_extra=500,
                        descuento_imss=500,
                    ),
                    PayrollLine(
                        payroll_id=draft_payroll.id,
                        employee_id=employee.id,
                        budget_item_id=sub_one.id,
                        nombre_trabajador=employee.nombre_completo,
                        puesto=employee.puesto,
                        salario_semanal=9000,
                        monto_devengado=9000,
                    ),
                ]
            )
            db.session.commit()

            values = project_consumption(project)
            self.assertEqual(float(values["presupuesto_base"]), 100000.0)
            self.assertEqual(float(values["presupuesto_partidas"]), 80000.0)
            self.assertEqual(float(values["total_comprometido"]), 45000.0)
            self.assertEqual(float(values["total"]), 11000.0)
            self.assertEqual(float(values["disponible_partidas"]), 35000.0)
            self.assertEqual(float(values["sin_asignar_partidas"]), 20000.0)
            self.assertEqual(float(values["restante_total"]), 55000.0)
            self.assertEqual(float(values["restante_comprometido"]), 34000.0)
            self.assertEqual(
                values["presupuesto_base"],
                values["restante_total"] + values["total_comprometido"],
            )
            self.assertEqual(
                values["total_comprometido"],
                values["total"] + values["restante_comprometido"],
            )
            project_id = project.id

        page = self.client.get(f"/obras/{project_id}")
        text = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn("Suma de subpartidas", text)
        self.assertIn("Subpartidas menos nóminas cerradas", text)
        self.assertIn("$55,000.00", text)
        self.assertIn("$34,000.00", text)

    def test_bulk_imports_and_multiple_administrators(self):
        """Valida plantillas, altas masivas, duplicados y permisos de un segundo admin."""

        self.register_admin()
        self.post(
            "/obras/nueva",
            {
                "nombre": "Obra Importaciones",
                "codigo": "IMP01",
                "tipo": "OBRA",
                "presupuesto_total": "500000",
                "presupuesto_mano_obra": "200000",
                "activa": "on",
            },
        )
        with app.app_context():
            project_id = Project.query.filter_by(codigo="IMP01").one().id
        for code, name, category, budget in (
            ("MO-IMP", "Mano de obra", "MANO_OBRA", "200000"),
            ("ADI-IMP", "Pagos adicionales", "ADICIONAL", "100000"),
            ("SUB-IMP", "Subcontratos", "SUBCONTRATO", "200000"),
        ):
            self.post(
                f"/obras/{project_id}/partidas/nueva",
                {
                    "codigo": code,
                    "nombre": name,
                    "categoria": category,
                    "presupuesto": budget,
                    "parent_id": "",
                    "activa": "on",
                },
            )

        with app.app_context():
            imported_project = Project.query.filter_by(codigo="IMP01").one()
            imported_item = BudgetItem.query.filter_by(
                project_id=imported_project.id, codigo="ADI-IMP"
            ).one()
            imported_supply = SupplyItem(
                clave="INS-IMP-01",
                descripcion="INSUMO PARA PAGO IMPORTADO",
                tipo="MATERIAL",
                unidad="PZA",
                clave_sat="00000000",
                moneda="MXN",
            )
            db.session.add(imported_supply)
            db.session.flush()
            db.session.add(
                BudgetExplosionItem(
                    project_id=imported_project.id,
                    budget_item_id=imported_item.id,
                    supply_item_id=imported_supply.id,
                    cantidad_presupuestada=10,
                    precio_unitario_sin_iva=2500,
                    importe_presupuestado=25000,
                    created_by_id=User.query.filter_by(correo="grecia@example.com").one().id,
                )
            )
            db.session.commit()

        # El administrador inicial crea otro administrador con el mismo acceso.
        created_admin = self.post(
            "/registro",
            {
                "nombre_completo": "Administradora Dos",
                "username": "admin.dos",
                "email": "admin2@example.com",
                "password": "Password123!",
                "password_confirm": "Password123!",
                "role": "administrador",
            },
        )
        self.assertIn("Usuario creado correctamente.", created_admin.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(User.query.filter_by(correo="admin2@example.com").one().role, "administrador")

        self.logout()
        self.login("admin.dos", "Password123!")
        self.assertEqual(self.client.get("/usuarios").status_code, 200)

        # Cada descarga debe ser un Excel legible con los encabezados definidos.
        for kind, definition in IMPORT_DEFINITIONS.items():
            response = self.client.get(
                f"/administracion/importar/{kind}/plantilla.xlsx"
            )
            self.assertEqual(response.status_code, 200)
            workbook = load_workbook(BytesIO(response.data), read_only=True)
            headers = [cell.value for cell in next(workbook["Datos"].iter_rows())]
            self.assertEqual(headers, [column[0] for column in definition["columns"]])
        self.assertNotIn(
            "PARTIDA_CODIGO",
            [column[0] for column in IMPORT_DEFINITIONS["trabajadores"]["columns"]],
        )

        worker = {
            "NOMBRE_COMPLETO": "María López Sánchez",
            "FECHA_INGRESO": "2026-07-20",
            "PUESTO": "Ayudante general",
            "OBRA_CODIGO": "IMP01",
            "SALARIO_SEMANAL": 4500,
            "ACTIVO": "SI",
            "TIENE_IMSS": "NO",
            "EMPRESA_EFECTIVO": "BCH",
        }
        invalid_worker = dict(worker, NOMBRE_COMPLETO="Persona Con Error", OBRA_CODIGO="NO-EXISTE")
        import_workers = self.post(
            "/administracion/importar/trabajadores",
            {
                "archivo": (
                    self.excel_file("trabajadores", [worker, invalid_worker]),
                    "trabajadores.xlsx",
                ),
                "existing_action": "OMITIR",
            },
        )
        worker_page = import_workers.get_data(as_text=True)
        self.assertIn("1 creados", worker_page)
        self.assertIn("1 errores", worker_page)
        self.assertIn("NO-EXISTE", worker_page)
        with app.app_context():
            self.assertEqual(float(Employee.query.one().salario_semanal), 4500.0)
            self.assertIsNone(Employee.query.one().budget_item_id)

        # La opción OMITIR conserva el registro; ACTUALIZAR reemplaza sus datos.
        changed_worker = dict(worker, SALARIO_SEMANAL=4800)
        self.post(
            "/administracion/importar/trabajadores",
            {
                "archivo": (self.excel_file("trabajadores", [changed_worker]), "trabajadores.xlsx"),
                "existing_action": "OMITIR",
            },
        )
        with app.app_context():
            self.assertEqual(float(Employee.query.one().salario_semanal), 4500.0)
        self.post(
            "/administracion/importar/trabajadores",
            {
                "archivo": (self.excel_file("trabajadores", [changed_worker]), "trabajadores.xlsx"),
                "existing_action": "ACTUALIZAR",
            },
        )
        with app.app_context():
            self.assertEqual(float(Employee.query.one().salario_semanal), 4800.0)

        provider = {
            "FECHA": "2026-07-22",
            "OBRA_CODIGO": "IMP01",
            "PARTIDA_CODIGO": "ADI-IMP",
            "INSUMO_CLAVE": "INS-IMP-01",
            "PROVEEDOR_CODIGO": "PROV-IMP",
            "PROVEEDOR": "Proveedor Importado",
            "CONCEPTO": "Trabajo extraordinario",
            "MONTO": 2500,
            "TIPO_MONTO": "SIN_IVA",
            "METODO_PAGO": "TRANSFERENCIA",
            "EMPRESA_CODIGO": "BCH",
        }
        subcontract = {
            "CONTRATISTA": "Contratista Importado",
            "ESPECIALIDAD": "Instalación hidráulica",
            "OBRA_CODIGO": "IMP01",
            "PARTIDA_CODIGO": "SUB-IMP",
            "PRESUPUESTO_SIN_IVA": 100000,
            "AVANCE_PORCENTAJE": 20,
            "UMBRAL_ALERTA_PORCENTAJE": 15,
            "ACTIVO": "SI",
        }
        provider_response = self.post(
            "/administracion/importar/proveedores",
            {
                "archivo": (self.excel_file("proveedores", [provider]), "proveedores.xlsx"),
                "existing_action": "OMITIR",
            },
        )
        provider_page = provider_response.get_data(as_text=True)
        self.assertIn("1 errores", provider_page)
        self.assertIn("carga directa de pagos de obra quedó cerrada", provider_page)

        subcontract_response = self.post(
            "/administracion/importar/subcontratistas",
            {
                "archivo": (
                    self.excel_file("subcontratistas", [subcontract]),
                    "subcontratistas.xlsx",
                ),
                "existing_action": "OMITIR",
            },
        )
        self.assertIn("1 creados", subcontract_response.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(AdditionalPayment.query.count(), 0)
            self.assertEqual(Contractor.query.count(), 1)
            self.assertEqual(Subcontract.query.count(), 1)

    def test_shared_permissions_and_module_csrf(self):
        """Comprueba autenticación compartida, alcance por obra y CSRF real."""

        self.register_admin()
        with app.app_context():
            assigned = Project(
                nombre="OBRA ASIGNADA",
                codigo="ASIG",
                tipo="obra",
                estado="activa",
            )
            hidden = Project(
                nombre="OBRA NO ASIGNADA",
                codigo="OCULTA",
                tipo="obra",
                estado="activa",
            )
            db.session.add_all([assigned, hidden])
            db.session.flush()
            capturista = User(
                nombre_completo="Capturista Seguridad",
                correo="seguridad@example.com",
                rol="capturista",
                centro_costo_id=assigned.id,
                activo=True,
            )
            capturista.set_password("Password123!")
            capturista.projects = [assigned]
            db.session.add(capturista)
            db.session.commit()
            capturista_id = capturista.id
            admin_id = User.query.filter_by(rol="admin").one().id

        with self.client.session_transaction() as flask_session:
            flask_session["_user_id"] = str(capturista_id)
            flask_session["_fresh"] = True

        self.assertEqual(self.client.get("/usuarios").status_code, 403)
        projects_page = self.client.get("/obras").get_data(as_text=True)
        self.assertIn("OBRA ASIGNADA", projects_page)
        self.assertNotIn("OBRA NO ASIGNADA", projects_page)

        with self.client.session_transaction() as flask_session:
            flask_session["_user_id"] = str(admin_id)
            flask_session["_fresh"] = True

        app.config["WTF_CSRF_ENABLED"] = True
        try:
            companies_page = self.client.get("/empresas").get_data(as_text=True)
            rejected = self.client.post(
                "/empresas",
                data={"codigo": "SIN", "nombre": "Sin token"},
            )
            self.assertEqual(rejected.status_code, 400)

            token_match = re.search(
                r'name="csrf_token"[^>]*value="([^"]+)"',
                companies_page,
            )
            self.assertIsNotNone(token_match)
            accepted = self.client.post(
                "/empresas",
                data={
                    "csrf_token": token_match.group(1),
                    "codigo": "SEG",
                    "nombre": "Empresa segura",
                },
                follow_redirects=True,
            )
            self.assertEqual(accepted.status_code, 200)
            with app.app_context():
                self.assertIsNotNone(
                    Company.query.filter_by(codigo="SEG").first()
                )
        finally:
            app.config["WTF_CSRF_ENABLED"] = False

    def test_logout_uses_rendered_flask_wtf_token_and_clears_session(self):
        """El formulario visible debe cerrar sesión y rechazar POST sin token."""

        self.register_admin()
        app.config["WTF_CSRF_ENABLED"] = True
        try:
            page = self.client.get("/dashboard", follow_redirects=True)
            self.assertEqual(page.status_code, 200)
            token_match = re.search(
                r'<form[^>]+action="/logout"[^>]*>.*?'
                r'name="csrf_token"[^>]*value="([^"]+)"',
                page.get_data(as_text=True),
                re.DOTALL,
            )
            self.assertIsNotNone(token_match)

            rejected = self.client.post("/logout")
            self.assertEqual(rejected.status_code, 400)
            self.assertIn(
                "El formulario ya no es válido",
                rejected.get_data(as_text=True),
            )

            accepted = self.client.post(
                "/logout",
                data={"csrf_token": token_match.group(1)},
                follow_redirects=False,
            )
            self.assertEqual(accepted.status_code, 302)
            self.assertEqual(accepted.headers["Location"], "/login")
            with self.client.session_transaction() as flask_session:
                self.assertNotIn("_user_id", flask_session)
        finally:
            app.config["WTF_CSRF_ENABLED"] = False

    def test_every_post_form_declares_a_csrf_control(self):
        """Evita agregar formularios POST sin token en cualquier módulo."""

        templates_root = Path(__file__).resolve().parents[1] / "templates"
        post_form = re.compile(
            r'<form\b[^>]*\bmethod\s*=\s*["\']post["\'][^>]*>'
            r'(.*?)</form\s*>',
            re.IGNORECASE | re.DOTALL,
        )
        csrf_control = re.compile(
            r"hidden_tag\s*\("
            r'|name\s*=\s*["\']csrf_token["\']'
            r"|\{\{\s*csrf\s*\(\s*\)\s*\}\}",
            re.IGNORECASE,
        )
        missing = []
        for template in sorted(templates_root.rglob("*.html")):
            source = template.read_text(encoding="utf-8")
            for match in post_form.finditer(source):
                if csrf_control.search(match.group(1)):
                    continue
                line = source.count("\n", 0, match.start()) + 1
                missing.append(f"{template.relative_to(templates_root)}:{line}")

        self.assertEqual(missing, [], f"Formularios POST sin CSRF: {missing}")

    def test_supervisor_cannot_reopen_payroll_from_another_project(self):
        self.register_admin()
        with app.app_context():
            assigned = Project(
                nombre="OBRA DEL SUPERVISOR",
                codigo="SUP-01",
                tipo="obra",
                estado="activa",
            )
            foreign = Project(
                nombre="OBRA AJENA",
                codigo="AJENA-01",
                tipo="obra",
                estado="activa",
            )
            db.session.add_all([assigned, foreign])
            db.session.flush()
            supervisor = User(
                nombre_completo="Supervisor Seguridad",
                correo="supervisor.seguridad@example.com",
                rol="supervisor",
                centro_costo_id=assigned.id,
                activo=True,
            )
            supervisor.set_password("Password123!")
            db.session.add(supervisor)
            supervisor.projects = [assigned]
            supervisor.asignar_permisos_predeterminados()
            admin = User.query.filter_by(rol="admin").one()
            db.session.flush()
            payroll = Payroll(
                project_id=foreign.id,
                semana_inicio=date(2026, 7, 13),
                semana_fin=date(2026, 7, 17),
                estado="aprobada",
                created_by_id=admin.id,
                closed_by_id=admin.id,
            )
            db.session.add(payroll)
            db.session.commit()
            payroll_id = payroll.id

        self.logout()
        self.login("supervisor.seguridad@example.com", "Password123!")
        response = self.client.post(
            f"/nominas/{payroll_id}/reabrir",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 404)
        with app.app_context():
            self.assertEqual(db.session.get(Payroll, payroll_id).estado, "aprobada")


if __name__ == "__main__":
    unittest.main()
